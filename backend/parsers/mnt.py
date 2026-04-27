import openpyxl
import re
import json
from typing import List, Dict, Optional, Union, Any
import logging
import traceback
from datetime import datetime, timedelta

logging.basicConfig(level=logging.DEBUG, format="%(asctime)s - %(levelname)s - %(message)s")
logging.getLogger("multipart.multipart").setLevel(logging.WARNING)
logger = logging.getLogger(__name__)

# Matches both MNT and MN in sheet/identifier names
_MNT_ID = r'WP(\d+)_MN[T]?_(\d+)([a-zA-Z])R(\d+)'


class MNTParser:
    def __init__(self, file_path, sheet_name="Test_conditions"):
        self.file_path = file_path
        try:
            self.wb = openpyxl.load_workbook(file_path, data_only=True)
            if sheet_name in self.wb.sheetnames:
                self.ws = self.wb[sheet_name]
            elif "Test Information" in self.wb.sheetnames:
                self.ws = self.wb["Test Information"]
            else:
                self.ws = self.wb[self.wb.sheetnames[0]]
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}"); raise
        self.email_regex = r"^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$"

    def _nk(self, key):
        if not key: return None
        return re.sub(r"_+", "_", re.sub(r"[^a-z0-9]+", "_", str(key).strip().lower())).strip("_")

    def _sf(self, v):
        if v in (None, "", " "): return None
        try:
            if isinstance(v, str): v = v.replace(",", ".").strip()
            return float(v)
        except: return None

    def _date(self, v):
        try:
            if isinstance(v, datetime): return v.strftime("%Y-%m-%d")
            if isinstance(v, (int, float)): return (datetime(1899, 12, 30) + timedelta(days=float(v))).strftime("%Y-%m-%d")
            if isinstance(v, str):
                for f in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
                    try: return datetime.strptime(v.strip(), f).strftime("%Y-%m-%d")
                    except: pass
            return str(v).strip() if v else None
        except: return str(v).strip() if v else None

    def _str(self, v):
        if v is None: return None
        return str(v).strip()

    def _kv(self, max_col=6):
        data = []
        for ri, row in enumerate(self.ws.iter_rows(min_row=1, max_col=max_col), start=1):
            k = row[0].value
            if not k: continue
            key = self._nk(k)
            if not key: continue
            val = None
            for c in range(1, min(len(row), max_col)):
                if row[c].value is not None: val = row[c].value; break
            email = row[3].value if len(row) > 3 else None
            data.append({"row": ri, "key": key, "value": val, "email": email})
        return data

    def _g(self, rows, *keys):
        return next((r["value"] for r in rows if any(k in r["key"] for k in keys) and r["value"] is not None), None)

    def _find_sheets(self, prefix):
        """Find sheets by prefix, normalizing double underscores."""
        p = prefix.lower(); seen = set(); result = []
        for s in sorted(self.wb.sheetnames):
            sl = s.lower()
            # Normalize double underscores for matching
            sl_norm = re.sub(r'_+', '_', sl)
            if sl_norm in seen: continue
            if sl.startswith(p) or sl.startswith(p.replace(" ", "_")) or sl.startswith(p.replace("_", " ")):
                seen.add(sl_norm); result.append(s)
            # Also match with double underscore variants
            elif re.sub(r'_+', '_', sl).startswith(p):
                seen.add(sl_norm); result.append(s)
        return result

    # ====== TEST DETAILS ======
    def extract_work_package(self):
        rows = self._kv()
        lead, assay = [], []
        for r in rows:
            if "lead_scientist" in r["key"]:
                em = r["email"]; lead.append({"name": r["value"], "email": em if em and re.match(self.email_regex, str(em)) else None})
            if "assay_test_work_conducted_by" in r["key"]:
                em = r["email"]; assay.append({"name": r["value"], "email": em if em and re.match(self.email_regex, str(em)) else None})
        # Also pick up rows 17-18 (additional assay scientists without key label)
        ws = self.ws
        for ri in range(17, 20):
            name = ws.cell(row=ri, column=2).value
            email_label = ws.cell(row=ri, column=3).value
            email = ws.cell(row=ri, column=4).value
            if name and email_label and "mail" in str(email_label).lower():
                assay.append({"name": str(name), "email": str(email) if email else None})
        g = lambda *k: self._g(rows, *k)
        return {"wp_name": g("project_work_package"), "partner": g("partner_conducting_test_assay"), "laboratory_name": g("test_facility_laboratory_name"), "full_test_name": g("full_name_of_test_assay"), "test_acronym": g("short_name_or_acronym"), "test_type": g("type_or_class_of_experimental"), "endpoint": g("end_point_being_investigated"), "endpoint_outcome": g("metric_s_used_to_assess"), "sop": g("sop_s_for_test"), "path": g("path_link_to_sop"), "lead_scientists": lead, "assay_scientists": assay}

    def extract_material(self):
        rows = self._kv(); g = lambda *k: self._g(rows, *k)
        return {"material_identifier": g("sample_cms_internal_identifier"), "erm_id": g("erm_identifier"), "material_name": g("material_name"), "core_chemistry": g("core_chemistry"), "cas_no": g("cas_no"), "cas_for_core": g("cas_for_core"), "material_supplier": g("material_supplier"), "material_state": g("material_state"), "batch": g("batch"), "vial": g("vial"), "preparation_date": self._date(g("date_of_preparation")), "endotoxin": g("endotoxin"), "stock_concentration": g("stock_oncentration", "stock_concentration"), "molecular_weight": g("molecular_weight"), "particles_in_stock": g("no_of_particles_in_stock")}

    def extract_cell_line(self):
        rows = self._kv(); ws = self.ws; g = lambda *k: self._g(rows, *k)
        passages = [ws.cell(row=63, column=c).value for c in range(2, 8) if ws.cell(row=63, column=c).value is not None]
        return {"cell_type": g("detailed_cell_type"), "cell_line_short": g("cell_line_short_name"), "supplier": g("supplier"), "passage_numbers": passages, "plate_details": g("plate_details"), "cells_per_well": g("number_of_cells_per_well"), "volume_per_well": g("total_volume_per_well"), "medium": g("medium"), "serum": g("serum_inc"), "serum_concentration_culture": g("serum_concentration_in_culture"), "serum_concentration_treatment": g("serum_concentration_in_treatment"), "serum_heat_inactivated": g("was_serum_heat_inactivated"), "antibiotics": g("antibiotics"), "complete_growth_medium": g("complete_growth_medium"), "culture_conditions": g("cell_culture_conditions")}

    def extract_dispersion(self):
        rows = self._kv(max_col=10); g = lambda *k: self._g(rows, *k); ws = self.ws
        aids = {}
        for cl, cv in [(3, 4), (5, 6), (7, 8), (9, 10)]:
            lbl = ws.cell(row=55, column=cl).value; val = ws.cell(row=55, column=cv).value
            if lbl: aids[str(lbl).strip().rstrip(":")] = str(val) if val else None
        return {"dispersion_protocol": g("standard_dispersion_protocol", "specify_standard_dispersion"), "dispersion_technique": g("dispersion_technique", "otherwise_specify_dispersion"), "dispersion_agent": g("dispersion_agent"), "additives": g("additives_used"), "dispersed_in_medium": g("dispersed_in_cell_culture"), "aids": aids, "time_duration": g("specify_time_duration"), "energy": g("energy")}

    def extract_treatment(self):
        ws = self.ws
        concentrations = []
        for c in range(2, 12):
            lbl = ws.cell(row=84, column=c).value
            if lbl is None: break
            concentrations.append({"label": str(lbl), "ug_ml": self._sf(ws.cell(row=85, column=c).value)})
        plate_series = [ws.cell(row=87, column=c).value for c in range(2, 12) if ws.cell(row=87, column=c).value]
        g = lambda *k: self._g(self._kv(), *k)
        return {"time_unit": g("time_point_unit"), "time_labels": [ws.cell(row=80, column=c).value for c in range(2, 8) if ws.cell(row=80, column=c).value], "time_points": [ws.cell(row=81, column=c).value for c in range(2, 8) if ws.cell(row=81, column=c).value], "concentration_unit": g("treatment_concentration_series_unit"), "concentrations": concentrations, "plate_series": plate_series, "positive_control_abbr": g("positive_controls_abbreviation"), "positive_control_desc": g("positive_controls_description"), "negative_control_abbr": g("negative_controls_abbreviation"), "negative_control_desc": g("negative_controls_description"), "num_experiments": g("number_of_experiments")}

    def extract_replication_metadata(self):
        ws = self.ws; meta = []
        for ri in range(38, 50):
            tid = ws.cell(row=ri, column=2).value
            if tid is None: continue
            ts = str(tid).strip()
            if not re.search(r'R\d+', ts, re.IGNORECASE): continue
            meta.append({"test_identifier_number": ts, "test_start_date": self._date(ws.cell(row=ri, column=3).value), "test_end_date": self._date(ws.cell(row=ri, column=4).value), "replicate_label": self._str(ws.cell(row=ri, column=5).value)})
        return meta

    # ====== RAW DATA ======
    def _extract_raw_block(self, sn):
        ws = self.wb[sn]
        m = re.search(_MNT_ID, re.sub(r"_+", "_", sn), re.IGNORECASE)
        run_label = f"R{m.group(4)}" if m else sn
        experiment_label = self._str(ws.cell(row=6, column=1).value)

        # Group structure: NC(3,4) C1(5,6) C2(7,8) C3(9,10) PC0.5(11,12) PC0.75(14,15)
        groups = [
            {"label": self._str(ws.cell(row=5, column=3).value) or "NC", "n_col": 3, "un_col": 4},
            {"label": self._str(ws.cell(row=5, column=5).value) or "C1", "n_col": 5, "un_col": 6},
            {"label": self._str(ws.cell(row=5, column=7).value) or "C2", "n_col": 7, "un_col": 8},
            {"label": self._str(ws.cell(row=5, column=9).value) or "C3", "n_col": 9, "un_col": 10},
            {"label": f"PC {self._str(ws.cell(row=4, column=11).value) or '0.5 µg'}", "n_col": 11, "un_col": 12},
            {"label": f"PC {self._str(ws.cell(row=4, column=14).value) or '0.75 µg'}", "n_col": 14, "un_col": 15},
        ]

        # Read 36 fields of view (rows 7-42)
        fields = []
        for r in range(7, 43):
            fov = ws.cell(row=r, column=2).value
            if fov is None: break
            row_data = {"field": int(fov)}
            for g in groups:
                row_data[f"{g['label']}_N"] = self._sf(ws.cell(row=r, column=g["n_col"]).value)
                row_data[f"{g['label']}_uN"] = self._sf(ws.cell(row=r, column=g["un_col"]).value)
            fields.append(row_data)

        # Summary row (row 43) — can be "Sum" or "Avg" depending on experiment
        summary_label = self._str(ws.cell(row=43, column=2).value) or "Sum"
        summary = {}
        for g in groups:
            summary[f"{g['label']}_N"] = self._sf(ws.cell(row=43, column=g["n_col"]).value)
            summary[f"{g['label']}_uN"] = self._sf(ws.cell(row=43, column=g["un_col"]).value)

        # SD row (row 44)
        sds = {}
        for g in groups:
            sds[f"{g['label']}_N"] = self._sf(ws.cell(row=44, column=g["n_col"]).value)
            sds[f"{g['label']}_uN"] = self._sf(ws.cell(row=44, column=g["un_col"]).value)

        # RICC (rows 46-48)
        ricc_headers = [self._str(ws.cell(row=47, column=c).value) for c in range(3, 11) if ws.cell(row=47, column=c).value]
        ricc_values = [self._sf(ws.cell(row=48, column=c).value) for c in range(3, 11) if ws.cell(row=48, column=c).value is not None]
        ricc_data = dict(zip(ricc_headers, ricc_values)) if ricc_headers else {}

        return {
            "run_label": run_label,
            "raw_sheet_name": sn,
            "experiment_label": experiment_label,
            "group_labels": [g["label"] for g in groups],
            "fields": fields,
            "summary_label": summary_label,
            "summary": summary,
            "sds": sds,
            "field_count": len(fields),
            "ricc_data": ricc_data,
        }

    def extract_raw_data(self):
        blocks = []; seen = set()
        for sn in self._find_sheets("raw_data_"):
            sn_norm = re.sub(r'_+', '_', sn)
            if not re.search(_MNT_ID, sn_norm, re.IGNORECASE): continue
            sl_norm = sn_norm.lower()
            if sl_norm in seen: continue
            seen.add(sl_norm)
            blocks.append(self._extract_raw_block(sn))
        return blocks

    # ====== PROCESSED DATA ======
    def _extract_processed_block(self, sn):
        ws = self.wb[sn]
        m = re.search(_MNT_ID, re.sub(r"_+", "_", sn), re.IGNORECASE)
        run_label = f"R{m.group(4)}" if m else sn

        # Dynamically discover groups from row 5 headers
        groups = []
        for c in range(3, min(ws.max_column + 1, 30)):
            label = self._str(ws.cell(row=5, column=c).value)
            if label is not None:
                groups.append({"label": label, "n_col": c, "un_col": c + 1, "pct_col": c + 2})

        # Read fields (rows 7-42)
        fields = []
        for r in range(7, 43):
            fov = ws.cell(row=r, column=2).value
            if fov is None: break
            row_data = {"field": int(fov)}
            for g in groups:
                row_data[f"{g['label']}_N"] = self._sf(ws.cell(row=r, column=g["n_col"]).value)
                row_data[f"{g['label']}_uN"] = self._sf(ws.cell(row=r, column=g["un_col"]).value)
                row_data[f"{g['label']}_%uN/N"] = self._sf(ws.cell(row=r, column=g["pct_col"]).value)
            fields.append(row_data)

        # Avg row (row 43)
        avg = {}
        for g in groups:
            avg[f"{g['label']}_N"] = self._sf(ws.cell(row=43, column=g["n_col"]).value)
            avg[f"{g['label']}_uN"] = self._sf(ws.cell(row=43, column=g["un_col"]).value)
            avg[f"{g['label']}_%uN/N"] = self._sf(ws.cell(row=43, column=g["pct_col"]).value)

        # SD row (row 44)
        sd = {}
        for g in groups:
            sd[f"{g['label']}_N"] = self._sf(ws.cell(row=44, column=g["n_col"]).value)
            sd[f"{g['label']}_uN"] = self._sf(ws.cell(row=44, column=g["un_col"]).value)
            sd[f"{g['label']}_%uN/N"] = self._sf(ws.cell(row=44, column=g["pct_col"]).value)

        # Sum of N (row 45)
        sum_n = {g["label"]: self._sf(ws.cell(row=45, column=g["n_col"]).value) for g in groups}

        # Acceptance criteria
        accept_1_text = self._str(ws.cell(row=48, column=2).value)
        accept_1_result = self._str(ws.cell(row=48, column=9).value)
        accept_2_text = self._str(ws.cell(row=51, column=2).value)
        accept_2_result = self._str(ws.cell(row=51, column=9).value)

        # RICC (rows 54-64)
        ricc_headers = [self._str(ws.cell(row=59, column=c).value) for c in range(3, 11) if ws.cell(row=59, column=c).value]
        ricc_values = [self._sf(ws.cell(row=60, column=c).value) for c in range(3, 11) if ws.cell(row=60, column=c).value is not None]
        ricc_data = dict(zip(ricc_headers, ricc_values)) if ricc_headers else {}
        ricc_formula = self._str(ws.cell(row=56, column=3).value)
        ricc_results = {}
        for c in [5, 7, 9]:
            lbl = self._str(ws.cell(row=58, column=c).value)
            val = self._sf(ws.cell(row=61, column=c).value)
            if lbl: ricc_results[lbl] = val
        ricc_accept_text = self._str(ws.cell(row=64, column=2).value)
        ricc_accept_result = self._str(ws.cell(row=64, column=9).value)

        # Pairwise comparisons: scan row 41 for "p-value" header to find column offset
        pairwise = []
        pw_alpha = None
        pw_col_offset = None
        for c in range(19, 30):
            v = self._str(ws.cell(row=41, column=c).value)
            if v and "p-value" in v.lower():
                pw_col_offset = c
                pw_alpha = self._sf(ws.cell(row=41, column=c + 1).value)
                break
        if pw_col_offset:
            # Comparisons start at row 42, 2 cols before the p-value col
            g1_col = pw_col_offset - 2
            g2_col = pw_col_offset - 1
            for r in range(42, 52):
                g1 = self._str(ws.cell(row=r, column=g1_col).value)
                g2 = self._str(ws.cell(row=r, column=g2_col).value)
                pv = self._sf(ws.cell(row=r, column=pw_col_offset).value)
                sig = self._str(ws.cell(row=r, column=pw_col_offset + 1).value)
                if g1 and g2:
                    pairwise.append({"group_1": g1, "group_2": g2, "p_value": pv, "significant": sig})

        return {
            "run_label": run_label,
            "processed_sheet_name": sn,
            "group_labels": [g["label"] for g in groups],
            "fields": fields,
            "field_count": len(fields),
            "avg": avg,
            "sd": sd,
            "sum_n": sum_n,
            "acceptance_1": {"text": accept_1_text, "result": accept_1_result},
            "acceptance_2": {"text": accept_2_text, "result": accept_2_result},
            "ricc_formula": ricc_formula,
            "ricc_data": ricc_data,
            "ricc_results": ricc_results,
            "ricc_acceptance": {"text": ricc_accept_text, "result": ricc_accept_result},
            "pairwise": pairwise,
            "pairwise_alpha": pw_alpha,
        }

    def extract_processed_data(self):
        blocks = []; seen = set()
        for sn in self._find_sheets("processed_data_"):
            sn_norm = re.sub(r'_+', '_', sn)
            if not re.search(_MNT_ID, sn_norm, re.IGNORECASE): continue
            sl_norm = sn_norm.lower()
            if sl_norm in seen: continue
            seen.add(sl_norm)
            blocks.append(self._extract_processed_block(sn))
        return blocks

    # ====== FINAL RESULTS ======
    def extract_final_results(self):
        sheets = self._find_sheets("final_results") + self._find_sheets("final results")
        if not sheets: return {"available": False}
        ws = self.wb[sheets[0]]

        # Group headers: row 5, cols 4,7,10,13,16
        fr_groups = []
        for start_col in [4, 7, 10, 13, 16]:
            label = self._str(ws.cell(row=5, column=start_col).value)
            if label: fr_groups.append({"label": label, "n_col": start_col, "un_col": start_col + 1, "pct_col": start_col + 2})

        # Per-experiment data (rows 7-14, pairs of mean+sum)
        experiments = []
        for r in range(7, 15, 2):
            exp_label = self._str(ws.cell(row=r, column=2).value)
            row_type = self._str(ws.cell(row=r, column=3).value)
            if not exp_label: continue
            mean_data = {}; sum_data = {}
            for g in fr_groups:
                mean_data[f"{g['label']}_N"] = self._sf(ws.cell(row=r, column=g["n_col"]).value)
                mean_data[f"{g['label']}_uN"] = self._sf(ws.cell(row=r, column=g["un_col"]).value)
                mean_data[f"{g['label']}_%uN/N"] = self._sf(ws.cell(row=r, column=g["pct_col"]).value)
                sum_data[f"{g['label']}_N"] = self._sf(ws.cell(row=r + 1, column=g["n_col"]).value)
                sum_data[f"{g['label']}_uN"] = self._sf(ws.cell(row=r + 1, column=g["un_col"]).value)
            experiments.append({"label": exp_label, "mean": mean_data, "sum": sum_data})

        # Overall mean (row 15), SD (row 16)
        overall_mean = {}; overall_sd = {}
        for g in fr_groups:
            overall_mean[f"{g['label']}_N"] = self._sf(ws.cell(row=15, column=g["n_col"]).value)
            overall_mean[f"{g['label']}_uN"] = self._sf(ws.cell(row=15, column=g["un_col"]).value)
            overall_mean[f"{g['label']}_%uN/N"] = self._sf(ws.cell(row=15, column=g["pct_col"]).value)
            overall_sd[f"{g['label']}_N"] = self._sf(ws.cell(row=16, column=g["n_col"]).value)
            overall_sd[f"{g['label']}_uN"] = self._sf(ws.cell(row=16, column=g["un_col"]).value)
            overall_sd[f"{g['label']}_%uN/N"] = self._sf(ws.cell(row=16, column=g["pct_col"]).value)

        # Summary table (rows 39-44)
        summary = []
        for r in range(40, 45):
            label = self._str(ws.cell(row=r, column=2).value)
            if label is None: continue
            summary.append({"label": label, "conc_ug": self._str(ws.cell(row=r, column=3).value), "conc_part": self._str(ws.cell(row=r, column=4).value), "pct_uN_N": self._sf(ws.cell(row=r, column=5).value), "sd": self._sf(ws.cell(row=r, column=6).value)})

        # Acceptance criteria
        accept_1 = {"text": self._str(ws.cell(row=48, column=2).value), "result": self._str(ws.cell(row=48, column=9).value)}
        accept_2 = {"text": self._str(ws.cell(row=51, column=2).value), "result": self._str(ws.cell(row=51, column=9).value)}

        return {
            "available": True,
            "group_labels": [g["label"] for g in fr_groups],
            "experiments": experiments,
            "overall_mean": overall_mean,
            "overall_sd": overall_sd,
            "summary": summary,
            "acceptance_1": accept_1,
            "acceptance_2": accept_2,
        }

    # ====== STATISTICAL ANALYSIS ======
    def extract_statistical_analysis(self):
        sheets = self._find_sheets("statistical_analysis") + self._find_sheets("statistical analysis")
        if not sheets: return {"available": False}
        ws = self.wb[sheets[0]]

        # Experiment data (rows 4-9)
        group_headers = [self._str(ws.cell(row=4, column=c).value) for c in range(3, 10) if ws.cell(row=4, column=c).value is not None]
        exp_data = []
        for r in range(5, 10):
            label = self._str(ws.cell(row=r, column=2).value)
            if not label: continue
            vals = [self._sf(ws.cell(row=r, column=c).value) for c in range(3, 3 + len(group_headers))]
            exp_data.append({"label": label, "values": dict(zip(group_headers, vals))})

        # ANOVA summary (rows 15-21, Polish labels preserved)
        anova_summary = []
        for r in range(17, 25):
            label = self._str(ws.cell(row=r, column=2).value)
            if not label: break
            anova_summary.append({"group": label, "count": self._sf(ws.cell(row=r, column=3).value), "sum": self._sf(ws.cell(row=r, column=4).value), "mean": self._sf(ws.cell(row=r, column=5).value), "variance": self._sf(ws.cell(row=r, column=6).value)})

        # ANOVA table (rows 25-29, Polish labels preserved)
        anova_table = []
        for r in range(26, 30):
            src = self._str(ws.cell(row=r, column=2).value)
            if not src or src.lower() in ('razem', 'total'): continue
            anova_table.append({"source": src, "ss": self._sf(ws.cell(row=r, column=3).value), "df": self._sf(ws.cell(row=r, column=4).value), "ms": self._sf(ws.cell(row=r, column=5).value), "f_stat": self._sf(ws.cell(row=r, column=6).value), "p_value": self._sf(ws.cell(row=r, column=7).value), "f_crit": self._sf(ws.cell(row=r, column=8).value)})

        # Total row (row 29)
        total_ss = self._sf(ws.cell(row=29, column=3).value)
        total_df = self._sf(ws.cell(row=29, column=4).value)

        # p-value significance
        alpha = self._sf(ws.cell(row=31, column=4).value)
        is_significant = self._str(ws.cell(row=32, column=2).value)

        # Post-hoc pairwise comparisons (rows 40-50)
        posthoc = []
        for r in range(41, 55):
            groups = self._str(ws.cell(row=r, column=2).value)
            if not groups or not groups.startswith("Group"): continue
            posthoc.append({"comparison": groups, "p_value": self._sf(ws.cell(row=r, column=3).value), "significant": self._str(ws.cell(row=r, column=4).value)})

        # Post-hoc alpha levels
        posthoc_alpha = {"anova": self._sf(ws.cell(row=42, column=7).value), "bonferroni": self._sf(ws.cell(row=43, column=7).value)}

        # Post-hoc acceptance
        posthoc_accept = {"text": self._str(ws.cell(row=54, column=2).value), "result": self._str(ws.cell(row=54, column=9).value)}

        return {
            "available": True,
            "group_headers": group_headers,
            "experiment_data": exp_data,
            "anova_summary_label": self._str(ws.cell(row=15, column=2).value),
            "anova_summary": anova_summary,
            "anova_table_label": self._str(ws.cell(row=24, column=2).value),
            "anova_table": anova_table,
            "total_ss": total_ss,
            "total_df": total_df,
            "alpha": alpha,
            "is_significant": is_significant,
            "posthoc": posthoc,
            "posthoc_alpha": posthoc_alpha,
            "posthoc_acceptance": posthoc_accept,
        }

    # ====== PARSE ALL ======
    def parse_all_data(self):
        try:
            return {
                "test_details": {"work_package": self.extract_work_package(), "material": self.extract_material(), "cell_line": self.extract_cell_line(), "dispersion": self.extract_dispersion(), "treatment": self.extract_treatment()},
                "replication_metadata": self.extract_replication_metadata(),
                "replications": self.extract_raw_data(),
                "processed_data": self.extract_processed_data(),
                "final_results": self.extract_final_results(),
                "statistical_analysis": self.extract_statistical_analysis(),
            }
        except Exception as e:
            logger.error(f"Error: {e}\n{traceback.format_exc()}"); raise


def _fix_degree_symbols(obj):
    if isinstance(obj, str): return obj.replace("oC", "\u00b0C")
    if isinstance(obj, dict): return {(k.replace("oC", "\u00b0C") if isinstance(k, str) else k): _fix_degree_symbols(v) for k, v in obj.items()}
    if isinstance(obj, list): return [_fix_degree_symbols(v) for v in obj]
    return obj

def parse_excel_mnt(file_path, sheet_name="Test_conditions"):
    return _fix_degree_symbols(MNTParser(file_path, sheet_name).parse_all_data())

if __name__ == "__main__":
    import sys
    fp = sys.argv[1] if len(sys.argv) > 1 else "/mnt/user-data/uploads/CMS_MNT_26b_CH__CIT.xlsx"
    d = parse_excel_mnt(fp)
    print(json.dumps(d, indent=2, default=str)[:12000])