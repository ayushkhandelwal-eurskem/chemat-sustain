import openpyxl
import re
from dataclasses import dataclass, asdict, field
from typing import List, Dict, Optional, Union
import traceback
@dataclass
class Scientist:
    name: Optional[str] = None
    email: Optional[str] = None

@dataclass
class WorkPackageData:
    wp_name: Optional[str] = None
    partner: Optional[str] = None
    full_test_name: Optional[str] = None
    test_acronym: Optional[str] = None
    test_type: Optional[str] = None
    endpoint: Optional[str] = None
    sop: Optional[str] = None
    path: Optional[str] = None
    
    lead_scientists: List[Scientist] = None
    assay_scientists: List[Scientist] = None

@dataclass
class MaterialData:
    material_identifier: Optional[str] = None
    erm_id : Optional[str] = None
    core_chemistry: Optional[str] = None
    material_state: Optional[str] = None
    batch: Optional[str] = None
    preparation_date: Optional[str] = None
    endotoxin_absent: Optional[str] = None
    particles_stock: Optional[str] = None
    aids_to_disperse: Optional[str] = None
    time_point_unit: Optional[str] = None
    treatment_concentration_unit: Optional[str] = None
    positive_controls_abbr: Optional[str] = None

@dataclass
class TreatmentConcentration:
    labels: Optional[int] = None
    c_ugml: Optional[int] = None
    no_of_particles: Optional[float] = None
    treatment_type_series: Optional[int] = None

@dataclass
class ReplicationData:
    test_identifier_number: Optional[str] = None
    test_start_date: Optional[str] = None
    test_end_date: Optional[str] = None
    no_of_replicate: Optional[int] = None

@dataclass
class ProtocolData:
    protocol_name: Optional[str] = None
    protocol_number: Optional[str] = None
    name_of_the_plate_type: Optional[str] = None
    number_of_repeats: Optional[int] = None
    delay_between_repeats: Optional[str] = None
    measurement_height: Optional[str] = None
    shaking_duration: Optional[str] = None
    shaking_speed: Optional[str] = None
    shaking_diameter: Optional[str] = None
    shaking_type: Optional[str] = None
    repeated_operation: Optional[str] = None

@dataclass
class LabelData:
    name: Optional[str] = None
    technology: Optional[str] = None
    clamp_filter_name: Optional[str] = None
    clamp_filter_slot: Optional[str] = None
    measurement_time: Optional[str] = None
    absorbance_mode: Optional[str] = None
    excitation_aperture: Optional[str] = None

@dataclass
class RawData:
    plate: Optional[str] = None
    repeat: Optional[str] = None
    well: Optional[str] = None
    type: Optional[str] = None
    time_1: Optional[str] = None
    p570: Optional[float] = None
    time_2: Optional[str] = None
    p650: Optional[float] = None

@dataclass
class PlateMapData:
    column_1: List[str] = None
    column_2: List[str] = None
    column_3: List[str] = None
    column_4: List[str] = None
    column_5: List[str] = None
    column_6: List[str] = None
    column_7: List[str] = None
    column_8: List[str] = None

@dataclass
class AcceptanceCriteria:
    criteria_type: str  # CV, NC percentage, OD value, etc.
    threshold: Union[float, str]  # Threshold value or description
    status: str  # PASSED or FAILED

@dataclass
class AbsorbanceMeasurement:
    wavelength: int
    concentrations: List[float]  # NPs concentrations (10, 7.5, 5, etc.)
    readings: Dict[str, List[float]]  # NC, 10, 7.5, etc. as keys with corresponding readings
    mean_values: Dict[str, float] = None
    std_dev: Optional[Dict[str, float]] = None
    cv_values: Optional[Dict[str, float]] = None
    acceptance_criteria: Optional[List[AcceptanceCriteria]] = None
    mean_nc: Optional[float] = None  # Mean value for NC (negative control)

@dataclass
class ViabilityData:
    concentrations: List[float]
    readings: Dict[str, List[float]]
    mean_values: Dict[str, float]
    std_dev: Dict[str, float]
    percentage_values: Dict[str, float]
    percentage_std: Dict[str, float]

@dataclass
class ExperimentData:
    processed_sheet_name: str
    experiment_id: str
    absorbance_570: AbsorbanceMeasurement
    absorbance_650: AbsorbanceMeasurement
    background_subtraction: Dict[str, List[float]]
    mean_values: Dict[str, float]
    nc_status: Dict[str, str]
    od_criteria: AcceptanceCriteria
    viability_data: ViabilityData

@dataclass
class PercentViabilityVsNC:
    concentrations: List[Union[float, str]] = field(default_factory=list)
    mean: List[float] = field(default_factory=list)
    std_dev: List[float] = field(default_factory=list)
    reverse_mean: List[float] = field(default_factory=list)
    reverse_std_dev: List[float] = field(default_factory=list)
    reverse_mean_without_pc: List[float] = field(default_factory=list)
    reverse_std_dev_without_pc: List[float] = field(default_factory=list)

@dataclass
class FinalResults:
    excluded_sheets: List[str] = None
    percent_viability_vs_nc: PercentViabilityVsNC = None
    reverse_concentrations: List[Union[float, str]] = field(default_factory=list)
    concentrations_dash: List[Union[float, str]] = field(default_factory=list)
    reverse_concentrations_dash: List[Union[float, str]] = field(default_factory=list)
    log_dose : List[Union[float, str]] = field(default_factory=list)
    dose : List[Union[float, str]] = field(default_factory=list)
    log_dose_dash : List[Union[float, str]] = field(default_factory=list)
    dose_dash : List[Union[float, str]] = field(default_factory=list)
    r_squared : float = None
    r : float = None
    slope : float = None
    intercept : float = None
    slope_dash : float = None
    intercept_dash : float = None

    ec10: List[float] = field(default_factory=list)
    ec25: List[float] = field(default_factory=list)
    ec50: List[float] = field(default_factory=list)

    ec10_dash: List[float] = field(default_factory=list)
    ec25_dash: List[float] = field(default_factory=list)
    ec50_dash: List[float] = field(default_factory=list)

    final_table : List[Dict[str, Union[str, float]]] = field(default_factory=list)

class MTTParser:
    def __init__(self, file_path, sheet_name="Test_conditions"):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.wb = openpyxl.load_workbook(file_path, data_only=True)
        self.ws = self.wb[sheet_name]
        self.email_regex = r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$'

    def normalize_key(self, key):
        """Convert key to lowercase, replace spaces/special characters with underscores."""
        if key:
            normalized = key.strip().lower()
            normalized = re.sub(r'[^a-zA-Z0-9]', '_', normalized)
            normalized = re.sub(r'_+', '_', normalized)
            return normalized
        return None

    def split_value_unit(self, value):
        """Split a string value into its numeric part and unit."""
        if isinstance(value, str) and " " in value:
            parts = value.split(" ", 1)
            return parts[0], parts[1]  # Return the numeric value and the unit
        return value, None

    def extract_work_package_data(self):
        data = []
        lead_scientists = []
        assay_scientists = []

        for row in self.ws.iter_rows():
            key_cell = self.ws.cell(row=row[0].row, column=1).value
            value_cell = self.ws.cell(row=row[0].row, column=2).value
            email_cell = self.ws.cell(row=row[0].row, column=4).value
            comment = row[0].comment

            if comment:
                key = self.normalize_key(key_cell)
                entry = {"Key": key, "Value": value_cell}

                if email_cell and re.match(self.email_regex, str(email_cell)):
                    entry["Email"] = email_cell

                if key and "lead_scientist_contact_for_test_" in key:
                    lead_scientists.append(asdict(Scientist(name=value_cell, email=email_cell)))

                if key and "assay_test_work_conducted_by_" in key:
                    assay_scientists.append(asdict(Scientist(name=value_cell, email=email_cell)))

                data.append(entry)

        wp_data = asdict(WorkPackageData(
            wp_name=next((d["Value"] for d in data if d["Key"] == "project_work_package_"), None),
            partner=next((d["Value"] for d in data if d["Key"] == "partner_conducting_test_assay_"), None),
            full_test_name=next((d["Value"] for d in data if d["Key"] == "full_name_of_test_assay_add_oecd_test_ref_id_if_app_"), None),
            test_acronym=next((d["Value"] for d in data if d["Key"] == "short_name_or_acronym_for_test_assay_"), None),
            test_type=next((d["Value"] for d in data if d["Key"] == "type_or_class_of_experimental_test_as_used_here_"), None),
            endpoint=next((d["Value"] for d in data if d["Key"] == "end_point_being_investigated_assessed_by_the_test_"), None),
            sop=next((d["Value"] for d in data if d["Key"] == "sop_s_for_test_ref_project_or_other_doc_title_id_"), None),
            path=next((d["Value"] for d in data if d["Key"] == "path_link_to_sop_protocol_on_proj_server_web_where_applic_"), None),
            lead_scientists=lead_scientists,
            assay_scientists=assay_scientists
        ))

        return wp_data

    def extract_material_data(self):
        data = []
        for row in self.ws.iter_rows():
            key_cell = self.ws.cell(row=row[0].row, column=1).value
            value_cell = self.ws.cell(row=row[0].row, column=2).value
            comment = row[0].comment

            
            key = self.normalize_key(key_cell)
            data.append({"Key": key, "Value": value_cell})

        material_data = asdict(MaterialData(
            material_identifier=next((d["Value"] for d in data if d["Key"] == "test_material_details"), None),
            erm_id=next((d["Value"] for d in data if d["Key"] == "erm_identifier_number_"), None),
            core_chemistry=next((d["Value"] for d in data if d["Key"] == "core_chemistry_"), None),
            material_state=next((d["Value"] for d in data if d["Key"] == "material_state_"), None),
            batch=next((d["Value"] for d in data if d["Key"] == "batch"), None),
            preparation_date=next((d["Value"] for d in data if d["Key"] == "date_of_preparation_"), None),
            endotoxin_absent=next((d["Value"] for d in data if d["Key"] == "endotoxin_confirmed_as_absent_"), None),
            particles_stock=next((d["Value"].split()[0] for d in data if d["Key"] == "no_of_particles_in_stock_"), None),
            aids_to_disperse=next((d["Value"] for d in data if d["Key"] == "aids_used_to_disperse_"), None),
            time_point_unit=next((d["Value"] for d in data if d["Key"] == "time_point_unit_"), None),
            treatment_concentration_unit=next((d["Value"] for d in data if d["Key"] == "treatment_concentration_series_unit_"), None),
            positive_controls_abbr=next((d["Value"] for d in data if d["Key"] == "positive_controls_abbreviations_"), None)
        ))

        return material_data
    
    def extract_treatment_concentration_data(self):
        self.ws = self.wb["Test_conditions"]
        treatment_data = {}  # Initialize as a dictionary
        for current_row in range(83, 92):
            row_data = [self.ws.cell(row=current_row, column=col).value for col in range(2, self.ws.max_column + 1)]
            # if all(value is None for value in row_data):
            #     break  # Stop if the row is empty
            trimmed_row_data = [value for value in row_data if value is not None]
            
            # Store the data in the dictionary using the normalized key from column 1
            treatment_data[self.normalize_key(self.ws.cell(row=current_row, column=1).value)] = trimmed_row_data
            
            current_row += 1
        return treatment_data

    def extract_replications(self):
        test_conditions_sheet = "Test_conditions"
        ws = self.wb[test_conditions_sheet]

        replication_start_row = 39
        replication_columns = [2, 3, 4, 5]

        replications = []
        for row in ws.iter_rows(min_row=replication_start_row, min_col=replication_columns[0], max_col=replication_columns[-1]):
            if all(cell.value is None for cell in row):
                break
            if row[0].value is None or not str(row[0].value).startswith("WP"):
                break
            
            replication = asdict(ReplicationData(
                test_identifier_number=row[0].value,
                test_start_date=row[1].value,
                test_end_date=row[2].value,
                no_of_replicate=row[3].value
            ))
            replications.append(replication)

        return replications

    def extract_raw_data(self, test_identifier_number):
        sheet_name = f"Raw_data_{test_identifier_number}"
        raw_sheet = self.wb[sheet_name]

        raw_data_start_row = 4
        raw_data_columns = [3, 4, 5, 6, 7, 8, 9, 10]
        raw_data = []

        for row in raw_sheet.iter_rows(min_row=raw_data_start_row, min_col=raw_data_columns[0], max_col=raw_data_columns[-1]):
            if all(cell.value is None for cell in row):
                break
            
            raw_record = asdict(RawData(
                plate=row[0].value,
                repeat=row[1].value,
                well=row[2].value,
                type=row[3].value,
                time_1=row[4].value,
                p570=row[5].value,
                time_2=row[6].value,
                p650=row[7].value
            ))
            raw_data.append(raw_record)

        return raw_data
    
    def extract_protocol_data(self, sheet_name):
        raw_sheet = self.wb[sheet_name]
        protocol_start_row = 100
        protocol_end_row = 149
        
        def extract_label_and_value(text):
            pattern = r"^(.*?)\s*\.+\s*(.*?)$"
            match = re.search(pattern, text)
            
            if match:
                label = match.group(1).strip()
                value = match.group(2).strip()
                return [label, value]
            else:
                return None

        protocol_data = []
        plate_map = []

        for i in range(protocol_start_row, protocol_end_row + 1):
            row_value = raw_sheet.cell(row=i, column=3).value  # Column C
            if row_value:
                row_content = str(row_value).strip()

                # Check if the row starts with "Plate map"
                if "plate map" in row_content.lower():
                    for j in range(1, 9):  # Next 8 rows for plate map
                        plate_row = raw_sheet.cell(row=i + j, column=3).value  # Column C
                        if plate_row:
                            # Split plate row into columns
                            plate_row_cleaned = plate_row.split("|", 1)[-1].strip().split()
                            plate_map.append(plate_row_cleaned)
                    break

                # Extract protocol description
                parts = extract_label_and_value(row_content)
                if len(parts) == 2:
                    protocol_data.append({
                        "label": self.normalize_key(parts[0]),
                        "value": parts[1]
                    })

        # Extract protocol details
        protocol_desc_columns = [
            "protocol_name", "protocol_number", "name_of_the_plate_type",
            "number_of_repeats", "delay_between_repeats", "measurement_height",
            "shaking_duration", "shaking_speed", "shaking_diameter", "shaking_type",
            "repeated_operation"
        ]

        protocol_desc_data = {col: None for col in protocol_desc_columns}
        for item in protocol_data:
            if item["label"] in protocol_desc_columns:
                split_value = re.split(r'[ /]', item["value"])
                if len(split_value) == 2 and split_value[0].isdigit():
                    try:
                        protocol_desc_data[item["label"]] = (
                            float(split_value[0]) if '.' in split_value[0] 
                            else int(split_value[0])
                        )
                    except ValueError:
                        protocol_desc_data[item["label"]] = split_value[0]
                else:
                    protocol_desc_data[item["label"]] = item["value"]

        # Extract Label Data
        label_columns = [
            "name_of_the_label", "label_technology", "cw_lamp_filter_name", 
            "cw_lamp_filter_slot", "measurement_time", "absorbance_mode", 
            "excitation_aperture"
        ]

        labels = []
        current_label = None

        for item in protocol_data:
            if item["label"] == "name_of_the_label":
                if current_label:
                    labels.append(current_label)
                current_label = {}
            if item["label"] in label_columns:
                # Handle splitting of value and unit for specific fields
                if item["label"] == "measurement_time":
                    value, unit = self.split_value_unit(item["value"])
                    current_label["measurement_time_value"] = value
                    current_label["measurement_time_unit"] = unit
                else:
                    current_label[item["label"]] = item["value"]

        if current_label:
            labels.append(current_label)

        # Convert labels to LabelData objects
        label_objects = []
        for label in labels:
            label_obj = asdict(LabelData(
                name=label.get("name_of_the_label"),
                technology=label.get("label_technology"),
                clamp_filter_name=label.get("cw_lamp_filter_name"),
                clamp_filter_slot=label.get("cw_lamp_filter_slot"),
                measurement_time=label.get("measurement_time_value"),
                absorbance_mode=label.get("absorbance_mode"),
                excitation_aperture=label.get("excitation_aperture")
            ))
            label_objects.append(label_obj)

        # Convert plate map to PlateMapData
        plate_map_data = asdict(PlateMapData(
            column_1=plate_map[0] if len(plate_map) > 0 else None,
            column_2=plate_map[1] if len(plate_map) > 1 else None,
            column_3=plate_map[2] if len(plate_map) > 2 else None,
            column_4=plate_map[3] if len(plate_map) > 3 else None,
            column_5=plate_map[4] if len(plate_map) > 4 else None,
            column_6=plate_map[5] if len(plate_map) > 5 else None,
            column_7=plate_map[6] if len(plate_map) > 6 else None,
            column_8=plate_map[7] if len(plate_map) > 7 else None
        ))

        protocol_details = {
            'protocol': asdict(ProtocolData(
                protocol_name=protocol_desc_data["protocol_name"],
                protocol_number=protocol_desc_data["protocol_number"],
                name_of_the_plate_type=protocol_desc_data["name_of_the_plate_type"],
                number_of_repeats=protocol_desc_data["number_of_repeats"],
                delay_between_repeats=protocol_desc_data["delay_between_repeats"],
                measurement_height=protocol_desc_data["measurement_height"],
                shaking_duration=protocol_desc_data["shaking_duration"],
                shaking_speed=protocol_desc_data["shaking_speed"],
                shaking_diameter=protocol_desc_data["shaking_diameter"],
                shaking_type=protocol_desc_data["shaking_type"],
                repeated_operation=protocol_desc_data["repeated_operation"]
            )),
            'labels': label_objects,
            'plate_map': plate_map_data
        }

        return protocol_details

    def extract_processed_data(self):
        processed_sheets = [
            sheet for sheet in self.wb.sheetnames if sheet.strip().lower().startswith("processed")
        ]
        processed_data = []
        for sheet_name in processed_sheets:
            sheet = self.wb[sheet_name]
            # Extract experiment ID
            experiment_id = sheet.cell(row=3, column=1).value
            
            # Get concentrations (column headers)
            concentrations = []
            for col in range(2, 12):  # Assuming columns B to K contain the concentrations
                header = sheet.cell(row=5, column=col).value
                if header is not None and header != "":
                    if header not in concentrations:
                        concentrations.append(header)
                    else:
                        concentrations.append(header + "'")

            
            # Parse 570nm absorbance data
            absorbance_570_readings = {}
            for col_idx, conc in enumerate(concentrations, start=2):
                readings = []
                for row in range(6, 12):  # Rows 6-11 contain the readings
                    readings.append(sheet.cell(row=row, column=col_idx).value)
                absorbance_570_readings[str(conc)] = readings
            absorbance_570_readings["mean_nc"] = sheet["J15"].value
            # Get mean, SD, and CV for 570nm
            mean_570 = {str(conc): sheet.cell(row=12, column=idx+2).value 
                        for idx, conc in enumerate(concentrations)}
            sd_570 = {str(conc): sheet.cell(row=13, column=idx+2).value 
                    for idx, conc in enumerate(concentrations)}
            cv_570 = {str(conc): sheet.cell(row=14, column=idx+2).value 
                    for idx, conc in enumerate(concentrations)}
            
            # Get acceptance status for CV
            cv_acceptance = asdict(AcceptanceCriteria(
                criteria_type="CV",
                threshold=20.0,
                status=sheet.cell(row=17, column=7).value
            ))
            
            # Parse 650nm absorbance data
            absorbance_650_readings = {}
            for col_idx, conc in enumerate(concentrations, start=2):
                readings = []
                for row in range(21, 27):  # Rows 21-26 contain the readings
                    readings.append(sheet.cell(row=row, column=col_idx).value)
                absorbance_650_readings[str(conc)] = readings
            
            # Background subtraction data
            background_subtraction = {}
            background_subtraction["concentrations"] = concentrations
            for col_idx, conc in enumerate(concentrations, start=2):
                readings = []
                for row in range(31, 37):  # Rows 31-36 contain the readings
                    readings.append(sheet.cell(row=row, column=col_idx).value)
                background_subtraction[str(conc)] = readings
            background_subtraction["mean_nc"] = sheet["B38"].value
            background_subtraction["mean_nc_dash"] = sheet["J38"].value
            background_subtraction["mean_all_nc"] = sheet["B39"].value
            # Mean values
            mean_values = {
                'background': sheet.cell(row=38, column=3).value,
                'all_nc': sheet.cell(row=39, column=3).value,
                'mean_650': sheet.cell(row=38, column=11).value
            }
            
            
            nc_status = {
                'NC1': sheet.cell(row=42, column=2).value,
                'NC2': sheet.cell(row=43, column=2).value,
                'NC1_greater_85': sheet.cell(row=42, column=3).value,
                'NC1_less_115': sheet.cell(row=42, column=4).value,
                'NC2_greater_85': sheet.cell(row=43, column=3).value,
                'NC2_less_115': sheet.cell(row=43, column=4).value
            }
            
            # OD criteria
            od_criteria = asdict(AcceptanceCriteria(
                criteria_type="OD570",
                threshold=">0.2 for NC",
                status=sheet.cell(row=42, column=9).value
            ))
            
            # Viability data
            viability_readings = {}
            for col_idx, conc in enumerate(concentrations, start=2):
                readings = []
                for row in range(50, 56):  # Rows 50-55 contain the viability readings
                    readings.append(sheet.cell(row=row, column=col_idx).value)
                viability_readings[str(conc)] = readings
            
            viability_means = {str(conc): sheet.cell(row=56, column=idx+2).value 
                            for idx, conc in enumerate(concentrations)}
            
            viability_std = {str(conc): sheet.cell(row=57, column=idx+2).value 
                            for idx, conc in enumerate(concentrations)}
            
            # Percentage of viability
            viability_percentages = {}
            viability_percent_std = {}
            
            concentrations_without_nc = [c for c in concentrations if c != "NC'"]
            for col_idx, conc in enumerate(concentrations_without_nc, start=2):
                if col_idx <= 11:  # Ensure we don't go out of bounds
                    viability_percentages[str(conc)] = sheet.cell(row=60, column=col_idx).value
                    viability_percent_std[str(conc)] = sheet.cell(row=61, column=col_idx).value
            
            # Create data objects
            absorbance_570 = asdict(AbsorbanceMeasurement(
                wavelength=570,
                concentrations=[float(c) if isinstance(c, (int, float)) else c for c in concentrations],
                readings=absorbance_570_readings,
                mean_values=mean_570,
                std_dev=sd_570,
                cv_values=cv_570,
                acceptance_criteria=[cv_acceptance],
                mean_nc=absorbance_570_readings["mean_nc"],
            ))
            
            absorbance_650 = asdict(AbsorbanceMeasurement(
                wavelength=650,
                concentrations=[float(c) if isinstance(c, (int, float)) else c for c in concentrations],
                readings=absorbance_650_readings,
                std_dev={},  # These values aren't visible in the image
                cv_values={},  # These values aren't visible in the image
            ))
            #print([float(c) if isinstance(c, (int, float)) else c for c in concentrations])
            viability_data = asdict(ViabilityData(
                concentrations=[float(c) if isinstance(c, (int, float)) else c for c in concentrations],
                readings=viability_readings,
                mean_values=viability_means,
                std_dev=viability_std,
                percentage_values=viability_percentages,
                percentage_std=viability_percent_std
            ))
            
            experiment_data = asdict(ExperimentData(
                processed_sheet_name=sheet_name,
                experiment_id=experiment_id,
                absorbance_570=absorbance_570,
                absorbance_650=absorbance_650,
                background_subtraction=background_subtraction,
                mean_values=mean_values,
                nc_status=nc_status,
                od_criteria=od_criteria,
                viability_data=viability_data
            ))
            processed_data.append(experiment_data)
        return processed_data

    def extract_final_results(self):
        processed_sheets = [
            sheet for sheet in self.wb.sheetnames if sheet.strip().lower().startswith("processed")
        ]
        sheet = self.wb["Final_results"]
        excluded_replicates = ["B7", "B18", "B29", "B40", "B51", "B62"]
        excluded_sheets = []
        for idx, sheet_name in enumerate(processed_sheets, 0):
            if sheet[excluded_replicates[idx]].value == "excluded":
                excluded_sheets.append(sheet_name)

        self.ws = self.wb["Final_results"]

        concentrations = []
        for cell in self.ws[75][3:12]:  # 0-based indexing, so column D is index 3
            if cell.value is not None:
                concentrations.append(cell.value)
        
        # Extract mean values (row 77, columns D to L)
        mean_values = []
        for cell in self.ws[76][3:12]:
            if cell.value is not None:
                mean_values.append(cell.value)
        
        # Extract standard deviation values (row 78, columns D to L)
        std_dev_values = []
        for cell in self.ws[77][3:12]:
            if cell.value is not None:
                std_dev_values.append(cell.value)
        
        
        percent_viability_vs_nc = asdict(PercentViabilityVsNC(
            concentrations=concentrations,
            mean=mean_values,
            std_dev=std_dev_values,
            reverse_mean=[mean_values[0]] + mean_values[-2:0:-1] + [mean_values[-1]],  # Reverse middle elements only
            reverse_std_dev=[std_dev_values[0]] + std_dev_values[-2:0:-1] + [std_dev_values[-1]],  # Reverse middle elements only  
            reverse_mean_without_pc=[mean_values[0]] + mean_values[-2:0:-1],  # Reverse middle elements only
            reverse_std_dev_without_pc=[std_dev_values[0]] + std_dev_values[-2:0:-1]  # Reverse middle elements only
        ))

        reverse_concentrations = []
        for cell in self.ws[80][3:12]:  # 0-based indexing, so column D is index 3
            if cell.value is not None:
                reverse_concentrations.append(cell.value)

        concentrations_dash = []
        for cell in self.ws[75][17:25]:  # 0-based indexing, so column D is index 3
            if cell.value is not None:
                concentrations_dash.append(cell.value)

        reverse_concentrations_dash = []
        for cell in self.ws[80][17:25]:  # 0-based indexing, so column D is index 3
            if cell.value is not None:
                reverse_concentrations_dash.append(cell.value)

        log_dose = []
        for cell in self.ws[127][3:11]:  # 0-based indexing, so column D is index 3
            if cell.value is not None:
                log_dose.append(cell.value)

        dose = []
        for cell in self.ws[128][3:11]:
            if cell.value is not None:
                dose.append(cell.value)

        log_dose_dash = []
        for cell in self.ws[127][17:25]:
            if cell.value is not None:
                log_dose_dash.append(cell.value)

        dose_dash = []
        for cell in self.ws[128][17:25]:
            if cell.value is not None:
                dose_dash.append(cell.value)

        r_squared = sheet["L118"].value
        r = sheet["L119"].value
        slope = sheet["D133"].value
        intercept = sheet["D134"].value
        slope_dash = sheet["R133"].value
        intercept_dash = sheet["R134"].value

        ec10 = []
        ec25 = []
        ec50 = []

        ec10.extend([sheet["D136"].value, sheet["E136"].value])
        ec25.extend([sheet["D137"].value, sheet["E137"].value])
        ec50.extend([sheet["D138"].value, sheet["E138"].value])


        ec10_dash = []
        ec25_dash = []
        ec50_dash = []

        ec10_dash.extend([sheet["R136"].value, sheet["S136"].value])
        ec25_dash.extend([sheet["R137"].value, sheet["S137"].value])
        ec50_dash.extend([sheet["R138"].value, sheet["S138"].value])

        final_table = {}
        headers = []
        for header in self.ws[144][1:6]:
            if header.value is not None:
                headers.append(header.value)
        
        final_table["headers"] = headers
        final_table["rows"] = []
        self
        for rows in self.ws.iter_rows(min_row=145, max_row=148, min_col=2, max_col=6):
            row_data = []
            for cell in rows:
                if cell.value is not None:
                    if cell.number_format.endswith('%') or 'percent' in cell.number_format.lower():  # Check if cell is in column 2
                        formatted = str((100 * cell.value)) + "%"
                    else:
                        formatted = cell.value
                    row_data.append(formatted)
            final_table["rows"].append(row_data)  



        return asdict(FinalResults(excluded_sheets=excluded_sheets, 
                            percent_viability_vs_nc=percent_viability_vs_nc, 
                            reverse_concentrations=reverse_concentrations,
                            concentrations_dash=concentrations_dash,
                            reverse_concentrations_dash=reverse_concentrations_dash,
                            r_squared=r_squared,
                            r=r,
                            log_dose=log_dose,
                            dose=dose,
                            log_dose_dash=log_dose_dash,
                            dose_dash=dose_dash,
                            slope=slope,
                            intercept=intercept,
                            slope_dash=slope_dash,
                            intercept_dash=intercept_dash,
                            ec10=ec10,
                            ec25=ec25,
                            ec50=ec50,
                            ec10_dash=ec10_dash,
                            ec25_dash=ec25_dash,
                            ec50_dash=ec50_dash,
                            final_table=final_table
                            ))
            



    def parse_all_data(self):
        """
        Parse all data from the Excel file and return a comprehensive dataset.
        
        Returns:
            dict: A dictionary containing parsed data for work package, materials, 
                  replications, raw data, and protocol details.
        """
        # Extract work package and material data from the main sheet
        work_package_data = self.extract_work_package_data()
        material_data = self.extract_material_data()
        # Extract replications
        replications = self.extract_replications()
        processed_data = self.extract_processed_data()

        # Prepare to store raw data and protocol details for each replication
        treatment_concentration_data = self.extract_treatment_concentration_data()
        final_resluts = self.extract_final_results()
        parsed_data = {
            'test_details' : {
                'work_package': work_package_data,
                'material': material_data,
                'treatment_concentration_data': treatment_concentration_data,
            },
            'replications': [],
            'processed_data': processed_data,
            'final_results': final_resluts
        }
        
        # Process each replication
        for replication in replications:
            test_identifier = replication["test_identifier_number"]
            
            # Extract raw data for this replication
            raw_data = self.extract_raw_data(test_identifier)
            
            # Extract protocol details for this replication
            protocol_details = self.extract_protocol_data(f"Raw_data_{test_identifier}")

            # Store replication data
            replication_data = {
                'replication': replication,
                'raw_data': raw_data,
                'protocol_details': protocol_details
            }

            parsed_data['replications'].append(replication_data)
        return parsed_data

def parse_excel_mtt(file_path, sheet_name="Test_conditions"):
    """
    Convenience function to parse an Excel file and return all extracted data.
    
    Args:
        file_path (str): Path to the Excel file
        sheet_name (str, optional): Name of the main sheet. Defaults to "Test_conditions"
    
    Returns:
        dict: Parsed data from the Excel file
    """
    parser = MTTParser(file_path, sheet_name)
    return parser.parse_all_data()

# Example usage
if __name__ == "__main__":
    file_path = "uploads/CMS_WP3_MTT_1a_FINAL.xlsx"
    parsed_data = parse_excel_mtt(file_path)
    
    # Optional: Print out some parsed data for verification
    print("Work Package Name:", parsed_data['work_package'].wp_name)
    print("Material Identifier:", parsed_data['material'].material_identifier)
    print("Number of Replications:", len(parsed_data['replications']))
    # Print full work package details
    print("\nFull Work Package Details:")
    for attr, value in vars(parsed_data['work_package']).items():
        print(f"{attr}: {value}")

    # Print full material details
    print("\nFull Material Details:")
    for attr, value in vars(parsed_data['material']).items():
        print(f"{attr}: {value}")

    print("\nTreatment Concentration Data:")
    # cnt = 0
    # for treatment in parsed_data['treatment_concentration_data']:
    #     print(f"Available keys in treatment:{cnt}", treatment.keys())
    #     cnt += 1
        # for key, values in treatment.items():
        #     print(f"{key}: {values}")
        # print()
    # for key, value in parsed_data["treatment_concentration_data"].items():
    #     print(f"{key}: {value}")
    print(len(parsed_data["treatment_concentration_data"]["treatment_concentration_series_labels"]))
    print("\n--- Replication Data ---")
    print(f"Total Replications: {len(parsed_data['replications'])}")

    for i, replication_entry in enumerate(parsed_data['replications'], 1):
        print(f"\n--- Replication {i} ---")
        
        # Print Replication Details
        print("Replication Object Details:")
        replication = replication_entry['replication']
        for attr, value in vars(replication).items():
            print(f"{attr}: {value}")
        
        # Print Raw Data
        print("\nRaw Data:")
        print(replication_entry['raw_data'])
        
        # Print Protocol Details
        print("\nProtocol Details:")
        if isinstance(replication_entry['protocol_details'], dict):
            for key, value in replication_entry['protocol_details'].items():
                print(f"{key}: {value}")
        else:
            print(replication_entry['protocol_details'])