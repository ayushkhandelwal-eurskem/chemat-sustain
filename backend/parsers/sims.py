import openpyxl
import re
from dataclasses import dataclass, asdict, field
from typing import List, Dict, Optional, Union, Any, Tuple
import logging
import traceback
from datetime import datetime, timedelta
from difflib import SequenceMatcher

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logging.getLogger("multipart.multipart").setLevel(logging.WARNING)
logger = logging.getLogger(__name__)

# ============================ Dataclasses ============================
@dataclass
class Scientist:
    name: Optional[str] = None
    email: Optional[str] = None

@dataclass
class WorkPackageData:
    wp_name: Optional[str] = None
    partner: Optional[str] = None
    laboratory_name: Optional[str] = None
    full_test_name: Optional[str] = None
    test_acronym: Optional[str] = None
    test_type: Optional[str] = None
    endpoint: Optional[str] = None
    endpoint_outcome: Optional[str] = None
    sop: Optional[str] = None
    path: Optional[str] = None
    lead_scientists: List[Scientist] = field(default_factory=list)
    assay_scientists: List[Scientist] = field(default_factory=list)

@dataclass
class MaterialData:
    material_identifier: Optional[str] = None
    erm_id: Optional[str] = None
    material_name: Optional[str] = None
    core_chemistry: Optional[str] = None
    cas: Optional[str] = None
    cas_for_core: Optional[str] = None
    supplier: Optional[str] = None
    material_state: Optional[str] = None
    batch: Optional[str] = None
    preparation_date: Optional[str] = None
    molar_concentration: Optional[str] = None
    particles_stock: Optional[str] = None

@dataclass
class ReplicationData:
    test_identifier_number: Optional[str] = None
    test_start_date: Optional[str] = None
    test_end_date: Optional[str] = None
    replication_count: Optional[int] = None

@dataclass
class SamplePreparationData:
    dispersion_protocol: Optional[str] = None
    dispersion_technique: Optional[str] = None
    dispersion_medium: Optional[str] = None
    sonicator_type: Optional[str] = None
    power: Optional[str] = None
    sonication_time: Optional[str] = None
    tip_thickness: Optional[str] = None
    tip_composition: Optional[str] = None
    ultrasonic_bath_size: Optional[str] = None
    sample_volume: Optional[str] = None
    final_concentration: Optional[str] = None
    additional_info: Optional[str] = None

@dataclass
class SIMSInstrumentationData:
    instrument_specs: Optional[str] = None
    primary_ions: Optional[str] = None
    detector: Optional[str] = None
    measurement_technique: Optional[str] = None
    mass_resolution: Optional[str] = None
    mass_range: Optional[str] = None
    scan_area: Optional[str] = None

@dataclass
class SIMSRawIon:
    channel: Optional[int] = None
    mass: Optional[float] = None
    intensity: Optional[int] = None

@dataclass
class SIMSRawData:
    negative_ions: List[SIMSRawIon] = field(default_factory=list)
    positive_ions: List[SIMSRawIon] = field(default_factory=list)

@dataclass
class SIMSProcessedIon:
    mass: Optional[float] = None
    counts: Optional[int] = None

@dataclass
class SIMSProcessedData:
    negative_ions: List[SIMSProcessedIon] = field(default_factory=list)
    positive_ions: List[SIMSProcessedIon] = field(default_factory=list)
    total_negative_counts: Optional[int] = None
    total_positive_counts: Optional[int] = None

@dataclass
class SIMSFinalIon:
    mass: Optional[float] = None
    fragment: Optional[str] = None

@dataclass
class SIMSFinalResults:
    negative_ions: List[SIMSFinalIon] = field(default_factory=list)
    positive_ions: List[SIMSFinalIon] = field(default_factory=list)


# ============================ Parser ============================
class SIMSParser:
    """
    SIMS Excel parser.

    Uses openpyxl in read_only mode for memory-safe parsing of large
    raw-ion sheets (300k+ rows). The small Test Information sheet is
    cached in memory once because we scan it for several different
    sets of keys.
    """

    EMAIL_REGEX = r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$'

    def __init__(self, file_path: str, sheet_name: str = "Test Information"):
        self.file_path = file_path
        self.sheet_name = sheet_name
        try:
            # read_only=True streams the workbook from disk instead of
            # loading every cell as a Python object. This is the single
            # most important change for large SIMS files.
            self.wb = openpyxl.load_workbook(
                file_path, data_only=True, read_only=True
            )
            logger.info(f"Successfully loaded Excel file: {file_path}")
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            raise

        # Cache the Test Information sheet once. It's tiny (~70 rows)
        # and is scanned by 5 different extract methods.
        self._info_rows: List[Tuple[Any, ...]] = self._cache_info_sheet()

    # ---------------------- helpers ----------------------
    def _cache_info_sheet(self) -> List[Tuple[Any, ...]]:
        if self.sheet_name not in self.wb.sheetnames:
            logger.error(f"Sheet '{self.sheet_name}' not found")
            return []
        ws = self.wb[self.sheet_name]
        # Pad short rows so column access is uniform
        rows = []
        for row in ws.iter_rows(min_row=1, max_col=5, values_only=True):
            padded = tuple(row) + (None,) * (5 - len(row))
            rows.append(padded[:5])
        return rows

    def normalize_key(self, key: Optional[str]) -> Optional[str]:
        if not key:
            return None
        normalized = str(key).strip().lower()
        normalized = re.sub(r'[^a-z0-9]', '_', normalized)
        normalized = re.sub(r'_+', '_', normalized).strip('_')
        return normalized or None

    def excel_date_to_string(self, value: Optional[Union[float, str, datetime]]) -> Optional[str]:
        try:
            if value is None:
                return None
            if isinstance(value, datetime):
                return value.strftime("%Y-%m-%d")
            if isinstance(value, (int, float)):
                base_date = datetime(1899, 12, 30)
                return (base_date + timedelta(days=float(value))).strftime("%Y-%m-%d")
            if isinstance(value, str):
                for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"]:
                    try:
                        return datetime.strptime(value, fmt).strftime("%Y-%m-%d")
                    except ValueError:
                        pass
                return value
            return str(value)
        except Exception as e:
            logger.warning(f"Failed to convert date {value!r}: {e}")
            return None

    def _scan_info_kv(self) -> List[Dict[str, Any]]:
        """
        Scan the cached Test Information sheet and return a list of
        {Key, Value, Email, RawKey} entries. The first non-None value
        in columns 2..5 (index 1..4) is taken as the value, matching
        the original behavior.
        """
        out = []
        for row in self._info_rows:
            key_cell = row[0]
            if not key_cell:
                continue
            raw_key = str(key_cell)
            key = self.normalize_key(raw_key)
            if not key:
                continue

            value = next((row[c] for c in range(1, 5) if row[c] is not None), None)
            email_cell = row[3]

            entry: Dict[str, Any] = {"Key": key, "Value": value, "RawKey": raw_key}
            if email_cell and re.match(self.EMAIL_REGEX, str(email_cell)):
                entry["Email"] = email_cell
            out.append(entry)
        return out

    @staticmethod
    def _find_value(scanned: List[Dict[str, Any]], key: str) -> Any:
        return next((d["Value"] for d in scanned if d["Key"] == key), None)

    @staticmethod
    def _fuzzy_resolve(scanned: List[Dict[str, Any]], expected_keys: List[str]) -> Dict[str, Any]:
        """
        Build a {expected_key: value} dict, resolving each expected key
        with exact match first, then a fuzzy (>0.85) fallback.
        """
        resolved: Dict[str, Any] = {k: None for k in expected_keys}
        # Exact pass
        for d in scanned:
            if d["Key"] in resolved and resolved[d["Key"]] is None:
                resolved[d["Key"]] = d["Value"]
        # Fuzzy pass for any still-missing
        unmatched = [k for k, v in resolved.items() if v is None]
        for missing in unmatched:
            best_ratio = 0.85
            best_value = None
            for d in scanned:
                ratio = SequenceMatcher(None, d["Key"], missing).ratio()
                if ratio > best_ratio:
                    best_ratio = ratio
                    best_value = d["Value"]
            if best_value is not None:
                resolved[missing] = best_value
        return resolved

    # ---------------------- Test Information extractors ----------------------
    def extract_work_package_data(self):
        scanned = self._scan_info_kv()
        lead_scientists = []
        assay_scientists = []
        for d in scanned:
            raw_lower = d["RawKey"].lower()
            if "lead scientist" in raw_lower:
                lead_scientists.append(Scientist(name=d["Value"], email=d.get("Email")))
            if "assay/test work" in raw_lower:
                assay_scientists.append(Scientist(name=d["Value"], email=d.get("Email")))

        wp = WorkPackageData(
            wp_name=self._find_value(scanned, "project_work_package"),
            partner=self._find_value(scanned, "partner_conducting_test_assay"),
            laboratory_name=self._find_value(scanned, "test_facility_laboratory_name"),
            full_test_name=self._find_value(scanned, "full_name_of_test_assay_add_oecd_test_ref_id_if_app"),
            test_acronym=self._find_value(scanned, "short_name_or_acronym_for_test_assay"),
            test_type=self._find_value(scanned, "type_or_class_of_experimental_test_as_used_here"),
            endpoint=self._find_value(scanned, "end_point_being_investigated_assessed_by_the_test"),
            endpoint_outcome=self._find_value(scanned, "metric_s_used_to_assess_end_point_outcome_response"),
            sop=self._find_value(scanned, "sop_s_for_test_ref_project_or_other_doc_title_id"),
            path=self._find_value(scanned, "path_link_to_sop_protocol_on_proj_server_web_where_applic"),
            lead_scientists=lead_scientists,
            assay_scientists=assay_scientists,
        )
        return asdict(wp)

    def extract_material_data(self):
        scanned = self._scan_info_kv()
        expected = [
            "sample_cms_internal_identifier", "erm_identifier_number", "material_name",
            "core_chemistry", "cas_no", "cas_for_core", "material_supplier",
            "material_state", "batch", "date_of_sample_preparation_for_tests",
            "molar_concentration", "number_of_particles_in_stock",
        ]
        r = self._fuzzy_resolve(scanned, expected)
        material = MaterialData(
            material_identifier=r["sample_cms_internal_identifier"],
            erm_id=r["erm_identifier_number"],
            material_name=r["material_name"],
            core_chemistry=r["core_chemistry"],
            cas=r["cas_no"],
            cas_for_core=r["cas_for_core"],
            supplier=r["material_supplier"],
            material_state=r["material_state"],
            batch=r["batch"],
            preparation_date=self.excel_date_to_string(r["date_of_sample_preparation_for_tests"]),
            molar_concentration=r["molar_concentration"],
            particles_stock=r["number_of_particles_in_stock"],
        )
        return asdict(material)

    def extract_sample_preparation_data(self):
        scanned = self._scan_info_kv()
        expected = [
            "specify_standard_dispersion_protocol_used",
            "or_otherwise_specify_dispersion_technique_used",
            "dispersion_dilution_medium", "sonicator_type", "power_w",
            "sonication_time_secs", "tip_thickness_mm", "tip_composition",
            "size_of_ultrasonic_bath_water_volume_dm3", "sample_volume",
            "final_sample_concentration_mg_l_or_ppm", "additional_information",
        ]
        r = self._fuzzy_resolve(scanned, expected)
        sp = SamplePreparationData(
            dispersion_protocol=r["specify_standard_dispersion_protocol_used"],
            dispersion_technique=r["or_otherwise_specify_dispersion_technique_used"],
            dispersion_medium=r["dispersion_dilution_medium"],
            sonicator_type=r["sonicator_type"],
            power=r["power_w"],
            sonication_time=r["sonication_time_secs"],
            tip_thickness=r["tip_thickness_mm"],
            tip_composition=r["tip_composition"],
            ultrasonic_bath_size=r["size_of_ultrasonic_bath_water_volume_dm3"],
            sample_volume=r["sample_volume"],
            final_concentration=r["final_sample_concentration_mg_l_or_ppm"],
            additional_info=r["additional_information"],
        )
        return asdict(sp)

    def extract_instrumentation_data(self):
        scanned = self._scan_info_kv()
        expected = [
            "sims_instrumentation_model_and_company", "primary_ions", "detector",
            "measurement_technique", "mass_resolution", "mass_range", "scan_area",
        ]
        r = self._fuzzy_resolve(scanned, expected)
        instr = SIMSInstrumentationData(
            instrument_specs=r["sims_instrumentation_model_and_company"],
            primary_ions=r["primary_ions"],
            detector=r["detector"],
            measurement_technique=r["measurement_technique"],
            mass_resolution=r["mass_resolution"],
            mass_range=r["mass_range"],
            scan_area=r["scan_area"],
        )
        return asdict(instr)

    def extract_replication(self):
        test_identifier = None
        start_date = None
        end_date = None
        replication_count = None
        for row in self._info_rows:
            key_cell = row[0]
            if not key_cell:
                continue
            key_lower = str(key_cell).lower()
            if 'test identifier number' in key_lower:
                test_identifier = row[1]
                start_date = self.excel_date_to_string(row[2])
                end_date = self.excel_date_to_string(row[3])
            if 'replication' in key_lower:
                try:
                    replication_count = int(row[1]) if row[1] is not None else None
                except (TypeError, ValueError):
                    replication_count = None
        if test_identifier is None:
            logger.warning("Could not find test identifier number in Test Information sheet.")
        return ReplicationData(
            test_identifier_number=test_identifier,
            test_start_date=start_date,
            test_end_date=end_date,
            replication_count=replication_count,
        )

    # ---------------------- Raw / Processed / Final ----------------------
    def extract_raw_data(self, raw_sheet_name: str) -> SIMSRawData:
        if raw_sheet_name not in self.wb.sheetnames:
            logger.error(f"Raw data sheet {raw_sheet_name} not found")
            return SIMSRawData()

        raw_ws = self.wb[raw_sheet_name]
        negative_ions: List[SIMSRawIon] = []
        positive_ions: List[SIMSRawIon] = []

        # Single streamed pass. Cols A,B,C = neg (idx 0,1,2); Q,R,S = pos (idx 16,17,18).
        # values_only=True returns plain tuples, much faster than cell objects.
        neg_done = False
        pos_done = False
        for row in raw_ws.iter_rows(min_row=3, max_col=19, values_only=True):
            # Pad if the row is shorter than expected
            if len(row) < 19:
                row = tuple(row) + (None,) * (19 - len(row))

            # Negative ion
            if not neg_done:
                neg_channel = row[0]
                if neg_channel is None:
                    neg_done = True
                else:
                    try:
                        negative_ions.append(SIMSRawIon(
                            channel=int(neg_channel),
                            mass=round(float(row[1]), 6) if row[1] is not None else None,
                            intensity=int(row[2]) if row[2] is not None else None,
                        ))
                    except (ValueError, TypeError):
                        pass

            # Positive ion
            if not pos_done:
                pos_channel = row[16]
                if pos_channel is None:
                    pos_done = True
                else:
                    try:
                        positive_ions.append(SIMSRawIon(
                            channel=int(pos_channel),
                            mass=round(float(row[17]), 6) if row[17] is not None else None,
                            intensity=int(row[18]) if row[18] is not None else None,
                        ))
                    except (ValueError, TypeError):
                        pass

            if neg_done and pos_done:
                break

        return SIMSRawData(negative_ions=negative_ions, positive_ions=positive_ions)

    def extract_processed_data(self, processed_sheet_name: str) -> SIMSProcessedData:
        if processed_sheet_name not in self.wb.sheetnames:
            logger.error(f"Processed data sheet {processed_sheet_name} not found")
            return SIMSProcessedData()

        proc_ws = self.wb[processed_sheet_name]
        # Pull rows 1..26 once into memory (tiny). row[0] is row 1.
        rows = list(proc_ws.iter_rows(min_row=1, max_row=26, max_col=3, values_only=True))
        # Pad short rows
        rows = [tuple(r) + (None,) * (3 - len(r)) if len(r) < 3 else r[:3] for r in rows]

        negative_ions: List[SIMSProcessedIon] = []
        positive_ions: List[SIMSProcessedIon] = []
        total_negative_counts = None
        total_positive_counts = None

        # Negative ions: rows 2..11 (Excel) -> rows[1..10] (0-indexed)
        for r in rows[1:11]:
            mass, counts = r[1], r[2]
            if mass is not None and mass != '':
                try:
                    negative_ions.append(SIMSProcessedIon(
                        mass=round(float(mass), 2),
                        counts=int(counts) if counts is not None else None,
                    ))
                except (ValueError, TypeError):
                    pass

        # Total negative counts at Excel row 13 -> rows[12]
        if len(rows) > 12 and rows[12][2] is not None:
            try:
                total_negative_counts = int(rows[12][2])
            except (ValueError, TypeError):
                total_negative_counts = None

        # Positive ions: Excel rows 17..24 -> rows[16..23]
        for r in rows[16:24]:
            mass, counts = r[1], r[2]
            if mass is not None and mass != '':
                try:
                    positive_ions.append(SIMSProcessedIon(
                        mass=round(float(mass), 2),
                        counts=int(counts) if counts is not None else None,
                    ))
                except (ValueError, TypeError):
                    pass

        # Total positive counts at Excel row 26 -> rows[25]
        if len(rows) > 25 and rows[25][2] is not None:
            try:
                total_positive_counts = int(rows[25][2])
            except (ValueError, TypeError):
                total_positive_counts = None

        return SIMSProcessedData(
            negative_ions=negative_ions,
            positive_ions=positive_ions,
            total_negative_counts=total_negative_counts,
            total_positive_counts=total_positive_counts,
        )

    def extract_final_results(self) -> SIMSFinalResults:
        final_sheet_name = next(
            (n for n in self.wb.sheetnames if "Final results" in n), None
        )
        if not final_sheet_name:
            logger.error("Final Results sheet not found")
            return SIMSFinalResults()

        final_ws = self.wb[final_sheet_name]
        # Stream rows 4..8, columns B..J. iter_rows works in read_only;
        # range syntax (final_ws['B4:J4']) does NOT.
        rows = list(final_ws.iter_rows(
            min_row=4, max_row=8, min_col=2, max_col=10, values_only=True
        ))
        # rows[0] = Excel row 4 (neg masses)
        # rows[1] = Excel row 5 (neg fragments)
        # rows[3] = Excel row 7 (pos masses)
        # rows[4] = Excel row 8 (pos fragments)
        if len(rows) < 5:
            return SIMSFinalResults()

        negative_ions: List[SIMSFinalIon] = []
        for mass, fragment in zip(rows[0], rows[1]):
            if mass is not None and fragment is not None:
                try:
                    negative_ions.append(SIMSFinalIon(
                        mass=round(float(mass), 2),
                        fragment=str(fragment),
                    ))
                except (ValueError, TypeError):
                    pass

        positive_ions: List[SIMSFinalIon] = []
        for mass, fragment in zip(rows[3], rows[4]):
            if mass is not None and fragment is not None:
                try:
                    positive_ions.append(SIMSFinalIon(
                        mass=round(float(mass), 2),
                        fragment=str(fragment),
                    ))
                except (ValueError, TypeError):
                    pass

        return SIMSFinalResults(negative_ions=negative_ions, positive_ions=positive_ions)

    # ---------------------- Orchestrator ----------------------
    def parse_all_data(self) -> Dict[str, Union[Dict, List]]:
        try:
            parsed_data = {
                'test_details': {
                    'work_package': self.extract_work_package_data(),
                    'material': self.extract_material_data(),
                    'sample_preparation': self.extract_sample_preparation_data(),
                    'instrumentation': self.extract_instrumentation_data(),
                },
                'replication': asdict(self.extract_replication()),
                'replications': [],
                'processed_data': [],
                'final_results': asdict(self.extract_final_results()),
            }

            raw_sheets = [n for n in self.wb.sheetnames if n.startswith("Raw data_WP2_SIMS_")]
            processed_sheets = [n for n in self.wb.sheetnames if n.startswith("Processed data_WP2_SIMS_")]

            for raw_name in raw_sheets:
                parsed_data['replications'].append(asdict(self.extract_raw_data(raw_name)))
            for proc_name in processed_sheets:
                parsed_data['processed_data'].append(asdict(self.extract_processed_data(proc_name)))

            return parsed_data
        except Exception as e:
            logger.error(f"Error parsing Excel file: {e}\n{traceback.format_exc()}")
            raise
        finally:
            try:
                self.wb.close()
            except Exception:
                pass


def parse_excel_sims(file_path: str, sheet_name: str = "Test Information") -> Dict[str, Union[Dict, List]]:
    try:
        parser = SIMSParser(file_path, sheet_name)
        return parser.parse_all_data()
    except Exception as e:
        logger.error(f"Error in parse_excel_sims: {e}\n{traceback.format_exc()}")
        raise


if __name__ == "__main__":
    import sys
    file_path = sys.argv[1] if len(sys.argv) > 1 else "/Users/ayushkhandelwal/Documents/chemat-sustain/backend/data/WP2_SIMS_21aR1.xlsx"
    parsed = parse_excel_sims(file_path)
    print(parsed['test_details'])