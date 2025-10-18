import openpyxl
import re
from dataclasses import dataclass, asdict, field
from typing import List, Dict, Optional, Union
import logging
import traceback
from datetime import datetime, timedelta
from difflib import SequenceMatcher

# Set up logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
logging.getLogger("multipart.multipart").setLevel(logging.WARNING)
logger = logging.getLogger(__name__)

# Dataclasses tailored for UV-Vis
@dataclass
class Scientist:
    name: Optional[str] = None
    email: Optional[str] = None

@dataclass
class WorkPackageData:
    wp_name: Optional[str] = None
    partner: Optional[str] = None
    full_test_name: Optional[str] = None
    test_acronym: Optional[str] = None
    test_type: Optional[str] = None
    endpoint: Optional[str] = None
    endpoint_outcome: Optional[str] = None
    sop: Optional[str] = None
    path: Optional[str] = None
    lead_scientists: List[Scientist] = field(default_factory=list)
    assay_scientists: List[Scientist] = field(default_factory=list)

@dataclass
class MaterialData:
    material_identifier: Optional[str] = None
    erm_id: Optional[str] = None
    core_chemistry: Optional[str] = None
    material_name: Optional[str] = None
    material_state: Optional[str] = None
    cas: Optional[str] = None
    casforcore: Optional[str] = None
    batch: Optional[str] = None
    preparation_date: Optional[str] = None
    particles_stock: Optional[str] = None
    molar_concentration: Optional[str] = None

@dataclass
class ReplicationData:
    test_identifier_number: Optional[str] = None
    test_start_date: Optional[str] = None
    test_end_date: Optional[str] = None

@dataclass
class SamplePreparationData:
    dispersion_protocol: Optional[str] = None
    dispersion_technique: Optional[str] = None
    dispersion_medium: Optional[str] = None
    sonicator_type: Optional[str] = None
    power: Optional[str] = None
    sonication_time: Optional[str] = None
    tip_thickness: Optional[str] = None
    tip_composition: Optional[str] = None
    ultrasonic_bath_size: Optional[str] = None
    sample_volume: Optional[str] = None
    final_concentration: Optional[str] = None
    additional_info: Optional[str] = None

@dataclass
class UVVisInstrumentationData:
    instrument_specs: Optional[str] = None
    software: Optional[str] = None
    display_model: Optional[str] = None
    cell_model: Optional[str] = None
    optical_path_length: Optional[str] = None
    start_wavelength: Optional[str] = None
    end_wavelength: Optional[str] = None
    wavelength_interval: Optional[str] = None
    background: Optional[str] = None

@dataclass
class UVVisRawMeasurement:
    no: Optional[int] = None
    wavelength: Optional[float] = None
    absorbance: Optional[float] = None

@dataclass
class UVVisRawData:
    measurements: List[UVVisRawMeasurement] = field(default_factory=list)

@dataclass
class UVVisPeak:
    max_absorbance: Optional[float] = None
    wavelength: Optional[int] = None
    identified_compound: Optional[str] = None

@dataclass
class UVVisResultsData:
    peaks: List[UVVisPeak] = field(default_factory=list)

class UVVisParser:
    def __init__(self, file_path: str, sheet_name: str = "Test Information"):
        self.file_path = file_path
        self.sheet_name = sheet_name
        try:
            self.wb = openpyxl.load_workbook(file_path, data_only=True)
            self.ws = self.wb[sheet_name]
            logger.info(f"Successfully loaded Excel file: {file_path}")
        except Exception as e:
            logger.error(f"Failed to load workbook or sheet {sheet_name}: {e}")
            raise
        self.email_regex = r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$'

    def normalize_key(self, key: Optional[str]) -> Optional[str]:
        if key:
            normalized = str(key).strip().lower() if key is not None else ""
            normalized = re.sub(r'[^a-z0-9]', '_', normalized)
            normalized = re.sub(r'_+', '_', normalized).strip('_')
            return normalized
        return None

    def split_value_unit(self, value: Optional[str]) -> tuple[Optional[Union[str, float]], Optional[str]]:
        if isinstance(value, str) and " " in value:
            parts = value.split(" ", 1)
            try:
                numeric = float(parts[0]) if '.' in parts[0] else int(parts[0])
                return numeric, parts[1]
            except (ValueError, IndexError):
                return parts[0], parts[1] if len(parts) > 1 else None
        return value, None

    def excel_date_to_string(self, value: Optional[Union[float, str]]) -> Optional[str]:
        try:
            if isinstance(value, (int, float)):
                base_date = datetime(1899, 12, 30)
                delta = timedelta(days=float(value))
                return (base_date + delta).strftime("%Y-%m-%d")
            elif isinstance(value, str):
                for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"]:
                    try:
                        return datetime.strptime(value, fmt).strftime("%Y-%m-%d")
                    except ValueError:
                        pass
            return str(value) if value else None
        except Exception as e:
            logger.warning(f"Failed to convert date {value}: {e}")
            return None

    def extract_work_package_data(self):
        data = []
        lead_scientists = []
        assay_scientists = []

        for row in self.ws.iter_rows():
            key_cell = row[0].value
            value_cell = row[1].value if len(row) > 1 else None
            email_cell = row[3].value if len(row) > 3 else None
            comment = row[0].comment

            if comment:
                key = self.normalize_key(key_cell)
                if not key:
                    continue
                entry = {"Key": key, "Value": value_cell}

                if email_cell and re.match(self.email_regex, str(email_cell)):
                    entry["Email"] = email_cell

                if key and "lead_scientist_contact_for_test" in key:
                    lead_scientists.append(asdict(Scientist(name=value_cell, email=email_cell)))
                if key and "assay_test_work_conducted_by" in key:
                    assay_scientists.append(asdict(Scientist(name=value_cell, email=email_cell)))

                data.append(entry)

        wp_data = asdict(WorkPackageData(
            wp_name=next((d["Value"] for d in data if d["Key"] == "project_work_package"), None),
            partner=next((d["Value"] for d in data if d["Key"] == "partner_conducting_test_assay"), None),
            full_test_name=next((d["Value"] for d in data if d["Key"] == "full_name_of_test_assay_add_oecd_test_ref_id_if_app"), None),
            test_acronym=next((d["Value"] for d in data if d["Key"] == "short_name_or_acronym_for_test_assay"), None),
            test_type=next((d["Value"] for d in data if d["Key"] == "type_or_class_of_experimental_test_as_used_here"), None),
            endpoint=next((d["Value"] for d in data if d["Key"] == "end_point_being_investigated_assessed_by_the_test"), None),
            endpoint_outcome=next((d["Value"] for d in data if d["Key"] == "metric_s_used_to_assess_end_point_outcome_response"), None),
            sop=next((d["Value"] for d in data if d["Key"] == "sop_s_for_test_ref_project_or_other_doc_title_id"), None),
            path=next((d["Value"] for d in data if d["Key"] == "path_link_to_sop_protocol_on_proj_server_web_where_applic"), None),
            lead_scientists=lead_scientists,
            assay_scientists=assay_scientists
        ))
        return wp_data

    def extract_material_data(self):
        data = []
        expected_keys = [
            "sample_cms_internal_identifier",
            "erm_identifier_number",
            "core_chemistry",
            "material_name",
            "material_state",
            "cas_no",
            "cas_for_core",
            "batch",
            "date_of_preparation",
            "no_of_particles_in_stock",
            "molar_concentration"
        ]
        unmatched_keys = set(expected_keys)
        potential_matches = []

        logger.debug("Test Information sheet (columns A:E, rows 1:50) for Material Data:")
        for row_idx in range(1, min(self.ws.max_row + 1, 51)):
            row = self.ws[row_idx]
            values = [str(cell.value)[:30] if cell.value else "None" for cell in row[:5]]
            has_comment = "Y" if row[0].comment else "N"
            logger.debug(f"Row {row_idx} [Comment: {has_comment}]: {values}")

        for row_idx, row in enumerate(self.ws.iter_rows(min_row=1, max_col=5), start=1):
            key_cell = row[0].value
            if not key_cell:
                continue
            raw_key = str(key_cell)
            key = self.normalize_key(raw_key)
            if not key:
                continue

            value = None
            value_col = None
            for col_idx in range(1, 5):
                if col_idx < len(row) and row[col_idx].value is not None:
                    value = row[col_idx].value
                    value_col = col_idx + 1
                    break

            logger.debug(f"Row {row_idx}: Raw Key='{raw_key}', Normalized Key='{key}', Value='{value}', Value Col={value_col}")

            if key in expected_keys:
                unmatched_keys.discard(key)
                data.append({"Key": key, "Value": value})
                logger.debug(f"Exact match: Key='{key}', Value='{value}' at row {row_idx}, col {value_col}")
            else:
                for expected_key in expected_keys:
                    similarity = SequenceMatcher(None, key, expected_key).ratio()
                    if similarity > 0.85:
                        unmatched_keys.discard(expected_key)
                        data.append({"Key": expected_key, "Value": value})
                        logger.debug(f"Fuzzy match: Key='{key}' matched '{expected_key}' (similarity={similarity:.2f}), Value='{value}' at row {row_idx}, col {value_col}")
                    elif similarity > 0.5:
                        potential_matches.append(f"Potential match for '{expected_key}': '{raw_key}' at row {row_idx}, col 1, similarity={similarity:.2f}")

        if potential_matches:
            logger.debug("Potential key matches for Material Data:")
            for match in potential_matches:
                logger.debug(match)
        if unmatched_keys:
            logger.warning(f"Unmatched material data keys: {unmatched_keys}")

        material_data = asdict(MaterialData(
            material_identifier=next((d["Value"] for d in data if d["Key"] == "sample_cms_internal_identifier"), None),
            erm_id=next((d["Value"] for d in data if d["Key"] == "erm_identifier_number"), None),
            core_chemistry=next((d["Value"] for d in data if d["Key"] == "core_chemistry"), None),
            material_name=next((d["Value"] for d in data if d["Key"] == "material_name"), None),
            material_state=next((d["Value"] for d in data if d["Key"] == "material_state"), None),
            cas=next((d["Value"] for d in data if d["Key"] == "cas_no"), None),
            casforcore=next((d["Value"] for d in data if d["Key"] == "cas_for_core"), None),
            batch=next((d["Value"] for d in data if d["Key"] == "batch"), None),
            preparation_date=self.excel_date_to_string(next((d["Value"] for d in data if d["Key"] == "date_of_preparation"), None)),
            particles_stock=next((d["Value"] for d in data if d["Key"] == "no_of_particles_in_stock"), None),
            molar_concentration=next((d["Value"] for d in data if d["Key"] == "molar_concentration"), None)
        ))
        logger.debug(f"Material data extracted: {material_data}")
        return material_data

    def extract_sample_preparation_data(self):
        data = []
        expected_keys = [
            "specify_standard_dispersion_protocol_used",
            "or_otherwise_specify_dispersion_technique_used",
            "dispersion_dilution_medium",
            "sonicator_type",
            "power_w",
            "sonication_time_secs",
            "tip_thickness_mm",
            "tip_composition",
            "size_of_ultrasonic_bath_water_volume_dm3",
            "sample_volume",
            "final_sample_concentration_mg_l_or_ppm",
            "additional_information"
        ]
        unmatched_keys = set(expected_keys)
        potential_matches = []

        logger.debug("Test Information sheet (columns A:E, rows 1:50) for Sample Preparation Data:")
        for row_idx in range(1, min(self.ws.max_row + 1, 51)):
            row = self.ws[row_idx]
            values = [str(cell.value)[:30] if cell.value else "None" for cell in row[:5]]
            has_comment = "Y" if row[0].comment else "N"
            logger.debug(f"Row {row_idx} [Comment: {has_comment}]: {values}")

        for row_idx, row in enumerate(self.ws.iter_rows(min_row=1, max_col=5), start=1):
            key_cell = row[0].value
            if not key_cell:
                continue
            raw_key = str(key_cell)
            key = self.normalize_key(raw_key)
            if not key:
                continue

            value = None
            value_col = None
            for col_idx in range(1, 5):
                if col_idx < len(row) and row[col_idx].value is not None:
                    value = row[col_idx].value
                    value_col = col_idx + 1
                    break

            logger.debug(f"Row {row_idx}: Raw Key='{raw_key}', Normalized Key='{key}', Value='{value}', Value Col={value_col}")

            if key in expected_keys:
                unmatched_keys.discard(key)
                data.append({"Key": key, "Value": value})
                logger.debug(f"Exact match: Key='{key}', Value='{value}' at row {row_idx}, col {value_col}")
            else:
                for expected_key in expected_keys:
                    similarity = SequenceMatcher(None, key, expected_key).ratio()
                    if similarity > 0.85:
                        unmatched_keys.discard(expected_key)
                        data.append({"Key": expected_key, "Value": value})
                        logger.debug(f"Fuzzy match: Key='{key}' matched '{expected_key}' (similarity={similarity:.2f}), Value='{value}' at row {row_idx}, col {value_col}")
                    elif similarity > 0.5:
                        potential_matches.append(f"Potential match for '{expected_key}': '{raw_key}' at row {row_idx}, col 1, similarity={similarity:.2f}")

        if potential_matches:
            logger.debug("Potential key matches for Sample Preparation Data:")
            for match in potential_matches:
                logger.debug(match)
        if unmatched_keys:
            logger.warning(f"Unmatched sample preparation data keys: {unmatched_keys}")

        sample_preparation_data = asdict(SamplePreparationData(
            dispersion_protocol=next((d["Value"] for d in data if d["Key"] == "specify_standard_dispersion_protocol_used"), None),
            dispersion_technique=next((d["Value"] for d in data if d["Key"] == "or_otherwise_specify_dispersion_technique_used"), None),
            dispersion_medium=next((d["Value"] for d in data if d["Key"] == "dispersion_dilution_medium"), None),
            sonicator_type=next((d["Value"] for d in data if d["Key"] == "sonicator_type"), None),
            power=next((d["Value"] for d in data if d["Key"] == "power_w"), None),
            sonication_time=next((d["Value"] for d in data if d["Key"] == "sonication_time_secs"), None),
            tip_thickness=next((d["Value"] for d in data if d["Key"] == "tip_thickness_mm"), None),
            tip_composition=next((d["Value"] for d in data if d["Key"] == "tip_composition"), None),
            ultrasonic_bath_size=next((d["Value"] for d in data if d["Key"] == "size_of_ultrasonic_bath_water_volume_dm3"), None),
            sample_volume=next((d["Value"] for d in data if d["Key"] == "sample_volume"), None),
            final_concentration=next((d["Value"] for d in data if d["Key"] == "final_sample_concentration_mg_l_or_ppm"), None),
            additional_info=next((d["Value"] for d in data if d["Key"] == "additional_information"), None)
        ))
        logger.debug(f"Sample preparation data extracted: {sample_preparation_data}")
        return sample_preparation_data

    def extract_instrumentation_data(self):
        data = []
        expected_keys = [
            "uv_vis_instrument_specifications",
            "software",
            "display_model",
            "cell_model",
            "optical_path_length",
            "start_wavelength",
            "end_wavelength",
            "wavelength_interval_nm",
            "background"
        ]
        unmatched_keys = set(expected_keys)
        potential_matches = []

        logger.debug("Test Information sheet (columns A:E, rows 1:70) for Instrumentation Data:")
        for row_idx in range(1, min(self.ws.max_row + 1, 71)):
            row = self.ws[row_idx]
            values = [str(cell.value)[:30] if cell.value else "None" for cell in row[:5]]
            has_comment = "Y" if row[0].comment else "N"
            logger.debug(f"Row {row_idx} [Comment: {has_comment}]: {values}")

        for row_idx, row in enumerate(self.ws.iter_rows(min_row=1, max_col=5), start=1):
            key_cell = row[0].value
            if not key_cell:
                continue
            raw_key = str(key_cell)
            key = self.normalize_key(raw_key)
            if not key:
                continue

            value = None
            value_col = None
            for col_idx in range(1, 5):
                if col_idx < len(row) and row[col_idx].value is not None:
                    value = row[col_idx].value
                    value_col = col_idx + 1
                    break

            logger.debug(f"Row {row_idx}: Raw Key='{raw_key}', Normalized Key='{key}', Value='{value}', Value Col={value_col}")

            if key in expected_keys:
                unmatched_keys.discard(key)
                data.append({"Key": key, "Value": value})
                logger.debug(f"Exact match: Key='{key}', Value='{value}' at row {row_idx}, col {value_col}")
            else:
                for expected_key in expected_keys:
                    similarity = SequenceMatcher(None, key, expected_key).ratio()
                    if similarity > 0.85:
                        unmatched_keys.discard(expected_key)
                        data.append({"Key": expected_key, "Value": value})
                        logger.debug(f"Fuzzy match: Key='{key}' matched '{expected_key}' (similarity={similarity:.2f}), Value='{value}' at row {row_idx}, col {value_col}")
                    elif similarity > 0.5:
                        potential_matches.append(f"Potential match for '{expected_key}': '{raw_key}' at row {row_idx}, col 1, similarity={similarity:.2f}")

        if potential_matches:
            logger.debug("Potential key matches for Instrumentation Data:")
            for match in potential_matches:
                logger.debug(match)
        if unmatched_keys:
            logger.warning(f"Unmatched instrumentation data keys: {unmatched_keys}")

        instrumentation_data = asdict(UVVisInstrumentationData(
            instrument_specs=next((d["Value"] for d in data if d["Key"] == "uv_vis_instrument_specifications"), None),
            software=next((d["Value"] for d in data if d["Key"] == "software"), None),
            display_model=next((d["Value"] for d in data if d["Key"] == "display_model"), None),
            cell_model=next((d["Value"] for d in data if d["Key"] == "cell_model"), None),
            optical_path_length=next((d["Value"] for d in data if d["Key"] == "optical_path_length"), None),
            start_wavelength=next((d["Value"] for d in data if d["Key"] == "start_wavelength"), None),
            end_wavelength=next((d["Value"] for d in data if d["Key"] == "end_wavelength"), None),
            wavelength_interval=next((d["Value"] for d in data if d["Key"] == "wavelength_interval_nm"), None),
            background=next((d["Value"] for d in data if d["Key"] == "background"), None)
        ))
        logger.debug(f"Instrumentation data extracted: {instrumentation_data}")
        return instrumentation_data

    def extract_replication(self):
        start_date = None
        end_date = None
        test_identifier = None
        for row in self.ws.iter_rows():
            key_cell = row[0].value
            if key_cell and 'test identifier number' in key_cell.lower():
                test_identifier = row[1].value
                start_date = self.excel_date_to_string(row[2].value)
                end_date = self.excel_date_to_string(row[3].value)
                break
        if test_identifier is None:
            logger.warning("Could not find test identifier number in Test Information sheet.")
        return asdict(ReplicationData(
            test_identifier_number=test_identifier,
            test_start_date=start_date,
            test_end_date=end_date
        ))

    def extract_raw_data(self, raw_sheet_name: str):
        logger.debug(f"Extracting raw data for sheet: {raw_sheet_name}")
        
        measurements = []

        if raw_sheet_name not in self.wb.sheetnames:
            logger.error(f"Raw data sheet {raw_sheet_name} not found")
            return asdict(UVVisRawData(measurements=measurements))

        raw_ws = self.wb[raw_sheet_name]
        logger.debug(f"Processing raw data sheet: {raw_sheet_name}")

        # Log first 5 rows for debugging
        logger.debug(f"Raw sheet {raw_sheet_name} first 5 rows:")
        for row_idx, row in enumerate(raw_ws.iter_rows(min_row=1, max_row=5, max_col=10), start=1):
            values = [str(cell.value)[:30] if cell.value else "None" for cell in row]
            logger.debug(f"Row {row_idx}: {values}")

        # Find value columns based on headers
        no_col_idx = None
        wavelength_col_idx = None
        absorbance_col_idx = None
        header_row = 1
        for col in range(1, raw_ws.max_column + 1):
            cell_value = raw_ws.cell(row=header_row, column=col).value
            if cell_value:
                h = self.normalize_key(cell_value)
                if 'no' in h:
                    no_col_idx = col
                    logger.debug(f"Found no at col {col}: {cell_value}")
                if 'wavelength' in h:
                    wavelength_col_idx = col
                    logger.debug(f"Found wavelength at col {col}: {cell_value}")
                if 'absorbance' in h:
                    absorbance_col_idx = col
                    logger.debug(f"Found absorbance at col {col}: {cell_value}")

        if no_col_idx is None or wavelength_col_idx is None or absorbance_col_idx is None:
            logger.warning(f"Could not find all required columns in {raw_sheet_name}")
            return asdict(UVVisRawData(measurements=measurements))

        # Extract measurement data
        max_rows = min(raw_ws.max_row, 1000)
        empty_row_count = 0
        for row_idx in range(header_row + 1, max_rows + 1):
            no_val = raw_ws.cell(row=row_idx, column=no_col_idx).value
            wavelength = raw_ws.cell(row=row_idx, column=wavelength_col_idx).value
            absorbance = raw_ws.cell(row=row_idx, column=absorbance_col_idx).value
            logger.debug(f"Measurement row {row_idx}: no={no_val}, wavelength={wavelength}, absorbance={absorbance}")
            
            if no_val is None and wavelength is None and absorbance is None:
                empty_row_count += 1
                if empty_row_count >= 5:
                    logger.debug(f"Reached end of measurement data at row {row_idx}")
                    break
                continue
            
            empty_row_count = 0
            try:
                no_int = int(no_val) if no_val is not None else None
                wavelength_float = float(wavelength) if wavelength is not None else None
                absorbance_float = float(absorbance) if absorbance is not None else None
                
                if wavelength_float is not None or absorbance_float is not None:
                    measurements.append(UVVisRawMeasurement(
                        no=no_int,
                        wavelength=wavelength_float,
                        absorbance=absorbance_float
                    ))
            except (ValueError, TypeError) as e:
                logger.warning(f"Invalid measurement data at row {row_idx}: {e}")
                continue
        
        logger.debug(f"Extracted {len(measurements)} measurements")
        return asdict(UVVisRawData(measurements=measurements))

    def extract_final_results(self):
        logger.debug("Extracting final results")
        final_sheet_name = [name for name in self.wb.sheetnames if "Final results" in name][0] if any("Final results" in name for name in self.wb.sheetnames) else None
        if not final_sheet_name:
            logger.error("Final Results sheet not found")
            return asdict(UVVisResultsData())

        ws = self.wb[final_sheet_name]
        logger.debug(f"Processing final results sheet: {final_sheet_name}")

        results_data = UVVisResultsData()

        # Extract peaks from row 4, columns A-L grouped by 3
        data_row = 4
        peak_cols = [
            (1, 2, 3),  # A:B:C for peak1
            (4, 5, 6),  # D:E:F for peak2
            (7, 8, 9),  # G:H:I for peak3
            (10, 11, 12)  # J:K:L for peak4
        ]
        for max_abs_col, wl_col, compound_col in peak_cols:
            max_abs = ws.cell(row=data_row, column=max_abs_col).value
            wavelength = ws.cell(row=data_row, column=wl_col).value
            compound = ws.cell(row=data_row, column=compound_col).value

            if max_abs is not None or wavelength is not None or compound is not None:
                try:
                    max_abs_float = float(max_abs) if max_abs else None
                    wl_int = int(wavelength) if wavelength else None
                    compound_str = str(compound) if compound else None
                    results_data.peaks.append(UVVisPeak(
                        max_absorbance=max_abs_float,
                        wavelength=wl_int,
                        identified_compound=compound_str
                    ))
                except (ValueError, TypeError) as e:
                    logger.warning(f"Invalid peak data at row {data_row}, cols {max_abs_col}-{compound_col}: {e}")

        logger.debug(f"Extracted {len(results_data.peaks)} peaks")
        return asdict(results_data)

    def parse_all_data(self) -> Dict[str, Union[Dict, List]]:
        try:
            work_package_data = self.extract_work_package_data()
            material_data = self.extract_material_data()
            sample_preparation_data = self.extract_sample_preparation_data()
            instrumentation_data = self.extract_instrumentation_data()
            replication = self.extract_replication()

            parsed_data = {
                'test_details': {
                    'work_package': work_package_data,
                    'material': material_data,
                    'sample_preparation': sample_preparation_data,
                    'instrumentation': instrumentation_data
                },
                'replication': replication,
                'replications': {},
                'final_results': {}
            }

            test_identifier = replication['test_identifier_number']
            if test_identifier:
                raw_sheet = f"Raw data_{test_identifier}"
                final_sheet = f"Final results_{test_identifier}"

                if raw_sheet in self.wb.sheetnames:
                    parsed_data['replications'] = self.extract_raw_data(raw_sheet)
                else:
                    logger.warning(f"Raw data sheet {raw_sheet} not found")

                parsed_data['final_results'] = self.extract_final_results()

            return parsed_data
        except Exception as e:
            logger.error(f"Error parsing Excel file: {e}\n{traceback.format_exc()}")
            raise

def parse_excel_uv_vis(file_path: str, sheet_name: str = "Test Information") -> Dict[str, Union[Dict, List]]:
    try:
        parser = UVVisParser(file_path, sheet_name)
        return parser.parse_all_data()
    except Exception as e:
        logger.error(f"Error in parse_excel_uv_vis: {e}\n{traceback.format_exc()}")
        raise

if __name__ == "__main__":
    file_path = "backend/data/1760762791_WP2_UV-Vis_1aR1.xlsx"
    try:
        parsed_data = parse_excel_uv_vis(file_path)
        print("Parsed Data:")
        #print("Test Details:", parsed_data['test_details'])
        #print("Replication:", parsed_data['replication'])
        print("Raw Data:", parsed_data['raw_data'])
        #print("Final Results:", parsed_data['final_results'])
    except Exception as e:
        print(f"Error: {e}")