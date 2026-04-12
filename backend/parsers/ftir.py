import openpyxl
import re
from dataclasses import dataclass, asdict, field
from typing import List, Dict, Optional, Union
import logging
import traceback
from datetime import datetime, timedelta
from difflib import SequenceMatcher

# Set up logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
logging.getLogger("multipart.multipart").setLevel(logging.WARNING)
logger = logging.getLogger(__name__)

# =========================================================
# Universal FTIR identifier pattern — accepts any letter
# (a, b, c, …) between the number and R.
#
#   Group 1 = WP number        (e.g. "2")
#   Group 2 = FTIR number      (e.g. "16")
#   Group 3 = variant letter   (e.g. "a" or "b")
#   Group 4 = run number       (e.g. "1")
# =========================================================
_FTIR_ID = r'WP(\d+)_FTIR_(\d+)([a-zA-Z])R(\d+)'

# Dataclasses (adapted for FTIR)
@dataclass
class Scientist:
    name: Optional[str] = None
    email: Optional[str] = None

@dataclass
class WorkPackageData:
    wp_name: Optional[str] = None
    partner: Optional[str] = None
    full_test_name: Optional[str] = None
    test_acronym: Optional[str] = None
    test_type: Optional[str] = None
    endpoint: Optional[str] = None
    endpoint_outcome: Optional[str] = None
    sop: Optional[str] = None
    path: Optional[str] = None
    lead_scientists: List[Scientist] = field(default_factory=list)
    assay_scientists: List[Scientist] = field(default_factory=list)

@dataclass
class MaterialData:
    material_identifier: Optional[str] = None
    erm_id: Optional[str] = None
    core_chemistry: Optional[str] = None
    material_name: Optional[str] = None
    material_state: Optional[str] = None
    cas: Optional[str] = None
    casforcore: Optional[str] = None
    batch: Optional[str] = None
    preparation_date: Optional[str] = None
    particles_stock: Optional[str] = None
    molar_concentration: Optional[str] = None

@dataclass
class ReplicationData:
    test_identifier_number: Optional[str] = None
    test_start_date: Optional[str] = None
    test_end_date: Optional[str] = None

@dataclass
class SamplePreparationData:
    dispersion_protocol: Optional[str] = None
    dispersion_technique: Optional[str] = None
    dispersion_medium: Optional[str] = None
    sonicator_type: Optional[str] = None
    power: Optional[str] = None
    sonication_time: Optional[str] = None
    tip_thickness: Optional[str] = None
    tip_composition: Optional[str] = None
    ultrasonic_bath_size: Optional[str] = None
    sample_volume: Optional[str] = None
    final_concentration: Optional[str] = None
    additional_info: Optional[str] = None

@dataclass
class InstrumentationData:
    ftir_model: Optional[str] = None
    beamsplitter: Optional[str] = None
    detector: Optional[str] = None
    measurement_technique: Optional[str] = None
    accessory_model: Optional[str] = None
    atr_crystal_material: Optional[str] = None
    resolution: Optional[str] = None
    number_of_scans: Optional[int] = None
    replication: Optional[int] = None
    spectral_range: Optional[str] = None

@dataclass
class WavelengthTransmittanceData:
    wavelengths: List[float] = field(default_factory=list)
    transmittances: List[float] = field(default_factory=list)

@dataclass
class PeakTransmittanceData:
    peaks: List[float] = field(default_factory=list)
    transmittances: List[float] = field(default_factory=list)

@dataclass
class RawData:
    run_number: int
    wavelength_transmittance: WavelengthTransmittanceData

@dataclass
class ProcessedData:
    run_number: int
    peak_transmittance: PeakTransmittanceData

@dataclass
class FunctionalGroup:
    group_name: str  # Maximum length: 100 characters
    peaks: List[float]

@dataclass
class ResultsData:
    functional_groups: List[FunctionalGroup] = field(default_factory=list)

class FTIRParser:
    def __init__(self, file_path: str, sheet_name: str = "Test Information"):
        self.file_path = file_path
        self.sheet_name = sheet_name
        try:
            self.wb = openpyxl.load_workbook(file_path, data_only=True)
            self.ws = self.wb[sheet_name]
            logger.info(f"Successfully loaded Excel file: {file_path}")
            logger.debug(f"Available sheet names: {self.wb.sheetnames}")
        except Exception as e:
            logger.error(f"Failed to load workbook or sheet {sheet_name}: {e}")
            raise
        self.email_regex = r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$'

    def normalize_key(self, key: Optional[str]) -> Optional[str]:
        if key:
            normalized = key.strip().lower()
            normalized = re.sub(r'[^a-zA-Z0-9]', '_', normalized)
            normalized = re.sub(r'_+', '_', normalized)
            return normalized
        return None

    def split_value_unit(self, value: Optional[str]) -> tuple[Optional[Union[str, float]], Optional[str]]:
        if isinstance(value, str) and " " in value:
            parts = value.split(" ", 1)
            try:
                numeric = float(parts[0]) if '.' in parts[0] else int(parts[0])
                return numeric, parts[1]
            except (ValueError, IndexError):
                return parts[0], parts[1] if len(parts) > 1 else None
        return value, None

    def excel_date_to_string(self, value: Optional[Union[float, str]]) -> Optional[str]:
        try:
            if isinstance(value, (int, float)):
                base_date = datetime(1899, 12, 30)
                delta = timedelta(days=float(value))
                return (base_date + delta).strftime("%Y-%m-%d")
            elif isinstance(value, str):
                for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"]:
                    try:
                        return datetime.strptime(value, fmt).strftime("%Y-%m-%d")
                    except ValueError:
                        continue
            return str(value) if value else None
        except Exception as e:
            logger.warning(f"Failed to convert date {value}: {e}")
            return None

    def extract_work_package_data(self):
        data = []
        lead_scientists = []
        assay_scientists = []

        for row in self.ws.iter_rows():
            key_cell = row[0].value
            value_cell = row[1].value if len(row) > 1 else None
            email_cell = row[3].value if len(row) > 3 else None

            key = self.normalize_key(key_cell)
            if not key:
                continue
            entry = {"Key": key, "Value": value_cell}

            if email_cell and re.match(self.email_regex, str(email_cell)):
                entry["Email"] = email_cell

            if key and "lead_scientist" in key:
                lead_scientists.append(asdict(Scientist(name=value_cell, email=email_cell)))
            if key and "assay_test_work_conducted_by" in key:
                assay_scientists.append(asdict(Scientist(name=value_cell, email=email_cell)))

            data.append(entry)

        wp_data = asdict(WorkPackageData(
            wp_name=next((d["Value"] for d in data if d["Key"] == "project_work_package_"), None),
            partner=next((d["Value"] for d in data if d["Key"] == "partner_conducting_test_assay_"), None),
            full_test_name=next((d["Value"] for d in data if d["Key"] == "full_name_of_test_assay_add_oecd_test_ref_id_if_app_"), None),
            test_acronym=next((d["Value"] for d in data if d["Key"] == "short_name_or_acronym_for_test_assay_"), None),
            test_type=next((d["Value"] for d in data if d["Key"] == "type_or_class_of_experimental_test_as_used_here_"), None),
            endpoint=next((d["Value"] for d in data if d["Key"] == "end_point_being_investigated_assessed_by_the_test_"), None),
            endpoint_outcome=next((d["Value"] for d in data if d["Key"] == "metric_s_used_to_assess_end_point_outcome_response_"), None),
            sop=next((d["Value"] for d in data if d["Key"] == "sop_s_for_test_ref_project_or_other_doc_title_id_"), None),
            path=next((d["Value"] for d in data if d["Key"] == "path_link_to_sop_protocol_on_proj_server_web_where_applic_"), None),
            lead_scientists=lead_scientists,
            assay_scientists=assay_scientists
        ))
        logger.debug(f"Extracted work package data: {wp_data}")
        return wp_data

    def extract_material_data(self):
        data = []
        expected_keys = [
            "sample_cms_internal_identifier",
            "erm_identifier_number_",
            "core_chemistry_",
            "material_name",
            "material_state_",
            "cas_no_",
            "cas_for_core_",
            "batch",
            "date_of_sample_preparation_for_tests",
            "number_of_particles_in_stock_",
            "molar_concentration"
        ]
        unmatched_keys = set(expected_keys)

        for row_idx, row in enumerate(self.ws.iter_rows(min_row=1, max_col=5), start=1):
            key_cell = row[0].value
            if not key_cell:
                continue
            key = self.normalize_key(str(key_cell))
            value = row[1].value if len(row) > 1 else None

            if key in expected_keys:
                unmatched_keys.discard(key)
                data.append({"Key": key, "Value": value})
            else:
                for expected_key in list(unmatched_keys):
                    similarity = SequenceMatcher(None, key, expected_key).ratio()
                    if similarity > 0.8:
                        unmatched_keys.discard(expected_key)
                        data.append({"Key": expected_key, "Value": value})

        if unmatched_keys:
            logger.warning(f"Unmatched material data keys: {unmatched_keys}")

        material_data = asdict(MaterialData(
            material_identifier=next((d["Value"] for d in data if d["Key"] == "sample_cms_internal_identifier"), None),
            erm_id=next((d["Value"] for d in data if d["Key"] == "erm_identifier_number_"), None),
            core_chemistry=next((d["Value"] for d in data if d["Key"] == "core_chemistry_"), None),
            material_name=next((d["Value"] for d in data if d["Key"] == "material_name"), None),
            material_state=next((d["Value"] for d in data if d["Key"] == "material_state_"), None),
            cas=next((d["Value"] for d in data if d["Key"] == "cas_no_"), None),
            casforcore=next((d["Value"] for d in data if d["Key"] == "cas_for_core_"), None),
            batch=next((d["Value"] for d in data if d["Key"] == "batch"), None),
            preparation_date=self.excel_date_to_string(next((d["Value"] for d in data if d["Key"] == "date_of_sample_preparation_for_tests"), None)),
            particles_stock=next((d["Value"] for d in data if d["Key"] == "number_of_particles_in_stock_"), None),
            molar_concentration=next((d["Value"] for d in data if d["Key"] == "molar_concentration"), None)
        ))
        logger.debug(f"Extracted material data: {material_data}")
        return material_data

    def extract_sample_preparation_data(self):
        data = []
        expected_keys = [
            "specify_standard_dispersion_protocol_used_",
            "or_otherwise_specify_dispersion_technique_used_",
            "dispersion_dilution_medium",
            "sonicator_type",
            "power_w_",
            "sonication_time_secs_",
            "tip_thickness_mm_",
            "tip_composition",
            "size_of_ultrasonic_bath_water_volume_dm3_",
            "sample_volume",
            "final_sample_concentration_mg_l_or_ppm_",
            "additional_information"
        ]
        unmatched_keys = set(expected_keys)

        for row_idx, row in enumerate(self.ws.iter_rows(min_row=1, max_col=5), start=1):
            key_cell = row[0].value
            if not key_cell:
                continue
            key = self.normalize_key(str(key_cell))
            value = row[1].value if len(row) > 1 else None

            if key in expected_keys:
                unmatched_keys.discard(key)
                data.append({"Key": key, "Value": value})
            else:
                for expected_key in list(unmatched_keys):
                    similarity = SequenceMatcher(None, key, expected_key).ratio()
                    if similarity > 0.8:
                        unmatched_keys.discard(expected_key)
                        data.append({"Key": expected_key, "Value": value})

        if unmatched_keys:
            logger.warning(f"Unmatched sample preparation keys: {unmatched_keys}")

        sample_preparation_data = asdict(SamplePreparationData(
            dispersion_protocol=next((d["Value"] for d in data if d["Key"] == "specify_standard_dispersion_protocol_used_"), None),
            dispersion_technique=next((d["Value"] for d in data if d["Key"] == "or_otherwise_specify_dispersion_technique_used_"), None),
            dispersion_medium=next((d["Value"] for d in data if d["Key"] == "dispersion_dilution_medium"), None),
            sonicator_type=next((d["Value"] for d in data if d["Key"] == "sonicator_type"), None),
            power=next((d["Value"] for d in data if d["Key"] == "power_w_"), None),
            sonication_time=next((d["Value"] for d in data if d["Key"] == "sonication_time_secs_"), None),
            tip_thickness=next((d["Value"] for d in data if d["Key"] == "tip_thickness_mm_"), None),
            tip_composition=next((d["Value"] for d in data if d["Key"] == "tip_composition"), None),
            ultrasonic_bath_size=next((d["Value"] for d in data if d["Key"] == "size_of_ultrasonic_bath_water_volume_dm3_"), None),
            sample_volume=next((d["Value"] for d in data if d["Key"] == "sample_volume"), None),
            final_concentration=next((d["Value"] for d in data if d["Key"] == "final_sample_concentration_mg_l_or_ppm_"), None),
            additional_info=next((d["Value"] for d in data if d["Key"] == "additional_information"), None)
        ))
        logger.debug(f"Extracted sample preparation data: {sample_preparation_data}")
        return sample_preparation_data

    def extract_instrumentation_data(self):
        data = []
        expected_keys = [
            "ftir_model",
            "beamsplitter",
            "detector",
            "measurment_technique",
            "accessory_model",
            "atr_crystal_material",
            "resolution",
            "number_of_scans",
            "replication",
            "spectrral_range"
        ]
        unmatched_keys = set(expected_keys)

        for row_idx, row in enumerate(self.ws.iter_rows(min_row=1, max_col=5), start=1):
            key_cell = row[0].value
            if not key_cell:
                continue
            key = self.normalize_key(str(key_cell))
            value = row[1].value if len(row) > 1 else None

            if key in expected_keys:
                unmatched_keys.discard(key)
                data.append({"Key": key, "Value": value})
            else:
                for expected_key in list(unmatched_keys):
                    similarity = SequenceMatcher(None, key, expected_key).ratio()
                    if similarity > 0.8:
                        unmatched_keys.discard(expected_key)
                        data.append({"Key": expected_key, "Value": value})

        if unmatched_keys:
            logger.warning(f"Unmatched instrumentation keys: {unmatched_keys}")

        try:
            number_of_scans = int(next((d["Value"] for d in data if d["Key"] == "number_of_scans"), None))
        except (ValueError, TypeError):
            number_of_scans = None

        try:
            replication = int(next((d["Value"] for d in data if d["Key"] == "replication"), None))
        except (ValueError, TypeError):
            replication = None

        instrumentation_data = asdict(InstrumentationData(
            ftir_model=next((d["Value"] for d in data if d["Key"] == "ftir_model"), None),
            beamsplitter=next((d["Value"] for d in data if d["Key"] == "beamsplitter"), None),
            detector=next((d["Value"] for d in data if d["Key"] == "detector"), None),
            measurement_technique=next((d["Value"] for d in data if d["Key"] == "measurment_technique"), None),
            accessory_model=next((d["Value"] for d in data if d["Key"] == "accessory_model"), None),
            atr_crystal_material=next((d["Value"] for d in data if d["Key"] == "atr_crystal_material"), None),
            resolution=next((d["Value"] for d in data if d["Key"] == "resolution"), None),
            number_of_scans=number_of_scans,
            replication=replication,
            spectral_range=next((d["Value"] for d in data if d["Key"] == "spectrral_range"), None)
        ))
        logger.debug(f"Extracted instrumentation data: {instrumentation_data}")
        return instrumentation_data

    def extract_replications(self):
        replications = []
        # FIX: accept any letter (a, b, c, …) before R
        raw_sheets = [name for name in self.wb.sheetnames if re.match(r'Raw data_' + _FTIR_ID, name, re.IGNORECASE)]
        logger.debug(f"Found raw data sheets: {raw_sheets}")
        
        for raw_sheet in raw_sheets:
            match = re.search(_FTIR_ID, raw_sheet, re.IGNORECASE)
            if not match:
                logger.warning(f"Raw sheet name {raw_sheet} does not match expected pattern")
                continue
            wp_number = match.group(1)
            ftir_number = match.group(2)
            letter = match.group(3)        # ← captured letter
            run_number = match.group(4)    # ← shifted from group(3)
            test_identifier = f"WP{wp_number}_FTIR_{ftir_number}{letter}R{run_number}"
            logger.debug(f"Processing replication for sheet {raw_sheet}: test_identifier={test_identifier}")
            
            start_date = None
            end_date = None
            for row in self.ws.iter_rows(min_row=1, max_col=5):
                key = self.normalize_key(row[0].value)
                if key and "test_start_date" in key:
                    start_date = self.excel_date_to_string(row[1].value if len(row) > 1 else None)
                if key and "test_end_date" in key:
                    end_date = self.excel_date_to_string(row[1].value if len(row) > 1 else None)

            replication = asdict(ReplicationData(
                test_identifier_number=test_identifier,
                test_start_date=start_date,
                test_end_date=end_date
            ))
            replications.append(replication)

        return replications

    def extract_raw_data(self, raw_sheet_name: str):
        run_number = 1  # Default
        match = re.search(_FTIR_ID, raw_sheet_name, re.IGNORECASE)
        if match:
            run_number = int(match.group(4))   # ← group(4) now
        logger.debug(f"Extracting raw data for sheet {raw_sheet_name}, run_number={run_number}")

        wavelength_transmittance = WavelengthTransmittanceData()

        if raw_sheet_name not in self.wb.sheetnames:
            logger.warning(f"Raw data sheet {raw_sheet_name} not found in workbook")
            return asdict(RawData(run_number=run_number, wavelength_transmittance=wavelength_transmittance))

        raw_ws = self.wb[raw_sheet_name]

        header_row = None
        for row_idx, row in enumerate(raw_ws.iter_rows(min_row=1, max_col=2), start=1):
            wav_header = str(row[0].value).lower() if row[0].value else ""
            trans_header = str(row[1].value).lower() if row[1].value else ""
            if "wavelength" in wav_header and "transmitance" in trans_header:
                header_row = row_idx
                logger.debug(f"Found raw data header at row {row_idx}: {wav_header}, {trans_header}")
                break

        if header_row:
            for row in raw_ws.iter_rows(min_row=header_row + 1, min_col=1, max_col=2):
                wavelength = row[0].value
                transmittance = row[1].value
                if wavelength is None or transmittance is None:
                    continue
                try:
                    wavelength_transmittance.wavelengths.append(float(wavelength))
                    wavelength_transmittance.transmittances.append(float(transmittance))
                except (ValueError, TypeError) as e:
                    logger.warning(f"Failed to parse raw data at row {row[0].row}: {e}")
                    continue
            logger.debug(f"Extracted {len(wavelength_transmittance.wavelengths)} raw data points for run {run_number}")
        else:
            logger.warning(f"No valid header found in raw data sheet {raw_sheet_name}")

        return asdict(RawData(run_number=run_number, wavelength_transmittance=wavelength_transmittance))

    def extract_processed_data(self, processed_sheet_name: str):
        run_number = 1  # Default
        match = re.search(_FTIR_ID, processed_sheet_name, re.IGNORECASE)
        if match:
            run_number = int(match.group(4))   # ← group(4) now
        logger.debug(f"Extracting processed data for sheet {processed_sheet_name}, run_number={run_number}")

        peak_transmittance = PeakTransmittanceData()

        if processed_sheet_name not in self.wb.sheetnames:
            logger.warning(f"Processed data sheet {processed_sheet_name} not found in workbook")
            return asdict(ProcessedData(run_number=run_number, peak_transmittance=peak_transmittance))

        processed_ws = self.wb[processed_sheet_name]
        logger.debug(f"Opened processed data sheet: {processed_sheet_name}")

        header_row = None
        for row_idx, row in enumerate(processed_ws.iter_rows(min_row=2, max_col=3), start=2):
            peak_header = str(row[1].value).lower() if len(row) > 1 and row[1].value else ""
            trans_header = str(row[2].value).lower() if len(row) > 2 and row[2].value else ""
            if "peak position [cm-1]" in peak_header and "transmitance [%]" in trans_header:
                header_row = row_idx
                logger.debug(f"Found processed data header at row {row_idx}")
                break

        if header_row:
            for row in processed_ws.iter_rows(min_row=header_row + 1, min_col=2, max_col=3):
                if len(row) < 2:
                    continue
                peak = row[0].value
                transmittance = row[1].value
                if peak is None or transmittance is None:
                    continue
                try:
                    peak_transmittance.peaks.append(float(peak))
                    peak_transmittance.transmittances.append(float(transmittance))
                except (ValueError, TypeError) as e:
                    logger.warning(f"Failed to parse processed data at row {row[0].row}: {e}")
                    continue
            logger.debug(f"Extracted {len(peak_transmittance.peaks)} processed data points for run {run_number}")
        else:
            logger.warning(f"No valid header found in processed data sheet {processed_sheet_name}")

        return asdict(ProcessedData(run_number=run_number, peak_transmittance=peak_transmittance))

    def extract_final_results(self):
        final_sheet_names = [name for name in self.wb.sheetnames if "Final results" in name]
        logger.debug(f"Found final results sheets: {final_sheet_names}")
        if not final_sheet_names:
            logger.warning("No final results sheet found")
            return asdict(ResultsData())

        ws = self.wb[final_sheet_names[0]]
        functional_groups = []

        header_row = 2
        first_cell = ws.cell(row=header_row, column=1).value
        if not first_cell or "functional group" not in str(first_cell).lower().strip():
            logger.warning(f"Header row {header_row} does not contain 'functional group': found '{first_cell}'")
            return asdict(ResultsData())

        data_row = 4
        if data_row > ws.max_row:
            logger.warning(f"Data row {data_row} exceeds sheet's maximum row {ws.max_row}")
            return asdict(ResultsData())

        max_col = 30
        row = ws[data_row]
        for i in range(0, min(len(row), max_col), 2):
            group_name = row[i].value
            peaks_str = row[i + 1].value if i + 1 < min(len(row), max_col) else None
            if group_name and peaks_str:
                peaks = []
                try:
                    for p in str(peaks_str).split(','):
                        p = p.strip()
                        if p:
                            peaks.append(float(p))
                    if len(group_name) > 100:
                        logger.warning(f"Group name '{group_name}' exceeds 100 characters, truncating")
                        group_name = group_name[:100]
                    functional_groups.append(FunctionalGroup(group_name=group_name, peaks=peaks))
                    logger.debug(f"Extracted functional group: {group_name}, peaks={peaks}")
                except (ValueError, TypeError) as e:
                    logger.warning(f"Error parsing peaks for group '{group_name}': {peaks_str}, error: {e}")
                    continue

        if not functional_groups:
            logger.warning("No valid functional groups extracted from data row")

        return asdict(ResultsData(functional_groups=functional_groups))

    def parse_all_data(self) -> Dict[str, Union[Dict, List]]:
        try:
            logger.debug("Starting parse_all_data")
            work_package_data = self.extract_work_package_data()
            material_data = self.extract_material_data()
            sample_preparation_data = self.extract_sample_preparation_data()
            instrumentation_data = self.extract_instrumentation_data()

            parsed_data = {
                'test_details': {
                    'work_package': work_package_data,
                    'material': material_data,
                    'sample_preparation': sample_preparation_data,
                    'instrumentation': instrumentation_data
                },
                'replication_metadata': self.extract_replications(),
                'replications': [],
                'processed_data': [],
                'final_results': {}
            }

            # FIX: accept any letter (a, b, c, …) before R
            raw_sheets = [name for name in self.wb.sheetnames if re.match(r'Raw data_' + _FTIR_ID, name, re.IGNORECASE)]
            processed_sheets = [name for name in self.wb.sheetnames if re.match(r'Processed data_' + _FTIR_ID, name, re.IGNORECASE)]
            logger.debug(f"Found raw data sheets: {raw_sheets}")
            logger.debug(f"Found processed data sheets: {processed_sheets}")

            for raw_sheet in raw_sheets:
                match = re.search(_FTIR_ID, raw_sheet, re.IGNORECASE)
                if not match:
                    logger.warning(f"Raw sheet name {raw_sheet} does not match expected pattern")
                    continue
                wp_number = match.group(1)
                ftir_number = match.group(2)
                letter = match.group(3)        # ← captured letter
                run_number = match.group(4)    # ← shifted from group(3)
                logger.debug(f"Processing raw sheet {raw_sheet}: wp={wp_number}, ftir={ftir_number}, letter={letter}, run={run_number}")

                # Match processed sheet using the SAME letter
                processed_sheet = next((name for name in processed_sheets 
                                    if re.search(rf'WP{wp_number}_FTIR_{ftir_number}{letter}R{run_number}', name, re.IGNORECASE)), None)
                logger.debug(f"Matched processed sheet for run {run_number}: {processed_sheet}")

                raw_data = self.extract_raw_data(raw_sheet)
                parsed_data['replications'].append(raw_data)

                if processed_sheet:
                    processed_data = self.extract_processed_data(processed_sheet)
                    parsed_data['processed_data'].append(processed_data)
                else:
                    logger.warning(f"No matching processed data sheet found for raw sheet {raw_sheet}")

            parsed_data['final_results'] = self.extract_final_results()
            logger.debug(f"Final parsed data: {parsed_data}")
            return parsed_data
        except Exception as e:
            logger.error(f"Error parsing Excel file: {e}\n{traceback.format_exc()}")
            raise

def _fix_degree_symbols(obj):
    """Recursively replace oC with °C in all string values."""
    if isinstance(obj, str):
        return obj.replace("oC", "°C")
    if isinstance(obj, dict):
        return {(k.replace("oC", "°C") if isinstance(k, str) else k): _fix_degree_symbols(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_fix_degree_symbols(v) for v in obj]
    return obj

def parse_excel_ftir(file_path: str) -> Dict[str, Union[Dict, List]]:
    """
    Parse FTIR Excel file and return data in a format compatible with TestCreate schema.
    """
    try:
        logger.debug(f"Starting parse_excel_ftir with file: {file_path}")
        parser = FTIRParser(file_path)
        parsed_data = parser.parse_all_data()
        logger.debug(f"Successfully parsed data: {parsed_data}")
        return _fix_degree_symbols(parsed_data)
    except Exception as e:
        logger.error(f"Error in parse_excel_ftir: {e}\n{traceback.format_exc()}")
        raise

if __name__ == "__main__":
    file_path = "backend/data/WP2/CMS_16a_TMR/FT-IR/WP2_FTIR_16aR1.xlsx"
    try:
        parsed_data = parse_excel_ftir(file_path)
        print("Final Results:", parsed_data['final_results'])
    except Exception as e:
        print(f"Error: {e}")