import openpyxl
import re
import json
from dataclasses import dataclass, asdict, field
from typing import List, Dict, Optional, Union, Any
import logging
import traceback
from datetime import datetime, timedelta

logging.basicConfig(level=logging.DEBUG, format="%(asctime)s - %(levelname)s - %(message)s")
logging.getLogger("multipart.multipart").setLevel(logging.WARNING)
logger = logging.getLogger(__name__)


# =========================================================
# Dataclasses
# =========================================================

@dataclass
class Scientist:
    name: Optional[str] = None
    email: Optional[str] = None


@dataclass
class WorkPackageData:
    wp_name: Optional[str] = None
    partner: Optional[str] = None
    laboratory_name: Optional[str] = None
    full_test_name: Optional[str] = None
    test_acronym: Optional[str] = None
    test_type: Optional[str] = None
    endpoint: Optional[str] = None
    endpoint_outcome: Optional[str] = None
    sop: Optional[str] = None
    path: Optional[str] = None
    lead_scientists: List[Scientist] = field(default_factory=list)
    assay_scientists: List[Scientist] = field(default_factory=list)


@dataclass
class MaterialData:
    material_identifier: Optional[str] = None
    erm_id: Optional[str] = None
    material_name: Optional[str] = None
    core_chemistry: Optional[str] = None
    cas_no: Optional[str] = None
    cas_for_core: Optional[str] = None
    material_supplier: Optional[str] = None
    material_state: Optional[str] = None
    batch: Optional[str] = None
    vial: Optional[str] = None
    preparation_date: Optional[str] = None
    stock_concentration: Optional[str] = None
    stock_concentration_secondary: Optional[str] = None
    molecular_weight: Optional[str] = None
    particles_stock: Optional[str] = None
    endotoxin_status: Optional[str] = None


@dataclass
class DispersionData:
    dispersion_protocol: Optional[str] = None
    dispersion_technique: Optional[str] = None
    dispersion_agent: Optional[str] = None
    agent_concentration: Optional[str] = None
    additives: Optional[str] = None
    dispersed_in_culture_medium: Optional[str] = None
    aids_used_to_disperse: Optional[str] = None
    sonication_bath: Optional[str] = None
    sonication_tip: Optional[str] = None
    time_duration: Optional[str] = None
    energy: Optional[str] = None


@dataclass
class CellLineData:
    cell_type: Optional[str] = None
    cell_line_short_name: Optional[str] = None
    supplier: Optional[str] = None
    passage_numbers: List[Optional[Union[int, str]]] = field(default_factory=list)
    plate_details: Optional[str] = None
    number_of_cells_per_chamber: Optional[Union[int, float, str]] = None
    total_volume_per_chamber: Optional[str] = None
    medium: Optional[str] = None
    serum: Optional[str] = None
    serum_concentration_culture_medium: Optional[Union[int, float, str]] = None
    serum_concentration_treatment_medium: Optional[Union[int, float, str]] = None
    serum_heat_inactivated: Optional[str] = None
    antibiotics: Optional[str] = None
    complete_growth_medium: Optional[str] = None
    cell_culture_conditions: Optional[str] = None
    trypan_blue_solution: Optional[str] = None
    incubation_time_with_tb: Optional[str] = None
    tb_volume: Optional[str] = None


@dataclass
class TreatmentData:
    time_point_unit: Optional[str] = None
    time_point_labels: List[str] = field(default_factory=list)
    time_points: List[Union[int, float, str]] = field(default_factory=list)
    concentration_unit: Optional[str] = None
    concentration_labels: List[str] = field(default_factory=list)
    concentrations_ug_ml: List[Union[int, float, str]] = field(default_factory=list)
    concentrations_particles: List[Union[int, float, str]] = field(default_factory=list)
    controls_abbreviations: List[str] = field(default_factory=list)
    controls_description: List[str] = field(default_factory=list)
    number_of_experiments: Optional[Union[int, str]] = None
    notes_a: Optional[str] = None
    notes_b: Optional[str] = None
    notes_c: Optional[str] = None


@dataclass
class ReplicationMetadata:
    test_identifier_number: Optional[str] = None
    test_start_date: Optional[str] = None
    test_end_date: Optional[str] = None
    replicate_label: Optional[str] = None
    raw_sheet_name: Optional[str] = None
    processed_sheet_name: Optional[str] = None


@dataclass
class TBMeasurement:
    chamber_label: Optional[str] = None
    percent_cell_death: Optional[float] = None


@dataclass
class TBRawDataBlock:
    run_label: Optional[str] = None
    plate_label: Optional[str] = None
    test_identifier: Optional[str] = None
    metric_name: Optional[str] = None
    raw_sheet_name: Optional[str] = None
    processed_sheet_name: Optional[str] = None
    protocol_name: Optional[str] = None
    protocol_number: Optional[str] = None
    plate_type: Optional[str] = None
    number_of_repeats: Optional[str] = None
    protocol_created_by: Optional[str] = None
    instrument_serial_number: Optional[str] = None
    measurements: List[TBMeasurement] = field(default_factory=list)
    assay_notes: List[str] = field(default_factory=list)


@dataclass
class TBProcessedDataBlock:
    run_label: Optional[str] = None
    processed_sheet_name: Optional[str] = None
    identifier: Optional[str] = None
    metric_name: Optional[str] = None
    condition_labels: List[str] = field(default_factory=list)
    values: List[Optional[float]] = field(default_factory=list)
    has_data: Optional[bool] = None


@dataclass
class TBFinalResults:
    sample_identifier: Optional[str] = None
    endpoint: Optional[str] = None
    condition_labels: List[str] = field(default_factory=list)
    replicate_rows: List[Dict[str, Any]] = field(default_factory=list)
    mean_row: Dict[str, Any] = field(default_factory=dict)
    sd_row: Dict[str, Any] = field(default_factory=dict)


# =========================================================
# Parser
# =========================================================

class TBMParser:
    def __init__(self, file_path: str, sheet_name: str = "Test_conditions"):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.parser_warnings: List[Dict[str, Any]] = []

        try:
            self.wb = openpyxl.load_workbook(file_path, data_only=True)
            self.ws = self.wb[sheet_name]
            logger.info("Successfully loaded TB workbook: %s", file_path)
        except Exception as e:
            logger.error("Failed to load workbook or sheet '%s': %s", sheet_name, e)
            raise

        self.email_regex = r"^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$"

    # -----------------------------------------------------
    # Helpers
    # -----------------------------------------------------
    def normalize_key(self, key: Optional[str]) -> Optional[str]:
        if key is None:
            return None
        normalized = str(key).strip().lower()
        normalized = re.sub(r"[^a-z0-9]+", "_", normalized)
        normalized = re.sub(r"_+", "_", normalized).strip("_")
        return normalized

    def excel_date_to_string(self, value) -> Optional[str]:
        try:
            if isinstance(value, datetime):
                return value.strftime("%Y-%m-%d")

            if isinstance(value, (int, float)):
                base_date = datetime(1899, 12, 30)
                return (base_date + timedelta(days=float(value))).strftime("%Y-%m-%d")

            if isinstance(value, str):
                for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d", "%d.%m.%Y"):
                    try:
                        return datetime.strptime(value.strip(), fmt).strftime("%Y-%m-%d")
                    except ValueError:
                        pass

            return str(value).strip() if value not in (None, "") else None
        except Exception as e:
            logger.warning("Failed to convert date '%s': %s", value, e)
            return str(value).strip() if value not in (None, "") else None

    def _safe_float(self, value) -> Optional[float]:
        if value in (None, ""):
            return None
        try:
            if isinstance(value, str):
                value = value.replace(",", ".").strip()
            return float(value)
        except (ValueError, TypeError):
            return None

    def _safe_int(self, value) -> Optional[int]:
        if value in (None, ""):
            return None
        try:
            if isinstance(value, str):
                value = value.replace(",", ".").strip()
            return int(float(value))
        except (ValueError, TypeError):
            return None

    def _safe_str(self, value) -> Optional[str]:
        return str(value).strip() if value not in (None, "") else None

    def _safe_excel_value(self, value):
        if value in (None, ""):
            return None
        if isinstance(value, str):
            return value.strip()
        return value

    def _sheet_key_values(self) -> List[Dict[str, Any]]:
        data = []
        for row_idx, row in enumerate(self.ws.iter_rows(min_row=1, max_col=6), start=1):
            key_cell = row[0].value
            if not key_cell:
                continue

            raw_key = str(key_cell).strip()
            key = self.normalize_key(raw_key)
            if not key:
                continue

            value = None
            for col_idx in range(2, 7):
                candidate = row[col_idx - 1].value
                if candidate is not None:
                    value = candidate
                    break

            email = row[3].value if len(row) > 3 else None

            data.append({
                "row": row_idx,
                "raw_key": raw_key,
                "key": key,
                "value": value,
                "email": email,
            })
        return data

    def _find_matching_processed_sheet(self, raw_sheet_name: str) -> Optional[str]:
        suffix = raw_sheet_name.replace("Raw_data_", "")
        candidate = f"Processed data_{suffix}"
        return candidate if candidate in self.wb.sheetnames else None

    def _extract_notes_after_marker(self, ws, marker: str = "Notes for the assay run") -> List[str]:
        notes = []
        marker_row = None

        for row_idx in range(1, ws.max_row + 1):
            for col_idx in range(1, min(ws.max_column, 8) + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if isinstance(val, str) and marker.lower() in val.lower():
                    marker_row = row_idx
                    break
            if marker_row:
                break

        if marker_row:
            for r in range(marker_row + 1, min(marker_row + 12, ws.max_row) + 1):
                row_values = [ws.cell(row=r, column=c).value for c in range(1, min(ws.max_column, 8) + 1)]
                parts = [str(v).strip() for v in row_values if v not in (None, "")]
                if parts:
                    notes.append(" | ".join(parts))

        return notes

    def _extract_protocol_value(self, ws, prefix: str) -> Optional[str]:
        for row_idx in range(1, min(ws.max_row, 60) + 1):
            for col_idx in range(1, min(ws.max_column, 6) + 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                if isinstance(value, str) and prefix.lower() in value.lower():
                    parts = re.split(r"\.{2,}", value)
                    if len(parts) > 1:
                        return parts[-1].strip() or None
                    cleaned = value.lower().replace(prefix.lower(), "").strip(" .:")
                    return cleaned or None
        return None

    def _extract_key_value_rows(self, ws, start_row: int, end_row: int) -> List[Dict[str, Any]]:
        rows: List[Dict[str, Any]] = []
        for row_idx in range(start_row, end_row + 1):
            label = self._safe_str(ws.cell(row_idx, 1).value)
            value = self._safe_excel_value(ws.cell(row_idx, 2).value)
            if label is None and value is None:
                continue
            rows.append({"label": label, "value": value})
        return rows

    def _extract_table(self, ws, header_row: int, data_start_row: int, data_end_row: int, start_col: int, end_col: int) -> Dict[str, Any]:
        headers = [self._safe_excel_value(ws.cell(header_row, col).value) for col in range(start_col, end_col + 1)]
        rows = []
        for row_idx in range(data_start_row, data_end_row + 1):
            values = [self._safe_excel_value(ws.cell(row_idx, col).value) for col in range(start_col, end_col + 1)]
            if any(v is not None for v in values):
                rows.append(values)
        return {"headers": headers, "rows": rows}

    # -----------------------------------------------------
    # Test details
    # -----------------------------------------------------
    def extract_work_package_data(self) -> Dict[str, Any]:
        rows = self._sheet_key_values()
        lead_scientists: List[Scientist] = []
        assay_scientists: List[Scientist] = []

        lead_name = self.ws.cell(row=15, column=2).value
        lead_email = self.ws.cell(row=15, column=4).value
        if lead_name:
            lead_scientists.append(
                Scientist(
                    name=lead_name,
                    email=lead_email if lead_email and re.match(self.email_regex, str(lead_email)) else None,
                )
            )

        for r in [16, 17]:
            assay_name = self.ws.cell(row=r, column=2).value
            assay_email = self.ws.cell(row=r, column=4).value
            if assay_name:
                assay_scientists.append(
                    Scientist(
                        name=assay_name,
                        email=assay_email if assay_email and re.match(self.email_regex, str(assay_email)) else None,
                    )
                )

        result = WorkPackageData(
            wp_name=next((r["value"] for r in rows if r["key"] == "project_work_package"), None),
            partner=next((r["value"] for r in rows if r["key"] == "partner_conducting_test_assay"), None),
            laboratory_name=next((r["value"] for r in rows if r["key"] == "test_facility_laboratory_name"), None),
            full_test_name=next((r["value"] for r in rows if "full_name_of_test_assay" in r["key"]), None),
            test_acronym=next((r["value"] for r in rows if "short_name_or_acronym_for_test_assay" in r["key"]), None),
            test_type=next((r["value"] for r in rows if "type_or_class_of_experimental_test_as_used_here" in r["key"]), None),
            endpoint=next((r["value"] for r in rows if "end_point_being_investigated" in r["key"]), None),
            endpoint_outcome=next((r["value"] for r in rows if "metric_s_used_to_assess" in r["key"]), None),
            sop=next((r["value"] for r in rows if "sop_s_for_test" in r["key"]), None),
            path=next((r["value"] for r in rows if "path_link_to_sop" in r["key"]), None),
            lead_scientists=lead_scientists,
            assay_scientists=assay_scientists,
        )
        return asdict(result)

    def extract_material_data(self) -> Dict[str, Any]:
        result = MaterialData(
            material_identifier=self.ws.cell(26, 2).value,
            erm_id=self.ws.cell(27, 2).value,
            material_name=self.ws.cell(28, 2).value,
            core_chemistry=self.ws.cell(29, 2).value,
            cas_no=self.ws.cell(30, 2).value,
            cas_for_core=self.ws.cell(31, 2).value,
            material_supplier=self.ws.cell(32, 2).value,
            material_state=self.ws.cell(33, 2).value,
            batch=self.ws.cell(34, 2).value,
            vial=self.ws.cell(35, 2).value,
            preparation_date=self.excel_date_to_string(self.ws.cell(36, 2).value),
            endotoxin_status=self.ws.cell(40, 2).value,
            stock_concentration=self.ws.cell(41, 2).value,
            stock_concentration_secondary=self.ws.cell(41, 3).value,
            molecular_weight=self.ws.cell(42, 2).value,
            particles_stock=self.ws.cell(43, 2).value,
        )
        return asdict(result)

    def extract_dispersion_data(self) -> Dict[str, Any]:
        result = DispersionData(
            dispersion_protocol=self.ws.cell(45, 2).value,
            dispersion_technique=self.ws.cell(46, 2).value,
            dispersion_agent=self.ws.cell(47, 2).value,
            agent_concentration=self.ws.cell(47, 4).value,
            additives=self.ws.cell(48, 2).value,
            dispersed_in_culture_medium=self.ws.cell(49, 2).value,
            aids_used_to_disperse=self.ws.cell(50, 2).value,
            sonication_bath=self.ws.cell(50, 4).value,
            sonication_tip=self.ws.cell(50, 6).value,
            time_duration=self.ws.cell(51, 2).value,
            energy=self.ws.cell(52, 2).value,
        )
        return asdict(result)

    def extract_cell_line_data(self) -> Dict[str, Any]:
        passage_numbers = [
            self.ws.cell(58, 2).value,
            self.ws.cell(58, 3).value,
            self.ws.cell(58, 4).value,
        ]

        result = CellLineData(
            cell_type=self.ws.cell(54, 2).value,
            cell_line_short_name=self.ws.cell(55, 2).value,
            supplier=self.ws.cell(56, 2).value,
            passage_numbers=passage_numbers,
            plate_details=self.ws.cell(59, 2).value,
            number_of_cells_per_chamber=self.ws.cell(60, 2).value,
            total_volume_per_chamber=self.ws.cell(61, 2).value,
            medium=self.ws.cell(62, 2).value,
            serum=self.ws.cell(63, 2).value,
            serum_concentration_culture_medium=self.ws.cell(64, 2).value,
            serum_concentration_treatment_medium=self.ws.cell(65, 2).value,
            serum_heat_inactivated=self.ws.cell(66, 2).value,
            antibiotics=self.ws.cell(67, 2).value,
            complete_growth_medium=self.ws.cell(68, 2).value,
            cell_culture_conditions=self.ws.cell(69, 2).value,
            trypan_blue_solution=self.ws.cell(70, 2).value,
            incubation_time_with_tb=self.ws.cell(71, 2).value,
            tb_volume=self.ws.cell(72, 2).value,
        )
        return asdict(result)

    def extract_treatment_data(self) -> Dict[str, Any]:
        concentration_labels = [
            str(v).strip()
            for v in [self.ws.cell(79, 2).value, self.ws.cell(79, 3).value, self.ws.cell(79, 4).value, self.ws.cell(79, 5).value]
            if v not in (None, "")
        ]

        concentrations_ug_ml = [self.ws.cell(80, c).value for c in range(2, 6)]
        concentrations_particles = [self.ws.cell(81, c).value for c in range(2, 6)]
        time_point_labels = [str(v).strip() for v in [self.ws.cell(75, 2).value] if v not in (None, "")]
        time_points = [self.ws.cell(76, 2).value]
        controls_abbreviations = [str(v).strip() for v in [self.ws.cell(83, 2).value] if v not in (None, "")]
        controls_description = [str(v).strip() for v in [self.ws.cell(84, 2).value] if v not in (None, "")]

        result = TreatmentData(
            time_point_unit=self.ws.cell(74, 2).value,
            time_point_labels=time_point_labels,
            time_points=time_points,
            concentration_unit=self.ws.cell(78, 2).value,
            concentration_labels=concentration_labels,
            concentrations_ug_ml=concentrations_ug_ml,
            concentrations_particles=concentrations_particles,
            controls_abbreviations=controls_abbreviations,
            controls_description=controls_description,
            number_of_experiments=self.ws.cell(85, 2).value,
            notes_a=self.ws.cell(88, 2).value,
            notes_b=self.ws.cell(89, 2).value,
            notes_c=self.ws.cell(90, 2).value,
        )
        return asdict(result)

    # -----------------------------------------------------
    # Replication metadata
    # -----------------------------------------------------
    def extract_replication_metadata(self) -> List[Dict[str, Any]]:
        start_date = self.excel_date_to_string(self.ws.cell(36, 2).value)
        metadata: List[Dict[str, Any]] = []

        for row_idx in [37, 38, 39]:
            test_identifier = self.ws.cell(row_idx, 2).value
            replicate_label = self.ws.cell(row_idx, 3).value

            if test_identifier:
                raw_sheet_name = next(
                    (s for s in self.wb.sheetnames if s.startswith("Raw_data_") and str(test_identifier) in s),
                    None
                )

                processed_sheet_name = None
                if raw_sheet_name:
                    processed_sheet_name = self._find_matching_processed_sheet(raw_sheet_name)

                metadata.append(asdict(ReplicationMetadata(
                    test_identifier_number=test_identifier,
                    test_start_date=start_date,
                    test_end_date=None,
                    replicate_label=replicate_label,
                    raw_sheet_name=raw_sheet_name,
                    processed_sheet_name=processed_sheet_name,
                )))

        return metadata

    # -----------------------------------------------------
    # Raw data
    # -----------------------------------------------------
    def _extract_raw_block(self, raw_sheet_name: str, processed_sheet_name: Optional[str]) -> Dict[str, Any]:
        ws = self.wb[raw_sheet_name]

        test_identifier = self._safe_str(ws.cell(1, 1).value)
        plate_label = self._safe_str(ws.cell(1, 3).value)
        metric_name = self._safe_str(ws.cell(3, 3).value)

        measurements = []
        for col in range(4, 8):
            chamber_label = self._safe_str(ws.cell(2, col).value)
            value = self._safe_float(ws.cell(3, col).value)
            if chamber_label is not None or value is not None:
                measurements.append(TBMeasurement(
                    chamber_label=chamber_label,
                    percent_cell_death=value,
                ))

        result = TBRawDataBlock(
            run_label=plate_label,
            plate_label=plate_label,
            test_identifier=test_identifier,
            metric_name=metric_name,
            raw_sheet_name=raw_sheet_name,
            processed_sheet_name=processed_sheet_name,
            protocol_name=self._extract_protocol_value(ws, "Protocol name"),
            protocol_number=self._extract_protocol_value(ws, "Protocol number"),
            plate_type=self._extract_protocol_value(ws, "Name of the plate type"),
            number_of_repeats=self._extract_protocol_value(ws, "Number of repeats"),
            protocol_created_by=self._extract_protocol_value(ws, "Protocol created by"),
            instrument_serial_number=self._extract_protocol_value(ws, "Instrument serial number"),
            measurements=measurements,
            assay_notes=self._extract_notes_after_marker(ws),
        )
        return asdict(result)

    def extract_raw_data(self) -> List[Dict[str, Any]]:
        raw_sheets = sorted([s for s in self.wb.sheetnames if s.lower().startswith("raw_data")])
        raw_blocks = []

        for raw_sheet in raw_sheets:
            processed_sheet = self._find_matching_processed_sheet(raw_sheet)
            raw_blocks.append(self._extract_raw_block(raw_sheet, processed_sheet))

        return raw_blocks

    # -----------------------------------------------------
    # Processed data
    # -----------------------------------------------------
    def extract_processed_data(self) -> List[Dict[str, Any]]:
        processed_sheets = sorted([s for s in self.wb.sheetnames if s.lower().startswith("processed data")])
        blocks = []

        for sheet_name in processed_sheets:
            ws = self.wb[sheet_name]
            identifier = self._safe_str(ws.cell(1, 1).value)
            run_label = self._safe_str(ws.cell(4, 1).value)
            condition_labels = [self._safe_str(ws.cell(4, c).value) for c in range(2, 6)]
            values = [self._safe_float(ws.cell(5, c).value) for c in range(2, 6)]
            has_data = any(v is not None for v in values)

            blocks.append(asdict(TBProcessedDataBlock(
                run_label=run_label,
                processed_sheet_name=sheet_name,
                identifier=identifier,
                metric_name=self._safe_str(ws.cell(5, 1).value),
                condition_labels=[label for label in condition_labels if label is not None],
                values=values,
                has_data=has_data,
            )))

        return blocks

    # -----------------------------------------------------
    # Final results
    # -----------------------------------------------------
    def extract_final_results(self) -> Dict[str, Any]:
        if "Final results" not in self.wb.sheetnames:
            self.parser_warnings.append({
                "type": "missing_sheet",
                "sheet": "Final results",
                "note": "Final results sheet not found.",
            })
            return asdict(TBFinalResults())

        ws = self.wb["Final results"]

        sample_identifier = self._safe_str(ws.cell(1, 1).value)
        endpoint = self._safe_str(ws.cell(1, 2).value)
        condition_labels = [self._safe_str(ws.cell(3, c).value) for c in range(3, 7)]
        condition_labels = [label for label in condition_labels if label is not None]

        replicate_rows = []
        for row_idx in [4, 5, 6]:
            replicate_name = self._safe_str(ws.cell(row_idx, 2).value)
            if replicate_name:
                replicate_rows.append({
                    "replicate": replicate_name,
                    "values": {
                        label: self._safe_float(ws.cell(row_idx, 3 + idx).value)
                        for idx, label in enumerate(condition_labels)
                    }
                })

        mean_row = {
            "label": self._safe_str(ws.cell(7, 2).value),
            "values": {
                label: self._safe_float(ws.cell(7, 3 + idx).value)
                for idx, label in enumerate(condition_labels)
            }
        }

        sd_row = {
            "label": self._safe_str(ws.cell(8, 2).value),
            "values": {
                label: self._safe_float(ws.cell(8, 3 + idx).value)
                for idx, label in enumerate(condition_labels)
            }
        }

        return asdict(TBFinalResults(
            sample_identifier=sample_identifier,
            endpoint=endpoint,
            condition_labels=condition_labels,
            replicate_rows=replicate_rows,
            mean_row=mean_row,
            sd_row=sd_row,
        ))

    # -----------------------------------------------------
    # Statistical analysis
    # -----------------------------------------------------
    def extract_statistical_analysis(self) -> Dict[str, Any]:
        if "Statistics" not in self.wb.sheetnames:
            return {
                "available": False,
                "notes": "No statistics sheet found.",
            }

        ws = self.wb["Statistics"]

        tukey_rows = []
        for row_idx in range(46, 52):
            comparison = self._safe_str(ws.cell(row_idx, 1).value)
            if not comparison:
                continue
            tukey_rows.append({
                "comparison": comparison,
                "mean_diff": self._safe_excel_value(ws.cell(row_idx, 2).value),
                "ci_95_of_diff": self._safe_excel_value(ws.cell(row_idx, 3).value),
                "significant": self._safe_excel_value(ws.cell(row_idx, 4).value),
                "summary": self._safe_excel_value(ws.cell(row_idx, 5).value),
                "adjusted_p_value": self._safe_excel_value(ws.cell(row_idx, 6).value),
                "comparison_code": self._safe_excel_value(ws.cell(row_idx, 7).value),
            })

        anova_summary_rows = self._extract_key_value_rows(ws, 15, 19)
        brown_forsythe_rows = self._extract_key_value_rows(ws, 22, 25)

        # Backward-compatible maps while also preserving exact Excel labels.
        anova_summary_map = {
            "F": self._safe_excel_value(ws.cell(15, 2).value),
            "P value": self._safe_excel_value(ws.cell(16, 2).value),
            "P value summary": self._safe_excel_value(ws.cell(17, 2).value),
            "Significant diff. among means (P < 0.05)?": self._safe_excel_value(ws.cell(18, 2).value),
            "R squared": self._safe_excel_value(ws.cell(19, 2).value),
        }

        brown_forsythe_map = {
            "F (DFn, DFd)": self._safe_excel_value(ws.cell(22, 2).value),
            "P value": self._safe_excel_value(ws.cell(23, 2).value),
            "P value summary": self._safe_excel_value(ws.cell(24, 2).value),
            "Are SDs significantly different (P < 0.05)?": self._safe_excel_value(ws.cell(25, 2).value),
        }

        return {
            "available": True,
            "software": self._safe_str(ws.cell(10, 1).value),
            "test": self._safe_str(ws.cell(12, 1).value),
            "anova_summary_title": self._safe_str(ws.cell(14, 1).value),
            "anova_summary": anova_summary_rows,
            "anova_summary_map": anova_summary_map,
            "brown_forsythe_title": self._safe_str(ws.cell(21, 1).value),
            "brown_forsythe_summary": brown_forsythe_rows,
            "brown_forsythe_summary_map": brown_forsythe_map,
            "anova_table": {
                "title": self._safe_str(ws.cell(33, 1).value),
                **self._extract_table(ws, header_row=33, data_start_row=34, data_end_row=36, start_col=1, end_col=6),
            },
            "data_summary": {
                "title": self._safe_str(ws.cell(38, 1).value),
                "rows": self._extract_key_value_rows(ws, 39, 40),
            },
            "post_hoc_title": self._safe_str(ws.cell(43, 1).value),
            "tukey_table": {
                "title": self._safe_str(ws.cell(45, 1).value),
                "headers": [
                    "Comparison",
                    self._safe_excel_value(ws.cell(45, 2).value),
                    self._safe_excel_value(ws.cell(45, 3).value),
                    self._safe_excel_value(ws.cell(45, 4).value),
                    self._safe_excel_value(ws.cell(45, 5).value),
                    self._safe_excel_value(ws.cell(45, 6).value),
                ],
                "rows": tukey_rows,
            },
        }

    # -----------------------------------------------------
    # Parse all
    # -----------------------------------------------------
    def parse_all_data(self) -> Dict[str, Union[Dict, List]]:
        try:
            raw_data = self.extract_raw_data()
            parsed_data = {
                "test_details": {
                    "work_package": self.extract_work_package_data(),
                    "material": self.extract_material_data(),
                    "dispersion": self.extract_dispersion_data(),
                    "cell_line": self.extract_cell_line_data(),
                    "treatment": self.extract_treatment_data(),
                },
                "replication_metadata": self.extract_replication_metadata(),
                "raw_data": raw_data,
                "replications": raw_data,
                "processed_data": self.extract_processed_data(),
                "final_results": self.extract_final_results(),
                "statistical_analysis": self.extract_statistical_analysis(),
            }

            if self.parser_warnings:
                parsed_data["parser_warnings"] = self.parser_warnings

            logger.info("FINAL TB JSON:\n%s", json.dumps(parsed_data, indent=2, default=str))
            return parsed_data

        except Exception as e:
            logger.error("Error parsing TB file: %s\n%s", e, traceback.format_exc())
            raise


def parse_excel_tbm(file_path: str, sheet_name: str = "Test_conditions") -> Dict[str, Union[Dict, List]]:
    try:
        parser = TBMParser(file_path, sheet_name)
        return parser.parse_all_data()
    except Exception as e:
        logger.error("Error in parse_excel_tb: %s\n%s", e, traceback.format_exc())
        raise


if __name__ == "__main__":
    file_path = "backend/data/CMS_WP3_TB_10b_AgNP.xlsx"
    parsed_data = parse_excel_tbm(file_path)
    print(json.dumps(parsed_data, indent=2, default=str)[:12000])