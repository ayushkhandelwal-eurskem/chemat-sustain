import openpyxl
import re
import json
from typing import List, Dict, Optional, Union, Any
from dataclasses import dataclass, field, asdict
import logging
import traceback
from datetime import datetime, timedelta, time as dt_time
from difflib import SequenceMatcher

logging.basicConfig(level=logging.DEBUG, format="%(asctime)s - %(levelname)s - %(message)s")
logging.getLogger("multipart.multipart").setLevel(logging.WARNING)
logger = logging.getLogger(__name__)


# ------------------------------------------------------------------ #
# Dataclasses                                                         #
# ------------------------------------------------------------------ #
@dataclass
class Scientist:
    name: Optional[str] = None
    email: Optional[str] = None


@dataclass
class WorkPackageData:
    wp_name: Optional[str] = None
    partner: Optional[str] = None
    laboratory_name: Optional[str] = None
    full_test_name: Optional[str] = None
    oecd_iso_ref: Optional[str] = None
    test_type: Optional[str] = None
    endpoint: Optional[str] = None
    endpoint_outcome: Optional[str] = None
    sop: Optional[str] = None
    test_start_date: Optional[str] = None
    test_end_date: Optional[str] = None
    lead_scientists: List[Dict] = field(default_factory=list)
    assay_scientists: List[Dict] = field(default_factory=list)


@dataclass
class MaterialData:
    material_label: Optional[str] = None
    material_identifier: Optional[str] = None
    erm_id: Optional[str] = None
    material_name: Optional[str] = None
    core_chemistry: Optional[str] = None
    cas_no: Optional[str] = None
    cas_for_core: Optional[str] = None
    material_supplier: Optional[str] = None
    material_state: Optional[str] = None
    batch: Optional[str] = None
    vial: Optional[str] = None
    preparation_date: Optional[str] = None
    stock_concentration: Optional[str] = None
    molar_concentration: Optional[str] = None
    particles_in_stock: Optional[str] = None


@dataclass
class ApplicationData:
    nominal_start_algae_numbers: Optional[float] = None
    total_volume_per_replicate: Optional[float] = None
    concentrations_tested: Optional[str] = None
    replicates_per_concentration: Optional[float] = None


@dataclass
class TestConditionsData:
    culture_medium: Optional[str] = None
    light_cycle: Optional[str] = None
    light_intensity: Optional[str] = None
    illumination_direction: Optional[str] = None
    temperature: Optional[str] = None
    aeration: Optional[str] = None
    salinity: Optional[str] = None
    total_incubation_time: Optional[str] = None


@dataclass
class AnalysisConditionsData:
    type_of_measurement: Optional[str] = None
    measurement_device: Optional[str] = None
    chlorophyll_extraction: Optional[str] = None
    chlorophyll_extraction_protocol: Optional[str] = None
    excitation_nm: Optional[float] = None
    emission_nm: Optional[float] = None
    bandwidth_nm: Optional[float] = None
    ht_voltage_v: Optional[float] = None
    measurement_replicates: Optional[str] = None
    timepoints: List[Any] = field(default_factory=list)


@dataclass
class CultureConditionsData:
    supplier: Optional[str] = None
    culture_medium: Optional[str] = None
    nutrients: Optional[str] = None
    nutrient_components: List[str] = field(default_factory=list)
    light_cycle: Optional[str] = None
    light_intensity: Optional[str] = None
    illumination_direction: Optional[str] = None
    temperature: Optional[str] = None
    aeration: Optional[str] = None
    salinity: Optional[str] = None


# ------------------------------------------------------------------ #
# Parser                                                              #
# ------------------------------------------------------------------ #
class AlgaeParser:
    def __init__(self, file_path, sheet_name="Test Details"):
        self.file_path = file_path
        self.spelling_issues: List[str] = []
        try:
            self.wb = openpyxl.load_workbook(file_path, data_only=True)
            if sheet_name in self.wb.sheetnames:
                self.ws = self.wb[sheet_name]
            elif "Test_conditions" in self.wb.sheetnames:
                self.ws = self.wb["Test_conditions"]
            elif "Test Information" in self.wb.sheetnames:
                self.ws = self.wb["Test Information"]
            else:
                self.ws = self.wb[self.wb.sheetnames[0]]
            logger.info(f"Loaded workbook {file_path}; using details sheet '{self.ws.title}'")
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            raise
        self.email_regex = r"^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$"

    # ---------- helpers ---------- #
    def normalize_key(self, key: Optional[str]) -> Optional[str]:
        if key:
            normalized = str(key).strip().lower()
            normalized = re.sub(r"[^a-z0-9]", "_", normalized)
            normalized = re.sub(r"_+", "_", normalized).strip("_")
            return normalized
        return None

    _nk = normalize_key

    def _sf(self, v) -> Optional[float]:
        if v in (None, "", " "):
            return None
        try:
            if isinstance(v, str):
                v = v.replace(",", ".").strip()
            return float(v)
        except (ValueError, TypeError):
            return None

    def excel_date_to_string(self, value) -> Optional[str]:
        try:
            if isinstance(value, datetime):
                return value.strftime("%Y-%m-%d")
            if isinstance(value, (int, float)):
                return (datetime(1899, 12, 30) + timedelta(days=float(value))).strftime("%Y-%m-%d")
            if isinstance(value, str):
                for f in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d.%m.%Y %H:%M", "%d.%m.%Y"):
                    try:
                        return datetime.strptime(value.strip(), f).strftime("%Y-%m-%d")
                    except ValueError:
                        continue
                self.spelling_issues.append(f"Unrecognized date format: {value!r}")
                return str(value).strip()
            return str(value).strip() if value else None
        except Exception:
            self.spelling_issues.append(f"Malformed date value: {value!r}")
            return str(value).strip() if value else None

    def _time_str(self, v) -> Optional[str]:
        if v is None:
            return None
        if isinstance(v, dt_time):
            return f"{v.hour:02d}:{v.minute:02d}"
        return str(v)

    def _kv(self, max_col=6) -> List[Dict]:
        """Key-value scan of the details sheet."""
        data = []
        for ri, row in enumerate(self.ws.iter_rows(min_row=1, max_col=max_col), start=1):
            k = row[0].value
            if not k:
                continue
            key = self.normalize_key(k)
            if not key:
                continue
            val = None
            for c in range(1, min(len(row), max_col)):
                if row[c].value is not None:
                    val = row[c].value
                    break
            email = row[3].value if len(row) > 3 else None
            data.append({"row": ri, "key": key, "value": val, "email": email})
        return data

    def _g(self, rows, *keys):
        return next((r["value"] for r in rows if any(k in r["key"] for k in keys) and r["value"] is not None), None)

    def _fuzzy_get(self, rows, *targets, threshold=0.8):
        """Fuzzy label matching for instrument-style header typos."""
        best = None
        best_ratio = threshold
        for r in rows:
            for t in targets:
                ratio = SequenceMatcher(None, r["key"], t).ratio()
                if ratio >= best_ratio and r["value"] is not None:
                    best_ratio = ratio
                    best = r["value"]
        return best

    def _parse_linear_equation(self, eq: Optional[str]):
        """Parse 'y= 10892x - 5967,5' -> (slope, intercept)."""
        if not eq or not isinstance(eq, str):
            return None, None
        s = eq.replace(",", ".").replace(" ", "")
        m = re.search(r"y=([+-]?[\d.]+)x([+-][\d.]+)?", s, re.IGNORECASE)
        if not m:
            return None, None
        slope = self._sf(m.group(1))
        intercept = self._sf(m.group(2)) if m.group(2) else None
        if intercept is not None:
            intercept = abs(intercept) if "-" in (m.group(2) or "") else intercept
            # keep sign as written
            intercept = self._sf(m.group(2))
        return slope, intercept

    def _find_sheets(self, prefix) -> List[str]:
        p = prefix.lower()
        result = []
        for s in self.wb.sheetnames:
            sl = s.lower()
            if sl.startswith(p) or sl.replace(" ", "").startswith(p.replace(" ", "")):
                result.append(s)
        return result

    # ---------- TEST DETAILS ---------- #
    def extract_work_package_data(self) -> Dict:
        rows = self._kv()
        lead, assay = [], []
        for r in rows:
            if "lead_scientist" in r["key"]:
                em = r["email"]
                lead.append(asdict(Scientist(name=r["value"], email=em if em and re.match(self.email_regex, str(em)) else None)))
            if "assay_test_work_conducted_by" in r["key"]:
                em = r["email"]
                assay.append(asdict(Scientist(name=r["value"], email=em if em and re.match(self.email_regex, str(em)) else None)))
        g = lambda *k: self._g(rows, *k)
        wp = WorkPackageData(
            wp_name=g("project_work_package"),
            partner=g("partner_conducting_test_assay"),
            laboratory_name=g("test_facility_laboratory_name"),
            full_test_name=g("full_name_of_test"),
            oecd_iso_ref=g("oecd_iso_test_ref_id"),
            test_type=g("type_or_class_of_experimental"),
            endpoint=g("end_point_being_investigated"),
            endpoint_outcome=g("metric_s_used_to_assess"),
            sop=g("sop_s_for_test"),
            test_start_date=self.excel_date_to_string(g("test_start_date")),
            test_end_date=self.excel_date_to_string(g("test_end_date")),
            lead_scientists=lead,
            assay_scientists=assay,
        )
        return asdict(wp)

    def extract_material_data(self) -> Dict:
        rows = self._kv()
        g = lambda *k: self._g(rows, *k)
        mat = MaterialData(
            material_label=g("test_material_details"),
            material_identifier=g("sample_cms_internal_identifier"),
            erm_id=g("erm_identifier"),
            material_name=g("material_name"),
            core_chemistry=g("core_chemistry"),
            cas_no=g("cas_no"),
            cas_for_core=g("cas_for_core"),
            material_supplier=g("material_supplier"),
            material_state=g("material_state"),
            batch=g("batch"),
            vial=g("vial"),
            preparation_date=self.excel_date_to_string(g("date_of_preparation")),
            stock_concentration=g("stock_concentration"),
            molar_concentration=g("molar_concentration"),
            particles_in_stock=g("no_of_particles_in_stock"),
        )
        return asdict(mat)

    def extract_application_data(self) -> Dict:
        rows = self._kv()
        g = lambda *k: self._g(rows, *k)
        app = ApplicationData(
            nominal_start_algae_numbers=self._sf(g("nominal_start_algae_numbers")),
            total_volume_per_replicate=self._sf(g("total_volume_per_replicate")),
            concentrations_tested=str(g("concentrations_tested")) if g("concentrations_tested") is not None else None,
            replicates_per_concentration=self._sf(g("replicates_per_concentration")),
        )
        return asdict(app)

    def extract_test_conditions_data(self) -> Dict:
        """The APPLICATION-level TEST CONDITIONS block (rows ~38-46)."""
        ws = self.ws
        # locate the first TEST CONDITIONS header
        start = None
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=1).value
            if v and "test conditions" in str(v).strip().lower():
                start = r
                break
        tc = TestConditionsData()
        if start is None:
            return asdict(tc)
        block = {}
        for r in range(start + 1, min(start + 12, ws.max_row + 1)):
            k = ws.cell(row=r, column=1).value
            v = ws.cell(row=r, column=2).value
            if k and "analysis" in str(k).strip().lower():
                break
            if k:
                block[self.normalize_key(k)] = v
        gg = lambda *keys: next((block[bk] for bk in block for k in keys if k in bk and block[bk] is not None), None)
        tc.culture_medium = gg("culture_medium")
        tc.light_cycle = str(gg("light_cycle")) if gg("light_cycle") is not None else None
        tc.light_intensity = str(gg("light_intensity")) if gg("light_intensity") is not None else None
        tc.illumination_direction = gg("illumination_direction")
        tc.temperature = str(gg("temperature")) if gg("temperature") is not None else None
        tc.aeration = str(gg("aeration")) if gg("aeration") is not None else None
        tc.salinity = str(gg("salinity")) if gg("salinity") is not None else None
        tc.total_incubation_time = str(gg("total_incubation_time")) if gg("total_incubation_time") is not None else None
        return asdict(tc)

    def extract_analysis_conditions_data(self) -> Dict:
        rows = self._kv()
        g = lambda *k: self._g(rows, *k)
        ws = self.ws
        # timepoints: label at ANALYSIS 'Timepoints (h)' then values run down column 2
        timepoints = []
        for r in range(1, ws.max_row + 1):
            k = ws.cell(row=r, column=1).value
            if k and "timepoints" in str(k).strip().lower():
                v0 = ws.cell(row=r, column=2).value
                if v0 is not None:
                    timepoints.append(v0)
                rr = r + 1
                while rr <= ws.max_row and ws.cell(row=rr, column=1).value is None and ws.cell(row=rr, column=2).value is not None:
                    timepoints.append(ws.cell(row=rr, column=2).value)
                    rr += 1
                break
        # chlorophyll extraction protocol may sit on the row after "Chlorophyll extraction"
        chl_protocol = None
        for r in range(1, ws.max_row + 1):
            k = ws.cell(row=r, column=1).value
            if k and "chlorophyll extraction" in str(k).strip().lower():
                nxt = ws.cell(row=r + 1, column=2).value
                if ws.cell(row=r + 1, column=1).value is None and nxt is not None:
                    chl_protocol = str(nxt)
                break
        ac = AnalysisConditionsData(
            type_of_measurement=g("type_of_measurement"),
            measurement_device=g("measurement_device"),
            chlorophyll_extraction=str(g("chlorophyll_extraction")) if g("chlorophyll_extraction") is not None else None,
            chlorophyll_extraction_protocol=chl_protocol,
            excitation_nm=self._sf(g("excitation_nm")),
            emission_nm=self._sf(g("emission_nm")),
            bandwidth_nm=self._sf(g("bandwidth_nm")),
            ht_voltage_v=self._sf(g("ht_voltage_v")),
            measurement_replicates=g("measurement_replicates"),
            timepoints=timepoints,
        )
        return asdict(ac)

    def extract_culture_conditions_data(self) -> Dict:
        ws = self.ws
        start = None
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=1).value
            if v and "culture conditions and test medium" in str(v).strip().lower():
                start = r
                break
        cc = CultureConditionsData()
        if start is None:
            return asdict(cc)
        nutrient_components = []
        block = {}
        for r in range(start + 1, ws.max_row + 1):
            k = ws.cell(row=r, column=1).value
            v = ws.cell(row=r, column=2).value
            if k:
                block[self.normalize_key(k)] = v
            elif v is not None:
                # continuation rows: nutrient / vitamin component names
                nutrient_components.append(str(v).strip())
        gg = lambda *keys: next((block[bk] for bk in block for k in keys if k in bk and block[bk] is not None), None)
        cc.supplier = gg("supplier")
        cc.culture_medium = gg("culture_medium")
        cc.nutrients = gg("nutrients")
        cc.nutrient_components = nutrient_components
        cc.light_cycle = str(gg("light_cycle")) if gg("light_cycle") is not None else None
        cc.light_intensity = str(gg("light_intensity")) if gg("light_intensity") is not None else None
        cc.illumination_direction = gg("illumination_direction")
        cc.temperature = str(gg("temperature")) if gg("temperature") is not None else None
        cc.aeration = str(gg("aeration")) if gg("aeration") is not None else None
        cc.salinity = str(gg("salinity")) if gg("salinity") is not None else None
        return asdict(cc)

    def extract_replications(self) -> List[Dict]:
        """Replication metadata: one entry per RawData (timepoint) sheet."""
        reps = []
        for sn in self._find_sheets("rawdata"):
            ws = self.wb[sn]
            m = re.search(r"\(([^)]+)\)", sn)
            timepoint = m.group(1) if m else sn
            reps.append({
                "raw_sheet_name": sn,
                "timepoint": timepoint,
                "file_name": ws.cell(row=2, column=2).value,
                "sample_name": ws.cell(row=4, column=2).value,
                "comment": str(ws.cell(row=5, column=2).value) if ws.cell(row=5, column=2).value else None,
                "username": ws.cell(row=7, column=2).value,
                "organization": ws.cell(row=9, column=2).value,
                "creation_date": self.excel_date_to_string(ws.cell(row=11, column=2).value),
                "last_update": self.excel_date_to_string(ws.cell(row=12, column=2).value),
                "model_name": ws.cell(row=14, column=2).value,
                "serial_no": ws.cell(row=15, column=2).value,
            })
        return reps

    # ---------- RAW DATA ---------- #
    def extract_raw_data(self) -> List[Dict]:
        blocks = []
        for sn in self._find_sheets("rawdata"):
            ws = self.wb[sn]
            m = re.search(r"\(([^)]+)\)", sn)
            timepoint = m.group(1) if m else sn
            # instrument metadata (label:value from column1:column2 up to the readings header)
            instrument = {}
            header_row = None
            for r in range(1, ws.max_row + 1):
                k = ws.cell(row=r, column=1).value
                if k and str(k).strip() == "No.":
                    header_row = r
                    break
                v = ws.cell(row=r, column=2).value
                if k and v not in (None, "", " ") and str(k).strip() not in ("Column1",):
                    instrument[self.normalize_key(k)] = v
            wavelength_header = ws.cell(row=header_row, column=6).value if header_row else None
            readings = []
            if header_row:
                for r in range(header_row + 1, ws.max_row + 1):
                    no = ws.cell(row=r, column=1).value
                    if no is None:
                        continue
                    readings.append({
                        "no": no,
                        "mode": ws.cell(row=r, column=2).value,
                        "cell_no": ws.cell(row=r, column=3).value,
                        "sample_name": ws.cell(row=r, column=4).value,
                        "comment": ws.cell(row=r, column=5).value,
                        "value": self._sf(ws.cell(row=r, column=6).value),
                    })
            blocks.append({
                "timepoint": timepoint,
                "raw_sheet_name": sn,
                "wavelength_header": str(wavelength_header) if wavelength_header else None,
                "instrument_metadata": instrument,
                "reading_count": len(readings),
                "readings": readings,
            })
        return blocks

    # ---------- PROCESSED DATA ---------- #
    def extract_processed_data(self, sheet_name="Processed Data") -> Dict:
        sheets = self._find_sheets("processed data") or self._find_sheets("processed_data")
        if not sheets:
            return {"available": False}
        ws = self.wb[sheets[0]]
        title = ws.cell(row=1, column=1).value
        subtitle = ws.cell(row=2, column=1).value
        metric_label = ws.cell(row=5, column=2).value

        # calibration curve captured on the side of each block (col 9)
        calibration = {
            "equation": None,
            "slope": None,
            "intercept": None,
        }
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=9).value
            if v and isinstance(v, str) and v.strip().lower().startswith("y="):
                calibration["equation"] = v.strip()
                sl = self._sf(ws.cell(row=r + 1, column=9).value)
                ic = self._sf(ws.cell(row=r + 2, column=9).value)
                if sl is None or ic is None:
                    psl, pic = self._parse_linear_equation(v)
                    sl = sl if sl is not None else psl
                    ic = ic if ic is not None else pic
                calibration["slope"] = sl
                calibration["intercept"] = ic
                break

        # locate each timepoint block: a row in col1 matching T\d+
        block_starts = []
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=1).value
            if v and re.fullmatch(r"T\d+", str(v).strip()):
                block_starts.append((str(v).strip(), r))

        blocks = []
        for i, (tp, start) in enumerate(block_starts):
            end = block_starts[i + 1][1] - 1 if i + 1 < len(block_starts) else ws.max_row
            # Label row = the group/replicate-label row: col2 holds a non-numeric string
            # (e.g. "Control1", "CMS_1a-Np 1"). It may sit at start+1 or start+2 depending
            # on whether an extra "Replicate N" caption row precedes it.
            group_labels = None
            in_rfu = False
            rfu_rows = []
            rfu_mean = None
            cells_row = None
            mean_cells = {}
            for r in range(start + 1, end + 1):
                lbl = ws.cell(row=r, column=1).value
                lbl_s = str(lbl).strip().lower() if lbl else ""
                c2 = ws.cell(row=r, column=2).value
                vals = [self._sf(ws.cell(row=r, column=c).value) for c in range(2, 8)]

                if lbl and lbl_s.startswith("relative fluorescence"):
                    in_rfu = True
                    rfu_rows.append(vals)
                    continue
                if lbl and lbl_s.startswith("rfu (mean)"):
                    in_rfu = False
                    rfu_mean = vals
                    continue
                if lbl and lbl_s.startswith("cells per ml"):
                    cells_row = vals
                    continue
                if lbl and lbl_s.startswith("mean (cells"):
                    mean_cells = {
                        "control": self._sf(ws.cell(row=r, column=2).value),
                        "treatment": self._sf(ws.cell(row=r, column=5).value),
                    }
                    continue
                if in_rfu and lbl is None and any(v is not None for v in vals):
                    rfu_rows.append(vals)
                    continue
                # label row: numeric-free, contains strings in col2
                if group_labels is None and c2 is not None and self._sf(c2) is None:
                    group_labels = [ws.cell(row=r, column=c).value for c in range(2, 8)]

            blocks.append({
                "timepoint": tp,
                "group_labels": [str(x) if x is not None else None for x in (group_labels or [])],
                "rfu_replicates": rfu_rows,
                "rfu_mean": rfu_mean,
                "cells_per_ml": cells_row,
                "mean_cells_per_ml": mean_cells,
            })

        return {
            "available": True,
            "title": str(title) if title else None,
            "subtitle": str(subtitle) if subtitle else None,
            "metric_label": str(metric_label) if metric_label else None,
            "calibration_curve": calibration,
            "blocks": blocks,
        }

    # ---------- CALIBRATION CURVE ---------- #
    def extract_calibration_curve(self) -> Dict:
        sheets = self._find_sheets("calibration")
        if not sheets:
            return {"available": False}
        ws = self.wb[sheets[0]]
        concentrations = [self._sf(ws.cell(row=1, column=c).value) for c in range(2, ws.max_column + 1) if ws.cell(row=1, column=c).value is not None]
        replicate_rfu = []
        for r in range(2, ws.max_row + 1):
            vals = [self._sf(ws.cell(row=r, column=c).value) for c in range(2, 2 + len(concentrations))]
            if any(v is not None for v in vals):
                replicate_rfu.append(vals)
        equation = None
        slope = None
        intercept = None
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                if v and isinstance(v, str) and v.strip().lower().startswith("y="):
                    equation = v.strip()
                    slope, intercept = self._parse_linear_equation(v)
                    break
            if equation:
                break
        return {
            "available": True,
            "concentrations": concentrations,
            "replicate_rfu": replicate_rfu,
            "equation": equation,
            "slope": slope,
            "intercept": intercept,
        }

    # ---------- FINAL RESULTS ---------- #
    def extract_final_results(self) -> Dict:
        sheets = self._find_sheets("final results") or self._find_sheets("final_results")
        if not sheets:
            return {"available": False}
        ws = self.wb[sheets[0]]

        # locate the summary growth table header row ("Time (h)")
        hdr_row = None
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=2).value
            if v and "time (h)" in str(v).strip().lower():
                hdr_row = r
                break
        treatments_label = None
        headers = []
        growth = []
        if hdr_row:
            treatments_label = ws.cell(row=hdr_row - 1, column=3).value
            headers = [str(ws.cell(row=hdr_row, column=c).value) for c in range(2, 5) if ws.cell(row=hdr_row, column=c).value is not None]
            for r in range(hdr_row + 1, ws.max_row + 1):
                t = ws.cell(row=r, column=2).value
                if t is None or not isinstance(t, (int, float)):
                    break
                growth.append({
                    "time_h": self._sf(t),
                    "control": self._sf(ws.cell(row=r, column=3).value),
                    "treatment": self._sf(ws.cell(row=r, column=4).value),
                })

        def _find_label(*needles):
            for r in range(1, ws.max_row + 1):
                v = ws.cell(row=r, column=2).value
                if v and any(n in str(v).strip().lower() for n in needles):
                    return r
            return None

        # validity
        validity = {}
        vr = _find_label("validity of the test")
        if vr:
            validity = {
                "text": str(ws.cell(row=vr, column=2).value),
                "result": str(ws.cell(row=vr, column=4).value) if ws.cell(row=vr, column=4).value else None,
                "note": str(ws.cell(row=vr + 1, column=2).value) if ws.cell(row=vr + 1, column=2).value else None,
            }

        # statistics
        stats = {}
        sr = _find_label("significant difference to control")
        if sr:
            stats = {
                "significant_difference_label": str(ws.cell(row=sr, column=2).value),
                "significant_difference": str(ws.cell(row=sr + 1, column=2).value) if ws.cell(row=sr + 1, column=2).value else None,
                "equality_of_slopes_label": str(ws.cell(row=sr, column=5).value) if ws.cell(row=sr, column=5).value else None,
                "equality_of_slopes": str(ws.cell(row=sr + 1, column=5).value) if ws.cell(row=sr + 1, column=5).value else None,
            }

        # EC50
        ec50 = {}
        er = _find_label("calculated ec50")
        if er:
            ec50 = {
                "label": str(ws.cell(row=er, column=2).value),
                "description": str(ws.cell(row=er, column=4).value) if ws.cell(row=er, column=4).value else None,
                "value": str(ws.cell(row=er + 1, column=2).value) if ws.cell(row=er + 1, column=2).value else None,
            }

        return {
            "available": True,
            "treatments_label": str(treatments_label) if treatments_label else None,
            "headers": headers,
            "growth_curve": growth,
            "validity": validity,
            "statistics": stats,
            "ec50": ec50,
            "calibration_curve": self.extract_calibration_curve(),
        }

    # ---------- ALL ---------- #
    def parse_all_data(self) -> Dict:
        try:
            return {
                "test_details": {
                    "work_package": self.extract_work_package_data(),
                    "material": self.extract_material_data(),
                    "application": self.extract_application_data(),
                    "test_conditions": self.extract_test_conditions_data(),
                    "analysis_conditions": self.extract_analysis_conditions_data(),
                    "culture_conditions": self.extract_culture_conditions_data(),
                    "replications": self.extract_replications(),
                    "parser_warnings": self.spelling_issues,
                },
                "replications": self.extract_raw_data(),
                "processed_data": self.extract_processed_data(),
                "final_results": self.extract_final_results(),
            }
        except Exception as e:
            logger.error(f"Error parsing data: {e}\n{traceback.format_exc()}")
            raise


def _fix_degree_symbols(obj):
    if isinstance(obj, str):
        return obj.replace("oC", "\u00b0C")
    if isinstance(obj, dict):
        return {(k.replace("oC", "\u00b0C") if isinstance(k, str) else k): _fix_degree_symbols(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_fix_degree_symbols(v) for v in obj]
    return obj


def parse_excel_algae(file_path: str, sheet_name: str = "Test Details") -> Dict:
    return _fix_degree_symbols(AlgaeParser(file_path, sheet_name).parse_all_data())


if __name__ == "__main__":
    import sys
    fp = sys.argv[1] if len(sys.argv) > 1 else "backend/data/CMS_WP3_Algae_toxicity_1a_DB.xlsx"
    d = parse_excel_algae(fp)
    td = d["test_details"]
    print("=== ALGAE PARSE SUMMARY ===")
    print(f"Test: {td['work_package'].get('full_test_name')}")
    print(f"Material: {td['material'].get('material_name')}")
    print(f"Raw data blocks (timepoints): {len(d['replications'])}")
    print(f"Processed blocks: {len(d['processed_data'].get('blocks', []))}")
    print(f"Final growth points: {len(d['final_results'].get('growth_curve', []))}")
    print(f"Calibration points: {len(d['final_results'].get('calibration_curve', {}).get('concentrations', []))}")
    print(f"Parser warnings: {td['parser_warnings']}")
    print("\n=== JSON PREVIEW ===")
    print(json.dumps(d, indent=2, default=str)[:12000])