import openpyxl
import re
from dataclasses import dataclass, asdict, field
from typing import List, Dict, Optional, Union
import logging
import traceback
from datetime import datetime, timedelta
from difflib import SequenceMatcher

# Set up logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
logging.getLogger("multipart.multipart").setLevel(logging.WARNING)
logger = logging.getLogger(__name__)

# Dataclasses tailored for Zeta
@dataclass
class Scientist:
    name: Optional[str] = None
    email: Optional[str] = None

@dataclass
class WorkPackageData:
    wp_name: Optional[str] = None
    partner: Optional[str] = None
    full_test_name: Optional[str] = None
    test_acronym: Optional[str] = None
    test_type: Optional[str] = None
    endpoint: Optional[str] = None
    endpoint_outcome: Optional[str] = None
    sop: Optional[str] = None
    path: Optional[str] = None
    lead_scientists: List[Scientist] = field(default_factory=list)
    assay_scientists: List[Scientist] = field(default_factory=list)

@dataclass
class MaterialData:
    material_identifier: Optional[str] = None
    erm_id: Optional[str] = None
    core_chemistry: Optional[str] = None
    material_name: Optional[str] = None
    material_state: Optional[str] = None
    cas: Optional[str] = None
    casforcore: Optional[str] = None
    batch: Optional[str] = None
    preparation_date: Optional[str] = None
    particles_stock: Optional[str] = None
    molar_concentration: Optional[str] = None

@dataclass
class ReplicationData:
    test_identifier_number: Optional[str] = None
    test_start_date: Optional[str] = None
    test_end_date: Optional[str] = None

@dataclass
class SamplePreparationData:
    dispersion_protocol: Optional[str] = None
    dispersion_technique: Optional[str] = None
    dispersion_medium: Optional[str] = None
    sonicator_type: Optional[str] = None
    power: Optional[str] = None
    sonication_time: Optional[str] = None
    tip_thickness: Optional[str] = None
    tip_composition: Optional[str] = None
    ultrasonic_bath_size: Optional[str] = None
    sample_volume: Optional[str] = None
    final_concentration: Optional[str] = None
    additional_info: Optional[str] = None

@dataclass
class ZetaInstrumentationData:
    instrument_specs: Optional[str] = None
    cell_model: Optional[str] = None
    temperature: Optional[str] = None
    thermal_equilibrium_time: Optional[str] = None
    replication: Optional[int] = None
    approximation: Optional[str] = None
    henry_factor: Optional[float] = None
    adjustment_mode: Optional[str] = None
    max_voltage: Optional[str] = None
    quality: Optional[str] = None
    max_number_of_runs: Optional[int] = None
    refractive_index_medium: Optional[float] = None
    viscosity_medium: Optional[str] = None
    relative_permittivity_medium: Optional[float] = None

@dataclass
class ZetaRawPhase:
    time: Optional[float] = None
    phase_measured: Optional[float] = None
    phase_fitted: Optional[float] = None
    voltage: Optional[float] = None

@dataclass
class ZetaRawIntensity:
    time: Optional[float] = None
    monitor: Optional[float] = None
    detector: Optional[float] = None

@dataclass
class ZetaRawParameters:
    processed_runs: Optional[int] = None
    filter_optical_density: Optional[float] = None
    mean_intensity: Optional[float] = None
    adjusted_voltage: Optional[float] = None
    transmittance: Optional[float] = None

@dataclass
class ZetaRawData:
    run_number: int
    phase_data: List[ZetaRawPhase] = field(default_factory=list)
    intensity_data: List[ZetaRawIntensity] = field(default_factory=list)
    parameters: ZetaRawParameters = field(default_factory=ZetaRawParameters)

@dataclass
class ZetaDistribution:
    zeta_mv: Optional[float] = None
    frequency: Optional[float] = None

@dataclass
class ZetaProcessedResults:
    mean_zeta: Optional[float] = None
    std_dev: Optional[float] = None
    peak_max: Optional[float] = None
    mobility: Optional[float] = None
    conductivity: Optional[float] = None

@dataclass
class ZetaProcessedData:
    run_number: int
    distribution: List[ZetaDistribution] = field(default_factory=list)
    results: ZetaProcessedResults = field(default_factory=ZetaProcessedResults)

@dataclass
class ZetaFinalResults:
    mean_zeta: Optional[float] = None
    pooled_std_zeta: Optional[float] = None
    std_between_zeta: Optional[float] = None
    mobility: Optional[float] = None
    std_between_mobility: Optional[float] = None
    conductivity: Optional[float] = None
    std_between_conductivity: Optional[float] = None

class ZetaParser:
    def __init__(self, file_path: str, sheet_name: str = "Test Information"):
        self.file_path = file_path
        self.sheet_name = sheet_name
        try:
            self.wb = openpyxl.load_workbook(file_path, data_only=True)
            self.ws = self.wb[sheet_name]
            logger.info(f"Successfully loaded Excel file: {file_path}")
        except Exception as e:
            logger.error(f"Failed to load workbook or sheet {sheet_name}: {e}")
            raise
        self.email_regex = r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$'

    def normalize_key(self, key: Optional[str]) -> Optional[str]:
        if key:
            normalized = str(key).strip().lower() if key is not None else ""
            normalized = re.sub(r'[^a-z0-9]', '_', normalized)
            normalized = re.sub(r'_+', '_', normalized).strip('_')
            return normalized
        return None

    def split_value_unit(self, value: Optional[str]) -> tuple[Optional[Union[str, float]], Optional[str]]:
        if isinstance(value, str) and " " in value:
            parts = value.split(" ", 1)
            try:
                numeric = float(parts[0]) if '.' in parts[0] else int(parts[0])
                return numeric, parts[1] if len(parts) > 1 else None
            except (ValueError, IndexError):
                return parts[0], parts[1] if len(parts) > 1 else None
        return value, None

    def excel_date_to_string(self, value: Optional[Union[float, str]]) -> Optional[str]:
        try:
            if isinstance(value, (int, float)):
                base_date = datetime(1899, 12, 30)
                delta = timedelta(days=float(value))
                return (base_date + delta).strftime("%Y-%m-%d")
            elif isinstance(value, str):
                for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"]:
                    try:
                        return datetime.strptime(value, fmt).strftime("%Y-%m-%d")
                    except ValueError:
                        pass
            return str(value) if value else None
        except Exception as e:
            logger.warning(f"Failed to convert date {value}: {e}")
            return None

    def extract_work_package_data(self):
        data = []
        lead_scientists = []
        assay_scientists = []

        for row in self.ws.iter_rows():
            key_cell = row[0].value
            value_cell = row[1].value if len(row) > 1 else None
            email_cell = row[3].value if len(row) > 3 else None
            comment = row[0].comment

            if comment:
                key = self.normalize_key(key_cell)
                if not key:
                    continue
                entry = {"Key": key, "Value": value_cell}

                if email_cell and re.match(self.email_regex, str(email_cell)):
                    entry["Email"] = email_cell

                if key and "lead_scientist" in key:
                    lead_scientists.append(Scientist(name=value_cell, email=email_cell))
                if key and "assay_test_work" in key:
                    assay_scientists.append(Scientist(name=value_cell, email=email_cell))

                data.append(entry)

        wp_data = WorkPackageData(
            wp_name=next((d["Value"] for d in data if d["Key"] == "project_work_package"), None),
            partner=next((d["Value"] for d in data if d["Key"] == "partner_conducting_test_assay"), None),
            full_test_name=next((d["Value"] for d in data if d["Key"] == "full_name_of_test_assay_add_oecd_test_ref_id_if_app"), None),
            test_acronym=next((d["Value"] for d in data if d["Key"] == "short_name_or_acronym_for_test_assay"), None),
            test_type=next((d["Value"] for d in data if d["Key"] == "type_or_class_of_experimental_test_as_used_here"), None),
            endpoint=next((d["Value"] for d in data if d["Key"] == "end_point_being_investigated_assessed_by_the_test"), None),
            endpoint_outcome=next((d["Value"] for d in data if d["Key"] == "metric_s_used_to_assess_end_point_outcome_response"), None),
            sop=next((d["Value"] for d in data if d["Key"] == "sop_s_for_test_ref_project_or_other_doc_title_id"), None),
            path=next((d["Value"] for d in data if d["Key"] == "path_link_to_sop_protocol_on_proj_server_web_where_applic"), None),
            lead_scientists=lead_scientists,
            assay_scientists=assay_scientists
        )
        return asdict(wp_data)

    def extract_material_data(self):
        data = []
        expected_keys = [
            "sample_cms_internal_identifier",
            "erm_identifier_number",
            "core_chemistry",
            "material_name",
            "material_state",
            "cas_no",
            "cas_for_core",
            "batch",
            "date_of_preparation",
            "no_of_particles_in_stock",
            "molar_concentration"
        ]
        unmatched_keys = set(expected_keys)
        potential_matches = []

        for row_idx, row in enumerate(self.ws.iter_rows(min_row=1, max_col=5), start=1):
            key_cell = row[0].value
            if not key_cell:
                continue
            raw_key = str(key_cell)
            key = self.normalize_key(raw_key)
            if not key:
                continue

            value = None
            for col_idx in range(1, 5):
                if col_idx < len(row) and row[col_idx].value is not None:
                    value = row[col_idx].value
                    break

            if key in expected_keys:
                unmatched_keys.discard(key)
                data.append({"Key": key, "Value": value})
            else:
                for expected_key in expected_keys:
                    similarity = SequenceMatcher(None, key, expected_key).ratio()
                    if similarity > 0.85:
                        unmatched_keys.discard(expected_key)
                        data.append({"Key": expected_key, "Value": value})
                    elif similarity > 0.5:
                        potential_matches.append(f"Potential match for '{expected_key}': '{raw_key}' similarity={similarity:.2f}")

        if unmatched_keys:
            logger.warning(f"Unmatched material data keys: {unmatched_keys}")

        material_data = MaterialData(
            material_identifier=next((d["Value"] for d in data if d["Key"] == "sample_cms_internal_identifier"), None),
            erm_id=next((d["Value"] for d in data if d["Key"] == "erm_identifier_number"), None),
            core_chemistry=next((d["Value"] for d in data if d["Key"] == "core_chemistry"), None),
            material_name=next((d["Value"] for d in data if d["Key"] == "material_name"), None),
            material_state=next((d["Value"] for d in data if d["Key"] == "material_state"), None),
            cas=next((d["Value"] for d in data if d["Key"] == "cas_no"), None),
            casforcore=next((d["Value"] for d in data if d["Key"] == "cas_for_core"), None),
            batch=next((d["Value"] for d in data if d["Key"] == "batch"), None),
            preparation_date=self.excel_date_to_string(next((d["Value"] for d in data if d["Key"] == "date_of_preparation"), None)),
            particles_stock=next((d["Value"] for d in data if d["Key"] == "no_of_particles_in_stock"), None),
            molar_concentration=next((d["Value"] for d in data if d["Key"] == "molar_concentration"), None)
        )
        return asdict(material_data)

    def extract_sample_preparation_data(self):
        data = []
        expected_keys = [
            "specify_standard_dispersion_protocol_used",
            "or_otherwise_specify_dispersion_technique_used",
            "dispersion_dilution_medium",
            "sonicator_type",
            "power_w",
            "sonication_time_secs",
            "tip_thickness_mm",
            "tip_composition",
            "size_of_ultrasonic_bath_water_volume_dm3",
            "sample_volume",
            "final_sample_concentration_mg_l_or_ppm",
            "additional_information"
        ]
        unmatched_keys = set(expected_keys)
        potential_matches = []

        for row_idx, row in enumerate(self.ws.iter_rows(min_row=1, max_col=5), start=1):
            key_cell = row[0].value
            if not key_cell:
                continue
            raw_key = str(key_cell)
            key = self.normalize_key(raw_key)
            if not key:
                continue

            value = None
            for col_idx in range(1, 5):
                if col_idx < len(row) and row[col_idx].value is not None:
                    value = row[col_idx].value
                    break

            if key in expected_keys:
                unmatched_keys.discard(key)
                data.append({"Key": key, "Value": value})
            else:
                for expected_key in expected_keys:
                    similarity = SequenceMatcher(None, key, expected_key).ratio()
                    if similarity > 0.85:
                        unmatched_keys.discard(expected_key)
                        data.append({"Key": expected_key, "Value": value})
                    elif similarity > 0.5:
                        potential_matches.append(f"Potential match for '{expected_key}': '{raw_key}' similarity={similarity:.2f}")

        if unmatched_keys:
            logger.warning(f"Unmatched sample preparation data keys: {unmatched_keys}")

        sample_preparation_data = SamplePreparationData(
            dispersion_protocol=next((d["Value"] for d in data if d["Key"] == "specify_standard_dispersion_protocol_used"), None),
            dispersion_technique=next((d["Value"] for d in data if d["Key"] == "or_otherwise_specify_dispersion_technique_used"), None),
            dispersion_medium=next((d["Value"] for d in data if d["Key"] == "dispersion_dilution_medium"), None),
            sonicator_type=next((d["Value"] for d in data if d["Key"] == "sonicator_type"), None),
            power=next((d["Value"] for d in data if d["Key"] == "power_w"), None),
            sonication_time=next((d["Value"] for d in data if d["Key"] == "sonication_time_secs"), None),
            tip_thickness=next((d["Value"] for d in data if d["Key"] == "tip_thickness_mm"), None),
            tip_composition=next((d["Value"] for d in data if d["Key"] == "tip_composition"), None),
            ultrasonic_bath_size=next((d["Value"] for d in data if d["Key"] == "size_of_ultrasonic_bath_water_volume_dm3"), None),
            sample_volume=next((d["Value"] for d in data if d["Key"] == "sample_volume"), None),
            final_concentration=next((d["Value"] for d in data if d["Key"] == "final_sample_concentration_mg_l_or_ppm"), None),
            additional_info=next((d["Value"] for d in data if d["Key"] == "additional_information"), None)
        )
        return asdict(sample_preparation_data)

    def extract_instrumentation_data(self):
        data = []
        expected_keys = [
            "zeta_instrument_specifications",
            "cell_model",
            "temperature",
            "thermal_equilibrium_time",
            "replication",
            "approximation",
            "henry_factor",
            "adjustment_mode",
            "max_voltage",
            "quality",
            "max_number_of_runs",
            "refractive_index_of_the_medium",
            "viscosity_of_the_medium",
            "relative_permittivity_of_the_medium"
        ]
        unmatched_keys = set(expected_keys)
        potential_matches = []

        for row_idx, row in enumerate(self.ws.iter_rows(min_row=1, max_col=5), start=1):
            key_cell = row[0].value
            if not key_cell:
                continue
            raw_key = str(key_cell)
            key = self.normalize_key(raw_key)
            if not key:
                continue

            value = None
            for col_idx in range(1, 5):
                if col_idx < len(row) and row[col_idx].value is not None:
                    value = row[col_idx].value
                    break

            if key in expected_keys:
                unmatched_keys.discard(key)
                data.append({"Key": key, "Value": value})
            else:
                for expected_key in expected_keys:
                    similarity = SequenceMatcher(None, key, expected_key).ratio()
                    if similarity > 0.85:
                        unmatched_keys.discard(expected_key)
                        data.append({"Key": expected_key, "Value": value})
                    elif similarity > 0.5:
                        potential_matches.append(f"Potential match for '{expected_key}': '{raw_key}' similarity={similarity:.2f}")

        if unmatched_keys:
            logger.warning(f"Unmatched instrumentation data keys: {unmatched_keys}")

        instrumentation_data = ZetaInstrumentationData(
            instrument_specs=next((d["Value"] for d in data if d["Key"] == "zeta_instrument_specifications"), None),
            cell_model=next((d["Value"] for d in data if d["Key"] == "cell_model"), None),
            temperature=next((d["Value"] for d in data if d["Key"] == "temperature"), None),
            thermal_equilibrium_time=next((d["Value"] for d in data if d["Key"] == "thermal_equilibrium_time"), None),
            replication=next((int(d["Value"]) if d["Value"] is not None else None for d in data if d["Key"] == "replication"), None),
            approximation=next((d["Value"] for d in data if d["Key"] == "approximation"), None),
            henry_factor=next((float(d["Value"]) if d["Value"] is not None else None for d in data if d["Key"] == "henry_factor"), None),
            adjustment_mode=next((d["Value"] for d in data if d["Key"] == "adjustment_mode"), None),
            max_voltage=next((d["Value"] for d in data if d["Key"] == "max_voltage"), None),
            quality=next((d["Value"] for d in data if d["Key"] == "quality"), None),
            max_number_of_runs=next((int(d["Value"]) if d["Value"] is not None else None for d in data if d["Key"] == "max_number_of_runs"), None),
            refractive_index_medium=next((float(d["Value"]) if d["Value"] is not None else None for d in data if d["Key"] == "refractive_index_of_the_medium"), None),
            viscosity_medium=next((d["Value"] for d in data if d["Key"] == "viscosity_of_the_medium"), None),
            relative_permittivity_medium=next((float(d["Value"]) if d["Value"] is not None else None for d in data if d["Key"] == "relative_permittivity_of_the_medium"), None)
        )
        return asdict(instrumentation_data)

    def extract_replication(self):
        start_date = None
        end_date = None
        test_identifier = None
        for row in self.ws.iter_rows():
            key_cell = row[0].value
            if key_cell and 'test identifier number' in str(key_cell).lower():
                test_identifier = row[1].value
                start_date = self.excel_date_to_string(row[2].value)
                end_date = self.excel_date_to_string(row[3].value)
                break
        if test_identifier is None:
            logger.warning("Could not find test identifier number in Test Information sheet.")
        return ReplicationData(
            test_identifier_number=test_identifier,
            test_start_date=start_date,
            test_end_date=end_date
        )

    def extract_raw_data(self, raw_sheet_name: str):
        match = re.search(r'R(\d+)', raw_sheet_name)
        run_number = int(match.group(1)) if match else 0

        if raw_sheet_name not in self.wb.sheetnames:
            logger.error(f"Raw data sheet {raw_sheet_name} not found")
            return ZetaRawData(run_number=run_number)

        raw_ws = self.wb[raw_sheet_name]

        phase_data = []
        intensity_data = []
        parameters = ZetaRawParameters()

        # Parameters from row 4, columns J-N
        try:
            parameters.processed_runs = int(raw_ws.cell(row=4, column=10).value)
            parameters.filter_optical_density = float(raw_ws.cell(row=4, column=11).value)
            parameters.mean_intensity = float(raw_ws.cell(row=4, column=12).value)
            parameters.adjusted_voltage = float(raw_ws.cell(row=4, column=13).value)
            parameters.transmittance = float(raw_ws.cell(row=4, column=14).value)
        except (ValueError, TypeError):
            logger.warning(f"Invalid parameter values in {raw_sheet_name}")

        # Data from row 5
        for row_idx in range(5, raw_ws.max_row + 1):
            # Phase A-D
            p_time = raw_ws.cell(row=row_idx, column=1).value
            p_meas = raw_ws.cell(row=row_idx, column=2).value
            p_fit = raw_ws.cell(row=row_idx, column=3).value
            volt = raw_ws.cell(row=row_idx, column=4).value

            # Intensity F-H
            i_time = raw_ws.cell(row=row_idx, column=6).value
            mon = raw_ws.cell(row=row_idx, column=7).value
            det = raw_ws.cell(row=row_idx, column=8).value

            if p_time is None and i_time is None:
                break

            if p_time is not None:
                try:
                    phase_data.append(ZetaRawPhase(
                        time=round(float(p_time),2),
                        phase_measured=round(float(p_meas),2) if p_meas else None,
                        phase_fitted=round(float(p_fit),2) if p_fit else None,
                        voltage=round(float(volt),2) if volt else None
                    ))
                except (ValueError, TypeError):
                    pass

            if i_time is not None:
                try:
                    intensity_data.append(ZetaRawIntensity(
                        time=round(float(i_time),2),
                        monitor=round(float(mon),2) if mon else None,
                        detector=round(float(det),2) if det else None
                    ))
                except (ValueError, TypeError):
                    pass

        return ZetaRawData(
            run_number=run_number,
            phase_data=phase_data,
            intensity_data=intensity_data,
            parameters=parameters
        )

    def extract_processed_data(self, processed_sheet_name: str):
        match = re.search(r'R(\d+)', processed_sheet_name)
        run_number = int(match.group(1)) if match else 0

        if processed_sheet_name not in self.wb.sheetnames:
            logger.error(f"Processed data sheet {processed_sheet_name} not found")
            return ZetaProcessedData(run_number=run_number)

        proc_ws = self.wb[processed_sheet_name]

        distribution = []
        results = ZetaProcessedResults()

    # Results row 4, columns D-H
        try:
            results.mean_zeta = float(proc_ws.cell(row=4, column=4).value) if proc_ws.cell(row=4, column=4).value is not None else None
            results.std_dev = float(proc_ws.cell(row=4, column=5).value) if proc_ws.cell(row=4, column=5).value is not None else None
            results.peak_max = float(proc_ws.cell(row=4, column=6).value) if proc_ws.cell(row=4, column=6).value is not None else None
            results.mobility = float(proc_ws.cell(row=4, column=7).value) if proc_ws.cell(row=4, column=7).value is not None else None
            results.conductivity = float(proc_ws.cell(row=4, column=8).value) if proc_ws.cell(row=4, column=8).value is not None else None
        except (ValueError, TypeError):
            logger.warning(f"Invalid results values in {processed_sheet_name}")

    # Distribution from columns A:B
        for row_idx in range(4, proc_ws.max_row + 1):
            zeta_raw = proc_ws.cell(row=row_idx, column=1).value
            freq_raw = proc_ws.cell(row=row_idx, column=2).value

        # stop when the distribution block ends
            if zeta_raw is None:
                break

            try:
                zeta = round(float(zeta_raw), 2)
                freq = round(float(freq_raw), 2) if freq_raw is not None else 0.0

                distribution.append(ZetaDistribution(
                    zeta_mv=zeta,
                    frequency=freq
                ))
            except (ValueError, TypeError):
                logger.warning(
                    f"Skipping invalid distribution row {row_idx} in {processed_sheet_name}: "
                    f"zeta={zeta_raw}, freq={freq_raw}"
                )
                continue

        return ZetaProcessedData(
            run_number=run_number,
            distribution=distribution,
            results=results
        )

    def extract_final_results(self):
        final_sheet_name = [name for name in self.wb.sheetnames if "Final results" in name][0] if any("Final results" in name for name in self.wb.sheetnames) else None
        if not final_sheet_name:
            logger.error("Final Results sheet not found")
            return ZetaFinalResults()

        final_ws = self.wb[final_sheet_name]

        results = ZetaFinalResults()

        try:
            results.mean_zeta = float(final_ws.cell(row=4, column=1).value)
            results.pooled_std_zeta = float(final_ws.cell(row=4, column=2).value)
            results.std_between_zeta = float(final_ws.cell(row=4, column=3).value)
            results.mobility = float(final_ws.cell(row=4, column=4).value)
            results.std_between_mobility = float(final_ws.cell(row=4, column=5).value)
            results.conductivity = float(final_ws.cell(row=4, column=6).value)
            results.std_between_conductivity = float(final_ws.cell(row=4, column=7).value)
        except (ValueError, TypeError):
            logger.warning("Invalid final results values")

        return results

    def parse_all_data(self) -> Dict[str, Union[Dict, List]]:
        try:
            work_package_data = self.extract_work_package_data()
            material_data = self.extract_material_data()
            sample_preparation_data = self.extract_sample_preparation_data()
            instrumentation_data = self.extract_instrumentation_data()
            replication = self.extract_replication()

            parsed_data = {
                'test_details': {
                    'work_package': work_package_data,
                    'material': material_data,
                    'sample_preparation': sample_preparation_data,
                    'instrumentation': instrumentation_data
                },
                'replication': asdict(replication),
                'replications': [],
                'processed_data': [],
                'final_results': asdict(self.extract_final_results())
            }

            raw_sheets = [name for name in self.wb.sheetnames if name.startswith("Raw data_Zeta_")]
            processed_sheets = [name for name in self.wb.sheetnames if name.startswith("Processed data_Zeta_")]

            for raw_name in raw_sheets:
                parsed_data['replications'].append(asdict(self.extract_raw_data(raw_name)))

            for proc_name in processed_sheets:
                parsed_data['processed_data'].append(asdict(self.extract_processed_data(proc_name)))

            return parsed_data
        except Exception as e:
            logger.error(f"Error parsing Excel file: {e}\n{traceback.format_exc()}")
            raise

def parse_excel_zeta(file_path: str, sheet_name: str = "Test Information") -> Dict[str, Union[Dict, List]]:
    try:
        parser = ZetaParser(file_path, sheet_name)
        return parser.parse_all_data()
    except Exception as e:
        logger.error(f"Error in parse_excel_zeta: {e}\n{traceback.format_exc()}")
        raise

if __name__ == "__main__":
    file_path = "backend/data/WP2_Zeta_1aR1_R5.xlsx"
    try:
        parsed_data = parse_excel_zeta(file_path)
        print(parsed_data['test_details']['work_package'])
        print(parsed_data['test_details']['material'])
    except Exception as e:
        print(f"Error: {e}")