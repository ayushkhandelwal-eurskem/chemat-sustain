import openpyxl
import re
import json
from dataclasses import dataclass, asdict, field
from typing import List, Dict, Optional, Union, Any
import logging
import traceback
from datetime import datetime, timedelta

logging.basicConfig(level=logging.DEBUG, format="%(asctime)s - %(levelname)s - %(message)s")
logging.getLogger("multipart.multipart").setLevel(logging.WARNING)
logger = logging.getLogger(__name__)


# =========================================================
# Dataclasses
# =========================================================

@dataclass
class Scientist:
    name: Optional[str] = None
    email: Optional[str] = None


@dataclass
class WorkPackageData:
    wp_name: Optional[str] = None
    partner: Optional[str] = None
    laboratory_name: Optional[str] = None
    full_test_name: Optional[str] = None
    test_acronym: Optional[str] = None
    test_type: Optional[str] = None
    endpoint: Optional[str] = None
    endpoint_outcome: Optional[str] = None
    sop: Optional[str] = None
    path: Optional[str] = None
    lead_scientists: List[Scientist] = field(default_factory=list)
    assay_scientists: List[Scientist] = field(default_factory=list)


@dataclass
class MaterialData:
    material_identifier: Optional[str] = None
    erm_id: Optional[str] = None
    material_name: Optional[str] = None
    core_chemistry: Optional[str] = None
    cas_no: Optional[str] = None
    cas_for_core: Optional[str] = None
    material_supplier: Optional[str] = None
    catalog_number: Optional[str] = None
    material_state: Optional[str] = None
    batch: Optional[str] = None
    vial: Optional[str] = None
    preparation_date: Optional[str] = None
    molar_concentration: Optional[str] = None
    particles_stock: Optional[str] = None


@dataclass
class DispersionData:
    dispersion_protocol: Optional[str] = None
    dispersion_technique: Optional[str] = None
    dispersion_medium: Optional[str] = None
    sonicator_type: Optional[str] = None
    power_w: Optional[str] = None
    sonication_time_s: Optional[str] = None
    tip_thickness_mm: Optional[str] = None
    tip_composition: Optional[str] = None
    bath_volume_dm3: Optional[str] = None
    sample_volume: Optional[str] = None
    final_concentration: Optional[str] = None
    additional_info: Optional[str] = None


@dataclass
class InstrumentationData:
    instrument_model: Optional[str] = None
    crucible_type: Optional[str] = None
    replication_count: Optional[Union[int, str]] = None
    replicate_labels: List[str] = field(default_factory=list)
    sample_masses: List[Dict[str, Optional[str]]] = field(default_factory=list)
    protective_atmosphere: Optional[str] = None
    temperature_range: Optional[str] = None
    heating_speed: Optional[str] = None


@dataclass
class ReplicationMetadata:
    test_identifier_number: Optional[str] = None
    test_start_date: Optional[str] = None
    test_end_date: Optional[str] = None
    replicate_label: Optional[str] = None
    raw_sheet_name: Optional[str] = None
    processed_sheet_name: Optional[str] = None


@dataclass
class TGADataPoint:
    time_min: Optional[float] = None
    temperature_c: Optional[float] = None
    mass_mg: Optional[float] = None
    dtg_pct_per_min: Optional[float] = None
    mass_pct: Optional[float] = None


@dataclass
class TGARawDataBlock:
    metric_name: Optional[str] = None
    raw_sheet_name: Optional[str] = None
    time_unit: Optional[str] = "min"
    temperature_unit: Optional[str] = "°C"
    mass_unit: Optional[str] = "mg"
    dtg_unit: Optional[str] = "%/min"
    mass_pct_unit: Optional[str] = "%"
    point_count: Optional[int] = None
    min_time_min: Optional[float] = None
    max_time_min: Optional[float] = None
    min_temperature_c: Optional[float] = None
    max_temperature_c: Optional[float] = None
    min_mass_pct: Optional[float] = None
    max_mass_pct: Optional[float] = None
    min_dtg: Optional[float] = None
    max_dtg: Optional[float] = None
    data_points: List[TGADataPoint] = field(default_factory=list)


@dataclass
class TGADecompositionStage:
    replicate_label: Optional[str] = None
    t_start_c: Optional[float] = None
    t_end_c: Optional[float] = None
    t_peak_c: Optional[float] = None
    mass_loss_pct: Optional[float] = None
    mass_loss_at_final_temp_pct: Optional[float] = None


@dataclass
class TGAFinalResultEntry:
    metric_name: Optional[str] = None
    value: Optional[Union[float, str]] = None
    std_dev_pct: Optional[Union[float, str]] = None


# =========================================================
# Parser
# =========================================================

class TGAParser:
    def __init__(self, file_path: str, sheet_name: str = "Test Information"):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.parser_warnings: List[Dict[str, Any]] = []

        try:
            self.wb = openpyxl.load_workbook(file_path, data_only=True)
            self.ws = self.wb[sheet_name]
            logger.info("Successfully loaded TGA workbook: %s", file_path)
        except Exception as e:
            logger.error("Failed to load workbook or sheet '%s': %s", sheet_name, e)
            raise

        self.email_regex = r"^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$"

    # ---- Generic helpers (same as DSC/XRD) ----
    def normalize_key(self, key: Optional[str]) -> Optional[str]:
        if key is None:
            return None
        normalized = str(key).strip().lower()
        normalized = re.sub(r"[^a-z0-9]+", "_", normalized)
        normalized = re.sub(r"_+", "_", normalized).strip("_")
        return normalized

    def excel_date_to_string(self, value) -> Optional[str]:
        try:
            if isinstance(value, datetime):
                return value.strftime("%Y-%m-%d")
            if isinstance(value, (int, float)):
                return (datetime(1899, 12, 30) + timedelta(days=float(value))).strftime("%Y-%m-%d")
            if isinstance(value, str):
                for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d", "%d.%m.%Y"):
                    try:
                        return datetime.strptime(value.strip(), fmt).strftime("%Y-%m-%d")
                    except ValueError:
                        pass
            return str(value).strip() if value not in (None, "") else None
        except Exception as e:
            logger.warning("Failed to convert date '%s': %s", value, e)
            return str(value).strip() if value not in (None, "") else None

    def _safe_float(self, value) -> Optional[float]:
        if value in (None, ""):
            return None
        try:
            if isinstance(value, str):
                value = value.replace(",", ".").strip()
            return float(value)
        except (ValueError, TypeError):
            return None

    def _get_first_value_in_row(self, row, start_col: int = 2, end_col: int = 6):
        for col_idx in range(start_col, end_col + 1):
            if col_idx - 1 < len(row):
                value = row[col_idx - 1].value
                if value is not None:
                    return value
        return None

    def _sheet_key_values(self) -> List[Dict[str, Any]]:
        data = []
        for row_idx, row in enumerate(self.ws.iter_rows(min_row=1, max_col=6), start=1):
            key_cell = row[0].value
            if not key_cell:
                continue
            raw_key = str(key_cell).strip()
            key = self.normalize_key(raw_key)
            if not key:
                continue
            value = self._get_first_value_in_row(row, 2, 6)
            email = row[3].value if len(row) > 3 else None
            data.append({"row": row_idx, "raw_key": raw_key, "key": key, "value": value, "email": email})
        return data

    def _find_sheets(self, prefix: str) -> List[str]:
        prefix_lower = prefix.lower()
        return sorted([
            s for s in self.wb.sheetnames
            if s.lower().startswith(prefix_lower) or s.lower().startswith(prefix_lower.replace(" ", "_"))
        ])

    # ---- Test details ----
    def extract_work_package_data(self) -> Dict[str, Any]:
        rows = self._sheet_key_values()
        lead, assay = [], []
        for r in rows:
            if "lead_scientist" in r["key"]:
                lead.append(Scientist(name=r["value"], email=r["email"] if r["email"] and re.match(self.email_regex, str(r["email"])) else None))
            if "assay_test_work_conducted_by" in r["key"]:
                assay.append(Scientist(name=r["value"], email=r["email"] if r["email"] and re.match(self.email_regex, str(r["email"])) else None))

        return asdict(WorkPackageData(
            wp_name=next((r["value"] for r in rows if r["key"] == "project_work_package"), None),
            partner=next((r["value"] for r in rows if r["key"] == "partner_conducting_test_assay"), None),
            laboratory_name=next((r["value"] for r in rows if r["key"] == "test_facility_laboratory_name"), None),
            full_test_name=next((r["value"] for r in rows if "full_name_of_test_assay" in r["key"]), None),
            test_acronym=next((r["value"] for r in rows if "short_name_or_acronym" in r["key"]), None),
            test_type=next((r["value"] for r in rows if "type_or_class_of_experimental" in r["key"]), None),
            endpoint=next((r["value"] for r in rows if "end_point_being_investigated" in r["key"]), None),
            endpoint_outcome=next((r["value"] for r in rows if "metric_s_used_to_assess" in r["key"]), None),
            sop=next((r["value"] for r in rows if "sop_s_for_test" in r["key"]), None),
            path=next((r["value"] for r in rows if "path_link_to_sop" in r["key"]), None),
            lead_scientists=lead, assay_scientists=assay,
        ))

    def extract_material_data(self) -> Dict[str, Any]:
        rows = self._sheet_key_values()
        batch_val = next((r["value"] for r in rows if r["key"] == "batch"), None)
        return asdict(MaterialData(
            material_identifier=next((r["value"] for r in rows if r["key"] == "sample_cms_internal_identifier"), None),
            erm_id=next((r["value"] for r in rows if "erm_identifier" in r["key"]), None),
            material_name=next((r["value"] for r in rows if r["key"] == "material_name"), None),
            core_chemistry=next((r["value"] for r in rows if "core_chemistry" in r["key"]), None),
            cas_no=next((r["value"] for r in rows if r["key"] == "cas_no"), None),
            cas_for_core=next((r["value"] for r in rows if "cas_for_core" in r["key"] or "cas_no_for_core" in r["key"]), None),
            material_supplier=next((r["value"] for r in rows if r["key"] == "material_supplier"), None),
            catalog_number=next((r["value"] for r in rows if "catalog_number" in r["key"]), None),
            material_state=next((r["value"] for r in rows if r["key"] == "material_state"), None),
            batch=str(batch_val) if batch_val is not None else None,
            vial=next((r["value"] for r in rows if r["key"] == "vial"), None),
            preparation_date=self.excel_date_to_string(next((r["value"] for r in rows if "date_of_sample_preparation" in r["key"] or "date_of_preparation" in r["key"]), None)),
            molar_concentration=next((r["value"] for r in rows if "molar_concentration" in r["key"]), None),
            particles_stock=next((r["value"] for r in rows if "no_of_particles" in r["key"] or "number_of_particles" in r["key"]), None),
        ))

    def extract_dispersion_data(self) -> Dict[str, Any]:
        rows = self._sheet_key_values()
        return asdict(DispersionData(
            dispersion_protocol=next((r["value"] for r in rows if "standard_dispersion_protocol" in r["key"] or "specify_standard_dispersion" in r["key"]), None),
            dispersion_technique=next((r["value"] for r in rows if "dispersion_technique" in r["key"] or "otherwise_specify_dispersion" in r["key"]), None),
            dispersion_medium=next((r["value"] for r in rows if "dispersion_dilution_medium" in r["key"] or "dispersion_medium" in r["key"]), None),
            sonicator_type=next((r["value"] for r in rows if "sonicator" in r["key"] and "type" in r["key"]), None),
            power_w=next((r["value"] for r in rows if r["key"] == "power_w"), None),
            sonication_time_s=next((r["value"] for r in rows if "sonication_time" in r["key"]), None),
            tip_thickness_mm=next((r["value"] for r in rows if "tip_thickness" in r["key"]), None),
            tip_composition=next((r["value"] for r in rows if "tip_composition" in r["key"]), None),
            bath_volume_dm3=next((r["value"] for r in rows if "ultrasonic_bath" in r["key"] or "water_volume" in r["key"]), None),
            sample_volume=next((r["value"] for r in rows if r["key"] == "sample_volume"), None),
            final_concentration=next((r["value"] for r in rows if "final_sample_concentration" in r["key"]), None),
            additional_info=next((r["value"] for r in rows if "additional_information" in r["key"]), None),
        ))

    def extract_instrumentation_data(self) -> Dict[str, Any]:
        rows = self._sheet_key_values()
        ws = self.ws
        replicate_labels = [str(ws.cell(row=61, column=c).value) for c in range(3, 15) if ws.cell(row=61, column=c).value is not None]
        sample_masses = []
        for r in range(62, 65):
            label = ws.cell(row=r, column=1).value
            value = ws.cell(row=r, column=2).value
            if label is not None:
                sample_masses.append({"label": str(label).strip(), "value": str(value).strip() if value is not None else None})

        return asdict(InstrumentationData(
            instrument_model=next((r["value"] for r in rows if "instrumentation" in r["key"] and ("model" in r["key"] or "company" in r["key"])), None),
            crucible_type=next((r["value"] for r in rows if "crucible" in r["key"]), None),
            replication_count=ws.cell(row=61, column=2).value,
            replicate_labels=replicate_labels,
            sample_masses=sample_masses,
            protective_atmosphere=next((r["value"] for r in rows if "protective_atmosphere" in r["key"]), None),
            temperature_range=next((r["value"] for r in rows if "temperature_range" in r["key"]), None),
            heating_speed=next((r["value"] for r in rows if "heating_speed" in r["key"]), None),
        ))

    # ---- Replication metadata ----
    def extract_replication_metadata(self) -> List[Dict[str, Any]]:
        ws = self.ws
        raw_sheets = self._find_sheets("raw data")
        proc_sheets = self._find_sheets("processed data")
        metadata = []

        for row_idx in range(39, 42):
            test_id = ws.cell(row=row_idx, column=2).value
            if test_id is None:
                continue
            test_id_str = str(test_id).strip()
            raw_match = next((s for s in raw_sheets if test_id_str.lower() in s.lower()), raw_sheets[0] if raw_sheets else None)
            proc_match = next((s for s in proc_sheets if test_id_str.lower() in s.lower()), proc_sheets[0] if proc_sheets else None)

            metadata.append(asdict(ReplicationMetadata(
                test_identifier_number=test_id_str,
                test_start_date=self.excel_date_to_string(ws.cell(row=row_idx, column=3).value),
                test_end_date=self.excel_date_to_string(ws.cell(row=row_idx, column=4).value),
                replicate_label="TGA Thermogram",
                raw_sheet_name=raw_match,
                processed_sheet_name=proc_match,
            )))

        if not metadata and raw_sheets:
            metadata.append(asdict(ReplicationMetadata(
                replicate_label="TGA Thermogram", raw_sheet_name=raw_sheets[0],
                processed_sheet_name=proc_sheets[0] if proc_sheets else None,
            )))
        return metadata

    # ---- Raw data ----
    def extract_raw_data(self) -> List[Dict[str, Any]]:
        return [self._extract_raw_block(s) for s in self._find_sheets("raw data")]

    def _extract_raw_block(self, raw_sheet_name: str) -> Dict[str, Any]:
        ws = self.wb[raw_sheet_name]

        # Detect columns from header row
        col_map: Dict[str, int] = {}
        for c in range(1, ws.max_column + 1):
            h = ws.cell(row=1, column=c).value
            if h is None:
                continue
            hl = str(h).strip().lower()
            if "time" in hl:
                col_map["time"] = c
            elif "temp" in hl and "mass" not in hl:
                col_map["temp"] = c
            elif "dtg" in hl:
                col_map["dtg"] = c
            elif "mass" in hl and "%" in hl:
                col_map["mass_pct"] = c
            elif "mass" in hl:
                col_map["mass"] = c

        time_c = col_map.get("time", 1)
        temp_c = col_map.get("temp", 2)
        mass_c = col_map.get("mass", 3)
        dtg_c = col_map.get("dtg", 4)
        mass_pct_c = col_map.get("mass_pct", 7)

        pts: List[Dict] = []
        time_v, temp_v, mass_pct_v, dtg_v = [], [], [], []

        for r in range(2, ws.max_row + 1):
            t = self._safe_float(ws.cell(row=r, column=time_c).value)
            if t is None:
                continue
            temp = self._safe_float(ws.cell(row=r, column=temp_c).value)
            mass = self._safe_float(ws.cell(row=r, column=mass_c).value)
            dtg = self._safe_float(ws.cell(row=r, column=dtg_c).value)
            mpct = self._safe_float(ws.cell(row=r, column=mass_pct_c).value)

            pts.append(asdict(TGADataPoint(time_min=t, temperature_c=temp, mass_mg=mass, dtg_pct_per_min=dtg, mass_pct=mpct)))
            time_v.append(t)
            if temp is not None: temp_v.append(temp)
            if mpct is not None: mass_pct_v.append(mpct)
            if dtg is not None: dtg_v.append(dtg)

        return asdict(TGARawDataBlock(
            metric_name="TGA Thermogram (Temperature vs Mass %)",
            raw_sheet_name=raw_sheet_name,
            point_count=len(pts),
            min_time_min=min(time_v) if time_v else None,
            max_time_min=max(time_v) if time_v else None,
            min_temperature_c=min(temp_v) if temp_v else None,
            max_temperature_c=max(temp_v) if temp_v else None,
            min_mass_pct=min(mass_pct_v) if mass_pct_v else None,
            max_mass_pct=max(mass_pct_v) if mass_pct_v else None,
            min_dtg=min(dtg_v) if dtg_v else None,
            max_dtg=max(dtg_v) if dtg_v else None,
            data_points=pts,
        ))

    # ---- Processed data (decomposition stages) ----
    def extract_processed_data(self) -> Dict[str, Any]:
        sheets = self._find_sheets("processed data")
        if not sheets:
            return {"available": False, "stages": [], "notes": "No processed data sheet found."}

        stages = []
        for sname in sheets:
            ws = self.wb[sname]

            # Parse header row to build column map
            col_map: Dict[str, int] = {}
            for c in range(1, ws.max_column + 1):
                h = ws.cell(row=1, column=c).value
                if h is None:
                    continue
                hl = str(h).strip().lower()
                if "tstart" in hl or "t_start" in hl:
                    col_map["t_start"] = c
                elif "tend" in hl or "t_end" in hl:
                    col_map["t_end"] = c
                elif "tpeak" in hl or "t_peak" in hl:
                    col_map["t_peak"] = c
                elif "mass loss" in hl and "600" not in hl and "1000" not in hl:
                    col_map["mass_loss"] = c
                elif "mass loss" in hl and ("600" in hl or "1000" in hl):
                    col_map["mass_loss_final"] = c

            label_c = 1
            t_start_c = col_map.get("t_start", 2)
            t_end_c = col_map.get("t_end", 3)
            t_peak_c = col_map.get("t_peak", 4)
            ml_c = col_map.get("mass_loss", 5)
            mlf_c = col_map.get("mass_loss_final", 6)

            for r in range(2, ws.max_row + 1):
                label = ws.cell(row=r, column=label_c).value
                ts = self._safe_float(ws.cell(row=r, column=t_start_c).value)
                if ts is None and label is None:
                    continue

                stages.append(asdict(TGADecompositionStage(
                    replicate_label=str(label).strip() if label else None,
                    t_start_c=ts,
                    t_end_c=self._safe_float(ws.cell(row=r, column=t_end_c).value),
                    t_peak_c=self._safe_float(ws.cell(row=r, column=t_peak_c).value),
                    mass_loss_pct=self._safe_float(ws.cell(row=r, column=ml_c).value),
                    mass_loss_at_final_temp_pct=self._safe_float(ws.cell(row=r, column=mlf_c).value),
                )))

        return {"available": True, "stages": stages}

    # ---- Final results ----
    def extract_final_results(self) -> List[Dict[str, Any]]:
        sheets = self._find_sheets("final results")
        if not sheets:
            self.parser_warnings.append({"type": "missing_sheet", "sheet": "Final results", "note": "No Final results sheet found."})
            return []

        results = []
        for sname in sheets:
            ws = self.wb[sname]

            # Row 2 has headers in pairs: metric name, "Std,dev, (%)"
            # Row 3+ has data
            header_row = 2
            col = 1
            metric_cols: List[Dict[str, Any]] = []

            while col <= ws.max_column:
                h = ws.cell(row=header_row, column=col).value
                if h is None:
                    col += 1
                    continue
                hs = str(h).strip()
                if "std" in hs.lower() and "dev" in hs.lower():
                    col += 1
                    continue

                # This is a metric header; next col is its std dev
                metric_cols.append({"name": hs, "value_col": col, "sd_col": col + 1})
                col += 2

            for data_row in range(3, ws.max_row + 1):
                has_data = any(
                    ws.cell(row=data_row, column=mc["value_col"]).value is not None
                    for mc in metric_cols
                )
                if not has_data:
                    continue

                for mc in metric_cols:
                    val = ws.cell(row=data_row, column=mc["value_col"]).value
                    sd = ws.cell(row=data_row, column=mc["sd_col"]).value if mc["sd_col"] <= ws.max_column else None

                    results.append(asdict(TGAFinalResultEntry(
                        metric_name=mc["name"],
                        value=self._safe_float(val) if val is not None else None,
                        std_dev_pct=self._safe_float(sd) if sd is not None else None,
                    )))

        return results

    # ---- Statistical analysis (none for TGA) ----
    def extract_statistical_analysis(self) -> Dict[str, Any]:
        return {"available": False, "notes": "No dedicated statistical analysis section was found in this TGA workbook."}

    # ---- Parse all ----
    def parse_all_data(self) -> Dict[str, Union[Dict, List]]:
        try:
            parsed_data = {
                "test_details": {
                    "work_package": self.extract_work_package_data(),
                    "material": self.extract_material_data(),
                    "cell_line": {},
                    "dispersion": self.extract_dispersion_data(),
                    "instrumentation": self.extract_instrumentation_data(),
                },
                "replication_metadata": self.extract_replication_metadata(),
                "replications": self.extract_raw_data(),
                "processed_data": self.extract_processed_data(),
                "final_results": self.extract_final_results(),
                "statistical_analysis": self.extract_statistical_analysis(),
            }
            if self.parser_warnings:
                parsed_data["parser_warnings"] = self.parser_warnings

            logger.info("FINAL TGA JSON:\n%s", json.dumps(parsed_data, indent=2, default=str))
            return parsed_data
        except Exception as e:
            logger.error("Error parsing TGA file: %s\n%s", e, traceback.format_exc())
            raise


def parse_excel_tga(file_path: str, sheet_name: str = "Test Information") -> Dict[str, Union[Dict, List]]:
    try:
        parser = TGAParser(file_path, sheet_name)
        return parser.parse_all_data()
    except Exception as e:
        logger.error("Error in parse_excel_tga: %s\n%s", e, traceback.format_exc())
        raise


if __name__ == "__main__":
    import sys
    file_path = sys.argv[1] if len(sys.argv) > 1 else "/mnt/user-data/uploads/WP2_TGA_29aR1.xlsx"
    data = parse_excel_tga(file_path)

    print("=" * 70)
    print("TGA PARSER OUTPUT SUMMARY")
    print("=" * 70)

    wp = data["test_details"]["work_package"]
    print(f"\nWP: {wp['wp_name']}, Partner: {wp['partner']}")
    print(f"Test: {wp['test_acronym']} - {wp['full_test_name']}")

    mat = data["test_details"]["material"]
    print(f"Material: {mat['material_name']} ({mat['material_identifier']})")

    inst = data["test_details"]["instrumentation"]
    print(f"Instrument: {inst['instrument_model']}, Range: {inst['temperature_range']}")

    print(f"\nRaw Data Blocks: {len(data['replications'])}")
    for b in data["replications"]:
        print(f"  {b['point_count']} pts | Temp {b['min_temperature_c']:.0f}-{b['max_temperature_c']:.0f}°C | Mass {b['min_mass_pct']:.1f}-{b['max_mass_pct']:.1f}%")

    pd = data["processed_data"]
    print(f"\nProcessed Data: available={pd['available']}, stages={len(pd.get('stages', []))}")
    for s in pd.get("stages", []):
        print(f"  {s['replicate_label']}: TStart={s['t_start_c']}°C, TEnd={s['t_end_c']}°C, TPeak={s['t_peak_c']}°C, Loss={s['mass_loss_pct']}%, Final={s['mass_loss_at_final_temp_pct']}%")

    print(f"\nFinal Results: {len(data['final_results'])}")
    for fr in data["final_results"]:
        print(f"  {fr['metric_name']}: {fr['value']} (SD: {fr['std_dev_pct']})")

    preview = json.loads(json.dumps(data, default=str))
    for b in preview.get("replications", []):
        b["data_points"] = f"[{len(b['data_points'])} points]"
    print("\n--- JSON Preview ---")
    print(json.dumps(preview, indent=2, default=str)[:4000])