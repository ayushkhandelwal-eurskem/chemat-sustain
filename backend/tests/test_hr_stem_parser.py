from pathlib import Path

import pytest
from openpyxl import Workbook

from parsers.hr_stem import parse_excel_hr_stem


def _workbook(
    path: Path,
    *,
    test_identifier: str,
    raw_identifier: str,
    processed_prefix: str,
    length_header: str,
    final_identifier: str,
) -> None:
    workbook = Workbook()
    info = workbook.active
    info.title = "Test Information"
    info["A1"] = "Test Identifier Number:"
    info["B1"] = test_identifier

    raw = workbook.create_sheet(f"Raw data_{raw_identifier}")
    raw.append(["particle no.", "Feret min, Fmin [nm]", length_header, "Feret max, Fmax [nm]"])
    raw.append([1, 10, 12, 14])
    raw.append([2, 20, 24, 28])

    processed = workbook.create_sheet(f"{processed_prefix}_{raw_identifier}")
    processed.append([
        "Particles no.", "Aspect ratio", "ECD diameter [nm]", "radius ECD [nm]", "Roundness",
        None, "arithmetic mean", "standard deviation", "median",
    ])
    processed.append([1, 1.2, 11, 5.5, 0.8, None, 15, 5, 15])
    processed.append([2, 1.4, 22, 11, 0.7, None, 18, 6, 18])

    final = workbook.create_sheet(f"Final results_{final_identifier}")
    final["A3"] = "arithmetic mean"
    for column, value in enumerate(range(1, 19), start=1):
        final.cell(row=4, column=column, value=value)

    workbook.save(path)


@pytest.mark.parametrize(
    ("test_identifier", "raw_identifier", "processed_prefix", "length_header", "final_identifier"),
    [
        (
            "WP2_HR-STEM_2aR1",
            "WP2_STEM_2aR1",
            "Proceseed data",
            "Length, L [nm]",
            "WP2_STEM_2aR1",
        ),
        (
            "WP2_STEM_16aR1",
            "WP2_STEM_16aR1",
            "Processed data",
            "Length, LF [nm]",
            "WP2_HR-STEM_16aR1",
        ),
    ],
)
def test_hr_stem_parser_accepts_sheet_and_length_header_variants(
    tmp_path,
    test_identifier,
    raw_identifier,
    processed_prefix,
    length_header,
    final_identifier,
):
    path = tmp_path / "hr-stem.xlsx"
    _workbook(
        path,
        test_identifier=test_identifier,
        raw_identifier=raw_identifier,
        processed_prefix=processed_prefix,
        length_header=length_header,
        final_identifier=final_identifier,
    )

    parsed = parse_excel_hr_stem(str(path))

    assert len(parsed["replications"]["particles"]) == 2
    assert parsed["replications"]["particles"][0] == {
        "feret_min": 10.0,
        "length": 12.0,
        "feret_max": 14.0,
    }
    assert len(parsed["processed_data"]["processed_particles"]) == 2
    assert parsed["processed_data"]["metrics"]["feret_min_mean"] == 15.0
    assert parsed["final_results"]["feret_min_mean"] == 1.0