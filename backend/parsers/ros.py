import openpyxl
import re
import json
from typing import List, Dict, Optional, Union, Any
import logging
import traceback
from datetime import datetime, timedelta, time as dt_time

logging.basicConfig(level=logging.DEBUG, format="%(asctime)s - %(levelname)s - %(message)s")
logging.getLogger("multipart.multipart").setLevel(logging.WARNING)
logger = logging.getLogger(__name__)

_ROS_ID = r'WP(\d+)_ROS_(\d+)([a-zA-Z])R(\d+)'


class ROSParser:
    def __init__(self, file_path, sheet_name="Test_conditions"):
        self.file_path = file_path
        try:
            self.wb = openpyxl.load_workbook(file_path, data_only=True)
            if sheet_name in self.wb.sheetnames:
                self.ws = self.wb[sheet_name]
            elif "Test Information" in self.wb.sheetnames:
                self.ws = self.wb["Test Information"]
            else:
                self.ws = self.wb[self.wb.sheetnames[0]]
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}"); raise
        self.email_regex = r"^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$"

    def _nk(self, key):
        if not key: return None
        return re.sub(r"_+", "_", re.sub(r"[^a-z0-9]+", "_", str(key).strip().lower())).strip("_")

    def _sf(self, v):
        if v in (None, "", " "): return None
        try:
            if isinstance(v, str): v = v.replace(",", ".").strip()
            return float(v)
        except: return None

    def _date(self, v):
        try:
            if isinstance(v, datetime): return v.strftime("%Y-%m-%d")
            if isinstance(v, (int, float)): return (datetime(1899, 12, 30) + timedelta(days=float(v))).strftime("%Y-%m-%d")
            if isinstance(v, str):
                for f in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
                    try: return datetime.strptime(v.strip(), f).strftime("%Y-%m-%d")
                    except: pass
            return str(v).strip() if v else None
        except: return str(v).strip() if v else None

    def _time_str(self, v):
        if v is None: return None
        if isinstance(v, dt_time):
            parts = []
            if v.hour > 0: parts.append(f"{v.hour}h")
            if v.minute > 0: parts.append(f"{v.minute}m")
            ms = v.microsecond // 1000
            parts.append(f"{v.second}.{ms:03d}s")
            return " ".join(parts) if parts else "0s"
        return str(v)

    def _kv(self, max_col=6):
        data = []
        for ri, row in enumerate(self.ws.iter_rows(min_row=1, max_col=max_col), start=1):
            k = row[0].value
            if not k: continue
            key = self._nk(k)
            if not key: continue
            val = None
            for c in range(1, min(len(row), max_col)):
                if row[c].value is not None: val = row[c].value; break
            email = row[3].value if len(row) > 3 else None
            data.append({"row": ri, "key": key, "value": val, "email": email})
        return data

    def _g(self, rows, *keys):
        return next((r["value"] for r in rows if any(k in r["key"] for k in keys) and r["value"] is not None), None)

    def _find_sheets(self, prefix):
        p = prefix.lower(); seen = set(); result = []
        for s in sorted(self.wb.sheetnames):
            sl = s.lower()
            if sl in seen: continue
            if sl.startswith(p) or sl.startswith(p.replace(" ", "_")) or sl.startswith(p.replace("_", " ")):
                seen.add(sl); result.append(s)
        return result

    # ====== TEST DETAILS ======
    def extract_work_package(self):
        rows = self._kv()
        lead, assay = [], []
        for r in rows:
            if "lead_scientist" in r["key"]:
                em = r["email"]; lead.append({"name": r["value"], "email": em if em and re.match(self.email_regex, str(em)) else None})
            if "assay_test_work_conducted_by" in r["key"]:
                em = r["email"]; assay.append({"name": r["value"], "email": em if em and re.match(self.email_regex, str(em)) else None})
        g = lambda *k: self._g(rows, *k)
        return {"wp_name": g("project_work_package"), "partner": g("partner_conducting_test_assay"), "laboratory_name": g("test_facility_laboratory_name"), "full_test_name": g("full_name_of_test_assay"), "test_acronym": g("short_name_or_acronym"), "test_type": g("type_or_class_of_experimental"), "endpoint": g("end_point_being_investigated"), "endpoint_outcome": g("metric_s_used_to_assess"), "sop": g("sop_s_for_test"), "path": g("path_link_to_sop"), "lead_scientists": lead, "assay_scientists": assay}

    def extract_material(self):
        rows = self._kv(); g = lambda *k: self._g(rows, *k)
        return {"material_identifier": g("sample_cms_internal_identifier"), "erm_id": g("erm_identifier"), "material_name": g("material_name"), "core_chemistry": g("core_chemistry"), "cas_no": g("cas_no"), "cas_for_core": g("cas_for_core"), "material_supplier": g("material_supplier"), "material_state": g("material_state"), "batch": g("batch"), "vial": g("vial"), "preparation_date": self._date(g("date_of_preparation")), "size": g("size"), "endotoxin": g("endotoxin"), "stock_concentration": g("stock_oncentration", "stock_concentration"), "molecular_weight": g("molecular_weight"), "particles_in_stock": g("no_of_particles_in_stock")}

    def extract_cell_line(self):
        rows = self._kv(); ws = self.ws; g = lambda *k: self._g(rows, *k)
        passages = [ws.cell(row=63, column=c).value for c in range(2, 8) if ws.cell(row=63, column=c).value is not None]
        return {"cell_type": g("detailed_cell_type"), "cell_line_short": g("cell_line_short_name"), "supplier": g("supplier"), "passage_numbers": passages, "plate_details": g("plate_details"), "cells_per_well": g("number_of_cells_per_well"), "volume_per_well": g("total_volume_per_well"), "medium": g("medium"), "serum": g("serum_inc"), "serum_concentration_culture": g("serum_concentration_in_culture"), "serum_concentration_treatment": g("serum_concentration_in_treatment"), "serum_heat_inactivated": g("serum_heat_inactivated"), "antibiotics": g("antibiotics"), "complete_growth_medium": g("complete_growth_medium"), "culture_conditions": g("cell_culture_conditions"), "solvent_for_dcf": g("solvent_for_da_dfc"), "incubation_time_dcf": g("incubation_time_with_da_dfc"), "volume_of_solvent": g("volume_of_solvent")}

    def extract_dispersion(self):
        rows = self._kv(max_col=10); g = lambda *k: self._g(rows, *k); ws = self.ws
        aids = {}
        for cl, cv in [(3, 4), (5, 6), (7, 8), (9, 10)]:
            lbl = ws.cell(row=55, column=cl).value; val = ws.cell(row=55, column=cv).value
            if lbl: aids[str(lbl).strip().rstrip(":")] = str(val) if val else None
        return {"dispersion_protocol": g("standard_dispersion_protocol", "specify_standard_dispersion"), "dispersion_technique": g("dispersion_technique", "otherwise_specify_dispersion"), "dispersion_agent": g("dispersion_agent"), "agent_concentration": ws.cell(row=52, column=4).value, "additives": g("additives_used"), "dispersed_in_medium": g("dispersed_in_cell_culture"), "aids": aids, "time_duration": g("specify_time_duration"), "energy": g("energy")}

    def extract_treatment(self):
        ws = self.ws
        concentrations = []
        for c in range(2, 12):
            lbl = ws.cell(row=84, column=c).value
            if lbl is None: break
            concentrations.append({"label": str(lbl), "ug_ml": self._sf(ws.cell(row=85, column=c).value), "particles_x10_12_ml": self._sf(ws.cell(row=86, column=c).value)})
        plate_series = [ws.cell(row=87, column=c).value for c in range(2, 12) if ws.cell(row=87, column=c).value]
        g = lambda *k: self._g(self._kv(), *k)
        return {"time_unit": g("time_point_unit"), "time_labels": [ws.cell(row=80, column=c).value for c in range(2, 8) if ws.cell(row=80, column=c).value], "time_points": [ws.cell(row=81, column=c).value for c in range(2, 8) if ws.cell(row=81, column=c).value], "concentration_unit": g("treatment_concentration_series_unit"), "concentrations": concentrations, "plate_series": plate_series, "positive_control_abbr": g("positive_controls_abbreviation"), "positive_control_desc": g("positive_controls_description"), "negative_control_abbr": g("negative_controls_abbreviation"), "negative_control_desc": g("negative_controls_description"), "num_experiments": g("number_of_experiments")}

    # ====== RAW DATA ======
    def extract_raw_data(self):
        blocks = []; seen = set()
        # Pre-extract processed data group headers for mapping
        proc_headers_by_run = {}
        proc_sheets = [s for s in self._find_sheets("processed_data_") + self._find_sheets("processed data_") if re.search(_ROS_ID, s, re.IGNORECASE)]
        proc_seen = set()
        for sn in proc_sheets:
            if sn in proc_seen: continue
            proc_seen.add(sn)
            ws_p = self.wb[sn]
            m_p = re.search(_ROS_ID, sn, re.IGNORECASE)
            rl = f"R{m_p.group(4)}" if m_p else sn
            ghs = [str(ws_p.cell(row=5, column=c).value) for c in range(3, 20) if ws_p.cell(row=5, column=c).value is not None and "μg" not in str(ws_p.cell(row=5, column=c).value) and "ug" not in str(ws_p.cell(row=5, column=c).value).lower()]
            proc_headers_by_run[rl] = ghs

        for sn in [s for s in self._find_sheets("raw_data_") + self._find_sheets("raw data_") if re.search(_ROS_ID, s, re.IGNORECASE)]:
            if sn in seen: continue
            seen.add(sn); ws = self.wb[sn]
            m = re.search(_ROS_ID, sn, re.IGNORECASE)
            run_label = f"R{m.group(4)}" if m else sn
            readings = []
            for r in range(4, min(ws.max_row + 1, 20)):
                well = ws.cell(row=r, column=5).value
                if well is None: continue
                # Only accept actual well names (A01, B02, etc.)
                if not re.match(r'^[A-H]\d{1,2}$', str(well)): continue
                readings.append({"well": str(well), "type": str(ws.cell(row=r, column=6).value) if ws.cell(row=r, column=6).value else None, "time_r1": self._time_str(ws.cell(row=r, column=7).value), "fluorescein_r1": self._sf(ws.cell(row=r, column=8).value), "time_r2": self._time_str(ws.cell(row=r, column=17).value), "fluorescein_r2": self._sf(ws.cell(row=r, column=18).value), "time_r3": self._time_str(ws.cell(row=r, column=25).value), "fluorescein_r3": self._sf(ws.cell(row=r, column=26).value)})
            # Map concentration labels from processed data group headers
            group_labels = proc_headers_by_run.get(run_label, [])
            for i, rd in enumerate(readings):
                rd["concentration_group"] = group_labels[i] if i < len(group_labels) else None
            plate_meta = []
            for off, lbl in [(3, "reading 1"), (13, "reading 2"), (21, "reading 3")]:
                plate_meta.append({"reading": lbl, "plate": self._sf(ws.cell(row=12, column=off).value), "repeat": self._sf(ws.cell(row=12, column=off+1).value), "end_time": self._time_str(ws.cell(row=12, column=off+2).value), "start_temp": self._sf(ws.cell(row=12, column=off+3).value), "end_temp": self._sf(ws.cell(row=12, column=off+4).value), "barcode": str(ws.cell(row=12, column=off+5).value) if ws.cell(row=12, column=off+5).value else None})
            protocol = [str(ws.cell(row=r, column=3).value).strip() for r in range(19, min(ws.max_row+1, 66)) if ws.cell(row=r, column=3).value and str(ws.cell(row=r, column=3).value).strip()]
            blocks.append({"run_label": run_label, "raw_sheet_name": sn, "reading_count": len(readings), "readings": readings, "plate_metadata": plate_meta, "fluorescein_label": str(ws.cell(row=14, column=3).value) if ws.cell(row=14, column=3).value else None, "fluorescein_count": self._sf(ws.cell(row=15, column=3).value), "protocol_description": protocol})
        return blocks

    # ====== PROCESSED DATA ======
    def extract_processed_data(self):
        blocks = []; seen = set()
        for sn in [s for s in self._find_sheets("processed_data_") + self._find_sheets("processed data_") if re.search(_ROS_ID, s, re.IGNORECASE)]:
            if sn in seen: continue
            seen.add(sn); ws = self.wb[sn]
            m = re.search(_ROS_ID, sn, re.IGNORECASE)
            run_label = f"R{m.group(4)}" if m else sn
            title = ws.cell(row=4, column=3).value
            ghs = [{"col": c, "label": str(ws.cell(row=5, column=c).value)} for c in range(3, min(ws.max_column+1, 20)) if ws.cell(row=5, column=c).value is not None and "μg" not in str(ws.cell(row=5, column=c).value) and "ug" not in str(ws.cell(row=5, column=c).value).lower()]
            raw_values = [{gh["label"]: self._sf(ws.cell(row=r, column=gh["col"]).value) for gh in ghs} for r in range(6, 9)]
            mean_row = {gh["label"]: self._sf(ws.cell(row=32, column=gh["col"]).value) for gh in ghs}
            sd_row = {gh["label"]: self._sf(ws.cell(row=33, column=gh["col"]).value) for gh in ghs}
            cv_row = {gh["label"]: self._sf(ws.cell(row=34, column=gh["col"]).value) for gh in ghs}
            blocks.append({"run_label": run_label, "processed_sheet_name": sn, "title": str(title) if title else None, "experiment_label": str(ws.cell(row=6, column=1).value or ws.cell(row=29, column=2).value), "group_headers": [gh["label"] for gh in ghs], "raw_values": raw_values, "mean": mean_row, "sd": sd_row, "cv": cv_row, "acceptance_text": str(ws.cell(row=43, column=2).value) if ws.cell(row=43, column=2).value else None, "acceptance_result": str(ws.cell(row=43, column=8).value) if ws.cell(row=43, column=8).value else None})
        return blocks

    # ====== FINAL RESULTS ======
    def extract_final_results(self):
        sheets = self._find_sheets("final_results") + self._find_sheets("final results")
        if not sheets: return {"available": False}
        ws = self.wb[sheets[0]]

        def _sect(hdr_row, d_start, d_end, lbl_col, cols, alt_lbl_col=None):
            hdrs = {}
            for c in cols:
                v = ws.cell(row=hdr_row, column=c).value
                if v is not None: hdrs[c] = str(v)
            rows = []
            for r in range(d_start, d_end + 1):
                l = ws.cell(row=r, column=lbl_col).value
                if l is None and alt_lbl_col:
                    l = ws.cell(row=r, column=alt_lbl_col).value
                if l is None: continue
                ls = str(l).strip()
                if "acceptance" in ls.lower() or "criteria" in ls.lower(): break
                rows.append({"label": ls, "values": {hdrs[c]: self._sf(ws.cell(row=r, column=c).value) for c in hdrs}})
            return {"headers": list(hdrs.values()), "rows": rows}

        def _accept(text_row, result_col=8):
            return {"text": str(ws.cell(row=text_row, column=2).value) if ws.cell(row=text_row, column=2).value else None, "result": str(ws.cell(row=text_row, column=result_col).value) if ws.cell(row=text_row, column=result_col).value else None}

        return {
            "available": True,
            "material_id": str(ws.cell(row=3, column=1).value) if ws.cell(row=3, column=1).value else None,
            "fluorescence_ugml": _sect(6, 7, 15, 2, range(3, 7)),
            "fluorescence_ugml_acceptance": _accept(18),
            "reverse_ugml": _sect(23, 24, 32, 2, range(3, 7)),
            "reverse_ugml_acceptance": _accept(35),
            "fluorescence_ugml_chart": {**_sect(77, 78, 86, 2, range(4, 8), alt_lbl_col=3), "unit_label": str(ws.cell(row=77, column=8).value or "μg/mL")},
            "fluorescence_particles_chart": {**_sect(77, 78, 86, 19, range(21, 25), alt_lbl_col=20), "unit_label": str(ws.cell(row=77, column=25).value or "particles x10^12/mL")},
            "percentage_ugml": {**_sect(116, 117, 125, 2, range(4, 8), alt_lbl_col=3), "unit_label": str(ws.cell(row=116, column=8).value or "μg/mL")},
            "percentage_particles": {**_sect(116, 117, 125, 19, range(21, 25), alt_lbl_col=20), "unit_label": str(ws.cell(row=116, column=25).value or "particles x10^12/mL")},
            "data_summary": {"headers": [str(ws.cell(row=153, column=c).value) for c in range(2, 8) if ws.cell(row=153, column=c).value], "rows": [{str(ws.cell(row=153, column=c).value): (self._sf(ws.cell(row=r, column=c).value) if self._sf(ws.cell(row=r, column=c).value) is not None else str(ws.cell(row=r, column=c).value) if ws.cell(row=r, column=c).value else None) for c in range(2, 8) if ws.cell(row=153, column=c).value} for r in range(154, 165) if ws.cell(row=r, column=2).value is not None]},
        }

    # ====== STATISTICAL ANALYSIS ======
    def extract_statistical_analysis(self):
        sheets = self._find_sheets("statistical_analysis") + self._find_sheets("statistical analysis")
        if not sheets: return {"available": False}
        ws = self.wb[sheets[0]]
        group_labels = [str(ws.cell(row=5, column=c).value) for c in range(3, 7) if ws.cell(row=5, column=c).value is not None]
        exp_data = []
        for r in [6, 8, 10]:
            l = ws.cell(row=r, column=2).value
            if l: exp_data.append({"label": str(l), "values": {f"Group {c-2}": self._sf(ws.cell(row=r, column=c).value) for c in range(3, 7)}})
        grp_summary = []
        for r in range(26, 35):
            g = ws.cell(row=r, column=2).value
            if g is None: break
            grp_summary.append({"group": str(g), "count": self._sf(ws.cell(row=r, column=3).value), "sum": self._sf(ws.cell(row=r, column=4).value), "mean": self._sf(ws.cell(row=r, column=5).value), "variance": self._sf(ws.cell(row=r, column=6).value)})
        anova = []
        for r in range(34, 38):
            s = ws.cell(row=r, column=2).value
            if s is None: continue
            anova.append({"source": str(s), "ss": self._sf(ws.cell(row=r, column=3).value), "df": self._sf(ws.cell(row=r, column=4).value), "ms": self._sf(ws.cell(row=r, column=5).value), "f_stat": self._sf(ws.cell(row=r, column=6).value), "p_value": self._sf(ws.cell(row=r, column=7).value), "f_crit": self._sf(ws.cell(row=r, column=8).value)})
        return {"available": True, "group_labels": group_labels, "experiment_data": exp_data, "groups_summary": grp_summary, "anova_table": anova, "total_ss": self._sf(ws.cell(row=37, column=3).value), "total_df": self._sf(ws.cell(row=37, column=4).value), "alpha": self._sf(ws.cell(row=44, column=4).value), "is_significant": str(ws.cell(row=45, column=2).value) if ws.cell(row=45, column=2).value else None}

    def parse_all_data(self):
        try:
            return {"test_details": {"work_package": self.extract_work_package(), "material": self.extract_material(), "cell_line": self.extract_cell_line(), "dispersion": self.extract_dispersion(), "treatment": self.extract_treatment()}, "replications": self.extract_raw_data(), "processed_data": self.extract_processed_data(), "final_results": self.extract_final_results(), "statistical_analysis": self.extract_statistical_analysis()}
        except Exception as e:
            logger.error(f"Error: {e}\n{traceback.format_exc()}"); raise


def _fix_degree_symbols(obj):
    if isinstance(obj, str): return obj.replace("oC", "\u00b0C")
    if isinstance(obj, dict): return {(k.replace("oC", "\u00b0C") if isinstance(k, str) else k): _fix_degree_symbols(v) for k, v in obj.items()}
    if isinstance(obj, list): return [_fix_degree_symbols(v) for v in obj]
    return obj

def parse_excel_ros(file_path, sheet_name="Test_conditions"):
    return _fix_degree_symbols(ROSParser(file_path, sheet_name).parse_all_data())

if __name__ == "__main__":
    import sys
    fp = sys.argv[1] if len(sys.argv) > 1 else "backend/data/CMS_WP3_ROS_2a_n1_FINAL_DB.xlsx"
    d = parse_excel_ros(fp)
    print(json.dumps(d, indent=2, default=str)[:12000])