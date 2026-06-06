import openpyxl
import re
import json
import logging
import traceback
from datetime import datetime, date
from dataclasses import dataclass, asdict, field
from difflib import SequenceMatcher
from typing import List, Dict, Optional, Union

logging.basicConfig(level=logging.DEBUG)
logging.getLogger("multipart.multipart").setLevel(logging.WARNING)
logger = logging.getLogger("waterflea_parser")


# ---------------------------------------------------------------------------
# Shared dataclasses (reused naming across CheMat test types)
# ---------------------------------------------------------------------------
@dataclass
class Scientist:
    name: Optional[str] = None
    email: Optional[str] = None


@dataclass
class WorkPackageData:
    wp_name: Optional[str] = None
    partner: Optional[str] = None
    test_facility: Optional[str] = None
    full_test_name: Optional[str] = None
    oecd_iso_ref: Optional[str] = None
    test_type: Optional[str] = None
    endpoint: Optional[str] = None
    metric: Optional[str] = None
    sop: Optional[str] = None
    lead_scientists: List[Scientist] = field(default_factory=list)
    assay_scientists: List[Scientist] = field(default_factory=list)


@dataclass
class MaterialData:
    material_identifier: Optional[str] = None
    erm_id: Optional[str] = None
    material_name: Optional[str] = None
    core_chemistry: Optional[str] = None
    cas_no: Optional[str] = None
    cas_no_for_core: Optional[str] = None
    material_supplier: Optional[str] = None
    material_state: Optional[str] = None
    batch: Optional[str] = None
    vial: Optional[str] = None
    preparation_date: Optional[str] = None
    stock_concentration: Optional[str] = None
    molecular_weight: Optional[str] = None
    particles_stock: Optional[str] = None


# ---------------------------------------------------------------------------
# Test-specific dataclasses (Water flea / Daphtoxkit)
# ---------------------------------------------------------------------------
@dataclass
class WaterFleaApplication:
    start_daphnids_per_well: Optional[int] = None
    total_volume_per_replicate_ml: Optional[float] = None
    concentrations_tested: List[float] = field(default_factory=list)
    concentration_unit: Optional[str] = None
    replicates_per_concentration: Optional[int] = None


@dataclass
class WaterFleaTestConditions:
    test_medium: List[str] = field(default_factory=list)
    light_cycle_h: Optional[str] = None
    light_intensity_lux: Optional[str] = None
    illumination_direction: Optional[str] = None
    temperature_c: Optional[str] = None
    aeration: Optional[str] = None
    salinity_ppt: Optional[Union[float, str]] = None
    total_incubation_time_h: Optional[Union[float, str]] = None


@dataclass
class WaterFleaAnalysis:
    type_of_measurement: Optional[str] = None
    measurement_device: Optional[str] = None
    timepoints_h: List[Union[float, int]] = field(default_factory=list)


@dataclass
class WaterFleaReplicationMeta:
    test_start_date: Optional[str] = None
    test_end_date: Optional[str] = None
    replicates_per_concentration: Optional[int] = None


@dataclass
class WaterFleaWellObservation:
    concentration: Optional[str] = None
    time: Optional[str] = None
    well_1: Optional[str] = None
    well_2: Optional[str] = None
    well_3: Optional[str] = None
    well_4: Optional[str] = None
    total: Optional[str] = None
    mortality_percent: Optional[float] = None


@dataclass
class WaterFleaRawBlock:
    toxicant: Optional[str] = None
    well_headers: List[str] = field(default_factory=list)
    observations: List[WaterFleaWellObservation] = field(default_factory=list)


@dataclass
class WaterFleaMortalityGrid:
    concentrations: List[str] = field(default_factory=list)
    timepoints_h: List[Union[float, int]] = field(default_factory=list)
    # rows keyed by timepoint -> list of mortality % aligned to concentrations
    rows: List[Dict[str, Union[float, int, List[float]]]] = field(default_factory=list)


@dataclass
class WaterFleaFinalResults:
    mortality_grid: WaterFleaMortalityGrid = field(default_factory=WaterFleaMortalityGrid)
    validity: Optional[str] = None
    validity_criteria: Optional[str] = None
    significant_difference_to_control: Optional[str] = None
    ec50: Optional[str] = None
    ec50_description: Optional[str] = None
    reach_classification: Optional[str] = None
    reach_category: Optional[str] = None
    reach_threshold_label: Optional[str] = None
    reach_threshold_value: Optional[str] = None
    hazardous: Optional[str] = None


class WaterFleaParser:
    def __init__(self, file_path: str, sheet_name: str = "Test Details"):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.wb = openpyxl.load_workbook(file_path, data_only=True)
        self._sheet_lookup = {self._norm_sheet(s): s for s in self.wb.sheetnames}
        self.ws = self.wb[self._resolve_sheet(sheet_name)]
        self.email_regex = r"^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$"
        self.spelling_issues: List[str] = []
        logger.info("Loaded workbook %s with sheets %s", file_path, self.wb.sheetnames)

    # ------------------------------------------------------------------ utils
    @staticmethod
    def _norm_sheet(name: str) -> str:
        return re.sub(r"\s+", "", str(name)).strip().lower()

    def _resolve_sheet(self, requested: str) -> str:
        """Resolve a sheet name tolerant of trailing spaces / case / fuzzy typos."""
        key = self._norm_sheet(requested)
        if key in self._sheet_lookup:
            return self._sheet_lookup[key]
        best, best_score = None, 0.0
        for norm, actual in self._sheet_lookup.items():
            score = SequenceMatcher(None, key, norm).ratio()
            if score > best_score:
                best, best_score = actual, score
        if best is not None and best_score > 0.8:
            self.spelling_issues.append(
                f"Sheet name '{requested}' matched to '{best}' (similarity {best_score:.2f})"
            )
            logger.warning("Fuzzy-matched sheet '%s' -> '%s'", requested, best)
            return best
        raise KeyError(f"Sheet '{requested}' not found. Available: {self.wb.sheetnames}")

    def _ws(self, sheet_name: str):
        return self.wb[self._resolve_sheet(sheet_name)]

    def normalize_key(self, key: Optional[str]) -> Optional[str]:
        if key:
            normalized = str(key).strip().lower()
            normalized = re.sub(r"[^a-z0-9]", "_", normalized)
            normalized = re.sub(r"_+", "_", normalized).strip("_")
            return normalized
        return None

    def excel_date_to_string(self, value) -> Optional[str]:
        if value is None:
            return None
        try:
            if isinstance(value, (datetime, date)):
                return value.strftime("%Y-%m-%d")
            if isinstance(value, (int, float)):
                base = datetime(1899, 12, 30)
                return (base + __import__("datetime").timedelta(days=float(value))).strftime("%Y-%m-%d")
            text = str(value).strip()
            for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
                try:
                    return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
                except ValueError:
                    continue
            self.spelling_issues.append(f"Unrecognized date format: '{text}'")
            logger.warning("Unrecognized date format: %s", text)
            return text
        except Exception:
            logger.error("Date parse error for %r:\n%s", value, traceback.format_exc())
            self.spelling_issues.append(f"Malformed date value: '{value}'")
            return None

    def _kv_map(self, ws) -> Dict[str, object]:
        """Build a normalized {label_key: value(colB)} map from a label/value sheet."""
        result: Dict[str, object] = {}
        for r in range(1, ws.max_row + 1):
            label = ws.cell(row=r, column=1).value
            key = self.normalize_key(label)
            if key and key not in result:
                result[key] = ws.cell(row=r, column=2).value
        return result

    @staticmethod
    def _to_number(value):
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return value
        text = str(value).strip().replace(",", ".")
        try:
            return float(text) if "." in text else int(text)
        except ValueError:
            return value

    @staticmethod
    def _percent_from_fraction(fraction) -> Optional[float]:
        """Convert a 'n/m' deceased-fraction string into a mortality percentage."""
        if fraction is None:
            return None
        text = str(fraction).strip()
        m = re.match(r"^\s*(\d+(?:\.\d+)?)\s*/\s*(\d+(?:\.\d+)?)\s*$", text)
        if not m:
            return None
        num, den = float(m.group(1)), float(m.group(2))
        if den == 0:
            return None
        return round(num / den * 100, 4)

    # --------------------------------------------------------- column A lookup
    def _row_for_label(self, ws, expected_label: str, threshold: float = 0.8) -> Optional[int]:
        """Find the row whose column-A label matches expected_label (fuzzy)."""
        target = self.normalize_key(expected_label)
        best_row, best_score = None, 0.0
        for r in range(1, ws.max_row + 1):
            key = self.normalize_key(ws.cell(row=r, column=1).value)
            if not key:
                continue
            if key == target:
                return r
            score = SequenceMatcher(None, target, key).ratio()
            if score > best_score:
                best_row, best_score = r, score
        if best_row is not None and best_score >= threshold:
            return best_row
        return None

    def _collect_below(self, ws, start_row: int, col: int = 2, stop_col1: bool = True) -> List[object]:
        """Collect column values starting at start_row, stopping at the next labelled row."""
        values = []
        r = start_row
        while r <= ws.max_row:
            v = ws.cell(row=r, column=col).value
            label = ws.cell(row=r, column=1).value
            if r != start_row and stop_col1 and label not in (None, ""):
                break
            if v is not None:
                values.append(v)
            r += 1
        return values

    # ----------------------------------------------------------- extractions
    def extract_work_package_data(self) -> Dict:
        ws = self.ws
        kv = self._kv_map(ws)
        lead, assay = [], []
        for r in range(1, ws.max_row + 1):
            key = self.normalize_key(ws.cell(row=r, column=1).value)
            value = ws.cell(row=r, column=2).value
            email = ws.cell(row=r, column=4).value
            if not key:
                continue
            if "lead_scientist" in key and value:
                lead.append(asdict(Scientist(name=value, email=email)))
            if "assay_test_work_conducted_by" in key and value:
                assay.append(asdict(Scientist(name=value, email=email)))

        def g(*candidates):
            for c in candidates:
                if c in kv and kv[c] is not None:
                    return kv[c]
            return None

        wp = WorkPackageData(
            wp_name=g("project_work_package"),
            partner=g("partner_conducting_test_assay"),
            test_facility=g("test_facility_laboratory_name"),
            full_test_name=g("full_name_of_test"),
            oecd_iso_ref=g("oecd_iso_test_ref_id_if_app"),
            test_type=g("type_or_class_of_experimental_test_as_used_here"),
            endpoint=g("end_point_being_investigated_assessed_by_the_test"),
            metric=g("metric_s_used_to_assess_end_point_outcome_response"),
            sop=g("sop_s_for_test_ref_project_or_other_doc_title_id"),
            lead_scientists=lead,
            assay_scientists=assay,
        )
        return asdict(wp)

    def extract_material_data(self) -> Dict:
        kv = self._kv_map(self.ws)

        def g(*candidates):
            for c in candidates:
                if c in kv and kv[c] is not None:
                    return kv[c]
            return None

        material = MaterialData(
            material_identifier=g("sample_cms_internal_identifier"),
            erm_id=g("erm_identifier_number"),
            material_name=g("material_name"),
            core_chemistry=g("core_chemistry"),
            cas_no=g("cas_no"),
            cas_no_for_core=g("cas_no_for_core"),
            material_supplier=g("material_supplier"),
            material_state=g("material_state"),
            batch=g("batch"),
            vial=g("vial"),
            preparation_date=self.excel_date_to_string(g("date_of_preparation")),
            stock_concentration=g("stock_concentration"),
            molecular_weight=g("molecular_weight"),
            particles_stock=g("no_of_particles_in_stock"),
        )
        return asdict(material)

    def extract_application_data(self) -> Dict:
        ws = self.ws
        conc_row = self._row_for_label(ws, "Concentrations tested (mg·l−1)") \
            or self._row_for_label(ws, "Concentrations tested")
        concentrations = []
        unit = None
        if conc_row:
            label = ws.cell(row=conc_row, column=1).value or ""
            m = re.search(r"\(([^)]+)\)", str(label))
            if m:
                unit = m.group(1)
            concentrations = [self._to_number(v) for v in self._collect_below(ws, conc_row, col=2)]

        kv = self._kv_map(ws)
        app = WaterFleaApplication(
            start_daphnids_per_well=self._to_number(kv.get("start_number_of_daphnids_per_well")),
            total_volume_per_replicate_ml=self._to_number(kv.get("total_volume_per_replicate_ml")),
            concentrations_tested=concentrations,
            concentration_unit=unit,
            replicates_per_concentration=self._to_number(kv.get("replicates_per_concentration")),
        )
        return asdict(app)

    def extract_test_conditions_data(self) -> Dict:
        ws = self.ws
        medium_row = self._row_for_label(ws, "Test medium")
        medium = []
        if medium_row:
            medium = [str(v) for v in self._collect_below(ws, medium_row, col=2)]
        kv = self._kv_map(ws)
        cond = WaterFleaTestConditions(
            test_medium=medium,
            light_cycle_h=kv.get("light_cycle_h"),
            light_intensity_lux=kv.get("light_intensity_lux"),
            illumination_direction=kv.get("illumination_direction"),
            temperature_c=kv.get("temperature_c"),
            aeration=kv.get("aeration"),
            salinity_ppt=self._to_number(kv.get("salinity_ppt")),
            total_incubation_time_h=self._to_number(kv.get("total_incubation_time_h")),
        )
        return asdict(cond)

    def extract_analysis_data(self) -> Dict:
        ws = self.ws
        tp_row = self._row_for_label(ws, "Timepoints (h)")
        timepoints = []
        if tp_row:
            timepoints = [self._to_number(v) for v in self._collect_below(ws, tp_row, col=2)]
        kv = self._kv_map(ws)
        analysis = WaterFleaAnalysis(
            type_of_measurement=kv.get("type_of_measurement"),
            measurement_device=kv.get("measurement_device"),
            timepoints_h=timepoints,
        )
        return asdict(analysis)

    def extract_replication(self) -> Dict:
        kv = self._kv_map(self.ws)
        rep = WaterFleaReplicationMeta(
            test_start_date=self.excel_date_to_string(kv.get("test_start_date")),
            test_end_date=self.excel_date_to_string(kv.get("test_end_date")),
            replicates_per_concentration=self._to_number(kv.get("replicates_per_concentration")),
        )
        return asdict(rep)

    def extract_raw_data(self, sheet_name: str = "Raw Data") -> List[Dict]:
        ws = self._ws(sheet_name)
        # toxicant label
        tox_row = self._row_for_label(ws, "Toxicants tested:") or self._row_for_label(ws, "Toxicants tested")
        toxicant = ws.cell(row=tox_row, column=2).value if tox_row else None

        # locate header row: the row whose column C/onwards contains the well labels
        header_row = None
        for r in range(1, ws.max_row + 1):
            c = ws.cell(row=r, column=3).value
            if c and re.match(r"^\s*well", str(c), re.IGNORECASE):
                header_row = r
                break
        well_headers, data_start, max_data_col = [], None, 8
        if header_row:
            for col in range(3, ws.max_column + 1):
                v = ws.cell(row=header_row, column=col).value
                if v is not None:
                    well_headers.append(str(v).strip())
            data_start = header_row + 1

        blocks: Dict[str, WaterFleaRawBlock] = {}
        order: List[str] = []
        current_conc = None
        if data_start:
            for r in range(data_start, ws.max_row + 1):
                conc = ws.cell(row=r, column=1).value
                time = ws.cell(row=r, column=2).value
                if conc is not None:
                    current_conc = str(conc).strip()
                if time is None and conc is None:
                    continue
                obs = WaterFleaWellObservation(
                    concentration=current_conc,
                    time=str(time).strip() if time is not None else None,
                    well_1=self._cell_str(ws.cell(row=r, column=3).value),
                    well_2=self._cell_str(ws.cell(row=r, column=4).value),
                    well_3=self._cell_str(ws.cell(row=r, column=5).value),
                    well_4=self._cell_str(ws.cell(row=r, column=6).value),
                    total=self._cell_str(ws.cell(row=r, column=7).value),
                    mortality_percent=self._percent_from_fraction(ws.cell(row=r, column=7).value),
                )
                key = toxicant or "block"
                if key not in blocks:
                    blocks[key] = WaterFleaRawBlock(toxicant=key, well_headers=well_headers)
                    order.append(key)
                blocks[key].observations.append(obs)

        return [asdict(blocks[k]) for k in order]

    @staticmethod
    def _cell_str(value) -> Optional[str]:
        if value is None:
            return None
        return str(value).strip()

    def extract_final_results(self, sheet_name: str = "Final Results") -> Dict:
        ws = self._ws(sheet_name)

        # mortality grid: header row of concentrations + Time(h)/Mortality(%) rows
        conc_header_row = None
        for r in range(1, ws.max_row + 1):
            row_vals = [ws.cell(row=r, column=c).value for c in range(3, ws.max_column + 1)]
            if any(isinstance(v, str) and ("control" in v.lower() or "mg" in v.lower()) for v in row_vals):
                conc_header_row = r
                break

        concentrations: List[str] = []
        if conc_header_row:
            for c in range(3, ws.max_column + 1):
                v = ws.cell(row=conc_header_row, column=c).value
                if v is not None:
                    concentrations.append(str(v).strip())

        # find "Time (h)" anchor in column B
        time_anchor = self._find_in_col(ws, "Time (h)", col=2)
        timepoints: List[Union[float, int]] = []
        rows: List[Dict] = []
        if time_anchor:
            r = time_anchor + 1
            while r <= ws.max_row:
                t = ws.cell(row=r, column=2).value
                if t is None or not isinstance(t, (int, float)):
                    break
                values = []
                for c in range(3, 3 + len(concentrations)):
                    values.append(self._to_number(ws.cell(row=r, column=c).value))
                timepoints.append(self._to_number(t))
                rows.append({"time_h": self._to_number(t), "mortality_percent": values})
                r += 1

        grid = WaterFleaMortalityGrid(
            concentrations=concentrations,
            timepoints_h=timepoints,
            rows=rows,
        )

        # validity
        validity_row = self._find_in_col(ws, "Validity of the test", col=2)
        validity = ws.cell(row=validity_row, column=4).value if validity_row else None
        validity_criteria = ws.cell(row=validity_row + 1, column=2).value if validity_row else None

        # significant difference
        sig_row = self._find_in_col(ws, "Significant difference to control", col=2)
        significance = ws.cell(row=sig_row + 1, column=2).value if sig_row else None

        # EC50
        ec50_row = self._find_in_col(ws, "Calculated EC50", col=2)
        ec50_desc = ws.cell(row=ec50_row, column=4).value if ec50_row else None
        ec50 = ws.cell(row=ec50_row + 1, column=2).value if ec50_row else None

        # REACH (search subsequent labels strictly after the REACH header)
        reach_row = self._find_in_col(ws, "Classification by REACH", col=2)
        reach_classification = ws.cell(row=reach_row, column=2).value if reach_row else None
        reach_category = ws.cell(row=reach_row + 1, column=2).value if reach_row else None
        thr_row = (
            self._find_in_col(ws, "EC50", col=2, start=reach_row + 1) if reach_row else None
        )
        threshold_label = ws.cell(row=thr_row, column=2).value if thr_row else None
        threshold_value = ws.cell(row=thr_row, column=4).value if thr_row else None
        haz_row = (
            self._find_in_col(ws, "HAZARDOUS?", col=2, start=reach_row + 1) if reach_row else None
        )
        hazardous = ws.cell(row=haz_row, column=4).value if haz_row else None

        results = WaterFleaFinalResults(
            mortality_grid=grid,
            validity=self._cell_str(validity),
            validity_criteria=self._cell_str(validity_criteria),
            significant_difference_to_control=self._cell_str(significance),
            ec50=self._cell_str(ec50),
            ec50_description=self._cell_str(ec50_desc),
            reach_classification=self._cell_str(reach_classification),
            reach_category=self._cell_str(reach_category),
            reach_threshold_label=self._cell_str(threshold_label),
            reach_threshold_value=self._cell_str(threshold_value),
            hazardous=self._cell_str(hazardous),
        )
        return asdict(results)

    def _find_in_col(self, ws, needle: str, col: int = 2, start: int = 1) -> Optional[int]:
        needle_low = needle.lower()
        for r in range(start, ws.max_row + 1):
            v = ws.cell(row=r, column=col).value
            if v is not None and needle_low in str(v).lower():
                return r
        return None

    # ------------------------------------------------------------ orchestrate
    def parse_all_data(self) -> Dict:
        work_package = self.extract_work_package_data()
        material = self.extract_material_data()
        application = self.extract_application_data()
        test_conditions = self.extract_test_conditions_data()
        analysis = self.extract_analysis_data()
        replication = self.extract_replication()
        raw_blocks = self.extract_raw_data()
        final_results = self.extract_final_results()

        parsed_data = {
            "test_details": {
                "work_package": work_package,
                "material": material,
                "application": application,
                "test_conditions": test_conditions,
                "analysis": analysis,
                "replications": [replication],
                "parser_warnings": self.spelling_issues,
            },
            "replications": raw_blocks,     # raw data blocks (API maps to raw_data)
            "processed_data": {},           # this test type has no processed-data stage
            "final_results": final_results,
        }
        return parsed_data


def parse_excel_waterflea(file_path: str, sheet_name: str = "Test Details") -> Dict:
    parser = WaterFleaParser(file_path, sheet_name)
    return parser.parse_all_data()


if __name__ == "__main__":
    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else "CMS_WP3_Water_flea_2a_DB.xlsx"
    parsed = parse_excel_waterflea(path)
    print("Parsed Data Keys:", list(parsed.keys()))
    print("\n--- Test Details (work_package) ---")
    print(json.dumps(parsed["test_details"]["work_package"], indent=2, default=str))
    print("\n--- Material ---")
    print(json.dumps(parsed["test_details"]["material"], indent=2, default=str))
    print("\n--- Application ---")
    print(json.dumps(parsed["test_details"]["application"], indent=2, default=str))
    print("\n--- Raw Data blocks:", len(parsed["replications"]))
    if parsed["replications"]:
        print(json.dumps(parsed["replications"][0]["observations"][:2], indent=2, default=str))
    print("\n--- Final Results ---")
    print(json.dumps(parsed["final_results"], indent=2, default=str))
    print("\n--- Parser warnings:", parsed["test_details"]["parser_warnings"])