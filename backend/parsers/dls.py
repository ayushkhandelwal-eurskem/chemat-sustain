import openpyxl
import re
from dataclasses import dataclass, asdict, field
from typing import List, Dict, Optional, Union
import logging
import traceback
from datetime import datetime, timedelta
from difflib import SequenceMatcher

# Set up logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
logging.getLogger("multipart.multipart").setLevel(logging.WARNING)
logger = logging.getLogger(__name__)

# Dataclasses (unchanged, included for completeness)
@dataclass
class Scientist:
    name: Optional[str] = None
    email: Optional[str] = None

@dataclass
class WorkPackageData:
    wp_name: Optional[str] = None
    partner: Optional[str] = None
    full_test_name: Optional[str] = None
    test_acronym: Optional[str] = None
    test_type: Optional[str] = None
    endpoint: Optional[str] = None
    endpoint_outcome: Optional[str] = None
    sop: Optional[str] = None
    path: Optional[str] = None
    lead_scientists: List[Scientist] = field(default_factory=list)
    assay_scientists: List[Scientist] = field(default_factory=list)

@dataclass
class MaterialData:
    material_identifier: Optional[str] = None
    erm_id: Optional[str] = None
    core_chemistry: Optional[str] = None
    material_name: Optional[str] = None
    material_state: Optional[str] = None
    cas: Optional[str] = None
    casforcore: Optional[str] = None
    batch: Optional[str] = None
    preparation_date: Optional[str] = None
    particles_stock: Optional[str] = None
    molar_concentration: Optional[str] = None

@dataclass
class ReplicationData:
    test_identifier_number: Optional[str] = None
    test_start_date: Optional[str] = None
    test_end_date: Optional[str] = None

@dataclass
class SamplePreparationData:
    dispersion_protocol: Optional[str] = None
    dispersion_technique: Optional[str] = None
    dispersion_medium: Optional[str] = None
    sonicator_type: Optional[str] = None
    power: Optional[str] = None
    sonication_time: Optional[str] = None
    tip_thickness: Optional[str] = None
    tip_composition: Optional[str] = None
    ultrasonic_bath_size: Optional[str] = None
    sample_volume: Optional[str] = None
    final_concentration: Optional[str] = None
    additional_info: Optional[str] = None

@dataclass
class InstrumentationData:
    instrument_model: Optional[str] = None
    cell_model: Optional[str] = None
    temperature: Optional[str] = None
    thermal_equilibrium_time: Optional[str] = None
    number_of_runs: Optional[int] = None
    sub_runs: Optional[str] = None
    delay_between_runs: Optional[str] = None
    run_duration: Optional[str] = None
    laser_focus_position: Optional[str] = None
    scattering_angle: Optional[str] = None
    data_analysis_model: Optional[str] = None
    laser_attenuation: Optional[str] = None
    refractive_index_nm: Optional[float] = None
    absorption_index_nm: Optional[float] = None
    refractive_index_medium: Optional[float] = None
    viscosity_medium: Optional[str] = None

@dataclass
class CorrelationData:
    time_us: List[float] = field(default_factory=list)
    correlation_coefficient: List[float] = field(default_factory=list)

@dataclass
class SizeDistributionData:
    size_nm: List[float] = field(default_factory=list)
    mean_intensity_percent: List[float] = field(default_factory=list)

@dataclass
class RunMetrics:
    z_ave_hydrodynamic_diameter: Optional[float] = None
    pdi: Optional[float] = None
    peak_1_diameter: Optional[float] = None
    std_dev_peak_1: Optional[float] = None
    peak_1_intensity: Optional[float] = None
    peak_2_diameter: Optional[float] = None
    std_dev_peak_2: Optional[float] = None
    peak_2_intensity: Optional[float] = None
    peak_3_diameter: Optional[float] = None
    std_dev_peak_3: Optional[float] = None
    peak_3_intensity: Optional[float] = None
    derived_count_rate: Optional[float] = None

@dataclass
class RawData:
    run_number: int
    correlation_data: CorrelationData
    size_distribution: SizeDistributionData
    metrics: RunMetrics

@dataclass
class StatisticTableEntry:
    size_nm: float
    mean_intensity_percent: float
    std_dev: float

@dataclass
class ResultsData:
    z_ave_hydrodynamic_diameter: Optional[float] = None
    uncertainty_hydrodynamic_diameter: Optional[float] = None
    pdi: Optional[float] = None
    uncertainty_pdi: Optional[float] = None
    mean_peak_1_diameter: Optional[float] = None
    pooled_std_dev_peak_1: Optional[float] = None
    std_dev_between_measurements_peak_1: Optional[float] = None
    mean_peak_1_intensity: Optional[float] = None
    mean_peak_2_diameter: Optional[float] = None
    pooled_std_dev_peak_2: Optional[float] = None
    std_dev_between_measurements_peak_2: Optional[float] = None
    mean_peak_2_intensity: Optional[float] = None
    mean_peak_3_diameter: Optional[float] = None
    pooled_std_dev_peak_3: Optional[float] = None
    std_dev_between_measurements_peak_3: Optional[float] = None
    mean_peak_3_intensity: Optional[float] = None
    derived_count_rate: Optional[float] = None
    statistic_table: List[StatisticTableEntry] = field(default_factory=list)

class DLSParser:
    def __init__(self, file_path: str, sheet_name: str = "Test Information"):
        self.file_path = file_path
        self.sheet_name = sheet_name
        try:
            self.wb = openpyxl.load_workbook(file_path, data_only=True)
            self.ws = self.wb[sheet_name]
            logger.info(f"Successfully loaded Excel file: {file_path}")
        except Exception as e:
            logger.error(f"Failed to load workbook or sheet {sheet_name}: {e}")
            raise
        self.email_regex = r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$'

    def normalize_key(self, key: Optional[str]) -> Optional[str]:
        if key:
            normalized = key.strip().lower()
            normalized = re.sub(r'[^a-zA-Z0-9]', '_', normalized)
            normalized = re.sub(r'_+', '_', normalized)
            return normalized
        return None

    def split_value_unit(self, value: Optional[str]) -> tuple[Optional[Union[str, float]], Optional[str]]:
        if isinstance(value, str) and " " in value:
            parts = value.split(" ", 1)
            try:
                numeric = float(parts[0]) if '.' in parts[0] else int(parts[0])
                return numeric, parts[1]
            except (ValueError, IndexError):
                return parts[0], parts[1] if len(parts) > 1 else None
        return value, None

    def excel_date_to_string(self, value: Optional[Union[float, str]]) -> Optional[str]:
        try:
            if isinstance(value, (int, float)):
                base_date = datetime(1899, 12, 30)
                delta = timedelta(days=float(value))
                return (base_date + delta).strftime("%Y-%m-%d")
            elif isinstance(value, str):
                for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"]:
                    try:
                        return datetime.strptime(value, fmt).strftime("%Y-%m-%d")
                    except ValueError:
                        continue
            return str(value) if value else None
        except Exception as e:
            logger.warning(f"Failed to convert date {value}: {e}")
            return None

    def extract_work_package_data(self):
        data = []
        lead_scientists = []
        assay_scientists = []

        for row in self.ws.iter_rows():
            key_cell = row[0].value
            value_cell = row[1].value if len(row) > 1 else None
            email_cell = row[3].value if len(row) > 3 else None
            comment = row[0].comment

            if comment:
                key = self.normalize_key(key_cell)
                if not key:
                    continue
                entry = {"Key": key, "Value": value_cell}

                if email_cell and re.match(self.email_regex, str(email_cell)):
                    entry["Email"] = email_cell

                if key and "lead_scientist_contact_for_test_" in key:
                    lead_scientists.append(asdict(Scientist(name=value_cell, email=email_cell)))
                if key and "assay_test_work_conducted_by_" in key:
                    assay_scientists.append(asdict(Scientist(name=value_cell, email=email_cell)))

                data.append(entry)

        wp_data = asdict(WorkPackageData(
            wp_name=next((d["Value"] for d in data if d["Key"] == "project_work_package_"), None),
            partner=next((d["Value"] for d in data if d["Key"] == "partner_conducting_test_assay_"), None),
            full_test_name=next((d["Value"] for d in data if d["Key"] == "full_name_of_test_assay_add_oecd_test_ref_id_if_app_"), None),
            test_acronym=next((d["Value"] for d in data if d["Key"] == "short_name_or_acronym_for_test_assay_"), None),
            test_type=next((d["Value"] for d in data if d["Key"] == "type_or_class_of_experimental_test_as_used_here_"), None),
            endpoint=next((d["Value"] for d in data if d["Key"] == "end_point_being_investigated_assessed_by_the_test_"), None),
            endpoint_outcome=next((d["Value"] for d in data if d["Key"] == "metric_s_used_to_assess_end_point_outcome_response_"), None),
            sop=next((d["Value"] for d in data if d["Key"] == "sop_s_for_test_ref_project_or_other_doc_title_id_"), None),
            path=next((d["Value"] for d in data if d["Key"] == "path_link_to_sop_protocol_on_proj_server_web_where_applic_"), None),
            lead_scientists=lead_scientists,
            assay_scientists=assay_scientists
        ))
        return wp_data

    def extract_material_data(self):
        data = []
        expected_keys = [
            "sample_cms_internal_identifier",
            "erm_identifier_number",
            "core_chemistry_",
            "material_name_",
            "material_state_",
            "cas_no",
            "cas_for_core_",
            "batch",
            "date_of_preparation_",
            "no_of_particles_in_stock_",
            "molar_concentration"
        ]
        unmatched_keys = set(expected_keys)
        potential_matches = []

        logger.debug("Test Information sheet (columns A:E, rows 1:50) for Material Data:")
        for row_idx in range(1, min(self.ws.max_row + 1, 51)):
            row = self.ws[row_idx]
            values = [str(cell.value)[:30] if cell.value else "None" for cell in row[:5]]
            has_comment = "Y" if row[0].comment else "N"
            logger.debug(f"Row {row_idx} [Comment: {has_comment}]: {values}")

        for row_idx, row in enumerate(self.ws.iter_rows(min_row=1, max_col=5), start=1):
            key_cell = row[0].value
            if not key_cell:
                continue
            raw_key = str(key_cell)
            key = self.normalize_key(raw_key)
            if not key:
                continue

            value = None
            value_col = None
            for col_idx in range(1, 5):
                if col_idx < len(row) and row[col_idx].value is not None:
                    value = row[col_idx].value
                    value_col = col_idx + 1
                    break

            logger.debug(f"Row {row_idx}: Raw Key='{raw_key}', Normalized Key='{key}', Value='{value}', Value Col={value_col}")

            if key in expected_keys:
                unmatched_keys.discard(key)
                data.append({"Key": key, "Value": value})
                logger.debug(f"Exact match: Key='{key}', Value='{value}' at row {row_idx}, col {value_col}")
            else:
                for expected_key in expected_keys:
                    similarity = SequenceMatcher(None, key, expected_key).ratio()
                    if similarity > 0.85:
                        unmatched_keys.discard(expected_key)
                        data.append({"Key": expected_key, "Value": value})
                        logger.debug(f"Fuzzy match: Key='{key}' matched '{expected_key}' (similarity={similarity:.2f}), Value='{value}' at row {row_idx}, col {value_col}")
                    elif similarity > 0.5:
                        potential_matches.append(f"Potential match for '{expected_key}': '{raw_key}' at row {row_idx}, col 1, similarity={similarity:.2f}")

        if potential_matches:
            logger.debug("Potential key matches for Material Data:")
            for match in potential_matches:
                logger.debug(match)
        if unmatched_keys:
            logger.warning(f"Unmatched material data keys: {unmatched_keys}")

        material_data = asdict(MaterialData(
            material_identifier=next((d["Value"] for d in data if d["Key"] == "sample_cms_internal_identifier"), None),
            erm_id=next((d["Value"] for d in data if d["Key"] == "erm_identifier_number"), None),
            core_chemistry=next((d["Value"] for d in data if d["Key"] == "core_chemistry_"), None),
            material_name=next((d["Value"] for d in data if d["Key"] == "material_name_"), None),
            material_state=next((d["Value"] for d in data if d["Key"] == "material_state_"), None),
            cas=next((d["Value"] for d in data if d["Key"] == "cas_no"), None),
            casforcore=next((d["Value"] for d in data if d["Key"] == "cas_for_core_"), None),
            batch=next((d["Value"] for d in data if d["Key"] == "batch"), None),
            preparation_date=self.excel_date_to_string(next((d["Value"] for d in data if d["Key"] == "date_of_preparation_"), None)),
            particles_stock=next((d["Value"] for d in data if d["Key"] == "no_of_particles_in_stock_"), None),
            molar_concentration=next((d["Value"] for d in data if d["Key"] == "molar_concentration"), None)
        ))
        logger.debug(f"Material data extracted: {material_data}")
        return material_data

    def extract_sample_preparation_data(self):
        data = []
        expected_keys = [
            "specify_standard_dispersion_protocol_used_",
            "or_otherwise_specify_dispersion_technique_used_",
            "dispersion_dilution_medium",
            "sonicator_type",
            "power_w_",
            "sonication_time_secs_",
            "tip_thickness_mm_",
            "tip_composition",
            "size_of_ultrasonic_bath_water_volume_dm3_",
            "sample_volume",
            "final_sample_concentration_mg_l_or_ppm_",
            "additional_information"
        ]
        unmatched_keys = set(expected_keys)
        potential_matches = []

        logger.debug("Test Information sheet (columns A:E, rows 1:50) for Sample Preparation Data:")
        for row_idx in range(1, min(self.ws.max_row + 1, 51)):
            row = self.ws[row_idx]
            values = [str(cell.value)[:30] if cell.value else "None" for cell in row[:5]]
            has_comment = "Y" if row[0].comment else "N"
            logger.debug(f"Row {row_idx} [Comment: {has_comment}]: {values}")

        for row_idx, row in enumerate(self.ws.iter_rows(min_row=1, max_col=5), start=1):
            key_cell = row[0].value
            if not key_cell:
                continue
            raw_key = str(key_cell)
            key = self.normalize_key(raw_key)
            if not key:
                continue

            value = None
            value_col = None
            for col_idx in range(1, 5):
                if col_idx < len(row) and row[col_idx].value is not None:
                    value = row[col_idx].value
                    value_col = col_idx + 1
                    break

            logger.debug(f"Row {row_idx}: Raw Key='{raw_key}', Normalized Key='{key}', Value='{value}', Value Col={value_col}")

            if key in expected_keys:
                unmatched_keys.discard(key)
                data.append({"Key": key, "Value": value})
                logger.debug(f"Exact match: Key='{key}', Value='{value}' at row {row_idx}, col {value_col}")
            else:
                for expected_key in expected_keys:
                    similarity = SequenceMatcher(None, key, expected_key).ratio()
                    if similarity > 0.85:
                        unmatched_keys.discard(expected_key)
                        data.append({"Key": expected_key, "Value": value})
                        logger.debug(f"Fuzzy match: Key='{key}' matched '{expected_key}' (similarity={similarity:.2f}), Value='{value}' at row {row_idx}, col {value_col}")
                    elif similarity > 0.5:
                        potential_matches.append(f"Potential match for '{expected_key}': '{raw_key}' at row {row_idx}, col 1, similarity={similarity:.2f}")

        if potential_matches:
            logger.debug("Potential key matches for Sample Preparation Data:")
            for match in potential_matches:
                logger.debug(match)
        if unmatched_keys:
            logger.warning(f"Unmatched sample preparation data keys: {unmatched_keys}")

        sample_preparation_data = asdict(SamplePreparationData(
            dispersion_protocol=next((d["Value"] for d in data if d["Key"] == "specify_standard_dispersion_protocol_used_"), None),
            dispersion_technique=next((d["Value"] for d in data if d["Key"] == "or_otherwise_specify_dispersion_technique_used_"), None),
            dispersion_medium=next((d["Value"] for d in data if d["Key"] == "dispersion_dilution_medium"), None),
            sonicator_type=next((d["Value"] for d in data if d["Key"] == "sonicator_type"), None),
            power=next((d["Value"] for d in data if d["Key"] == "power_w_"), None),
            sonication_time=next((d["Value"] for d in data if d["Key"] == "sonication_time_secs_"), None),
            tip_thickness=next((d["Value"] for d in data if d["Key"] == "tip_thickness_mm_"), None),
            tip_composition=next((d["Value"] for d in data if d["Key"] == "tip_composition"), None),
            ultrasonic_bath_size=next((d["Value"] for d in data if d["Key"] == "size_of_ultrasonic_bath_water_volume_dm3_"), None),
            sample_volume=next((d["Value"] for d in data if d["Key"] == "sample_volume"), None),
            final_concentration=next((d["Value"] for d in data if d["Key"] == "final_sample_concentration_mg_l_or_ppm_"), None),
            additional_info=next((d["Value"] for d in data if d["Key"] == "additional_information"), None)
        ))
        logger.debug(f"Sample preparation data extracted: {sample_preparation_data}")
        return sample_preparation_data

    def extract_instrumentation_data(self):
        data = []
        expected_keys = [
            "dls_instrument_specifications",
            "cell_model",
            "temperature",
            "thermal_equilibrium_time",
            "replication",
            "number_of_sub_runs",
            "delay_between_runs",
            "duration_of_the_run",
            "laser_focus_position",
            "scattering_angle",
            "data_analysis_model",
            "laser_attenuation",
            "refractive_index_of_the_nm",
            "absorption_index_of_the_nm",
            "refractive_index_of_the_medium",
            "viscosity_of_the_medium"
        ]
        unmatched_keys = set(expected_keys)
        potential_matches = []

        logger.debug("Test Information sheet (columns A:E, rows 1:50) for Instrumentation Data:")
        for row_idx in range(1, min(self.ws.max_row + 1, 51)):
            row = self.ws[row_idx]
            values = [str(cell.value)[:30] if cell.value else "None" for cell in row[:5]]
            has_comment = "Y" if row[0].comment else "N"
            logger.debug(f"Row {row_idx} [Comment: {has_comment}]: {values}")

        for row_idx, row in enumerate(self.ws.iter_rows(min_row=1, max_col=5), start=1):
            key_cell = row[0].value
            if not key_cell:
                continue
            raw_key = str(key_cell)
            key = self.normalize_key(raw_key)
            if not key:
                continue

            value = None
            value_col = None
            for col_idx in range(1, 5):
                if col_idx < len(row) and row[col_idx].value is not None:
                    value = row[col_idx].value
                    value_col = col_idx + 1
                    break

            logger.debug(f"Row {row_idx}: Raw Key='{raw_key}', Normalized Key='{key}', Value='{value}', Value Col={value_col}")

            if key in expected_keys:
                unmatched_keys.discard(key)
                data.append({"Key": key, "Value": value})
                logger.debug(f"Exact match: Key='{key}', Value='{value}' at row {row_idx}, col {value_col}")
            else:
                for expected_key in expected_keys:
                    similarity = SequenceMatcher(None, key, expected_key).ratio()
                    if similarity > 0.85:
                        unmatched_keys.discard(expected_key)
                        data.append({"Key": expected_key, "Value": value})
                        logger.debug(f"Fuzzy match: Key='{key}' matched '{expected_key}' (similarity={similarity:.2f}), Value='{value}' at row {row_idx}, col {value_col}")
                    elif similarity > 0.5:
                        potential_matches.append(f"Potential match for '{expected_key}': '{raw_key}' at row {row_idx}, col 1, similarity={similarity:.2f}")

        if potential_matches:
            logger.debug("Potential key matches for Instrumentation Data:")
            for match in potential_matches:
                logger.debug(match)
        if unmatched_keys:
            logger.warning(f"Unmatched instrumentation data keys: {unmatched_keys}")

        try:
            number_of_runs_value = next((d["Value"] for d in data if d["Key"] == "replication"), None)
            number_of_runs = int(number_of_runs_value) if number_of_runs_value is not None else None
        except (ValueError, TypeError) as e:
            logger.warning(f"Failed to convert number_of_runs '{number_of_runs_value}': {e}")
            number_of_runs = None

        try:
            ri_nm_value = next((d["Value"] for d in data if d["Key"] == "refractive_index_of_the_nm"), None)
            refractive_index_nm = float(self.split_value_unit(str(ri_nm_value))[0]) if ri_nm_value is not None else None
        except (ValueError, TypeError) as e:
            logger.warning(f"Failed to convert refractive_index_nm '{ri_nm_value}': {e}")
            refractive_index_nm = None

        try:
            ai_nm_value = next((d["Value"] for d in data if d["Key"] == "absorption_index_of_the_nm"), None)
            absorption_index_nm = float(self.split_value_unit(str(ai_nm_value))[0]) if ai_nm_value is not None else None
        except (ValueError, TypeError) as e:
            logger.warning(f"Failed to convert absorption_index_nm '{ai_nm_value}': {e}")
            absorption_index_nm = None

        try:
            ri_medium_value = next((d["Value"] for d in data if d["Key"] == "refractive_index_of_the_medium"), None)
            refractive_index_medium = float(self.split_value_unit(str(ri_medium_value))[0]) if ri_medium_value is not None else None
        except (ValueError, TypeError) as e:
            logger.warning(f"Failed to convert refractive_index_medium '{ri_medium_value}': {e}")
            refractive_index_medium = None

        instrumentation_data = asdict(InstrumentationData(
            instrument_model=next((d["Value"] for d in data if d["Key"] == "dls_instrument_specifications"), None),
            cell_model=next((d["Value"] for d in data if d["Key"] == "cell_model"), None),
            temperature=next((d["Value"] for d in data if d["Key"] == "temperature"), None),
            thermal_equilibrium_time=next((d["Value"] for d in data if d["Key"] == "thermal_equilibrium_time"), None),
            number_of_runs=number_of_runs,
            sub_runs=next((d["Value"] for d in data if d["Key"] == "number_of_sub_runs"), None),
            delay_between_runs=next((d["Value"] for d in data if d["Key"] == "delay_between_runs"), None),
            run_duration=next((d["Value"] for d in data if d["Key"] == "duration_of_the_run"), None),
            laser_focus_position=next((d["Value"] for d in data if d["Key"] == "laser_focus_position"), None),
            scattering_angle=next((d["Value"] for d in data if d["Key"] == "scattering_angle"), None),
            data_analysis_model=next((d["Value"] for d in data if d["Key"] == "data_analysis_model"), None),
            laser_attenuation=next((d["Value"] for d in data if d["Key"] == "laser_attenuation"), None),
            refractive_index_nm=refractive_index_nm,
            absorption_index_nm=absorption_index_nm,
            refractive_index_medium=refractive_index_medium,
            viscosity_medium=next((d["Value"] for d in data if d["Key"] == "viscosity_of_the_medium"), None)
        ))
        logger.debug(f"Instrumentation data extracted: {instrumentation_data}")
        return instrumentation_data

    def extract_replications(self):
        replications = []
        raw_sheets = [name for name in self.wb.sheetnames
          if re.match(r'WP(\d+)_DLS_(\d+)aR(\d+)', name, re.IGNORECASE)]
        if not raw_sheets:
            logger.warning("No raw data sheets found matching expected patterns.")
            return replications
        
        logger.debug(f"Found raw data sheets: {raw_sheets}")
        
        for raw_sheet in raw_sheets:
            match = re.search(r'WP(\d+)_DLS_(\d+)aR(\d+)', raw_sheet, re.IGNORECASE)
            if not match:
                logger.warning(f"Invalid raw sheet name format: {raw_sheet}")
                continue
            wp_number = match.group(1)  # Capture WP number (1-5)
            dls_number = match.group(2)  # Capture DLS number (1-20)
            run_number = match.group(3)  # Capture aR number (1-5)
            test_identifier = f"WP{wp_number}_DLS_{dls_number}aR{run_number}"
            
            start_date = None
            end_date = None
            for row_idx, row in enumerate(self.ws.iter_rows(min_row=1, max_col=5), start=1):
                key_cell = row[0].value
                if not key_cell:
                    continue
                key = self.normalize_key(str(key_cell))
                if key and "start_date" in key:
                    start_date = self.excel_date_to_string(row[1].value if len(row) > 1 else None)
                if key and "end_date" in key:
                    end_date = self.excel_date_to_string(row[1].value if len(row) > 1 else None)

            replication = asdict(ReplicationData(
                test_identifier_number=test_identifier,
                test_start_date=start_date,
                test_end_date=end_date
            ))
            replications.append(replication)
            logger.debug(f"Added replication for Run {run_number}: {replication}")

        if not replications:
            logger.warning("No replication entries found. Check if raw data sheets exist or are correctly named.")
        
        logger.debug(f"Extracted {len(replications)} replication entries")
        return replications

    def extract_raw_data(self, raw_sheet_name: str, processed_sheet_name: str):
        logger.debug(f"Extracting run data for raw sheet: {raw_sheet_name}, processed sheet: {processed_sheet_name}")
        
        match = re.search(r'WP(\d+)_DLS_(\d+)aR(\d+)', raw_sheet_name, re.IGNORECASE)
        run_number = int(match.group(3)) if match else 0
        if run_number == 0:
            logger.warning(f"Could not extract run number from {raw_sheet_name}, defaulting to 0")

        correlation_data = CorrelationData()
        size_distribution = SizeDistributionData()
        metrics = RunMetrics()

        if raw_sheet_name not in self.wb.sheetnames:
            logger.error(f"Raw data sheet {raw_sheet_name} not found")
            return asdict(RawData(run_number=run_number, correlation_data=correlation_data, size_distribution=size_distribution, metrics=metrics))

        raw_ws = self.wb[raw_sheet_name]
        logger.debug(f"Processing raw data sheet: {raw_sheet_name}")

        logger.debug(f"Raw sheet {raw_sheet_name} first 5 rows:")
        for row_idx, row in enumerate(raw_ws.iter_rows(min_row=1, max_row=5, max_col=5), start=1):
            values = [str(cell.value)[:30] if cell.value else "None" for cell in row]
            logger.debug(f"Row {row_idx}: {values}")

        correlation_start_row = None
        for row_idx, row in enumerate(raw_ws.iter_rows(min_row=1, min_col=1, max_col=2), start=1):
            time_header = str(row[0].value).lower() if row[0].value else ""
            corr_header = str(row[1].value).lower() if row[1].value else ""
            if ("time" in time_header and "μs" in time_header) and ("correlation" in corr_header and "coefficient" in corr_header):
                correlation_start_row = row_idx + 1
                logger.debug(f"Found correlation header at row {row_idx}: {row[0].value}, {row[1].value}")
                break

        if correlation_start_row:
            empty_row_count = 0
            for row in raw_ws.iter_rows(min_row=correlation_start_row, min_col=1, max_col=2):
                time_us = row[0].value
                corr_coeff = row[1].value
                if time_us is None and corr_coeff is None:
                    empty_row_count += 1
                    if empty_row_count >= 5:
                        logger.debug(f"Reached end of correlation data at row {row[0].row}")
                        break
                    continue
                empty_row_count = 0
                try:
                    time_us_float = float(time_us) if time_us is not None else 0.0
                    corr_coeff_float = float(corr_coeff) if corr_coeff is not None else 0.0
                    if time_us_float > 0:
                        correlation_data.time_us.append(time_us_float)
                        correlation_data.correlation_coefficient.append(corr_coeff_float)
                except (ValueError, TypeError) as e:
                    logger.warning(f"Invalid correlation data at row {row[0].row}: time={time_us}, coeff={corr_coeff}, error={e}")
                    continue
            logger.debug(f"Extracted {len(correlation_data.time_us)} correlation data points")
        else:
            logger.warning(f"No correlation data header found in {raw_sheet_name}")

        if processed_sheet_name not in self.wb.sheetnames:
            logger.error(f"Processed data sheet {processed_sheet_name} not found")
            return asdict(RawData(run_number=run_number, correlation_data=correlation_data, size_distribution=size_distribution, metrics=metrics))

        processed_ws = self.wb[processed_sheet_name]
        logger.debug(f"Processing processed data sheet: {processed_sheet_name}")

        logger.debug(f"Processed sheet {processed_sheet_name} first 5 rows:")
        for row_idx, row in enumerate(processed_ws.iter_rows(min_row=1, max_row=5, max_col=5), start=1):
            values = [str(cell.value)[:30] if cell.value else "None" for cell in row]
            logger.debug(f"Row {row_idx}: {values}")

        header_row = None
        for row_idx, row in enumerate(processed_ws.iter_rows(min_row=1, max_col=10), start=1):
            for col_idx in range(len(row) - 1):
                size_col = str(row[col_idx].value).lower() if row[col_idx].value else ""
                intensity_col = str(row[col_idx + 1].value).lower() if col_idx + 1 < len(row) and row[col_idx + 1].value else ""
                if "size" in size_col and "nm" in size_col and "intensity" in intensity_col:
                    header_row = row_idx
                    size_col_idx = col_idx + 1
                    intensity_col_idx = col_idx + 2
                    logger.debug(f"Found size distribution header at row {row_idx}, cols {size_col_idx},{intensity_col_idx}: {row[col_idx].value}, {row[col_idx + 1].value}")
                    break
            if header_row:
                break

        if header_row:
            max_rows = min(processed_ws.max_row, header_row + 1000)
            empty_row_count = 0
            for row_idx in range(header_row + 1, max_rows + 1):
                size_nm = processed_ws.cell(row=row_idx, column=size_col_idx).value
                mean_intensity = processed_ws.cell(row=row_idx, column=intensity_col_idx).value
                logger.debug(f"Size distribution row {row_idx}: size={size_nm}, intensity={mean_intensity}")
                if size_nm is None and mean_intensity is None:
                    empty_row_count += 1
                    if empty_row_count >= 5:
                        logger.debug(f"Reached end of size distribution data at row {row_idx}")
                        break
                    continue
                empty_row_count = 0
                try:
                    size_nm_float = float(size_nm) if size_nm is not None else 0.0
                    intensity_float = float(mean_intensity) if mean_intensity is not None else 0.0
                    if size_nm_float > 0:
                        size_distribution.size_nm.append(size_nm_float)
                        size_distribution.mean_intensity_percent.append(intensity_float)
                except (ValueError, TypeError) as e:
                    logger.warning(f"Invalid size distribution data at row {row_idx}: size={size_nm}, intensity={mean_intensity}, error={e}")
                    continue
            logger.debug(f"Extracted {len(size_distribution.size_nm)} size distribution data points")
        else:
            logger.warning(f"No size distribution header found in {processed_sheet_name}")

        metric_names = [
            "z_ave_hydrodynamic_diameter", "pdi", "peak_1_diameter", "standard_deviation_peak_1",
            "peak_1_relative", "peak_2_diameter", "standard_deviation_peak_2", "peak_2_relative",
            "peak_3_diameter", "standard_deviation_peak_3", "peak_3_relative", "derived_count_rate"
        ]
        for row_idx in range(1, processed_ws.max_row + 1):
            row = processed_ws[row_idx]
            for cell in row:
                if not cell.value:
                    continue
                cell_value = str(cell.value).lower().replace(" ", "_").replace("-", "_")
                for metric_name in metric_names:
                    if metric_name in cell_value:
                        for offset in [(0, 1), (0, 2), (1, 0), (1, 1)]:
                            try:
                                value_cell = processed_ws.cell(row=row_idx + offset[0], column=cell.column + offset[1])
                                if value_cell.value is not None:
                                    try:
                                        setattr(metrics, metric_name.replace("standard_deviation_", "std_dev_").replace("_relative", "_intensity"), float(value_cell.value))
                                        logger.debug(f"Metric {metric_name} = {value_cell.value} at row {row_idx + offset[0]}, col {value_cell.column}")
                                    except (ValueError, TypeError):
                                        setattr(metrics, metric_name.replace("standard_deviation_", "std_dev_").replace("_relative", "_intensity"), None)
                                        logger.warning(f"Invalid value for {metric_name} at row {row_idx + offset[0]}, col {value_cell.column}: {value_cell.value}")
                                    break
                            except:
                                continue
        logger.debug(f"Extracted metrics: {metrics}")
        raw_data = RawData(
            run_number=run_number,
            correlation_data=correlation_data,
            size_distribution=size_distribution,
            metrics=metrics
        )
        logger.debug(f"Run {run_number} extracted: Z-ave={raw_data.metrics.z_ave_hydrodynamic_diameter}, PDI={raw_data.metrics.pdi}")
        return asdict(raw_data)

    def extract_final_results(self):
        logger.debug("Extracting final results")
        final_sheet_name = "Final Results"
        if final_sheet_name not in self.wb.sheetnames:
            logger.error(f"Final Results sheet '{final_sheet_name}' not found")
            return asdict(ResultsData())

        ws = self.wb[final_sheet_name]
        logger.debug(f"Processing final results sheet: {final_sheet_name}")

        metrics = {}
        statistic_table = []

        stat_start_row = None
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_col=10), start=1):
            for col_idx in range(len(row) - 2):
                size_col = str(row[col_idx].value).lower() if row[col_idx].value else ""
                intensity_col = str(row[col_idx + 1].value).lower() if col_idx + 1 < len(row) and row[col_idx + 1].value else ""
                std_dev_col = str(row[col_idx + 2].value).lower() if col_idx + 2 < len(row) and row[col_idx + 2].value else ""
                if "size" in size_col and "nm" in size_col and "intensity" in intensity_col and "std" in std_dev_col:
                    stat_start_row = row_idx
                    size_col_idx = col_idx + 1
                    intensity_col_idx = col_idx + 2
                    std_dev_col_idx = col_idx + 3
                    logger.debug(f"Found statistic table header at row {row_idx}, cols {size_col_idx},{intensity_col_idx},{std_dev_col_idx}: {row[col_idx].value}, {row[col_idx + 1].value}, {row[col_idx + 2].value}")
                    break
            if stat_start_row:
                break

        if stat_start_row:
            max_rows = min(ws.max_row, stat_start_row + 1000)
            empty_row_count = 0
            for row_idx in range(stat_start_row + 1, max_rows + 1):
                size_nm = ws.cell(row=row_idx, column=size_col_idx).value
                mean_intensity = ws.cell(row=row_idx, column=intensity_col_idx).value
                std_dev = ws.cell(row=row_idx, column=std_dev_col_idx).value
                logger.debug(f"Statistic table row {row_idx}: size={size_nm}, intensity={mean_intensity}, std_dev={std_dev}")
                if size_nm is None and mean_intensity is None and std_dev is None:
                    empty_row_count += 1
                    if empty_row_count >= 5:
                        logger.debug(f"Reached end of statistic table at row {row_idx}")
                        break
                    continue
                empty_row_count = 0
                try:
                    size_nm_float = float(size_nm) if size_nm is not None else 0.0
                    intensity_float = float(mean_intensity) if mean_intensity is not None else 0.0
                    std_dev_float = float(std_dev) if std_dev is not None else 0.0
                    if size_nm_float > 0:
                        statistic_table.append(StatisticTableEntry(
                            size_nm=size_nm_float,
                            mean_intensity_percent=intensity_float,
                            std_dev=std_dev_float
                        ))
                except (ValueError, TypeError) as e:
                    logger.warning(f"Invalid statistic table entry at row {row_idx}: size={size_nm}, intensity={mean_intensity}, std_dev={std_dev}, error={e}")
                    continue
            logger.debug(f"Extracted {len(statistic_table)} statistic table entries")
        else:
            logger.warning(f"No statistic table header found in {final_sheet_name}")

        metric_names = [
            "z_ave_hydrodynamic_diameter", "uncertainity_hydrodynamic_diameter", "pdi",
            "uncertainity_pddi", "mean_peak_1_diameter_by_intensity", "pooled_standard_deviation_peak_1",
            "standard_deviation_beetwen_measurements_peak_1", "mean_peak_1_relative_intensity",
            "mean_peak_2_diameter_by_intensity", "pooled_standard_deviation_peak_2",
            "standard_deviation_beetwen_measurements_peak_2", "mean_peak_2_relative_intensity",
            "mean_peak_3_diameter_by_intensity", "pooled_standard_deviation_peak_3",
            "standard_deviation_beetwen_measurements_peak_3", "mean_peak_3_relative_intensity",
            "derived_count_rate"
        ]
        for row_idx in range(1, ws.max_row + 1):
            row = ws[row_idx]
            for cell in row:
                if not cell.value:
                    continue
                cell_value = str(cell.value).lower().replace(" ", "_").replace("-", "_")
                for metric_name in metric_names:
                    if metric_name in cell_value:
                        for offset in [(0, 1), (0, 2), (1, 0), (1, 1), (-1, 0), (-1, 1)]:
                            try:
                                value_cell = ws.cell(row=row_idx + offset[0], column=cell.column + offset[1])
                                if value_cell.value is not None:
                                    try:
                                        metrics[metric_name] = float(value_cell.value)
                                        logger.debug(f"Metric {metric_name} = {metrics[metric_name]} at row {row_idx + offset[0]}, col {value_cell.column}")
                                    except (ValueError, TypeError):
                                        metrics[metric_name] = None
                                        logger.warning(f"Invalid value for {metric_name} at row {row_idx + offset[0]}, col {value_cell.column}: {value_cell.value}")
                                    break
                            except:
                                continue
        logger.debug(f"Extracted metrics: {metrics}")

        results_data = ResultsData(
            z_ave_hydrodynamic_diameter=metrics.get("z_ave_hydrodynamic_diameter"),
            uncertainty_hydrodynamic_diameter=metrics.get("uncertainity_hydrodynamic_diameter"),
            pdi=metrics.get("pdi"),
            uncertainty_pdi=metrics.get("uncertainity_pddi"),
            mean_peak_1_diameter=metrics.get("mean_peak_1_diameter_by_intensity"),
            pooled_std_dev_peak_1=metrics.get("pooled_standard_deviation_peak_1"),
            std_dev_between_measurements_peak_1=metrics.get("standard_deviation_beetwen_measurements_peak_1"),
            mean_peak_1_intensity=metrics.get("mean_peak_1_relative_intensity"),
            mean_peak_2_diameter=metrics.get("mean_peak_2_diameter_by_intensity"),
            pooled_std_dev_peak_2=metrics.get("pooled_standard_deviation_peak_2"),
            std_dev_between_measurements_peak_2=metrics.get("standard_deviation_beetwen_measurements_peak_2"),
            mean_peak_2_intensity=metrics.get("mean_peak_2_relative_intensity"),
            mean_peak_3_diameter=metrics.get("mean_peak_3_diameter_by_intensity"),
            pooled_std_dev_peak_3=metrics.get("pooled_standard_deviation_peak_3"),
            std_dev_between_measurements_peak_3=metrics.get("standard_deviation_beetwen_measurements_peak_3"),
            mean_peak_3_intensity=metrics.get("mean_peak_3_relative_intensity"),
            derived_count_rate=metrics.get("derived_count_rate"),
            statistic_table=statistic_table
        )
        logger.debug(f"Final results extracted: Z-ave={results_data.z_ave_hydrodynamic_diameter}, PDI={results_data.pdi}, Stats table size={len(statistic_table)}")
        return asdict(results_data)

    def parse_all_data(self) -> Dict[str, Union[Dict, List]]:
        try:
            work_package_data = self.extract_work_package_data()
            material_data = self.extract_material_data()
            sample_preparation_data = self.extract_sample_preparation_data()
            instrumentation_data = self.extract_instrumentation_data()

            parsed_data = {
                'test_details': {
                    'work_package': work_package_data,
                    'material': material_data,
                    'sample_preparation': sample_preparation_data,
                    'instrumentation': instrumentation_data
                },
                'replications': [],
                'processed_data': [],
                'final_results': {}
            }
            logger.debug(f"-------------------------------------------------------------------------")
            logger.debug(f"Parsed test details here at parse all data: {parsed_data}")

            SHEET_PATTERN = r'^({prefix}_WP[1-5]_DLS_[1-30]aR[1-5])$'

             # Function to filter sheets based on prefix
            def filter_sheets(sheetnames, prefix):
                 pattern = SHEET_PATTERN.format(prefix=prefix)
                 return [name for name in sheetnames if re.match(pattern, name, re.IGNORECASE)]

            # Filter raw and processed sheets
            raw_sheets = filter_sheets(self.wb.sheetnames, "Raw data")
            processed_sheets = filter_sheets(self.wb.sheetnames, "Processed data")

            for raw_sheet in raw_sheets:
                match = re.search(r'WP(\d+)_DLS_(\d+)aR(\d+)', raw_sheet, re.IGNORECASE)
                if not match:
                    logger.warning(f"Invalid raw sheet name format: {raw_sheet}")
                    continue
                wp_number = match.group(1)
                dls_number = match.group(2)
                run_number = match.group(3)
                if not (1 <= int(wp_number) <= 5 and 1 <= int(dls_number) <= 30 and 1 <= int(run_number) <= 5):
                   logger.warning(f"Invalid number range in sheet name: {raw_sheet}")
                   continue
                processed_sheet = next((name for name in processed_sheets 
                if re.search(rf'^.*WP{wp_number}_DLS_{dls_number}aR{run_number}$', name, re.IGNORECASE)),
                    None)
                if not processed_sheet:
                    logger.warning(f"No matching processed data sheet for {raw_sheet}")
                    continue

                try:
                    raw_data = self.extract_raw_data(raw_sheet, processed_sheet)
                    
                    # Structure raw_data to match frontend expectations
                    raw_data_entry = {
                        'run_number': raw_data.get('run_number', 0),
                        'correlation_data': raw_data.get('correlation_data', {
                            'time_us': [],
                           'correlation_coefficient': []
                        })
                    }
                    #print(f"Raw data for run {run_number}: {raw_data_entry}")
                    #parsed_data['replications'].append(raw_data_entry)
                    parsed_data['replications'].append(raw_data)

                    # Structure processed_data to match frontend expectations
                    processed_data_entry = {
                        'run_number': raw_data.get('run_number', 0),
                        'size_distribution': raw_data.get('size_distribution', {
                            'size_nm': [],
                            'mean_intensity_percent': []
                        }),
                        'metrics': {
                            'z_ave_hydrodynamic_diameter': raw_data.get('metrics', {}).get('z_ave_hydrodynamic_diameter'),
                            'pdi': raw_data.get('metrics', {}).get('pdi'),
                            'peak_1_diameter': raw_data.get('metrics', {}).get('peak_1_diameter'),
                            'std_dev_peak_1': raw_data.get('metrics', {}).get('std_dev_peak_1'),
                            'peak_1_intensity': raw_data.get('metrics', {}).get('peak_1_intensity'),
                            'peak_2_diameter': raw_data.get('metrics', {}).get('peak_2_diameter'),
                            'std_dev_peak_2': raw_data.get('metrics', {}).get('std_dev_peak_2'),
                            'peak_2_intensity': raw_data.get('metrics', {}).get('peak_2_intensity'),
                            'peak_3_diameter': raw_data.get('metrics', {}).get('peak_3_diameter'),
                            'std_dev_peak_3': raw_data.get('metrics', {}).get('std_dev_peak_3'),
                            'peak_3_intensity': raw_data.get('metrics', {}).get('peak_3_intensity'),
                            'derived_count_rate': raw_data.get('metrics', {}).get('derived_count_rate')
                        }
                    }
                    parsed_data['processed_data'].append(processed_data_entry)
                    logger.debug(f"Processed run {run_number}: {processed_data_entry}")
                except Exception as e:
                    logger.error(f"Error processing run {run_number} from {raw_sheet}: {e}\n{traceback.format_exc()}")
                    continue

            parsed_data['final_results'] = self.extract_final_results()
            

            logger.info(f"Parsed data: {len(parsed_data['replications'])} raw data entries, {len(parsed_data['processed_data'])} processed data entries")
            return parsed_data
        except Exception as e:
            logger.error(f"Error parsing Excel file: {e}\n{traceback.format_exc()}")
            raise

def parse_excel_dls(file_path: str, sheet_name: str = "Test Information") -> Dict[str, Union[Dict, List]]:
    """
    Parse DLS Excel file and return data in a format compatible with TestCreate schema.
    """
    try:
        parser = DLSParser(file_path, sheet_name)
        return parser.parse_all_data()
    except Exception as e:
        logger.error(f"Error in parse_excel_dls: {e}\n{traceback.format_exc()}")
        raise

if __name__ == "__main__":
    file_path = "backend/data/WP2/CMS_2a_AuNP/DLS/WP2_DLS_2aR1_R5.xlsx"
    try:
        parsed_data = parse_excel_dls(file_path)
        print("Parsed Data:")
        print("Test Details:", parsed_data['test_details'])
        #print("Raw Data:", parsed_data['replications'])
        #print("Processed Data:", parsed_data['processed_data'])
        #print("Final Results:", parsed_data['final_results'])
    except Exception as e:
        print(f"Error: {e}")