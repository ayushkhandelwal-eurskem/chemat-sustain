import openpyxl
import re
import json
from dataclasses import dataclass, asdict, field
from typing import List, Dict, Optional, Union, Any
import logging
import traceback
from datetime import datetime, timedelta

logging.basicConfig(level=logging.DEBUG, format="%(asctime)s - %(levelname)s - %(message)s")
logging.getLogger("multipart.multipart").setLevel(logging.WARNING)
logger = logging.getLogger(__name__)

_DSC_ID = r'WP(\d+)_DSC_(\d+)([a-zA-Z])R(\d+)'

@dataclass
class Scientist:
    name: Optional[str] = None
    email: Optional[str] = None

@dataclass
class WorkPackageData:
    wp_name: Optional[str] = None
    partner: Optional[str] = None
    laboratory_name: Optional[str] = None
    full_test_name: Optional[str] = None
    test_acronym: Optional[str] = None
    test_type: Optional[str] = None
    endpoint: Optional[str] = None
    endpoint_outcome: Optional[str] = None
    sop: Optional[str] = None
    path: Optional[str] = None
    lead_scientists: List[Scientist] = field(default_factory=list)
    assay_scientists: List[Scientist] = field(default_factory=list)

@dataclass
class MaterialData:
    material_identifier: Optional[str] = None
    erm_id: Optional[str] = None
    material_name: Optional[str] = None
    core_chemistry: Optional[str] = None
    cas_no: Optional[str] = None
    cas_for_core: Optional[str] = None
    material_supplier: Optional[str] = None
    catalog_number: Optional[str] = None
    material_state: Optional[str] = None
    batch: Optional[str] = None
    vial: Optional[str] = None
    preparation_date: Optional[str] = None
    molar_concentration: Optional[str] = None
    particles_stock: Optional[str] = None

@dataclass
class DispersionData:
    dispersion_protocol: Optional[str] = None
    dispersion_technique: Optional[str] = None
    dispersion_medium: Optional[str] = None
    sonicator_type: Optional[str] = None
    power_w: Optional[str] = None
    sonication_time_s: Optional[str] = None
    tip_thickness_mm: Optional[str] = None
    tip_composition: Optional[str] = None
    bath_volume_dm3: Optional[str] = None
    sample_volume: Optional[str] = None
    final_concentration: Optional[str] = None
    additional_info: Optional[str] = None

@dataclass
class InstrumentationData:
    instrument_model: Optional[str] = None
    crucible_type: Optional[str] = None
    replication_count: Optional[Union[int, str]] = None
    replicate_labels: List[str] = field(default_factory=list)
    sample_masses: List[Dict[str, Optional[str]]] = field(default_factory=list)
    protective_atmosphere: Optional[str] = None
    temperature_range: Optional[str] = None
    heating_speed: Optional[str] = None

@dataclass
class ReplicationMetadata:
    test_identifier_number: Optional[str] = None
    test_start_date: Optional[str] = None
    test_end_date: Optional[str] = None
    replicate_label: Optional[str] = None
    raw_sheet_name: Optional[str] = None
    processed_sheet_name: Optional[str] = None

@dataclass
class DSCDataPoint:
    time_min: Optional[float] = None
    temperature_c: Optional[float] = None
    heat_flow_mw_per_mg: Optional[float] = None
    sensitivity_uv_per_mw: Optional[float] = None

@dataclass
class DSCRawDataBlock:
    metric_name: Optional[str] = None
    raw_sheet_name: Optional[str] = None
    run_label: Optional[str] = None
    time_unit: Optional[str] = "min"
    temperature_unit: Optional[str] = "oC"
    heat_flow_unit: Optional[str] = "mW/mg"
    sensitivity_unit: Optional[str] = "uV/mW"
    point_count: Optional[int] = None
    min_time_min: Optional[float] = None
    max_time_min: Optional[float] = None
    min_temperature_c: Optional[float] = None
    max_temperature_c: Optional[float] = None
    min_heat_flow: Optional[float] = None
    max_heat_flow: Optional[float] = None
    data_points: List[DSCDataPoint] = field(default_factory=list)

@dataclass
class DSCProcessedReplicateRow:
    replicate_label: Optional[str] = None
    values: Dict[str, Optional[float]] = field(default_factory=dict)

@dataclass
class DSCProcessedDataBlock:
    headers: List[str] = field(default_factory=list)
    replicates: List[DSCProcessedReplicateRow] = field(default_factory=list)
    mean_row: Optional[DSCProcessedReplicateRow] = None
    std_dev_row: Optional[DSCProcessedReplicateRow] = None
    legend: Dict[str, str] = field(default_factory=dict)

@dataclass
class DSCFinalResultEntry:
    metric_name: Optional[str] = None
    value: Optional[Union[float, str]] = None
    std_dev: Optional[Union[float, str]] = None
    std_dev_unit: Optional[str] = None
    character: Optional[str] = None


class DSCParser:
    def __init__(self, file_path: str, sheet_name: str = "Test Information"):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.parser_warnings: List[Dict[str, Any]] = []
        try:
            self.wb = openpyxl.load_workbook(file_path, data_only=True)
            self.ws = self.wb[sheet_name]
        except Exception as e:
            logger.error("Failed to load workbook: %s", e)
            raise
        self.email_regex = r"^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$"

    def normalize_key(self, key):
        if key is None: return None
        n = re.sub(r"[^a-z0-9]+", "_", str(key).strip().lower())
        return re.sub(r"_+", "_", n).strip("_")

    def excel_date_to_string(self, value):
        try:
            if isinstance(value, datetime): return value.strftime("%Y-%m-%d")
            if isinstance(value, (int, float)): return (datetime(1899, 12, 30) + timedelta(days=float(value))).strftime("%Y-%m-%d")
            if isinstance(value, str):
                for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d.%m.%Y"):
                    try: return datetime.strptime(value.strip(), fmt).strftime("%Y-%m-%d")
                    except ValueError: pass
            return str(value).strip() if value not in (None, "") else None
        except: return str(value).strip() if value not in (None, "") else None

    def _safe_float(self, value):
        if value in (None, ""): return None
        try:
            if isinstance(value, str): value = value.replace(",", ".").strip()
            return float(value)
        except (ValueError, TypeError): return None

    def _get_first_value(self, row, start=2, end=6):
        for c in range(start, end + 1):
            if c - 1 < len(row) and row[c - 1].value is not None: return row[c - 1].value
        return None

    def _kv(self):
        data = []
        for ri, row in enumerate(self.ws.iter_rows(min_row=1, max_col=6), start=1):
            k = row[0].value
            if not k: continue
            key = self.normalize_key(k)
            if not key: continue
            data.append({"row": ri, "key": key, "value": self._get_first_value(row, 2, 6), "email": row[3].value if len(row) > 3 else None})
        return data

    def _find_sheets(self, prefix):
        p = prefix.lower()
        return sorted([s for s in self.wb.sheetnames if s.lower().startswith(p) or s.lower().startswith(p.replace(" ", "_"))])

    def extract_work_package_data(self):
        rows = self._kv()
        lead, assay = [], []
        for r in rows:
            if "lead_scientist" in r["key"]:
                lead.append(Scientist(name=r["value"], email=r["email"] if r["email"] and re.match(self.email_regex, str(r["email"])) else None))
            if "assay_test_work_conducted_by" in r["key"]:
                assay.append(Scientist(name=r["value"], email=r["email"] if r["email"] and re.match(self.email_regex, str(r["email"])) else None))
        g = lambda *keys: next((r["value"] for r in rows if any(k in r["key"] for k in keys) and r["value"] is not None), None)
        return asdict(WorkPackageData(wp_name=g("project_work_package"), partner=g("partner_conducting_test_assay"), laboratory_name=g("test_facility_laboratory_name"), full_test_name=g("full_name_of_test_assay"), test_acronym=g("short_name_or_acronym"), test_type=g("type_or_class_of_experimental"), endpoint=g("end_point_being_investigated"), endpoint_outcome=g("metric_s_used_to_assess"), sop=g("sop_s_for_test"), path=g("path_link_to_sop"), lead_scientists=lead, assay_scientists=assay))

    def extract_material_data(self):
        rows = self._kv()
        g = lambda *keys: next((r["value"] for r in rows if any(k in r["key"] for k in keys) and r["value"] is not None), None)
        batch = g("batch"); batch = str(batch) if batch is not None else None
        return asdict(MaterialData(material_identifier=next((r["value"] for r in rows if r["key"] == "sample_cms_internal_identifier"), None), erm_id=g("erm_identifier"), material_name=next((r["value"] for r in rows if r["key"] == "material_name"), None), core_chemistry=g("core_chemistry"), cas_no=next((r["value"] for r in rows if r["key"] == "cas_no"), None), cas_for_core=g("cas_for_core"), material_supplier=next((r["value"] for r in rows if r["key"] == "material_supplier"), None), catalog_number=g("catalog_number"), material_state=next((r["value"] for r in rows if r["key"] == "material_state"), None), batch=batch, vial=g("vial"), preparation_date=self.excel_date_to_string(g("date_of_sample_preparation", "date_of_preparation")), molar_concentration=g("molar_concentration"), particles_stock=g("number_of_particles", "no_of_particles")))

    def extract_dispersion_data(self):
        rows = self._kv()
        g = lambda *keys: next((r["value"] for r in rows if any(k in r["key"] for k in keys) and r["value"] is not None), None)
        return asdict(DispersionData(dispersion_protocol=g("standard_dispersion_protocol"), dispersion_technique=g("dispersion_technique", "otherwise_specify_dispersion"), dispersion_medium=g("dispersion_dilution_medium"), sonicator_type=g("sonicator"), power_w=next((r["value"] for r in rows if r["key"] == "power_w"), None), sonication_time_s=g("sonication_time"), tip_thickness_mm=g("tip_thickness"), tip_composition=g("tip_composition"), bath_volume_dm3=g("ultrasonic_bath", "water_volume"), sample_volume=next((r["value"] for r in rows if r["key"] == "sample_volume"), None), final_concentration=g("final_sample_concentration"), additional_info=g("additional_information")))

    def extract_instrumentation_data(self):
        rows = self._kv()
        ws = self.ws
        g = lambda *keys: next((r["value"] for r in rows if any(k in r["key"] for k in keys) and r["value"] is not None), None)
        rep_labels = [str(ws.cell(row=61, column=c).value) for c in range(3, 15) if ws.cell(row=61, column=c).value is not None]
        masses = []
        for r in range(62, 65):
            lbl = ws.cell(row=r, column=1).value
            val = ws.cell(row=r, column=2).value
            if lbl is None: continue
            entry = {"label": str(lbl).strip(), "value": str(val).strip() if val else None}
            notes = [str(ws.cell(row=r, column=c).value).strip() for c in range(3, 7) if ws.cell(row=r, column=c).value is not None]
            if notes: entry["notes"] = "; ".join(notes)
            masses.append(entry)
        return asdict(InstrumentationData(instrument_model=g("instrumentation"), crucible_type=g("crucible"), replication_count=ws.cell(row=61, column=2).value, replicate_labels=rep_labels, sample_masses=masses, protective_atmosphere=g("protective_atmosphere"), temperature_range=g("temperature_range"), heating_speed=g("heating_speed")))

    def extract_replication_metadata(self):
        ws = self.ws; raw_sheets = self._find_sheets("raw data"); proc_sheets = self._find_sheets("processed data"); meta = []
        for ri in range(39, 42):
            tid = ws.cell(row=ri, column=2).value
            if tid is None: continue
            ts = str(tid).strip()
            raw = next((s for s in raw_sheets if ts.lower() in s.lower()), None)
            meta.append(asdict(ReplicationMetadata(test_identifier_number=ts, test_start_date=self.excel_date_to_string(ws.cell(row=ri, column=3).value), test_end_date=self.excel_date_to_string(ws.cell(row=ri, column=4).value), replicate_label="DSC Thermogram", raw_sheet_name=raw, processed_sheet_name=proc_sheets[0] if proc_sheets else None)))
        return meta

    def extract_raw_data(self):
        return [self._extract_raw_block(s) for s in self._find_sheets("raw data")]

    def _extract_raw_block(self, sn):
        ws = self.wb[sn]
        cm = {}
        for c in range(1, ws.max_column + 1):
            h = ws.cell(row=1, column=c).value
            if h is None: continue
            hl = str(h).strip().lower()
            if "time" in hl: cm["time"] = c
            elif "temp" in hl: cm["temp"] = c
            elif "dsc" in hl or "heat" in hl: cm["hf"] = c
            elif "sensit" in hl: cm["sens"] = c
        tc, tempc, hfc, sc = cm.get("time", 1), cm.get("temp", 2), cm.get("hf", 3), cm.get("sens")

        # Units
        hf_unit, sens_unit = "mW/mg", "uV/mW"
        hh = ws.cell(row=1, column=hfc).value
        if hh and "/" in str(hh):
            p = str(hh).split("/", 1)
            if len(p) > 1: hf_unit = p[1].strip().strip("()")
        if sc:
            sh = ws.cell(row=1, column=sc).value
            if sh and "/" in str(sh):
                p = str(sh).split("/", 1)
                if len(p) > 1: sens_unit = p[1].strip().strip("()")

        m = re.search(_DSC_ID, sn, re.IGNORECASE)
        run_label = f"R{m.group(4)}" if m else sn

        pts, tv, tempv, hfv = [], [], [], []
        for r in range(2, ws.max_row + 1):
            t = self._safe_float(ws.cell(row=r, column=tc).value)
            if t is None: continue
            temp = self._safe_float(ws.cell(row=r, column=tempc).value)
            hf = self._safe_float(ws.cell(row=r, column=hfc).value)
            sens = self._safe_float(ws.cell(row=r, column=sc).value) if sc else None
            pts.append(asdict(DSCDataPoint(time_min=t, temperature_c=temp, heat_flow_mw_per_mg=hf, sensitivity_uv_per_mw=sens)))
            tv.append(t)
            if temp is not None: tempv.append(temp)
            if hf is not None: hfv.append(hf)

        return asdict(DSCRawDataBlock(metric_name=f"DSC Thermogram {run_label}", raw_sheet_name=sn, run_label=run_label, heat_flow_unit=hf_unit, sensitivity_unit=sens_unit, point_count=len(pts), min_time_min=min(tv) if tv else None, max_time_min=max(tv) if tv else None, min_temperature_c=min(tempv) if tempv else None, max_temperature_c=max(tempv) if tempv else None, min_heat_flow=min(hfv) if hfv else None, max_heat_flow=max(hfv) if hfv else None, data_points=pts))

    def extract_processed_data(self):
        sheets = self._find_sheets("processed data")
        if not sheets: return {"available": False, "blocks": []}
        blocks = []
        for sn in sheets:
            ws = self.wb[sn]
            headers = []
            for c in range(2, ws.max_column + 1):
                h = ws.cell(row=1, column=c).value
                if h is not None: headers.append({"col": c, "name": str(h).strip()})
            replicates, mean_row, sd_row = [], None, None
            for r in range(2, ws.max_row + 1):
                label = ws.cell(row=r, column=1).value
                if label is None: continue
                ls = str(label).strip()
                col2 = ws.cell(row=r, column=2).value
                if col2 is not None and isinstance(col2, str) and self._safe_float(col2) is None: continue
                vals = {hd["name"]: self._safe_float(ws.cell(row=r, column=hd["col"]).value) for hd in headers}
                rd = asdict(DSCProcessedReplicateRow(replicate_label=ls, values=vals))
                if ls.lower() in ("xm", "mean"): mean_row = rd
                elif ls.lower() in ("δ", "sd", "sigma", "std"): sd_row = rd
                else: replicates.append(rd)
            legend = {}
            for r in range(max(2, len(replicates) + 4), ws.max_row + 1):
                k, v = ws.cell(row=r, column=1).value, ws.cell(row=r, column=2).value
                if k and v and isinstance(v, str): legend[str(k).strip()] = str(v).strip()
            blocks.append(asdict(DSCProcessedDataBlock(headers=[h["name"] for h in headers], replicates=replicates, mean_row=mean_row, std_dev_row=sd_row, legend=legend)))
        return {"available": True, "blocks": blocks}

    def extract_final_results(self):
        sheets = self._find_sheets("final results")
        if not sheets: return []
        events = []
        for sn in sheets:
            ws = self.wb[sn]
            hr = 2; col = 1; groups = []
            while col <= ws.max_column:
                h = ws.cell(row=hr, column=col).value
                if h is None: col += 1; continue
                hs = str(h).strip().lower()
                if "standard deviation" in hs or "std" in hs or "character" in hs: col += 1; continue
                mn = str(ws.cell(row=hr, column=col).value).strip()
                vc = col; nc = col + 1; sdc = None; sdu = None; cc = None
                if nc <= ws.max_column:
                    nh = ws.cell(row=hr, column=nc).value
                    if nh and ("standard deviation" in str(nh).lower() or "std" in str(nh).lower()):
                        sdc = nc; um = re.search(r'\(([^)]+)\)|\[([^\]]+)\]', str(nh)); sdu = (um.group(1) or um.group(2)) if um else None; nc += 1
                if nc <= ws.max_column:
                    ch = ws.cell(row=hr, column=nc).value
                    if ch and "character" in str(ch).lower(): cc = nc; nc += 1
                groups.append({"mn": mn, "vc": vc, "sdc": sdc, "sdu": sdu, "cc": cc}); col = nc
            for dr in range(3, ws.max_row + 1):
                if not any(ws.cell(row=dr, column=g["vc"]).value is not None for g in groups): continue
                for g in groups:
                    events.append(asdict(DSCFinalResultEntry(metric_name=g["mn"], value=self._safe_float(ws.cell(row=dr, column=g["vc"]).value), std_dev=self._safe_float(ws.cell(row=dr, column=g["sdc"]).value) if g["sdc"] else None, std_dev_unit=g["sdu"], character=str(ws.cell(row=dr, column=g["cc"]).value).strip() if g["cc"] and ws.cell(row=dr, column=g["cc"]).value else None)))
        return events

    def extract_statistical_analysis(self):
        return {"available": False, "notes": "No statistical analysis in this DSC workbook."}

    def parse_all_data(self):
        try:
            d = {"test_details": {"work_package": self.extract_work_package_data(), "material": self.extract_material_data(), "cell_line": {}, "dispersion": self.extract_dispersion_data(), "instrumentation": self.extract_instrumentation_data()}, "replication_metadata": self.extract_replication_metadata(), "replications": self.extract_raw_data(), "processed_data": self.extract_processed_data(), "final_results": self.extract_final_results(), "statistical_analysis": self.extract_statistical_analysis()}
            if self.parser_warnings: d["parser_warnings"] = self.parser_warnings
            return d
        except Exception as e:
            logger.error("Error: %s\n%s", e, traceback.format_exc()); raise

def _fix_degree_symbols(obj):
    """Recursively replace oC with °C in all string values."""
    if isinstance(obj, str):
        return obj.replace("oC", "°C")
    if isinstance(obj, dict):
        return {(k.replace("oC", "°C") if isinstance(k, str) else k): _fix_degree_symbols(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_fix_degree_symbols(v) for v in obj]
    return obj

def parse_excel_dsc(file_path, sheet_name="Test Information"):
    return _fix_degree_symbols(DSCParser(file_path, sheet_name).parse_all_data())

if __name__ == "__main__":
    import sys
    fp = sys.argv[1] if len(sys.argv) > 1 else "/mnt/user-data/uploads/WP2_DSC_27aR1_R3.xlsx"
    data = parse_excel_dsc(fp)
    print("=" * 70); print("DSC PARSER OUTPUT SUMMARY"); print("=" * 70)
    wp = data["test_details"]["work_package"]; mat = data["test_details"]["material"]; inst = data["test_details"]["instrumentation"]
    print(f"\nWP: {wp['wp_name']}, Partner: {wp['partner']}, Test: {wp['test_acronym']}")
    print(f"Material: {mat['material_name']} ({mat['material_identifier']})")
    print(f"Instrument: {inst['instrument_model']}, Reps: {inst['replication_count']}")
    for sm in inst["sample_masses"]: print(f"  {sm['label']}: {sm['value']}{' — ' + sm.get('notes','') if sm.get('notes') else ''}")
    print(f"\nReplication Metadata: {len(data['replication_metadata'])}")
    for rm in data["replication_metadata"]: print(f"  {rm['test_identifier_number']}: {rm['test_start_date']} → {rm['raw_sheet_name']}")
    print(f"\nRaw Blocks: {len(data['replications'])}")
    for b in data["replications"]: print(f"  {b['run_label']}: {b['point_count']} pts, T={b['min_temperature_c']}..{b['max_temperature_c']}°C")
    pd = data["processed_data"]; print(f"\nProcessed: available={pd['available']}")
    for blk in pd.get("blocks", []):
        print(f"  Headers: {blk['headers']}")
        for rep in blk["replicates"]: print(f"    {rep['replicate_label']}: { {k: v for k, v in rep['values'].items() if v is not None} }")
        if blk["mean_row"]: print(f"    Mean: {blk['mean_row']['values']}")
        if blk["std_dev_row"]: print(f"    SD:   {blk['std_dev_row']['values']}")
    print(f"\nFinal Results: {len(data['final_results'])}")
    for fr in data["final_results"]: print(f"  {fr['metric_name']}: {fr['value']} ± {fr['std_dev']} {fr.get('std_dev_unit','')}, {fr['character']}")