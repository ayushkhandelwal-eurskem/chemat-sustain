"""
parse_rotifier.py
================================================================================
Parser for the Rotifer Toxicity Screening Test (RotoxKit M with Brachionus
plicatilis) Excel data recording forms used in the CMS/CheMatSustain project.

This parser handles BOTH:
  - Limit Test files (1 concentration only — e.g. CMS_WP3_Rotifier_2a.xlsx)
  - Definitive Test files (up to 5 concentrations — e.g. CMS_WP3_Rotifier_7b.xlsx
    8b, 9b, 10b, 11b, 12b, 26a, 27a, 28a, etc.)

Key design:
  - Concentrations live in B35:B39 on the "Test Details" sheet. The number of
    concentrations varies (1-5), and ALL subsequent rows on that sheet shift
    accordingly. We therefore locate downstream rows by their A-column label
    (case-insensitive) rather than by hard-coded row numbers.
  - Raw Data sheet rows shift the same way: Control + N concentrations, each
    occupying 2 rows (24h, 48h) starting at row 7.
  - Processed Data column counts shift with concentration count.
  - Final Results has variable presence of the "Statistic analysis" / EC25 /
    EC10 sub-sections depending on test type.

Sheet names (note the trailing space on "Raw Data "):
  ['Test Details', 'Raw Data ', 'Processed Data', 'Final Results']
"""

import openpyxl
import re
import json
import logging
import traceback
from dataclasses import dataclass, asdict, field
from typing import List, Dict, Optional, Union, Any
from datetime import datetime, timedelta
from difflib import SequenceMatcher

# ----------------------------------------------------------------------------
# Logging
# ----------------------------------------------------------------------------
logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)s - %(levelname)s - %(message)s",
)
logging.getLogger("multipart.multipart").setLevel(logging.WARNING)
logger = logging.getLogger(__name__)


# ============================================================================
# Dataclasses
# ============================================================================
@dataclass
class Scientist:
    name: Optional[str] = None
    email: Optional[str] = None


@dataclass
class WorkPackageData:
    wp_name: Optional[str] = None
    partner: Optional[str] = None
    laboratory_name: Optional[str] = None
    full_test_name: Optional[str] = None
    oecd_iso_ref: Optional[str] = None
    test_type: Optional[str] = None
    endpoint: Optional[str] = None
    endpoint_outcome: Optional[str] = None
    sop: Optional[str] = None
    test_start_date: Optional[str] = None
    test_end_date: Optional[str] = None
    lead_scientists: List[Dict] = field(default_factory=list)
    assay_scientists: List[Dict] = field(default_factory=list)


@dataclass
class MaterialData:
    material_identifier: Optional[str] = None
    erm_id: Optional[str] = None
    material_name: Optional[str] = None
    core_chemistry: Optional[str] = None
    cas_for_core: Optional[str] = None
    material_supplier: Optional[str] = None
    material_state: Optional[str] = None
    batch: Optional[str] = None
    vial: Optional[str] = None
    preparation_date: Optional[str] = None
    stock_concentration: Optional[str] = None
    molecular_weight: Optional[str] = None
    particles_in_stock: Optional[str] = None


@dataclass
class ApplicationData:
    """Test application parameters (rotifers/well, volume, concentrations,
    replicates per concentration)."""
    start_rotifers_per_well: Optional[float] = None
    total_volume_per_replicate_ml: Optional[float] = None
    concentrations_tested: List[float] = field(default_factory=list)
    concentration_unit: Optional[str] = None
    replicates_per_concentration: Optional[int] = None


@dataclass
class TestConditionsData:
    """Environmental/test conditions section."""
    test_medium: List[str] = field(default_factory=list)
    light_cycle: Optional[str] = None
    light_intensity: Optional[str] = None
    illumination_direction: Optional[str] = None
    temperature: Optional[str] = None
    aeration: Optional[str] = None
    salinity_ppt: Optional[float] = None
    total_incubation_time_h: Optional[float] = None


@dataclass
class AnalysisData:
    """Analysis section: how readings are taken."""
    type_of_measurement: Optional[str] = None
    measurement_device: Optional[str] = None
    timepoints_h: List[float] = field(default_factory=list)


@dataclass
class RotifierRawRow:
    toxicant: Optional[str] = None
    time_label: Optional[str] = None
    wells: Dict[str, str] = field(default_factory=dict)
    wells_numeric: Dict[str, Optional[int]] = field(default_factory=dict)
    well_total: Optional[int] = None  # cohort size, e.g. 5
    mean_dead: Optional[float] = None


@dataclass
class RotifierProcessedRow:
    time_label: Optional[str] = None
    values: Dict[str, Any] = field(default_factory=dict)


@dataclass
class RotifierProcessedBlock:
    title: Optional[str] = None              # "Absolute mortality" / "Relative mortality (%)"
    section_label: Optional[str] = None      # "Toxicants (50 mg/L)" or "CMS_7b_Ag-Np"
    column_headers: List[str] = field(default_factory=list)
    rows: List[Dict] = field(default_factory=list)


@dataclass
class RotifierFinalResultsTable:
    section_label: Optional[str] = None
    column_headers: List[str] = field(default_factory=list)  # ['Control', '0,166 mg/L', ...]
    time_column: str = "Time (h)"
    rows: List[Dict] = field(default_factory=list)           # [{time:24, values:{...}}, ...]


@dataclass
class AcceptanceItem:
    label: Optional[str] = None
    detail: Optional[str] = None
    result: Optional[str] = None  # PASSED/FAILED/YES/NO etc.


@dataclass
class EcValue:
    name: str                    # "EC50" | "EC25" | "EC10"
    description: Optional[str] = None
    value: Optional[str] = None  # may be "0,146 mg/L" or "not defined (> 100 mg·l-1)"


# ============================================================================
# Parser
# ============================================================================
class RotifierParser:
    # Default primary sheet ("Test Details")
    DEFAULT_SHEET = "Test Details"

    # Maximum concentrations supported (user constraint)
    MAX_CONCENTRATIONS = 5

    # First row where concentrations live (B35..B39)
    CONC_START_ROW = 35

    EMAIL_REGEX = r"^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$"

    def __init__(self, file_path: str, sheet_name: str = DEFAULT_SHEET):
        self.file_path = file_path
        self.sheet_name = sheet_name
        try:
            self.wb = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            logger.error(f"Failed to load workbook '{file_path}': {e}")
            raise

        # Resolve Test Details sheet with fuzzy fallback
        self.ws = self._resolve_sheet(sheet_name) or self._resolve_sheet("Test Details") \
            or self.wb[self.wb.sheetnames[0]]
        logger.info(f"Primary sheet selected: '{self.ws.title}'")

        # Resolve secondary sheets (tolerant of trailing whitespace / case)
        self.raw_sheet = self._resolve_sheet("Raw Data")
        self.processed_sheet = self._resolve_sheet("Processed Data")
        self.final_sheet = self._resolve_sheet("Final Results")

        # Build label->row index for the Test Details sheet so we can locate
        # rows whose positions shift with the concentration count.
        self.label_rows: Dict[str, int] = self._index_label_rows(self.ws)

        # Discover concentration count up-front (used by downstream extraction)
        self.concentrations: List[float] = self._read_concentrations()
        self.concentration_count = len(self.concentrations)
        logger.info(f"Discovered {self.concentration_count} concentration(s): {self.concentrations}")

        self.parser_warnings: List[Dict[str, str]] = []

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------
    def _resolve_sheet(self, target: str):
        """Resolve sheet name tolerantly: exact, case-insensitive, whitespace,
        and fuzzy-match (>0.85). Returns the worksheet object or None."""
        if not target:
            return None
        if target in self.wb.sheetnames:
            return self.wb[target]
        t = target.strip().lower()
        for s in self.wb.sheetnames:
            if s.strip().lower() == t:
                return self.wb[s]
        # fuzzy fallback
        best_name, best_score = None, 0.0
        for s in self.wb.sheetnames:
            score = SequenceMatcher(None, s.strip().lower(), t).ratio()
            if score > best_score:
                best_score, best_name = score, s
        if best_score >= 0.85:
            logger.warning(f"Fuzzy-matched sheet '{target}' -> '{best_name}' "
                           f"(score={best_score:.2f})")
            return self.wb[best_name]
        logger.warning(f"Could not resolve sheet '{target}'. Available: {self.wb.sheetnames}")
        return None

    def normalize_key(self, key: Optional[str]) -> Optional[str]:
        if not key:
            return None
        normalized = str(key).strip().lower()
        normalized = re.sub(r"[^a-z0-9]", "_", normalized)
        normalized = re.sub(r"_+", "_", normalized).strip("_")
        return normalized

    def _safe_float(self, v) -> Optional[float]:
        if v is None or v == "" or (isinstance(v, str) and v.strip() == ""):
            return None
        try:
            if isinstance(v, str):
                v = v.replace(",", ".").strip()
            return float(v)
        except (ValueError, TypeError):
            return None

    def _safe_int(self, v) -> Optional[int]:
        f = self._safe_float(v)
        return int(f) if f is not None else None

    def _safe_str(self, v) -> Optional[str]:
        if v is None:
            return None
        s = str(v).strip()
        return s if s else None

    def excel_date_to_string(self, value) -> Optional[str]:
        """Convert various Excel date representations to YYYY-MM-DD."""
        if value is None or value == "":
            return None
        try:
            if isinstance(value, datetime):
                return value.strftime("%Y-%m-%d")
            if isinstance(value, (int, float)):
                # Excel serial date
                return (datetime(1899, 12, 30) + timedelta(days=float(value))).strftime("%Y-%m-%d")
            if isinstance(value, str):
                s = value.strip()
                if not s:
                    return None
                for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d"):
                    try:
                        return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
                    except ValueError:
                        continue
                self.parser_warnings.append({
                    "type": "date_format",
                    "value": s,
                    "message": f"Date string '{s}' could not be parsed to YYYY-MM-DD",
                })
                return s
            return str(value)
        except Exception as e:
            self.parser_warnings.append({
                "type": "date_error",
                "value": str(value),
                "message": str(e),
            })
            return None

    def _parse_fraction(self, s: Any) -> tuple[Optional[int], Optional[int]]:
        """Parse a string like '4/5' into (4, 5). Returns (None, None) on failure."""
        if s is None:
            return None, None
        if isinstance(s, (int, float)):
            try:
                return int(s), None
            except Exception:
                return None, None
        text = str(s).strip()
        m = re.match(r"^\s*(\d+(?:[.,]\d+)?)\s*/\s*(\d+(?:[.,]\d+)?)\s*$", text)
        if not m:
            # also try simple integer
            try:
                return int(text.replace(",", ".")), None
            except Exception:
                return None, None
        try:
            num = int(float(m.group(1).replace(",", ".")))
            den = int(float(m.group(2).replace(",", ".")))
            return num, den
        except Exception:
            return None, None

    def _index_label_rows(self, ws) -> Dict[str, int]:
        """Build a normalized A-column label -> row number map for the Test
        Details sheet. We use this to find rows whose absolute position is
        sensitive to concentration count."""
        out: Dict[str, int] = {}
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=1).value
            nk = self.normalize_key(v) if v else None
            if nk and nk not in out:
                out[nk] = r
        return out

    def _row_for_label(self, *candidates: str) -> Optional[int]:
        """Look up a row by label, trying multiple candidates (case-tolerant,
        with fuzzy fallback)."""
        if not self.label_rows:
            return None
        norm_cands = [self.normalize_key(c) for c in candidates if c]
        for nk in norm_cands:
            if nk in self.label_rows:
                return self.label_rows[nk]
        # fuzzy fallback
        best_row, best_score = None, 0.0
        for label_nk, row in self.label_rows.items():
            for cand_nk in norm_cands:
                if not cand_nk:
                    continue
                score = SequenceMatcher(None, label_nk, cand_nk).ratio()
                if score > best_score:
                    best_score, best_row = score, row
        if best_score >= 0.8 and best_row is not None:
            return best_row
        return None

    def _read_concentrations(self) -> List[float]:
        """Concentrations live at B35..B(35+MAX-1). Stop on first row whose
        A-column label is non-empty after row 35 (signals start of next
        section, e.g. 'Replicates per concentration')."""
        conc: List[float] = []
        for offset in range(self.MAX_CONCENTRATIONS):
            r = self.CONC_START_ROW + offset
            a_val = self.ws.cell(row=r, column=1).value
            b_val = self.ws.cell(row=r, column=2).value
            # On row 35 the A label is "Concentrations tested (mg·l-1)" — accept it.
            # On rows 36+, if A is non-empty AND not blank, we've hit the next field.
            if offset > 0 and a_val is not None and str(a_val).strip() != "":
                break
            f = self._safe_float(b_val)
            if f is None:
                if offset == 0:
                    # no concentration at all (shouldn't happen for valid file)
                    break
                # blank in the middle => stop
                break
            conc.append(f)
        return conc

    # ------------------------------------------------------------------
    # TEST DETAILS
    # ------------------------------------------------------------------
    def extract_work_package_data(self) -> Dict:
        ws = self.ws
        lead = ws.cell(row=self._row_for_label("Lead Scientist & contact for test:") or 6, column=2).value
        lead_email = ws.cell(row=self._row_for_label("Lead Scientist & contact for test:") or 6, column=4).value
        assay = ws.cell(row=self._row_for_label("Assay/Test work conducted by:") or 7, column=2).value
        assay_email = ws.cell(row=self._row_for_label("Assay/Test work conducted by:") or 7, column=4).value

        def _email(e):
            return e if e and re.match(self.EMAIL_REGEX, str(e)) else None

        lead_list = []
        if lead:
            lead_list.append(asdict(Scientist(name=self._safe_str(lead), email=_email(lead_email))))
        assay_list = []
        if assay:
            assay_list.append(asdict(Scientist(name=self._safe_str(assay), email=_email(assay_email))))

        def _at(*labels, col=2):
            row = self._row_for_label(*labels)
            return ws.cell(row=row, column=col).value if row else None

        wp = WorkPackageData(
            wp_name=self._safe_str(_at("Project Work Package:")),
            partner=self._safe_str(_at("Partner conducting test/assay:")),
            laboratory_name=self._safe_str(_at("Test facility - Laboratory name:")),
            full_test_name=self._safe_str(_at("Full name of test")),
            oecd_iso_ref=self._safe_str(_at("OECD/ISO Test ref-ID if app.:")),
            test_type=self._safe_str(_at("Type or class of experimental test as used here:")),
            endpoint=self._safe_str(_at(" End-Point being investigated/assessed by the test:",
                                        "End-Point being investigated/assessed by the test:")),
            endpoint_outcome=self._safe_str(_at("Metric(s) used to assess End-Point outcome/response:")),
            sop=self._safe_str(_at("SOP(s) for test -  ref. project or other doc. - Title/ID:",
                                    "SOP(s) for test")),
            test_start_date=self.excel_date_to_string(_at("Test start date:")),
            test_end_date=self.excel_date_to_string(_at("Test end date:")),
            lead_scientists=lead_list,
            assay_scientists=assay_list,
        )
        return asdict(wp)

    def extract_material_data(self) -> Dict:
        ws = self.ws

        def _at(*labels, col=2):
            row = self._row_for_label(*labels)
            return ws.cell(row=row, column=col).value if row else None

        mat = MaterialData(
            material_identifier=self._safe_str(_at("Sample CMS Internal Identifier")),
            erm_id=self._safe_str(_at("ERM Identifier Number:")),
            material_name=self._safe_str(_at("Material name")),
            core_chemistry=self._safe_str(_at(" Core chemistry:", "Core chemistry:")),
            cas_for_core=self._safe_str(_at("CAS for Core:")),
            material_supplier=self._safe_str(_at("Material supplier:")),
            material_state=self._safe_str(_at("Material state:")),
            batch=self._safe_str(_at("Batch")),
            vial=self._safe_str(_at("Vial:")),
            preparation_date=self.excel_date_to_string(_at("Date of preparation:")),
            stock_concentration=self._safe_str(_at("Stock concentration:")),
            molecular_weight=self._safe_str(_at("Molecular weight:", "Molecular concentration:")),
            particles_in_stock=self._safe_str(_at("No of particles in stock:")),
        )
        return asdict(mat)

    def extract_application_data(self) -> Dict:
        """APPLICATION section (rotifers/well, volume, concentrations, replicates)."""
        ws = self.ws
        # Concentration unit lives in the A column of row 35 (e.g.
        # "Concentrations tested (mg·l−1)"). Extract the unit hint.
        a35 = ws.cell(row=self.CONC_START_ROW, column=1).value
        unit_hint = None
        if a35:
            m = re.search(r"\(([^)]+)\)", str(a35))
            if m:
                unit_hint = m.group(1).strip()

        # "Replicates per concentration" row sits at 35 + concentration_count
        # (when concentration_count >= 1). We fall back to label lookup.
        rep_row = self._row_for_label("Replicates per concentration")
        replicates = self._safe_int(ws.cell(row=rep_row, column=2).value) if rep_row else None

        def _at(label):
            row = self._row_for_label(label)
            return ws.cell(row=row, column=2).value if row else None

        app = ApplicationData(
            start_rotifers_per_well=self._safe_float(_at("Start number of rotifers per well")),
            total_volume_per_replicate_ml=self._safe_float(_at("Total volume per replicate (ml)")),
            concentrations_tested=self.concentrations,
            concentration_unit=unit_hint,
            replicates_per_concentration=replicates,
        )
        return asdict(app)

    def extract_test_conditions_data(self) -> Dict:
        """TEST CONDITIONS section (test medium list, light, temp, salinity, ...)."""
        ws = self.ws

        # Test medium occupies several rows (multi-component). Find first row,
        # then collect contiguous B-values until A-column shows the next label.
        medium_start = self._row_for_label("Test medium")
        media: List[str] = []
        if medium_start:
            r = medium_start
            while r <= ws.max_row:
                a_val = ws.cell(row=r, column=1).value
                b_val = ws.cell(row=r, column=2).value
                # We're still inside test medium if A is empty OR A is the
                # first row's label.
                if r != medium_start and a_val is not None and str(a_val).strip() != "":
                    break
                if b_val is not None and str(b_val).strip() != "":
                    media.append(self._safe_str(b_val))
                r += 1

        def _at(*labels):
            row = self._row_for_label(*labels)
            return ws.cell(row=row, column=2).value if row else None

        tc = TestConditionsData(
            test_medium=media,
            light_cycle=self._safe_str(_at("Light cycle (h)")),
            light_intensity=self._safe_str(_at("Light intensity (lux)")),
            illumination_direction=self._safe_str(_at("Illumination direction")),
            temperature=self._safe_str(_at("Temperature (°C)", "Temperature (oC)", "Temperature")),
            aeration=self._safe_str(_at("Aeration")),
            salinity_ppt=self._safe_float(_at("Salinity (ppt)")),
            total_incubation_time_h=self._safe_float(_at("Total incubation time (h)")),
        )
        return asdict(tc)

    def extract_analysis_data(self) -> Dict:
        """ANALYSIS section (measurement type, device, timepoints list)."""
        ws = self.ws

        def _at(*labels):
            row = self._row_for_label(*labels)
            return ws.cell(row=row, column=2).value if row else None
        tp_start = self._row_for_label("Timepoints (h)")
        timepoints: List[float] = []
        if tp_start:
            r = tp_start
            while r <= ws.max_row:
                a_val = ws.cell(row=r, column=1).value
                b_val = ws.cell(row=r, column=2).value
                if r != tp_start and a_val is not None and str(a_val).strip() != "":
                    break
                f = self._safe_float(b_val)
                if f is not None:
                    timepoints.append(f)
                elif b_val is None:
                    # blank => stop
                    if r != tp_start:
                        break
                r += 1

        an = AnalysisData(
            type_of_measurement=self._safe_str(_at("Type of measurement")),
            measurement_device=self._safe_str(_at("Measurement device")),
            timepoints_h=timepoints,
        )
        return asdict(an)

    def extract_replication(self) -> Dict:
        """Build a single replication-metadata record (this test format has
        one replication per file)."""
        wp = self.extract_work_package_data()
        mat = self.extract_material_data()
        app = self.extract_application_data()
        return {
            "test_identifier_number": mat.get("material_identifier"),
            "test_start_date": wp.get("test_start_date"),
            "test_end_date": wp.get("test_end_date"),
            "replicates_per_concentration": app.get("replicates_per_concentration"),
            "no_of_concentrations": len(app.get("concentrations_tested") or []),
        }

    # ------------------------------------------------------------------
    # RAW DATA
    # ------------------------------------------------------------------
    def extract_raw_data(self) -> List[Dict]:
        """Extract Raw Data sheet into one block (single experiment file).

        Structure on the sheet:
          Row 1: title
          Row 3: 'Toxicant tested:' | <material name>
          Row 5: 'Toxicants' | 'Time' | 'Deceased rotifers per well'
          Row 6: blank | blank | 'Well 1' .. 'Well 6'
          Row 7: 'Control' | '24 h' | '0/5' x 6
          Row 8: blank     | '48 h' | '0/5' x 6
          Row 9: '<conc 1>' | '24 h' | '..' x 6
          Row 10: blank     | '48 h' | '..' x 6
          ... up to N concentrations.
        """
        if self.raw_sheet is None:
            return []
        ws = self.raw_sheet

        # Material / toxicant name from row 3
        toxicant_tested = self._safe_str(ws.cell(row=3, column=2).value)

        # Discover well headers from row 6 (cols 3..8 typically)
        well_headers: List[str] = []
        for c in range(3, ws.max_column + 1):
            v = ws.cell(row=6, column=c).value
            if v is None or str(v).strip() == "":
                continue
            well_headers.append(self._safe_str(v))
        if not well_headers:
            well_headers = [f"Well {i}" for i in range(1, 7)]

        # Number of toxicant groups = 1 (Control) + N concentrations.
        # Each group has 2 rows (24 h, 48 h). First group starts at row 7.
        num_groups = 1 + max(1, self.concentration_count)
        rows: List[Dict] = []
        for gi in range(num_groups):
            base = 7 + gi * 2
            tox_label = self._safe_str(ws.cell(row=base, column=1).value)
            # 24h row
            for off, tlabel in [(0, ws.cell(row=base, column=2).value),
                                (1, ws.cell(row=base + 1, column=2).value)]:
                if off == 1 and (tlabel is None or str(tlabel).strip() == ""):
                    continue  # second timepoint missing
                r = base + off
                wells: Dict[str, str] = {}
                wells_num: Dict[str, Optional[int]] = {}
                well_total: Optional[int] = None
                num_sum = 0
                count_parsed = 0
                for ci, wh in enumerate(well_headers):
                    v = ws.cell(row=r, column=3 + ci).value
                    if v is None:
                        wells[wh] = ""
                        wells_num[wh] = None
                        continue
                    s = self._safe_str(v)
                    wells[wh] = s if s is not None else ""
                    num, den = self._parse_fraction(s)
                    wells_num[wh] = num
                    if num is not None:
                        num_sum += num
                        count_parsed += 1
                    if den is not None and well_total is None:
                        well_total = den
                mean_dead = (num_sum / count_parsed) if count_parsed > 0 else None
                rows.append(asdict(RotifierRawRow(
                    toxicant=tox_label,
                    time_label=self._safe_str(tlabel),
                    wells=wells,
                    wells_numeric=wells_num,
                    well_total=well_total,
                    mean_dead=mean_dead,
                )))

        block = {
            "run_label": "R1",
            "raw_sheet_name": ws.title,
            "toxicant_tested": toxicant_tested,
            "well_headers": well_headers,
            "rows": rows,
        }
        return [block]

    # ------------------------------------------------------------------
    # PROCESSED DATA
    # ------------------------------------------------------------------
    def _extract_processed_block(self, title_row: int) -> Optional[Dict]:
        """A processed-data block is laid out as:
            Row N      : 'Absolute mortality' or 'Relative mortality (%)' | section_label
            Row N+1    : 'Time' | 'Control' | '0,166 mg/L' | ...
            Row N+2 .. : '24 h' | values...
                         '48 h' | values...
        """
        ws = self.processed_sheet
        if ws is None:
            return None
        title = self._safe_str(ws.cell(row=title_row, column=1).value)
        section_label = self._safe_str(ws.cell(row=title_row, column=2).value)
        header_row = title_row + 1
        headers: List[str] = []
        for c in range(2, ws.max_column + 1):
            v = ws.cell(row=header_row, column=c).value
            if v is None or str(v).strip() == "":
                if headers:
                    break
                continue
            headers.append(self._safe_str(v))
        rows = []
        r = header_row + 1
        while r <= ws.max_row:
            time_label = self._safe_str(ws.cell(row=r, column=1).value)
            if not time_label:
                break
            # Stop if this row looks like a new block title (e.g. 'Relative
            # mortality (%)' starting again — caller protects against this).
            values: Dict[str, Any] = {}
            for ci, h in enumerate(headers):
                v = ws.cell(row=r, column=2 + ci).value
                if v is None:
                    values[h] = None
                    continue
                if isinstance(v, (int, float)):
                    values[h] = v
                else:
                    s = self._safe_str(v)
                    f = self._safe_float(s)
                    values[h] = f if f is not None else s
            rows.append({"time_label": time_label, "values": values})
            r += 1
            if r - header_row > 6:
                break

        return {
            "title": title,
            "section_label": section_label,
            "column_headers": headers,
            "rows": rows,
        }

    def extract_processed_data(self) -> Dict:
        ws = self.processed_sheet
        if ws is None:
            return {"available": False}

        # Locate the title rows ("Absolute mortality" and "Relative mortality (%)")
        # — they are pinned at rows 7 and 13 in both sample files, but be defensive.
        abs_title_row, rel_title_row = None, None
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=1).value
            if v is None:
                continue
            s = str(v).strip().lower()
            if abs_title_row is None and s.startswith("absolute mortality"):
                abs_title_row = r
            elif rel_title_row is None and s.startswith("relative mortality"):
                rel_title_row = r
        if abs_title_row is None:
            abs_title_row = 7
        if rel_title_row is None:
            rel_title_row = 13

        absolute = self._extract_processed_block(abs_title_row)
        relative = self._extract_processed_block(rel_title_row)

        # Header row 5 captures "Toxicant tested:" + value
        toxicant_label = self._safe_str(ws.cell(row=5, column=2).value)

        return {
            "available": True,
            "toxicant_tested": toxicant_label,
            "absolute_mortality": absolute,
            "relative_mortality_percent": relative,
        }

    # ------------------------------------------------------------------
    # FINAL RESULTS
    # ------------------------------------------------------------------
    def extract_final_results(self) -> Dict:
        ws = self.final_sheet
        if ws is None:
            return {"available": False}

        # Material/section label sits at C3 (column 3) in both samples
        section_label = self._safe_str(ws.cell(row=3, column=3).value)

        # Header row is row 4: B4='Time (h)', C4='Control', D4..='<conc> mg/L'
        time_label = self._safe_str(ws.cell(row=4, column=2).value) or "Time (h)"
        column_headers: List[str] = []
        for c in range(3, ws.max_column + 1):
            v = ws.cell(row=4, column=c).value
            if v is None or str(v).strip() == "":
                if column_headers:
                    break
                continue
            column_headers.append(self._safe_str(v))

        # Data rows 5..(up to the validity acceptance row)
        data_rows: List[Dict] = []
        r = 5
        while r <= ws.max_row:
            t_val = ws.cell(row=r, column=2).value
            if t_val is None or str(t_val).strip() == "":
                # blank between sections - stop
                break
            # Validity / Statistic / Calculated / Classification sub-sections start when
            # column 2 contains alphabetic content rather than a number.
            t_f = self._safe_float(t_val)
            if t_f is None:
                break
            values: Dict[str, Any] = {}
            for ci, h in enumerate(column_headers):
                v = ws.cell(row=r, column=3 + ci).value
                if v is None:
                    values[h] = None
                elif isinstance(v, (int, float)):
                    values[h] = v
                else:
                    f = self._safe_float(v)
                    values[h] = f if f is not None else self._safe_str(v)
            data_rows.append({"time_h": t_f, "values": values})
            r += 1
            if r - 4 > 10:
                break

        mortality_table = {
            "section_label": section_label,
            "time_column": time_label,
            "column_headers": column_headers,
            "rows": data_rows,
        }

        # Discover sub-sections by walking the rest of column 2
        validity: Optional[Dict] = None
        statistic_text: Optional[Dict] = None
        ec_values: List[Dict] = []
        reach_classification: Optional[Dict] = None

        rr = r
        while rr <= ws.max_row:
            v = self._safe_str(ws.cell(row=rr, column=2).value)
            if v is None:
                rr += 1
                continue
            lv = v.lower()

            if lv.startswith("validity"):
                # Result at column D (col=4) of same row, detail string on next row
                result = self._safe_str(ws.cell(row=rr, column=4).value)
                detail = self._safe_str(ws.cell(row=rr + 1, column=2).value)
                validity = asdict(AcceptanceItem(label=v, detail=detail, result=result))
                rr += 2
                continue

            if "statistic" in lv and "analysis" in lv:
                # "Statistic analysis" / "Statistical analysis"
                lines: List[str] = []
                rrr = rr + 1
                while rrr <= ws.max_row:
                    nxt = self._safe_str(ws.cell(row=rrr, column=2).value)
                    if nxt is None:
                        break
                    nxt_l = nxt.lower()
                    if (nxt_l.startswith("calculated")
                            or "reach" in nxt_l
                            or nxt_l.startswith("hazardous")):
                        break
                    lines.append(nxt)
                    rrr += 1
                statistic_text = {"label": v, "lines": lines}
                rr = rrr
                continue

            if lv.startswith("calculated ec"):
                # e.g. "Calculated EC50" with description in D (col 4) and
                # numerical/text value on next row at B (col 2).
                m = re.search(r"ec\s*(\d+)", lv)
                name = f"EC{m.group(1)}" if m else v
                desc = self._safe_str(ws.cell(row=rr, column=4).value)
                value = self._safe_str(ws.cell(row=rr + 1, column=2).value)
                ec_values.append(asdict(EcValue(name=name, description=desc, value=value)))
                rr += 2
                continue

            if "reach" in lv and "classification" in lv:
                # Subsection title; the next several rows contain
                # "Category Acute X", "48 h EC50", "HAZARDOUS?".
                category = None
                ec_threshold_label = None
                ec_threshold_value = None
                hazardous_label = None
                hazardous_value = None
                rrr = rr + 1
                while rrr <= ws.max_row and rrr - rr <= 6:
                    nxt_b = self._safe_str(ws.cell(row=rrr, column=2).value)
                    nxt_d = self._safe_str(ws.cell(row=rrr, column=4).value)
                    if nxt_b is None and nxt_d is None:
                        rrr += 1
                        continue
                    if nxt_b:
                        nxt_b_l = nxt_b.lower()
                        if nxt_b_l.startswith("category"):
                            category = nxt_b
                        elif "ec" in nxt_b_l and ec_threshold_label is None:
                            ec_threshold_label = nxt_b
                            ec_threshold_value = nxt_d
                        elif nxt_b_l.startswith("hazardous"):
                            hazardous_label = nxt_b
                            hazardous_value = nxt_d
                    rrr += 1
                reach_classification = {
                    "title": v,
                    "category": category,
                    "ec_threshold_label": ec_threshold_label,
                    "ec_threshold_value": ec_threshold_value,
                    "hazardous_label": hazardous_label,
                    "hazardous_value": hazardous_value,
                }
                rr = rrr
                continue

            rr += 1

        return {
            "available": True,
            "section_label": section_label,
            "mortality_table": mortality_table,
            "validity": validity,
            "statistic_text": statistic_text,
            "ec_values": ec_values,
            "reach_classification": reach_classification,
        }

    # ------------------------------------------------------------------
    # Orchestrator
    # ------------------------------------------------------------------
    def parse_all_data(self) -> Dict:
        try:
            test_details = {
                "work_package": self.extract_work_package_data(),
                "material": self.extract_material_data(),
                "application": self.extract_application_data(),
                "test_conditions": self.extract_test_conditions_data(),
                "analysis": self.extract_analysis_data(),
                "replications": [self.extract_replication()],
            }
            raw_blocks = self.extract_raw_data()
            processed = self.extract_processed_data()
            final = self.extract_final_results()

            # Attach any accumulated parser warnings
            if self.parser_warnings:
                test_details["parser_warnings"] = list(self.parser_warnings)

            return {
                "test_details": test_details,
                "replications": raw_blocks,
                "processed_data": processed,
                "final_results": final,
            }
        except Exception as e:
            logger.error(f"Failed to parse: {e}\n{traceback.format_exc()}")
            raise


# ============================================================================
# Convenience entry point
# ============================================================================
def parse_excel_rotifier(file_path: str, sheet_name: str = "Test Details") -> Dict:
    parser = RotifierParser(file_path, sheet_name)
    return parser.parse_all_data()


# ============================================================================
# Main
# ============================================================================
if __name__ == "__main__":
    import sys

    fp = sys.argv[1] if len(sys.argv) > 1 else "/mnt/user-data/uploads/CMS_WP3_Rotifier_7b.xlsx"
    parsed = parse_excel_rotifier(fp)
    print("=" * 80)
    print("Rotifier parser — summary")
    print("=" * 80)
    td = parsed["test_details"]
    print(f"  WP                 : {td['work_package'].get('wp_name')}")
    print(f"  Material           : {td['material'].get('material_identifier')}")
    print(f"  Test type          : {td['work_package'].get('test_type')}")
    print(f"  Start date         : {td['work_package'].get('test_start_date')}")
    print(f"  End date           : {td['work_package'].get('test_end_date')}")
    print(f"  Concentrations     : {td['application'].get('concentrations_tested')}")
    print(f"  Replicates per conc: {td['application'].get('replicates_per_concentration')}")
    print(f"  Raw blocks         : {len(parsed.get('replications', []))}")
    if parsed["replications"]:
        print(f"  Raw rows           : {len(parsed['replications'][0]['rows'])}")
    print(f"  Processed avail    : {parsed['processed_data'].get('available')}")
    print(f"  Final results      : {parsed['final_results'].get('available')}")
    if td.get("parser_warnings"):
        print(f"  Warnings           : {len(td['parser_warnings'])}")
    print()
    print("First 6000 chars of JSON:")
    print("-" * 80)
    print(json.dumps(parsed, indent=2, default=str)[:6000])