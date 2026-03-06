import openpyxl
import re
import json
from dataclasses import dataclass, asdict, field
from typing import List, Dict, Optional, Union, Any
import logging
import traceback
from datetime import datetime, timedelta

logging.basicConfig(level=logging.DEBUG, format="%(asctime)s - %(levelname)s - %(message)s")
logging.getLogger("multipart.multipart").setLevel(logging.WARNING)
logger = logging.getLogger(__name__)


# =========================================================
# Dataclasses
# =========================================================

@dataclass
class Scientist:
    name: Optional[str] = None
    email: Optional[str] = None


@dataclass
class WorkPackageData:
    wp_name: Optional[str] = None
    partner: Optional[str] = None
    laboratory_name: Optional[str] = None
    full_test_name: Optional[str] = None
    test_acronym: Optional[str] = None
    test_type: Optional[str] = None
    endpoint: Optional[str] = None
    endpoint_outcome: Optional[str] = None
    sop: Optional[str] = None
    path: Optional[str] = None
    lead_scientists: List[Scientist] = field(default_factory=list)
    assay_scientists: List[Scientist] = field(default_factory=list)


@dataclass
class MaterialData:
    material_identifier: Optional[str] = None
    erm_id: Optional[str] = None
    material_name: Optional[str] = None
    core_chemistry: Optional[str] = None
    cas_no: Optional[str] = None
    cas_for_core: Optional[str] = None
    material_supplier: Optional[str] = None
    material_state: Optional[str] = None
    batch: Optional[str] = None
    vial: Optional[str] = None
    preparation_date: Optional[str] = None
    stock_concentration: Optional[str] = None
    molecular_weight: Optional[str] = None
    particles_stock: Optional[str] = None


@dataclass
class SamplePreparationData:
    dispersion_protocol: Optional[str] = None
    dispersion_technique: Optional[str] = None
    dispersion_agent: Optional[str] = None
    additives: Optional[str] = None
    dispersed_in_culture_medium: Optional[str] = None
    aids_used_to_disperse: Optional[str] = None
    sonication_bath: Optional[str] = None
    sonication_tip: Optional[str] = None
    time_duration: Optional[str] = None
    rcf: Optional[Union[int, float, str]] = None
    deposition: Optional[str] = None


@dataclass
class InstrumentationData:
    xray_source: Optional[str] = None
    analyser: Optional[str] = None
    pass_energy: Optional[str] = None
    take_off_angle: Optional[str] = None
    spot_size: Optional[str] = None
    charge_neutralizer: Optional[str] = None
    vacuum: Optional[str] = None
    notes_a: Optional[str] = None
    notes_b: Optional[str] = None
    notes_c: Optional[str] = None


@dataclass
class ReplicationMetadata:
    test_identifier_number: Optional[str] = None
    test_start_date: Optional[str] = None
    test_end_date: Optional[str] = None
    replicate_label: Optional[str] = None
    raw_sheet_name: Optional[str] = None
    processed_sheet_name: Optional[str] = None


@dataclass
class XPSSpectrumPoint:
    binding_energy_ev: Optional[float] = None
    intensity_counts_per_s: Optional[float] = None


@dataclass
class XPSRawDataBlock:
    metric_name: Optional[str] = None
    spectrum_label: Optional[str] = None
    raw_sheet_name: Optional[str] = None
    processed_sheet_name: Optional[str] = None
    point_count: Optional[int] = None
    binding_energy_unit: Optional[str] = "eV"
    intensity_unit: Optional[str] = "counts per s"
    min_binding_energy_ev: Optional[float] = None
    max_binding_energy_ev: Optional[float] = None
    min_intensity_counts_per_s: Optional[float] = None
    max_intensity_counts_per_s: Optional[float] = None
    spectrum_points: List[XPSSpectrumPoint] = field(default_factory=list)
    assay_notes: List[str] = field(default_factory=list)
    acquisition_mode: Optional[str] = None
    test_identifier: Optional[str] = None


@dataclass
class XPSProcessedDataBlock:
    metric_name: Optional[str] = None
    spectrum_label: Optional[str] = None
    processed_sheet_name: Optional[str] = None
    processed_sheet_has_numeric_data: Optional[bool] = None
    extracted_title: Optional[str] = None
    point_count_from_raw: Optional[int] = None
    intensity_min: Optional[float] = None
    intensity_max: Optional[float] = None
    identifier: Optional[str] = None


@dataclass
class XPSElectronicProperty:
    region_label: Optional[str] = None
    atomic_percent: Optional[float] = None


@dataclass
class XPSFinalResultRow:
    sample_identifier: Optional[str] = None
    result_type: Optional[str] = None
    spot_label: Optional[str] = None
    electronic_properties: List[XPSElectronicProperty] = field(default_factory=list)


# =========================================================
# Parser
# =========================================================

class XPSParser:
    def __init__(self, file_path: str, sheet_name: str = "Test_conditions"):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.parser_warnings: List[Dict[str, Any]] = []

        try:
            self.wb = openpyxl.load_workbook(file_path, data_only=True)
            self.ws = self.wb[sheet_name]
            logger.info("Successfully loaded XPS workbook: %s", file_path)
        except Exception as e:
            logger.error("Failed to load workbook or sheet '%s': %s", sheet_name, e)
            raise

        self.email_regex = r"^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$"

    # -----------------------------------------------------
    # Helpers
    # -----------------------------------------------------
    def normalize_key(self, key: Optional[str]) -> Optional[str]:
        if key is None:
            return None
        normalized = str(key).strip().lower()
        normalized = re.sub(r"[^a-z0-9]+", "_", normalized)
        normalized = re.sub(r"_+", "_", normalized).strip("_")
        return normalized

    def excel_date_to_string(self, value) -> Optional[str]:
        try:
            if isinstance(value, datetime):
                return value.strftime("%Y-%m-%d")

            if isinstance(value, (int, float)):
                base_date = datetime(1899, 12, 30)
                return (base_date + timedelta(days=float(value))).strftime("%Y-%m-%d")

            if isinstance(value, str):
                for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d", "%d.%m.%Y"):
                    try:
                        return datetime.strptime(value.strip(), fmt).strftime("%Y-%m-%d")
                    except ValueError:
                        pass

            return str(value).strip() if value not in (None, "") else None
        except Exception as e:
            logger.warning("Failed to convert date '%s': %s", value, e)
            return str(value).strip() if value not in (None, "") else None

    def _safe_float(self, value) -> Optional[float]:
        if value in (None, ""):
            return None
        try:
            if isinstance(value, str):
                value = value.replace(",", ".").strip()
            return float(value)
        except (ValueError, TypeError):
            return None

    def _get_first_value_in_row(self, row, start_col: int = 2, end_col: int = 6):
        for col_idx in range(start_col, end_col + 1):
            value = row[col_idx - 1].value
            if value is not None:
                return value
        return None

    def _sheet_key_values(self) -> List[Dict[str, Any]]:
        data = []

        for row_idx, row in enumerate(self.ws.iter_rows(min_row=1, max_col=6), start=1):
            key_cell = row[0].value
            if not key_cell:
                continue

            raw_key = str(key_cell).strip()
            key = self.normalize_key(raw_key)
            if not key:
                continue

            value = self._get_first_value_in_row(row, 2, 6)
            email = row[3].value if len(row) > 3 else None

            data.append({
                "row": row_idx,
                "raw_key": raw_key,
                "key": key,
                "value": value,
                "email": email,
            })

        return data

    def _extract_sheet_title(self, ws) -> Optional[str]:
        for row_idx in range(1, min(5, ws.max_row) + 1):
            for col_idx in range(1, min(5, ws.max_column) + 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                if isinstance(value, str) and value.strip():
                    return value.strip()
        return None

    def _extract_notes_after_marker(self, ws, marker: str = "Notes for the assay run") -> List[str]:
        notes: List[str] = []
        marker_row = None

        for row_idx in range(1, ws.max_row + 1):
            for col_idx in range(1, min(6, ws.max_column) + 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                if isinstance(value, str) and marker.lower() in value.lower():
                    marker_row = row_idx
                    break
            if marker_row:
                break

        if marker_row:
            for row_idx in range(marker_row + 1, min(marker_row + 10, ws.max_row) + 1):
                row_values = [
                    ws.cell(row=row_idx, column=col_idx).value
                    for col_idx in range(1, min(6, ws.max_column) + 1)
                ]
                parts = [str(v).strip() for v in row_values if v not in (None, "")]
                if parts:
                    notes.append(" | ".join(parts))

        return notes

    def _metric_name_from_sheet(self, sheet_name: str, title: Optional[str]) -> str:
        source = f"{sheet_name} {title or ''}".lower()

        if "survey" in source or "long range" in source:
            return "Survey Spectrum"
        if "na1s" in source:
            return "Na1s Core Level"
        if "o1s" in source:
            return "O1s Core Level"
        if "c1s" in source:
            return "C1s Core Level"
        if "au4f" in source:
            return "Au4f Core Level"
        if "si2p" in source:
            return "Si2p Core Level"

        return title or sheet_name

    def _find_matching_processed_sheet(self, raw_sheet_name: str) -> Optional[str]:
        raw_lower = raw_sheet_name.lower()

        candidates = {
            "survey": "Processed data_XPS_AuNPs1aU_Sur",
            "na1s": "Processed data_XPS_AuNPs1aU_Na",
            "o1s": "Processed data_XPS_AuNPs1aU_O",
            "c1s": "Processed data_XPS_AuNPs1aU_C",
            "au4f": "Processed data_XPS_AuNPs1aU_Au",
            "si2p": "Processed data_XPS_AuNPs1aU_Si",
        }

        for token, processed_sheet in candidates.items():
            if token in raw_lower and processed_sheet in self.wb.sheetnames:
                return processed_sheet

        return None

    # -----------------------------------------------------
    # Test details
    # -----------------------------------------------------
    def extract_work_package_data(self) -> Dict[str, Any]:
        rows = self._sheet_key_values()
        lead_scientists: List[Scientist] = []
        assay_scientists: List[Scientist] = []

        for row in rows:
            key = row["key"]
            value = row["value"]
            email = row["email"]

            if "lead_scientist" in key:
                lead_scientists.append(
                    Scientist(
                        name=value,
                        email=email if email and re.match(self.email_regex, str(email)) else None,
                    )
                )

            if "assay_test_work_conducted_by" in key:
                assay_scientists.append(
                    Scientist(
                        name=value,
                        email=email if email and re.match(self.email_regex, str(email)) else None,
                    )
                )

        result = WorkPackageData(
            wp_name=next((r["value"] for r in rows if r["key"] == "project_work_package"), None),
            partner=next((r["value"] for r in rows if r["key"] == "partner_conducting_test_assay"), None),
            laboratory_name=next((r["value"] for r in rows if r["key"] == "test_facility_laboratory_name"), None),
            full_test_name=next((r["value"] for r in rows if "full_name_of_test_assay" in r["key"]), None),
            test_acronym=next((r["value"] for r in rows if "short_name_or_acronym" in r["key"]), None),
            test_type=next((r["value"] for r in rows if "type_or_class_of_experimental" in r["key"]), None),
            endpoint=next((r["value"] for r in rows if "end_point_being_investigated" in r["key"]), None),
            endpoint_outcome=next((r["value"] for r in rows if "metric_s_used_to_assess" in r["key"]), None),
            sop=next((r["value"] for r in rows if "sop_s_for_test" in r["key"]), None),
            path=next((r["value"] for r in rows if "path_link_to_sop" in r["key"]), None),
            lead_scientists=lead_scientists,
            assay_scientists=assay_scientists,
        )
        return asdict(result)

    def extract_material_data(self) -> Dict[str, Any]:
        rows = self._sheet_key_values()

        result = MaterialData(
            material_identifier=next((r["value"] for r in rows if r["key"] == "sample_cms_internal_identifier"), None),
            erm_id=next((r["value"] for r in rows if "erm_identifier" in r["key"]), None),
            material_name=next((r["value"] for r in rows if r["key"] == "material_name"), None),
            core_chemistry=next((r["value"] for r in rows if "core_chemistry" in r["key"]), None),
            cas_no=next((r["value"] for r in rows if r["key"] == "cas_no"), None),
            cas_for_core=next((r["value"] for r in rows if "cas_no_for_core" in r["key"]), None),
            material_supplier=next((r["value"] for r in rows if r["key"] == "material_supplier"), None),
            material_state=next((r["value"] for r in rows if r["key"] == "material_state"), None),
            batch=next((r["value"] for r in rows if r["key"] == "batch"), None),
            vial=next((r["value"] for r in rows if r["key"] == "vial"), None),
            preparation_date=self.excel_date_to_string(
                next((r["value"] for r in rows if "date_of_preparation" in r["key"]), None)
            ),
            stock_concentration=next((r["value"] for r in rows if "stock_concentration" in r["key"]), None),
            molecular_weight=next((r["value"] for r in rows if "molecular_weight" in r["key"]), None),
            particles_stock=next((r["value"] for r in rows if "no_of_particles_in_stock" in r["key"]), None),
        )
        return asdict(result)

    def extract_sample_preparation_data(self) -> Dict[str, Any]:
        ws = self.ws
        result = SamplePreparationData(
            dispersion_protocol=ws.cell(row=48, column=2).value,
            dispersion_technique=ws.cell(row=49, column=2).value,
            dispersion_agent=ws.cell(row=50, column=2).value,
            additives=ws.cell(row=51, column=2).value,
            dispersed_in_culture_medium=ws.cell(row=52, column=2).value,
            aids_used_to_disperse=ws.cell(row=53, column=2).value,
            sonication_bath=ws.cell(row=53, column=4).value,
            sonication_tip=ws.cell(row=53, column=6).value,
            time_duration=ws.cell(row=54, column=2).value,
            rcf=ws.cell(row=55, column=2).value,
            deposition=ws.cell(row=56, column=2).value,
        )
        return asdict(result)

    def extract_instrumentation_data(self) -> Dict[str, Any]:
        ws = self.ws
        result = InstrumentationData(
            xray_source=ws.cell(row=58, column=2).value,
            analyser=ws.cell(row=59, column=2).value,
            pass_energy=ws.cell(row=60, column=2).value,
            take_off_angle=ws.cell(row=61, column=2).value,
            spot_size=ws.cell(row=62, column=2).value,
            charge_neutralizer=ws.cell(row=63, column=2).value,
            vacuum=ws.cell(row=64, column=2).value,
            notes_a=ws.cell(row=65, column=2).value,
            notes_b=ws.cell(row=66, column=2).value,
            notes_c=ws.cell(row=67, column=2).value,
        )
        return asdict(result)

    # -----------------------------------------------------
    # Run metadata
    # -----------------------------------------------------
    def extract_replication_metadata(self) -> List[Dict[str, Any]]:
        raw_sheets = sorted([s for s in self.wb.sheetnames if s.lower().startswith("raw_data")])
        test_start_date = self.excel_date_to_string(self.ws.cell(row=36, column=2).value)
        metadata: List[Dict[str, Any]] = []

        for raw_sheet in raw_sheets:
            raw_ws = self.wb[raw_sheet]
            processed_sheet = self._find_matching_processed_sheet(raw_sheet)
            title = self._extract_sheet_title(raw_ws)
            metric_name = self._metric_name_from_sheet(raw_sheet, title)

            test_identifier = raw_ws.cell(row=1, column=1).value
            if test_identifier is not None:
                test_identifier = str(test_identifier).strip()

            metadata.append(asdict(ReplicationMetadata(
                test_identifier_number=test_identifier,
                test_start_date=test_start_date,
                test_end_date=None,
                replicate_label=metric_name,
                raw_sheet_name=raw_sheet,
                processed_sheet_name=processed_sheet,
            )))

        return metadata

    # -----------------------------------------------------
    # Raw data
    # -----------------------------------------------------
    def _extract_raw_block(self, raw_sheet_name: str, processed_sheet_name: Optional[str]) -> Dict[str, Any]:
        ws = self.wb[raw_sheet_name]
        title = self._extract_sheet_title(ws)
        metric_name = self._metric_name_from_sheet(raw_sheet_name, title)

        test_identifier = ws.cell(row=1, column=1).value
        if test_identifier is not None:
            test_identifier = str(test_identifier).strip()

        acquisition_mode = ws.cell(row=1, column=3).value

        spectrum_points: List[Dict[str, Any]] = []
        binding_vals: List[float] = []
        intensity_vals: List[float] = []

        for row_idx in range(3, ws.max_row + 1):
            binding_energy = self._safe_float(ws.cell(row=row_idx, column=3).value)
            intensity = self._safe_float(ws.cell(row=row_idx, column=4).value)

            if binding_energy is not None and intensity is not None:
                spectrum_points.append(asdict(XPSSpectrumPoint(
                    binding_energy_ev=binding_energy,
                    intensity_counts_per_s=intensity,
                )))
                binding_vals.append(binding_energy)
                intensity_vals.append(intensity)

        assay_notes = self._extract_notes_after_marker(ws)

        result = XPSRawDataBlock(
            metric_name=metric_name,
            spectrum_label=title,
            raw_sheet_name=raw_sheet_name,
            processed_sheet_name=processed_sheet_name,
            point_count=len(spectrum_points),
            binding_energy_unit="eV",
            intensity_unit="counts per s",
            min_binding_energy_ev=min(binding_vals) if binding_vals else None,
            max_binding_energy_ev=max(binding_vals) if binding_vals else None,
            min_intensity_counts_per_s=min(intensity_vals) if intensity_vals else None,
            max_intensity_counts_per_s=max(intensity_vals) if intensity_vals else None,
            spectrum_points=spectrum_points,
            assay_notes=assay_notes,
            acquisition_mode=acquisition_mode,
            test_identifier=test_identifier,
        )
        return asdict(result)

    def extract_raw_data(self) -> List[Dict[str, Any]]:
        raw_sheets = sorted([s for s in self.wb.sheetnames if s.lower().startswith("raw_data")])
        raw_blocks: List[Dict[str, Any]] = []

        for raw_sheet in raw_sheets:
            processed_sheet = self._find_matching_processed_sheet(raw_sheet)
            raw_blocks.append(self._extract_raw_block(raw_sheet, processed_sheet))

        return raw_blocks

    # -----------------------------------------------------
    # Processed data
    # -----------------------------------------------------
    def extract_processed_data(self, raw_blocks: List[Dict[str, Any]]) -> Dict[str, Any]:
        processed_sheets = sorted([s for s in self.wb.sheetnames if s.lower().startswith("processed data")])
        blocks: List[Dict[str, Any]] = []

        for raw_block in raw_blocks:
            processed_sheet_name = raw_block.get("processed_sheet_name")
            extracted_title = None
            has_numeric_data = False
            identifier = None

            if processed_sheet_name and processed_sheet_name in self.wb.sheetnames:
                pws = self.wb[processed_sheet_name]
                extracted_title = self._extract_sheet_title(pws)
                identifier = pws.cell(row=1, column=1).value

                for row_idx in range(2, min(100, pws.max_row) + 1):
                    for col_idx in range(1, min(10, pws.max_column) + 1):
                        value = pws.cell(row=row_idx, column=col_idx).value
                        if isinstance(value, (int, float)):
                            has_numeric_data = True
                            break
                    if has_numeric_data:
                        break

                if not has_numeric_data:
                    self.parser_warnings.append({
                        "type": "empty_processed_sheet",
                        "sheet": processed_sheet_name,
                        "note": f"Processed sheet '{processed_sheet_name}' contains no numeric processed data beyond the title row.",
                    })

            blocks.append(asdict(XPSProcessedDataBlock(
                metric_name=raw_block.get("metric_name"),
                spectrum_label=raw_block.get("spectrum_label"),
                processed_sheet_name=processed_sheet_name,
                processed_sheet_has_numeric_data=has_numeric_data,
                extracted_title=extracted_title,
                point_count_from_raw=raw_block.get("point_count"),
                intensity_min=raw_block.get("min_intensity_counts_per_s"),
                intensity_max=raw_block.get("max_intensity_counts_per_s"),
                identifier=str(identifier).strip() if identifier not in (None, "") else None,
            )))

        return {
            "spectra": blocks,
            "processed_sheet_count": len(processed_sheets),
            "processed_numeric_data_available": any(
                block.get("processed_sheet_has_numeric_data") for block in blocks
            ),
        }

    # -----------------------------------------------------
    # Final results
    # -----------------------------------------------------
    def extract_final_results(self) -> List[Dict[str, Any]]:
        if "Final results" not in self.wb.sheetnames:
            self.parser_warnings.append({
                "type": "missing_sheet",
                "sheet": "Final results",
                "note": "Final results sheet not found.",
            })
            return []

        ws = self.wb["Final results"]
        sample_identifier = ws.cell(row=1, column=1).value
        result_type = ws.cell(row=1, column=2).value

        headers: List[tuple[int, str]] = []
        for col_idx in range(3, ws.max_column + 1):
            value = ws.cell(row=3, column=col_idx).value
            if value not in (None, ""):
                headers.append((col_idx, str(value).strip()))

        results: List[Dict[str, Any]] = []

        for row_idx in range(4, ws.max_row + 1):
            spot_label = ws.cell(row=row_idx, column=2).value
            if spot_label in (None, ""):
                continue

            electronic_properties: List[XPSElectronicProperty] = []

            for col_idx, label in headers:
                atomic_percent = self._safe_float(ws.cell(row=row_idx, column=col_idx).value)
                if atomic_percent is not None:
                    electronic_properties.append(XPSElectronicProperty(
                        region_label=label,
                        atomic_percent=atomic_percent,
                    ))

            results.append(asdict(XPSFinalResultRow(
                sample_identifier=str(sample_identifier).strip() if sample_identifier not in (None, "") else None,
                result_type=str(result_type).strip() if result_type not in (None, "") else None,
                spot_label=str(spot_label).strip(),
                electronic_properties=electronic_properties,
            )))

        return results

    # -----------------------------------------------------
    # Statistical analysis
    # -----------------------------------------------------
    def extract_statistical_analysis(self) -> Dict[str, Any]:
        return {
            "available": False,
            "notes": "No dedicated statistical analysis section was found in this XPS workbook.",
        }

    # -----------------------------------------------------
    # Parse all
    # -----------------------------------------------------
    def parse_all_data(self) -> Dict[str, Union[Dict, List]]:
        try:
            work_package_data = self.extract_work_package_data()
            material_data = self.extract_material_data()
            sample_preparation_data = self.extract_sample_preparation_data()
            instrumentation_data = self.extract_instrumentation_data()

            replication_metadata = self.extract_replication_metadata()
            raw_data = self.extract_raw_data()
            processed_data = self.extract_processed_data(raw_data)
            final_results = self.extract_final_results()
            statistical_analysis = self.extract_statistical_analysis()

            parsed_data = {
                "test_details": {
                    "work_package": work_package_data,
                    "material": material_data,
                    "sample_preparation": sample_preparation_data,
                    "instrumentation": instrumentation_data,
                },
                "replication_metadata": replication_metadata,
                "replications": raw_data,
                "processed_data": processed_data,
                "final_results": final_results,
                "statistical_analysis": statistical_analysis,
            }

            if self.parser_warnings:
                parsed_data["parser_warnings"] = self.parser_warnings

            logger.info("FINAL XPS JSON:\n%s", json.dumps(parsed_data, indent=2, default=str))
            return parsed_data

        except Exception as e:
            logger.error("Error parsing XPS file: %s\n%s", e, traceback.format_exc())
            raise


def parse_excel_xps(file_path: str, sheet_name: str = "Test_conditions") -> Dict[str, Union[Dict, List]]:
    try:
        parser = XPSParser(file_path, sheet_name)
        return parser.parse_all_data()
    except Exception as e:
        logger.error("Error in parse_excel_xps: %s\n%s", e, traceback.format_exc())
        raise


if __name__ == "__main__":
    file_path = "/mnt/data/CMS_WP4_1a_AuNP_XPS.xlsx"
    parsed_data = parse_excel_xps(file_path)

    print("=" * 70)
    print("XPS PARSER OUTPUT SUMMARY")
    print("=" * 70)

    wp = parsed_data["test_details"]["work_package"]
    print(f"\nWP: {wp['wp_name']}, Partner: {wp['partner']}")
    print(f"Test: {wp['test_acronym']} - {wp['full_test_name']}")

    print(f"\nReplication Metadata Count: {len(parsed_data['replication_metadata'])}")
    print(f"Raw Blocks Count: {len(parsed_data['replications'])}")
    print(f"Final Results Count: {len(parsed_data['final_results'])}")

    print("\nJSON Preview:")
    print(json.dumps(parsed_data, indent=2, default=str)[:4000])