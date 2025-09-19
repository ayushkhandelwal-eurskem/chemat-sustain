import openpyxl
import re
from dataclasses import dataclass, asdict, field
from typing import List, Dict, Optional, Union
import logging
import traceback
from datetime import datetime, timedelta
from difflib import SequenceMatcher

# Set up logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
logging.getLogger("multipart.multipart").setLevel(logging.WARNING)
logger = logging.getLogger(__name__)

# Dataclasses tailored for HR-STEM
@dataclass
class Scientist:
    name: Optional[str] = None
    email: Optional[str] = None

@dataclass
class WorkPackageData:
    wp_name: Optional[str] = None
    partner: Optional[str] = None
    full_test_name: Optional[str] = None
    test_acronym: Optional[str] = None
    test_type: Optional[str] = None
    endpoint: Optional[str] = None
    endpoint_outcome: Optional[str] = None
    sop: Optional[str] = None
    path: Optional[str] = None
    lead_scientists: List[Scientist] = field(default_factory=list)
    assay_scientists: List[Scientist] = field(default_factory=list)

@dataclass
class MaterialData:
    material_identifier: Optional[str] = None
    erm_id: Optional[str] = None
    core_chemistry: Optional[str] = None
    material_name: Optional[str] = None
    material_state: Optional[str] = None
    cas: Optional[str] = None
    casforcore: Optional[str] = None
    batch: Optional[str] = None
    preparation_date: Optional[str] = None
    particles_stock: Optional[str] = None
    molar_concentration: Optional[str] = None

@dataclass
class ReplicationData:
    test_identifier_number: Optional[str] = None
    test_start_date: Optional[str] = None
    test_end_date: Optional[str] = None

@dataclass
class SamplePreparationData:
    dispersion_protocol: Optional[str] = None
    dispersion_technique: Optional[str] = None
    dispersion_medium: Optional[str] = None
    sonicator_type: Optional[str] = None
    power: Optional[str] = None
    sonication_time: Optional[str] = None
    tip_thickness: Optional[str] = None
    tip_composition: Optional[str] = None
    ultrasonic_bath_size: Optional[str] = None
    sample_volume: Optional[str] = None
    final_concentration: Optional[str] = None
    additional_info: Optional[str] = None

@dataclass
class InstrumentationData:
    electron_microscope_type: Optional[str] = None
    sample_grid: Optional[str] = None
    acceleration_voltage: Optional[str] = None
    tilt_angle: Optional[str] = None
    spot_size: Optional[str] = None
    aperture: Optional[str] = None

@dataclass
class ParticleMeasurement:
    feret_min: Optional[float] = None
    length: Optional[float] = None
    feret_max: Optional[float] = None

@dataclass
class ProcessedParticle:
    aspect_ratio: Optional[float] = None
    ecd_diameter: Optional[float] = None
    radius_ecd: Optional[float] = None
    roundness: Optional[float] = None

@dataclass
class HistogramBin:
    bin_start: float
    bin_end: float
    count: int

@dataclass
class HistogramData:
    feret_min: List[HistogramBin] = field(default_factory=list)
    length: List[HistogramBin] = field(default_factory=list)
    feret_max: List[HistogramBin] = field(default_factory=list)

@dataclass
class RunMetrics:
    feret_min_mean: Optional[float] = None
    feret_min_std: Optional[float] = None
    feret_min_median: Optional[float] = None
    length_mean: Optional[float] = None
    length_std: Optional[float] = None
    length_median: Optional[float] = None
    feret_max_mean: Optional[float] = None
    feret_max_std: Optional[float] = None
    feret_max_median: Optional[float] = None
    ecd_mean: Optional[float] = None
    ecd_std: Optional[float] = None
    ecd_median: Optional[float] = None
    aspect_ratio_mean: Optional[float] = None
    aspect_ratio_std: Optional[float] = None
    aspect_ratio_median: Optional[float] = None
    roundness_mean: Optional[float] = None
    roundness_std: Optional[float] = None
    roundness_median: Optional[float] = None
    density: Optional[float] = None
    volume_np_nm3: Optional[float] = None
    volume_np_cm3: Optional[float] = None
    mass_np_g: Optional[float] = None
    mass_colloid: Optional[float] = None
    c_colloid_percent: Optional[float] = None
    no_particles_stock: Optional[float] = None

@dataclass
class RawData:
    run_number: int
    particles: List[ParticleMeasurement] = field(default_factory=list)

@dataclass
class ProcessedData:
    run_number: int
    processed_particles: List[ProcessedParticle] = field(default_factory=list)
    metrics: RunMetrics = field(default_factory=RunMetrics)
    histogram: HistogramData = field(default_factory=HistogramData)

@dataclass
class StatisticTableEntry:
    metric: str
    mean: float
    std_dev: float
    median: float

@dataclass
class ResultsData:
    feret_min_mean: Optional[float] = None
    feret_min_std: Optional[float] = None
    feret_min_median: Optional[float] = None
    length_mean: Optional[float] = None
    length_std: Optional[float] = None
    length_median: Optional[float] = None
    feret_max_mean: Optional[float] = None
    feret_max_std: Optional[float] = None
    feret_max_median: Optional[float] = None
    ecd_mean: Optional[float] = None
    ecd_std: Optional[float] = None
    ecd_median: Optional[float] = None
    aspect_ratio_mean: Optional[float] = None
    aspect_ratio_std: Optional[float] = None
    aspect_ratio_median: Optional[float] = None
    roundness_mean: Optional[float] = None
    roundness_std: Optional[float] = None
    roundness_median: Optional[float] = None
    statistic_table: List[StatisticTableEntry] = field(default_factory=list)

class HRSTEMParser:
    def __init__(self, file_path: str, sheet_name: str = "Test Information"):
        self.file_path = file_path
        self.sheet_name = sheet_name
        try:
            self.wb = openpyxl.load_workbook(file_path, data_only=True)
            self.ws = self.wb[sheet_name]
            logger.info(f"Successfully loaded Excel file: {file_path}")
        except Exception as e:
            logger.error(f"Failed to load workbook or sheet {sheet_name}: {e}")
            raise
        self.email_regex = r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$'

    def normalize_key(self, key: Optional[str]) -> Optional[str]:
        if key:
            normalized = str(key).strip().lower() if key is not None else ""
            normalized = re.sub(r'[^a-z0-9]', '_', normalized)
            normalized = re.sub(r'_+', '_', normalized).strip('_')
            return normalized
        return None

    def split_value_unit(self, value: Optional[str]) -> tuple[Optional[Union[str, float]], Optional[str]]:
        if isinstance(value, str) and " " in value:
            parts = value.split(" ", 1)
            try:
                numeric = float(parts[0]) if '.' in parts[0] else int(parts[0])
                return numeric, parts[1]
            except (ValueError, IndexError):
                return parts[0], parts[1] if len(parts) > 1 else None
        return value, None

    def excel_date_to_string(self, value: Optional[Union[float, str]]) -> Optional[str]:
        try:
            if isinstance(value, (int, float)):
                base_date = datetime(1899, 12, 30)
                delta = timedelta(days=float(value))
                return (base_date + delta).strftime("%Y-%m-%d")
            elif isinstance(value, str):
                for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"]:
                    try:
                        return datetime.strptime(value, fmt).strftime("%Y-%m-%d")
                    except ValueError:
                        pass
            return str(value) if value else None
        except Exception as e:
            logger.warning(f"Failed to convert date {value}: {e}")
            return None

    def extract_work_package_data(self):
        data = []
        lead_scientists = []
        assay_scientists = []

        for row in self.ws.iter_rows():
            key_cell = row[0].value
            value_cell = row[1].value if len(row) > 1 else None
            email_cell = row[3].value if len(row) > 3 else None
            comment = row[0].comment

            if comment:
                key = self.normalize_key(key_cell)
                if not key:
                    continue
                entry = {"Key": key, "Value": value_cell}

                if email_cell and re.match(self.email_regex, str(email_cell)):
                    entry["Email"] = email_cell

                if key and "lead_scientist_contact_for_test" in key:
                    lead_scientists.append(asdict(Scientist(name=value_cell, email=email_cell)))
                if key and "assay_test_work_conducted_by" in key:
                    assay_scientists.append(asdict(Scientist(name=value_cell, email=email_cell)))

                data.append(entry)

        wp_data = asdict(WorkPackageData(
            wp_name=next((d["Value"] for d in data if d["Key"] == "project_work_package"), None),
            partner=next((d["Value"] for d in data if d["Key"] == "partner_conducting_test_assay"), None),
            full_test_name=next((d["Value"] for d in data if d["Key"] == "full_name_of_test_assay_add_oecd_test_ref_id_if_app"), None),
            test_acronym=next((d["Value"] for d in data if d["Key"] == "short_name_or_acronym_for_test_assay"), None),
            test_type=next((d["Value"] for d in data if d["Key"] == "type_or_class_of_experimental_test_as_used_here"), None),
            endpoint=next((d["Value"] for d in data if d["Key"] == "end_point_being_investigated_assessed_by_the_test"), None),
            endpoint_outcome=next((d["Value"] for d in data if d["Key"] == "metric_s_used_to_assess_end_point_outcome_response"), None),
            sop=next((d["Value"] for d in data if d["Key"] == "sop_s_for_test_ref_project_or_other_doc_title_id"), None),
            path=next((d["Value"] for d in data if d["Key"] == "path_link_to_sop_protocol_on_proj_server_web_where_applic"), None),
            lead_scientists=lead_scientists,
            assay_scientists=assay_scientists
        ))
        return wp_data

    def extract_material_data(self):
        data = []
        expected_keys = [
            "sample_cms_internal_identifier",
            "erm_identifier_number",
            "core_chemistry",
            "material_name",
            "material_state",
            "cas_no",
            "cas_for_core",
            "batch",
            "date_of_preparation",
            "no_of_particles_in_stock",
            "molar_concentration"
        ]
        unmatched_keys = set(expected_keys)
        potential_matches = []

        logger.debug("Test Information sheet (columns A:E, rows 1:50) for Material Data:")
        for row_idx in range(1, min(self.ws.max_row + 1, 51)):
            row = self.ws[row_idx]
            values = [str(cell.value)[:30] if cell.value else "None" for cell in row[:5]]
            has_comment = "Y" if row[0].comment else "N"
            logger.debug(f"Row {row_idx} [Comment: {has_comment}]: {values}")

        for row_idx, row in enumerate(self.ws.iter_rows(min_row=1, max_col=5), start=1):
            key_cell = row[0].value
            if not key_cell:
                continue
            raw_key = str(key_cell)
            key = self.normalize_key(raw_key)
            if not key:
                continue

            value = None
            value_col = None
            for col_idx in range(1, 5):
                if col_idx < len(row) and row[col_idx].value is not None:
                    value = row[col_idx].value
                    value_col = col_idx + 1
                    break

            logger.debug(f"Row {row_idx}: Raw Key='{raw_key}', Normalized Key='{key}', Value='{value}', Value Col={value_col}")

            if key in expected_keys:
                unmatched_keys.discard(key)
                data.append({"Key": key, "Value": value})
                logger.debug(f"Exact match: Key='{key}', Value='{value}' at row {row_idx}, col {value_col}")
            else:
                for expected_key in expected_keys:
                    similarity = SequenceMatcher(None, key, expected_key).ratio()
                    if similarity > 0.85:
                        unmatched_keys.discard(expected_key)
                        data.append({"Key": expected_key, "Value": value})
                        logger.debug(f"Fuzzy match: Key='{key}' matched '{expected_key}' (similarity={similarity:.2f}), Value='{value}' at row {row_idx}, col {value_col}")
                    elif similarity > 0.5:
                        potential_matches.append(f"Potential match for '{expected_key}': '{raw_key}' at row {row_idx}, col 1, similarity={similarity:.2f}")

        if potential_matches:
            logger.debug("Potential key matches for Material Data:")
            for match in potential_matches:
                logger.debug(match)
        if unmatched_keys:
            logger.warning(f"Unmatched material data keys: {unmatched_keys}")

        material_data = asdict(MaterialData(
            material_identifier=next((d["Value"] for d in data if d["Key"] == "sample_cms_internal_identifier"), None),
            erm_id=next((d["Value"] for d in data if d["Key"] == "erm_identifier_number"), None),
            core_chemistry=next((d["Value"] for d in data if d["Key"] == "core_chemistry"), None),
            material_name=next((d["Value"] for d in data if d["Key"] == "material_name"), None),
            material_state=next((d["Value"] for d in data if d["Key"] == "material_state"), None),
            cas=next((d["Value"] for d in data if d["Key"] == "cas_no"), None),
            casforcore=next((d["Value"] for d in data if d["Key"] == "cas_for_core"), None),
            batch=next((d["Value"] for d in data if d["Key"] == "batch"), None),
            preparation_date=self.excel_date_to_string(next((d["Value"] for d in data if d["Key"] == "date_of_preparation"), None)),
            particles_stock=next((d["Value"] for d in data if d["Key"] == "no_of_particles_in_stock"), None),
            molar_concentration=next((d["Value"] for d in data if d["Key"] == "molar_concentration"), None)
        ))
        logger.debug(f"Material data extracted: {material_data}")
        return material_data

    def extract_sample_preparation_data(self):
        data = []
        expected_keys = [
            "specify_standard_dispersion_protocol_used",
            "or_otherwise_specify_dispersion_technique_used",
            "dispersion_dilution_medium",
            "sonicator_type",
            "power_w",
            "sonication_time_secs",
            "tip_thickness_mm",
            "tip_composition",
            "size_of_ultrasonic_bath_water_volume_dm3",
            "sample_volume",
            "final_sample_concentration_mg_l_or_ppm",
            "additional_information"
        ]
        unmatched_keys = set(expected_keys)
        potential_matches = []

        logger.debug("Test Information sheet (columns A:E, rows 1:50) for Sample Preparation Data:")
        for row_idx in range(1, min(self.ws.max_row + 1, 51)):
            row = self.ws[row_idx]
            values = [str(cell.value)[:30] if cell.value else "None" for cell in row[:5]]
            has_comment = "Y" if row[0].comment else "N"
            logger.debug(f"Row {row_idx} [Comment: {has_comment}]: {values}")

        for row_idx, row in enumerate(self.ws.iter_rows(min_row=1, max_col=5), start=1):
            key_cell = row[0].value
            if not key_cell:
                continue
            raw_key = str(key_cell)
            key = self.normalize_key(raw_key)
            if not key:
                continue

            value = None
            value_col = None
            for col_idx in range(1, 5):
                if col_idx < len(row) and row[col_idx].value is not None:
                    value = row[col_idx].value
                    value_col = col_idx + 1
                    break

            logger.debug(f"Row {row_idx}: Raw Key='{raw_key}', Normalized Key='{key}', Value='{value}', Value Col={value_col}")

            if key in expected_keys:
                unmatched_keys.discard(key)
                data.append({"Key": key, "Value": value})
                logger.debug(f"Exact match: Key='{key}', Value='{value}' at row {row_idx}, col {value_col}")
            else:
                for expected_key in expected_keys:
                    similarity = SequenceMatcher(None, key, expected_key).ratio()
                    if similarity > 0.85:
                        unmatched_keys.discard(expected_key)
                        data.append({"Key": expected_key, "Value": value})
                        logger.debug(f"Fuzzy match: Key='{key}' matched '{expected_key}' (similarity={similarity:.2f}), Value='{value}' at row {row_idx}, col {value_col}")
                    elif similarity > 0.5:
                        potential_matches.append(f"Potential match for '{expected_key}': '{raw_key}' at row {row_idx}, col 1, similarity={similarity:.2f}")

        if potential_matches:
            logger.debug("Potential key matches for Sample Preparation Data:")
            for match in potential_matches:
                logger.debug(match)
        if unmatched_keys:
            logger.warning(f"Unmatched sample preparation data keys: {unmatched_keys}")

        sample_preparation_data = asdict(SamplePreparationData(
            dispersion_protocol=next((d["Value"] for d in data if d["Key"] == "specify_standard_dispersion_protocol_used"), None),
            dispersion_technique=next((d["Value"] for d in data if d["Key"] == "or_otherwise_specify_dispersion_technique_used"), None),
            dispersion_medium=next((d["Value"] for d in data if d["Key"] == "dispersion_dilution_medium"), None),
            sonicator_type=next((d["Value"] for d in data if d["Key"] == "sonicator_type"), None),
            power=next((d["Value"] for d in data if d["Key"] == "power_w"), None),
            sonication_time=next((d["Value"] for d in data if d["Key"] == "sonication_time_secs"), None),
            tip_thickness=next((d["Value"] for d in data if d["Key"] == "tip_thickness_mm"), None),
            tip_composition=next((d["Value"] for d in data if d["Key"] == "tip_composition"), None),
            ultrasonic_bath_size=next((d["Value"] for d in data if d["Key"] == "size_of_ultrasonic_bath_water_volume_dm3"), None),
            sample_volume=next((d["Value"] for d in data if d["Key"] == "sample_volume"), None),
            final_concentration=next((d["Value"] for d in data if d["Key"] == "final_sample_concentration_mg_l_or_ppm"), None),
            additional_info=next((d["Value"] for d in data if d["Key"] == "additional_information"), None)
        ))
        logger.debug(f"Sample preparation data extracted: {sample_preparation_data}")
        return sample_preparation_data

    def extract_instrumentation_data(self):
        data = []
        expected_keys = [
            "type_of_the_electron_microscope",
            "sample_grid_sample_holder",
            "acceleration_voltage",
            "tilt_angle",
            "spot_size",
            "aperture"
        ]
        unmatched_keys = set(expected_keys)
        potential_matches = []

        logger.debug("Test Information sheet (columns A:E, rows 1:50) for Instrumentation Data:")
        for row_idx in range(1, min(self.ws.max_row + 1, 51)):
            row = self.ws[row_idx]
            values = [str(cell.value)[:30] if cell.value else "None" for cell in row[:5]]
            has_comment = "Y" if row[0].comment else "N"
            logger.debug(f"Row {row_idx} [Comment: {has_comment}]: {values}")

        for row_idx, row in enumerate(self.ws.iter_rows(min_row=1, max_col=5), start=1):
            key_cell = row[0].value
            if not key_cell:
                continue
            raw_key = str(key_cell)
            key = self.normalize_key(raw_key)
            if not key:
                continue

            value = None
            value_col = None
            for col_idx in range(1, 5):
                if col_idx < len(row) and row[col_idx].value is not None:
                    value = row[col_idx].value
                    value_col = col_idx + 1
                    break

            logger.debug(f"Row {row_idx}: Raw Key='{raw_key}', Normalized Key='{key}', Value='{value}', Value Col={value_col}")

            if key in expected_keys:
                unmatched_keys.discard(key)
                data.append({"Key": key, "Value": value})
                logger.debug(f"Exact match: Key='{key}', Value='{value}' at row {row_idx}, col {value_col}")
            else:
                for expected_key in expected_keys:
                    similarity = SequenceMatcher(None, key, expected_key).ratio()
                    if similarity > 0.85:
                        unmatched_keys.discard(expected_key)
                        data.append({"Key": expected_key, "Value": value})
                        logger.debug(f"Fuzzy match: Key='{key}' matched '{expected_key}' (similarity={similarity:.2f}), Value='{value}' at row {row_idx}, col {value_col}")
                    elif similarity > 0.5:
                        potential_matches.append(f"Potential match for '{expected_key}': '{raw_key}' at row {row_idx}, col 1, similarity={similarity:.2f}")

        if potential_matches:
            logger.debug("Potential key matches for Instrumentation Data:")
            for match in potential_matches:
                logger.debug(match)
        if unmatched_keys:
            logger.warning(f"Unmatched instrumentation data keys: {unmatched_keys}")

        instrumentation_data = asdict(InstrumentationData(
            electron_microscope_type=next((d["Value"] for d in data if d["Key"] == "type_of_the_electron_microscope"), None),
            sample_grid=next((d["Value"] for d in data if d["Key"] == "sample_grid_sample_holder"), None),
            acceleration_voltage=next((d["Value"] for d in data if d["Key"] == "acceleration_voltage"), None),
            tilt_angle=next((d["Value"] for d in data if d["Key"] == "tilt_angle"), None),
            spot_size=next((d["Value"] for d in data if d["Key"] == "spot_size"), None),
            aperture=next((d["Value"] for d in data if d["Key"] == "aperture"), None)
        ))
        logger.debug(f"Instrumentation data extracted: {instrumentation_data}")
        return instrumentation_data

    def extract_replication(self):
        start_date = None
        end_date = None
        test_identifier = None
        for row in self.ws.iter_rows():
            key_cell = row[0].value
            if key_cell and 'test identifier number' in key_cell.lower():
                test_identifier = row[1].value
                start_date = self.excel_date_to_string(row[2].value)
                end_date = self.excel_date_to_string(row[3].value)
                break
        if test_identifier is None:
            logger.warning("Could not find test identifier number in Test Information sheet.")
        return asdict(ReplicationData(
            test_identifier_number=test_identifier,
            test_start_date=start_date,
            test_end_date=end_date
        ))

    def extract_raw_data(self, raw_sheet_name: str):
        logger.debug(f"Extracting raw data for sheet: {raw_sheet_name}")
        
        # Extract run number from sheet name
        match = re.search(r'R(\d+)', raw_sheet_name, re.IGNORECASE)
        run_number = int(match.group(1)) if match else 1
        
        particles = []

        if raw_sheet_name not in self.wb.sheetnames:
            logger.error(f"Raw data sheet {raw_sheet_name} not found")
            return asdict(RawData(run_number=run_number, particles=particles))

        raw_ws = self.wb[raw_sheet_name]
        logger.debug(f"Processing raw data sheet: {raw_sheet_name}")

        # Log first 5 rows for debugging
        logger.debug(f"Raw sheet {raw_sheet_name} first 5 rows:")
        for row_idx, row in enumerate(raw_ws.iter_rows(min_row=1, max_row=5, max_col=10), start=1):
            values = [str(cell.value)[:30] if cell.value else "None" for cell in row]
            logger.debug(f"Row {row_idx}: {values}")

        # Find value columns based on [nm]
        feret_min_col_idx = None
        length_col_idx = None
        feret_max_col_idx = None
        header_row = 1
        for col in range(1, raw_ws.max_column + 1):
            cell_value = raw_ws.cell(row=header_row, column=col).value
            if cell_value:
                h = self.normalize_key(cell_value)
                if 'fmin_nm' in h:
                    feret_min_col_idx = col
                    logger.debug(f"Found feret_min at col {col}: {cell_value}")
                if 'l_nm' in h:
                    length_col_idx = col
                    logger.debug(f"Found length at col {col}: {cell_value}")
                if 'fmax_nm' in h:
                    feret_max_col_idx = col
                    logger.debug(f"Found feret_max at col {col}: {cell_value}")

        if feret_min_col_idx is None or length_col_idx is None or feret_max_col_idx is None:
            logger.warning(f"Could not find all required columns in {raw_sheet_name}")
            return asdict(RawData(run_number=run_number, particles=particles))

        # Extract particle data
        max_rows = raw_ws.max_row
        empty_row_count = 0
        for row_idx in range(header_row + 1, max_rows + 1):
            feret_min = raw_ws.cell(row=row_idx, column=feret_min_col_idx).value
            length = raw_ws.cell(row=row_idx, column=length_col_idx).value
            feret_max = raw_ws.cell(row=row_idx, column=feret_max_col_idx).value
            logger.debug(f"Particle row {row_idx}: feret_min={feret_min}, length={length}, feret_max={feret_max}")
            
            if feret_min is None and length is None and feret_max is None:
                empty_row_count += 1
                if empty_row_count >= 5:
                    logger.debug(f"Reached end of particle data at row {row_idx}")
                    break
                continue
            
            empty_row_count = 0
            try:
                feret_min_float = float(feret_min) if feret_min is not None else None
                length_float = float(length) if length is not None else None
                feret_max_float = float(feret_max) if feret_max is not None else None
                
                if feret_min_float is not None or length_float is not None or feret_max_float is not None:
                    particles.append(ParticleMeasurement(
                        feret_min=feret_min_float,
                        length=length_float,
                        feret_max=feret_max_float
                    ))
            except (ValueError, TypeError) as e:
                logger.warning(f"Invalid particle data at row {row_idx}: {e}")
                continue
        
        logger.debug(f"Extracted {len(particles)} particles")
        return asdict(RawData(run_number=run_number, particles=particles))

    def extract_processed_data(self, processed_sheet_name: str):
        if processed_sheet_name not in self.wb.sheetnames:
            logger.error(f"Processed data sheet {processed_sheet_name} not found")
            return asdict(ProcessedData(run_number=1))

        match = re.search(r'R(\d+)', processed_sheet_name, re.IGNORECASE)
        run_number = int(match.group(1)) if match else 1

        processed_ws = self.wb[processed_sheet_name]
        logger.debug(f"Processing processed data sheet: {processed_sheet_name}")

        processed_particles = []
        metrics = RunMetrics()
        histogram = HistogramData()

        # Log first 10 rows for debugging
        logger.debug(f"Processed sheet {processed_sheet_name} first 10 rows:")
        for row_idx, row in enumerate(processed_ws.iter_rows(min_row=1, max_row=10, max_col=30), start=1):
            values = [str(cell.value)[:30] if cell.value else "None" for cell in row]
            logger.debug(f"Row {row_idx}: {values}")

        # Find columns for processed particles
        aspect_col_idx = None
        ecd_col_idx = None
        radius_col_idx = None
        roundness_col_idx = None
        header_row = None
        for row_idx in range(1, 11):  # Search in first 10 rows
            for col in range(1, processed_ws.max_column + 1):
                cell_value = processed_ws.cell(row=row_idx, column=col).value
                if cell_value:
                    h = self.normalize_key(cell_value)
                    if 'aspect_ratio' in h and aspect_col_idx is None:
                        aspect_col_idx = col
                    if 'ecd_diameter_nm' in h and ecd_col_idx is None:
                        ecd_col_idx = col
                    if 'radius_ecd_nm' in h and radius_col_idx is None:
                        radius_col_idx = col
                    if 'roundness' in h and roundness_col_idx is None:
                        roundness_col_idx = col
            if aspect_col_idx and ecd_col_idx and radius_col_idx and roundness_col_idx:
                header_row = row_idx
                logger.debug(f"Found processed particle headers at row {row_idx}: aspect={aspect_col_idx}, ecd={ecd_col_idx}, radius={radius_col_idx}, roundness={roundness_col_idx}")
                break

        if header_row:
            max_rows = min(processed_ws.max_row, header_row + 1000)
            empty_row_count = 0
            for row_idx in range(header_row + 1, max_rows + 1):
                aspect = processed_ws.cell(row=row_idx, column=aspect_col_idx).value
                ecd = processed_ws.cell(row=row_idx, column=ecd_col_idx).value
                radius = processed_ws.cell(row=row_idx, column=radius_col_idx).value
                roundness = processed_ws.cell(row=row_idx, column=roundness_col_idx).value
                if aspect is None and ecd is None and radius is None and roundness is None:
                    empty_row_count += 1
                    if empty_row_count >= 5:
                        break
                    continue
                empty_row_count = 0
                try:
                    processed_particles.append(ProcessedParticle(
                        aspect_ratio=float(aspect) if aspect is not None else None,
                        ecd_diameter=float(ecd) if ecd is not None else None,
                        radius_ecd=float(radius) if radius is not None else None,
                        roundness=float(roundness) if roundness is not None else None
                    ))
                except (ValueError, TypeError) as e:
                    logger.warning(f"Invalid processed particle data at row {row_idx}: {e}")
                    continue
            logger.debug(f"Extracted {len(processed_particles)} processed particles")

        # Extract metrics from specific rows
        metric_mapping = {
            2: ('feret_min_mean', 'feret_min_std', 'feret_min_median'),
            3: ('length_mean', 'length_std', 'length_median'),
            4: ('feret_max_mean', 'feret_max_std', 'feret_max_median'),
            5: ('aspect_ratio_mean', 'aspect_ratio_std', 'aspect_ratio_median'),
            6: ('ecd_mean', 'ecd_std', 'ecd_median'),
            7: ('radius_ecd_mean', 'radius_ecd_std', 'radius_ecd_median'),  # Assuming ECR is radius_ecd
            8: ('roundness_mean', 'roundness_std', 'roundness_median')
        }
        for row_idx, (mean_attr, std_attr, median_attr) in metric_mapping.items():
            mean_val = processed_ws.cell(row=row_idx, column=7).value  # G
            std_val = processed_ws.cell(row=row_idx, column=8).value   # H
            median_val = processed_ws.cell(row=row_idx, column=9).value  # I
            try:
                setattr(metrics, mean_attr, float(mean_val) if mean_val else None)
                setattr(metrics, std_attr, float(std_val) if std_val else None)
                setattr(metrics, median_attr, float(median_val) if median_val else None)
            except ValueError:
                logger.warning(f"Invalid metric value at row {row_idx}")

        # Extract additional metrics using search
        metric_names = [
            "density", "volume_np_nm3", "volume_np_cm3", "mass_np_g",
            "mass_colloid", "c_colloid_percent", "no_particles_stock"
        ]
        for row_idx in range(1, processed_ws.max_row + 1):
            row = processed_ws[row_idx]
            for cell in row:
                if not cell.value:
                    continue
                cell_value = self.normalize_key(cell.value)
                for metric_name in metric_names:
                    if metric_name in cell_value:
                        for offset in [(0, 1), (0, 2), (1, 0), (1, 1)]:
                            try:
                                value_cell = processed_ws.cell(row=row_idx + offset[0], column=cell.column + offset[1])
                                if value_cell.value is not None:
                                    try:
                                        setattr(metrics, metric_name, float(value_cell.value))
                                        logger.debug(f"Metric {metric_name} = {value_cell.value}")
                                    except (ValueError, TypeError):
                                        logger.warning(f"Invalid value for {metric_name}: {value_cell.value}")
                                    break
                            except:
                                continue

        # Extract histogram data
        hist_start_row = 8  # From document
        hist_mapping = {
            'feret_min': (14, 15, 16),  # N start, O end, P count
            'length': (20, 21, 22),     # T start, U end, V count
            'feret_max': (26, 27, 28)   # Z start, AA end, AB count
        }
        for key, (start_col, end_col, count_col) in hist_mapping.items():
            hist_list = getattr(histogram, key)
            empty_count = 0
            for row_idx in range(hist_start_row, processed_ws.max_row + 1):
                bin_start = processed_ws.cell(row=row_idx, column=start_col).value
                bin_end = processed_ws.cell(row=row_idx, column=end_col).value
                count_val = processed_ws.cell(row=row_idx, column=count_col).value
                if bin_start is None and bin_end is None and count_val is None:
                    empty_count += 1
                    if empty_count >= 5:
                        break
                else:
                    empty_count = 0
                if bin_start is not None:
                    try:
                        start = float(bin_start)
                        end = float(bin_end) if bin_end else start
                        count = int(count_val) if count_val else 0
                        hist_list.append(HistogramBin(bin_start=start, bin_end=end, count=count))
                    except (ValueError, TypeError):
                        logger.warning(f"Invalid histogram bin at row {row_idx} for {key}")

        logger.debug(f"Extracted histograms: feret_min={len(histogram.feret_min)}, length={len(histogram.length)}, feret_max={len(histogram.feret_max)}")

        return asdict(ProcessedData(run_number=run_number, processed_particles=processed_particles, metrics=metrics, histogram=histogram))

    def extract_final_results(self):
        logger.debug("Extracting final results")
        final_sheet_name = [name for name in self.wb.sheetnames if "Final results" in name][0] if any("Final results" in name for name in self.wb.sheetnames) else None
        if not final_sheet_name:
            logger.error("Final Results sheet not found")
            return asdict(ResultsData())

        ws = self.wb[final_sheet_name]
        logger.debug(f"Processing final results sheet: {final_sheet_name}")

        results_data = ResultsData()
        statistic_table = []

        # Extract summary statistics
        for row_idx in range(1, ws.max_row + 1):
            if ws.cell(row_idx, 1).value == "arithmetic mean":
                feret_min_mean = ws.cell(row_idx + 1, 1).value
                feret_min_std = ws.cell(row_idx + 1, 2).value
                feret_min_median = ws.cell(row_idx + 1, 3).value
                length_mean = ws.cell(row_idx + 1, 4).value
                length_std = ws.cell(row_idx + 1, 5).value
                length_median = ws.cell(row_idx + 1, 6).value
                feret_max_mean = ws.cell(row_idx + 1, 7).value
                feret_max_std = ws.cell(row_idx + 1, 8).value
                feret_max_median = ws.cell(row_idx + 1, 9).value
                ecd_mean = ws.cell(row_idx + 1, 10).value
                ecd_std = ws.cell(row_idx + 1, 11).value
                ecd_median = ws.cell(row_idx + 1, 12).value
                aspect_ratio_mean = ws.cell(row_idx + 1, 13).value
                aspect_ratio_std = ws.cell(row_idx + 1, 14).value
                aspect_ratio_median = ws.cell(row_idx + 1, 15).value
                roundness_mean = ws.cell(row_idx + 1, 16).value
                roundness_std = ws.cell(row_idx + 1, 17).value
                roundness_median = ws.cell(row_idx + 1, 18).value

                results_data.feret_min_mean = float(feret_min_mean) if feret_min_mean else None
                results_data.feret_min_std = float(feret_min_std) if feret_min_std else None
                results_data.feret_min_median = float(feret_min_median) if feret_min_median else None
                results_data.length_mean = float(length_mean) if length_mean else None
                results_data.length_std = float(length_std) if length_std else None
                results_data.length_median = float(length_median) if length_median else None
                results_data.feret_max_mean = float(feret_max_mean) if feret_max_mean else None
                results_data.feret_max_std = float(feret_max_std) if feret_max_std else None
                results_data.feret_max_median = float(feret_max_median) if feret_max_median else None
                results_data.ecd_mean = float(ecd_mean) if ecd_mean else None
                results_data.ecd_std = float(ecd_std) if ecd_std else None
                results_data.ecd_median = float(ecd_median) if ecd_median else None
                results_data.aspect_ratio_mean = float(aspect_ratio_mean) if aspect_ratio_mean else None
                results_data.aspect_ratio_std = float(aspect_ratio_std) if aspect_ratio_std else None
                results_data.aspect_ratio_median = float(aspect_ratio_median) if aspect_ratio_median else None
                results_data.roundness_mean = float(roundness_mean) if roundness_mean else None
                results_data.roundness_std = float(roundness_std) if roundness_std else None
                results_data.roundness_median = float(roundness_median) if roundness_median else None
                break

        return asdict(results_data)

    def parse_all_data(self) -> Dict[str, Union[Dict, List]]:
        try:
            work_package_data = self.extract_work_package_data()
            material_data = self.extract_material_data()
            sample_preparation_data = self.extract_sample_preparation_data()
            instrumentation_data = self.extract_instrumentation_data()
            replication = self.extract_replication()

            parsed_data = {
                'test_details': {
                    'work_package': work_package_data,
                    'material': material_data,
                    'sample_preparation': sample_preparation_data,
                    'instrumentation': instrumentation_data
                },
                'replication': replication,
                'replications': {},
                'processed_data': {},
                'final_results': {}
            }

            test_identifier = replication['test_identifier_number']
            if test_identifier:
                raw_sheet = f"Raw data_{test_identifier}"
                processed_sheet = f"Processed data_{test_identifier}"
                final_sheet = f"Final results_{test_identifier}"

                if raw_sheet in self.wb.sheetnames:
                    parsed_data['replications'] = self.extract_raw_data(raw_sheet)
                else:
                    logger.warning(f"Raw data sheet {raw_sheet} not found")

                if processed_sheet in self.wb.sheetnames:
                    parsed_data['processed_data'] = self.extract_processed_data(processed_sheet)
                else:
                    logger.warning(f"Processed data sheet {processed_sheet} not found")

                parsed_data['final_results'] = self.extract_final_results()

            return parsed_data
        except Exception as e:
            logger.error(f"Error parsing Excel file: {e}\n{traceback.format_exc()}")
            raise

def parse_excel_hr_stem(file_path: str, sheet_name: str = "Test Information") -> Dict[str, Union[Dict, List]]:
    try:
        parser = HRSTEMParser(file_path, sheet_name)
        return parser.parse_all_data()
    except Exception as e:
        logger.error(f"Error in parse_excel_hr_stem: {e}\n{traceback.format_exc()}")
        raise

if __name__ == "__main__":
    file_path = "/Users/asuto/Documents/chemat filez/WP2_HR-STEM_1aR1.xlsx"
    try:
        parsed_data = parse_excel_hr_stem(file_path)
        print("Parsed Data:")
        #print("Test Details:", parsed_data['test_details'])
        #print("Replication:", parsed_data['replication'])
        #print("Raw Data:", parsed_data['replications'])
        print("Processed Data:", parsed_data['processed_data'])
       #print("Final Results:", parsed_data['final_results'])
    except Exception as e:
        print(f"Error: {e}")