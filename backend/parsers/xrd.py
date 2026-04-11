import openpyxl
import re
import json
from dataclasses import dataclass, asdict, field
from typing import List, Dict, Optional, Union, Any
import logging
import traceback
from datetime import datetime, timedelta

logging.basicConfig(level=logging.DEBUG, format="%(asctime)s - %(levelname)s - %(message)s")
logging.getLogger("multipart.multipart").setLevel(logging.WARNING)
logger = logging.getLogger(__name__)


# =========================================================
# Dataclasses
# =========================================================

@dataclass
class Scientist:
    name: Optional[str] = None
    email: Optional[str] = None


@dataclass
class WorkPackageData:
    wp_name: Optional[str] = None
    partner: Optional[str] = None
    laboratory_name: Optional[str] = None
    full_test_name: Optional[str] = None
    test_acronym: Optional[str] = None
    test_type: Optional[str] = None
    endpoint: Optional[str] = None
    endpoint_outcome: Optional[str] = None
    sop: Optional[str] = None
    path: Optional[str] = None
    lead_scientists: List[Scientist] = field(default_factory=list)
    assay_scientists: List[Scientist] = field(default_factory=list)


@dataclass
class MaterialData:
    material_identifier: Optional[str] = None
    erm_id: Optional[str] = None
    material_name: Optional[str] = None
    core_chemistry: Optional[str] = None
    cas_no: Optional[str] = None
    cas_for_core: Optional[str] = None
    material_supplier: Optional[str] = None
    material_state: Optional[str] = None
    batch: Optional[str] = None
    vial: Optional[str] = None
    preparation_date: Optional[str] = None
    molar_concentration: Optional[str] = None
    particles_stock: Optional[str] = None


@dataclass
class DispersionData:
    dispersion_protocol: Optional[str] = None
    dispersion_technique: Optional[str] = None
    dispersion_medium: Optional[str] = None
    sonicator_type: Optional[str] = None
    power_w: Optional[str] = None
    sonication_time_s: Optional[str] = None
    tip_thickness_mm: Optional[str] = None
    tip_composition: Optional[str] = None
    bath_volume_dm3: Optional[str] = None
    sample_volume: Optional[str] = None
    final_concentration: Optional[str] = None
    additional_info: Optional[str] = None


@dataclass
class InstrumentationData:
    diffractometer_model: Optional[str] = None
    xray_lamp: Optional[str] = None
    detector: Optional[str] = None
    measurement_technique: Optional[str] = None
    generator_voltage: Optional[str] = None
    scan_speed: Optional[str] = None
    resolution: Optional[str] = None
    number_of_scans: Optional[Union[int, str]] = None
    replication_count: Optional[Union[int, str]] = None
    replicate_labels: List[str] = field(default_factory=list)
    two_theta_range: Optional[str] = None


@dataclass
class ReplicationMetadata:
    test_identifier_number: Optional[str] = None
    test_start_date: Optional[str] = None
    test_end_date: Optional[str] = None
    replicate_label: Optional[str] = None
    raw_sheet_name: Optional[str] = None
    processed_sheet_name: Optional[str] = None


@dataclass
class XRDSpectrumPoint:
    two_theta_deg: Optional[float] = None
    counts_mean: Optional[float] = None
    counts_individual: List[Optional[float]] = field(default_factory=list)


@dataclass
class XRDRawDataBlock:
    metric_name: Optional[str] = None
    raw_sheet_name: Optional[str] = None
    processed_sheet_name: Optional[str] = None
    two_theta_unit: Optional[str] = "deg"
    counts_unit: Optional[str] = "counts"
    number_of_scans: Optional[int] = None
    point_count: Optional[int] = None
    min_two_theta_deg: Optional[float] = None
    max_two_theta_deg: Optional[float] = None
    min_counts_mean: Optional[float] = None
    max_counts_mean: Optional[float] = None
    spectrum_points: List[XRDSpectrumPoint] = field(default_factory=list)


@dataclass
class XRDPeakEntry:
    peak_number: Optional[int] = None
    position_2theta_deg: Optional[float] = None
    d_spacing_angstrom: Optional[float] = None
    height_counts: Optional[float] = None
    fwhm_left_2theta_deg: Optional[float] = None
    area_counts_2theta: Optional[float] = None


@dataclass
class XRDProcessedDataBlock:
    processed_sheet_name: Optional[str] = None
    peak_count: Optional[int] = None
    peaks: List[XRDPeakEntry] = field(default_factory=list)


@dataclass
class XRDFinalResultBlock:
    crystal_structure: Optional[str] = None
    other_crystal_forms: Optional[str] = None
    other_crystal_forms_concentration: Optional[str] = None


# =========================================================
# Parser
# =========================================================

class XRDParser:
    def __init__(self, file_path: str, sheet_name: str = "Test Information"):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.parser_warnings: List[Dict[str, Any]] = []

        try:
            self.wb = openpyxl.load_workbook(file_path, data_only=True)
            self.ws = self.wb[sheet_name]
            logger.info("Successfully loaded XRD workbook: %s", file_path)
        except Exception as e:
            logger.error("Failed to load workbook or sheet '%s': %s", sheet_name, e)
            raise

        self.email_regex = r"^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$"

    # -----------------------------------------------------
    # Generic helpers
    # -----------------------------------------------------
    def normalize_key(self, key: Optional[str]) -> Optional[str]:
        if key is None:
            return None
        normalized = str(key).strip().lower()
        normalized = re.sub(r"[^a-z0-9]+", "_", normalized)
        normalized = re.sub(r"_+", "_", normalized).strip("_")
        return normalized

    def excel_date_to_string(self, value) -> Optional[str]:
        try:
            if isinstance(value, datetime):
                return value.strftime("%Y-%m-%d")

            if isinstance(value, (int, float)):
                base_date = datetime(1899, 12, 30)
                return (base_date + timedelta(days=float(value))).strftime("%Y-%m-%d")

            if isinstance(value, str):
                for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d", "%d.%m.%Y"):
                    try:
                        return datetime.strptime(value.strip(), fmt).strftime("%Y-%m-%d")
                    except ValueError:
                        pass

            return str(value).strip() if value not in (None, "") else None
        except Exception as e:
            logger.warning("Failed to convert date '%s': %s", value, e)
            return str(value).strip() if value not in (None, "") else None

    def _safe_float(self, value) -> Optional[float]:
        if value in (None, ""):
            return None
        try:
            if isinstance(value, str):
                value = value.replace(",", ".").strip()
            return float(value)
        except (ValueError, TypeError):
            return None

    def _safe_int(self, value) -> Optional[int]:
        if value in (None, ""):
            return None
        try:
            return int(value)
        except (ValueError, TypeError):
            return None

    def _get_first_value_in_row(self, row, start_col: int = 2, end_col: int = 6):
        for col_idx in range(start_col, end_col + 1):
            if col_idx - 1 < len(row):
                value = row[col_idx - 1].value
                if value is not None:
                    return value
        return None

    def _sheet_key_values(self) -> List[Dict[str, Any]]:
        data = []

        for row_idx, row in enumerate(self.ws.iter_rows(min_row=1, max_col=6), start=1):
            key_cell = row[0].value
            if not key_cell:
                continue

            raw_key = str(key_cell).strip()
            key = self.normalize_key(raw_key)
            if not key:
                continue

            value = self._get_first_value_in_row(row, 2, 6)
            email = row[3].value if len(row) > 3 else None

            data.append({
                "row": row_idx,
                "raw_key": raw_key,
                "key": key,
                "value": value,
                "email": email,
            })

        return data

    def _find_raw_sheets(self) -> List[str]:
        return sorted([s for s in self.wb.sheetnames if s.lower().startswith("raw data") or s.lower().startswith("raw_data")])

    def _find_processed_sheets(self) -> List[str]:
        return sorted([s for s in self.wb.sheetnames if s.lower().startswith("processed data") or s.lower().startswith("processed_data")])

    def _find_final_results_sheets(self) -> List[str]:
        return sorted([s for s in self.wb.sheetnames if s.lower().startswith("final results") or s.lower().startswith("final_results")])

    def _match_processed_to_raw(self, raw_sheet_name: str) -> Optional[str]:
        """Match a processed sheet to a raw sheet by shared identifier suffix."""
        raw_lower = raw_sheet_name.lower()

        # Extract identifier after 'raw data_' or 'raw_data_'
        raw_id = re.sub(r"^raw[\s_]?data[\s_]?", "", raw_lower).strip("_").strip()

        for ps in self._find_processed_sheets():
            ps_lower = ps.lower()
            ps_id = re.sub(r"^processed[\s_]?data[\s_]?", "", ps_lower).strip("_").strip()
            if ps_id == raw_id:
                return ps

        return None

    # -----------------------------------------------------
    # Test details
    # -----------------------------------------------------
    def extract_work_package_data(self) -> Dict[str, Any]:
        rows = self._sheet_key_values()
        lead_scientists: List[Scientist] = []
        assay_scientists: List[Scientist] = []

        for row in rows:
            key = row["key"]
            value = row["value"]
            email = row["email"]

            if "lead_scientist" in key:
                lead_scientists.append(
                    Scientist(
                        name=value,
                        email=email if email and re.match(self.email_regex, str(email)) else None,
                    )
                )

            if "assay_test_work_conducted_by" in key:
                assay_scientists.append(
                    Scientist(
                        name=value,
                        email=email if email and re.match(self.email_regex, str(email)) else None,
                    )
                )

        result = WorkPackageData(
            wp_name=next((r["value"] for r in rows if r["key"] == "project_work_package"), None),
            partner=next((r["value"] for r in rows if r["key"] == "partner_conducting_test_assay"), None),
            laboratory_name=next((r["value"] for r in rows if r["key"] == "test_facility_laboratory_name"), None),
            full_test_name=next((r["value"] for r in rows if "full_name_of_test_assay" in r["key"]), None),
            test_acronym=next((r["value"] for r in rows if "short_name_or_acronym" in r["key"]), None),
            test_type=next((r["value"] for r in rows if "type_or_class_of_experimental" in r["key"]), None),
            endpoint=next((r["value"] for r in rows if "end_point_being_investigated" in r["key"]), None),
            endpoint_outcome=next((r["value"] for r in rows if "metric_s_used_to_assess" in r["key"]), None),
            sop=next((r["value"] for r in rows if "sop_s_for_test" in r["key"]), None),
            path=next((r["value"] for r in rows if "path_link_to_sop" in r["key"]), None),
            lead_scientists=lead_scientists,
            assay_scientists=assay_scientists,
        )
        return asdict(result)

    def extract_material_data(self) -> Dict[str, Any]:
        rows = self._sheet_key_values()

        result = MaterialData(
            material_identifier=next((r["value"] for r in rows if r["key"] == "sample_cms_internal_identifier"), None),
            erm_id=next((r["value"] for r in rows if "erm_identifier" in r["key"]), None),
            material_name=next((r["value"] for r in rows if r["key"] == "material_name"), None),
            core_chemistry=next((r["value"] for r in rows if "core_chemistry" in r["key"]), None),
            cas_no=next((r["value"] for r in rows if r["key"] == "cas_no"), None),
            cas_for_core=next((r["value"] for r in rows if "cas_for_core" in r["key"] or "cas_no_for_core" in r["key"]), None),
            material_supplier=next((r["value"] for r in rows if r["key"] == "material_supplier"), None),
            material_state=next((r["value"] for r in rows if r["key"] == "material_state"), None),
            batch=next((r["value"] for r in rows if r["key"] == "batch"), None),
            vial=next((r["value"] for r in rows if r["key"] == "vial"), None),
            preparation_date=self.excel_date_to_string(
                next((r["value"] for r in rows if "date_of_sample_preparation" in r["key"] or "date_of_preparation" in r["key"]), None)
            ),
            molar_concentration=next((r["value"] for r in rows if "molar_concentration" in r["key"]), None),
            particles_stock=next((r["value"] for r in rows if "no_of_particles" in r["key"] or "number_of_particles" in r["key"]), None),
        )
        return asdict(result)

    def extract_dispersion_data(self) -> Dict[str, Any]:
        rows = self._sheet_key_values()

        result = DispersionData(
            dispersion_protocol=next(
                (r["value"] for r in rows if "standard_dispersion_protocol" in r["key"] or "specify_standard_dispersion" in r["key"]),
                None,
            ),
            dispersion_technique=next(
                (r["value"] for r in rows if "dispersion_technique" in r["key"] or "otherwise_specify_dispersion" in r["key"]),
                None,
            ),
            dispersion_medium=next(
                (r["value"] for r in rows if "dispersion_dilution_medium" in r["key"] or "dispersion_medium" in r["key"]),
                None,
            ),
            sonicator_type=next((r["value"] for r in rows if "sonicator" in r["key"] and "type" in r["key"]), None),
            power_w=next((r["value"] for r in rows if "power_w" in r["key"] or r["key"] == "power_w"), None),
            sonication_time_s=next((r["value"] for r in rows if "sonication_time" in r["key"]), None),
            tip_thickness_mm=next((r["value"] for r in rows if "tip_thickness" in r["key"]), None),
            tip_composition=next((r["value"] for r in rows if "tip_composition" in r["key"]), None),
            bath_volume_dm3=next((r["value"] for r in rows if "ultrasonic_bath" in r["key"] or "water_volume" in r["key"]), None),
            sample_volume=next((r["value"] for r in rows if "sample_volume" in r["key"]), None),
            final_concentration=next((r["value"] for r in rows if "final_sample_concentration" in r["key"]), None),
            additional_info=next((r["value"] for r in rows if "additional_information" in r["key"]), None),
        )
        return asdict(result)

    def extract_instrumentation_data(self) -> Dict[str, Any]:
        rows = self._sheet_key_values()
        ws = self.ws

        # Extract replicate labels from row 64 (cols 3+)
        replicate_labels: List[str] = []
        for col_idx in range(3, 12):
            v = ws.cell(row=64, column=col_idx).value
            if v is not None:
                replicate_labels.append(str(v))

        result = InstrumentationData(
            diffractometer_model=next((r["value"] for r in rows if "diffractometer" in r["key"]), None),
            xray_lamp=next((r["value"] for r in rows if "x_ray_lamp" in r["key"] or "xray_lamp" in r["key"]), None),
            detector=next((r["value"] for r in rows if r["key"] == "detector"), None),
            measurement_technique=next(
                (r["value"] for r in rows if "measurment_technique" in r["key"] or "measurement_technique" in r["key"]),
                None,
            ),
            generator_voltage=next((r["value"] for r in rows if "generator_voltage" in r["key"]), None),
            scan_speed=next((r["value"] for r in rows if "scan_speed" in r["key"]), None),
            resolution=next((r["value"] for r in rows if r["key"] == "resolution"), None),
            number_of_scans=next((r["value"] for r in rows if "number_of_scans" in r["key"]), None),
            replication_count=ws.cell(row=64, column=2).value,
            replicate_labels=replicate_labels,
            two_theta_range=next((r["value"] for r in rows if "2theta_range" in r["key"] or "theta_range" in r["key"]), None),
        )
        return asdict(result)

    # -----------------------------------------------------
    # Replication metadata
    # -----------------------------------------------------
    def extract_replication_metadata(self) -> List[Dict[str, Any]]:
        ws = self.ws
        raw_sheets = self._find_raw_sheets()
        metadata: List[Dict[str, Any]] = []

        # Test identifier row 38
        test_id = ws.cell(row=38, column=2).value
        test_start = self.excel_date_to_string(ws.cell(row=38, column=3).value)
        test_end = self.excel_date_to_string(ws.cell(row=38, column=4).value)

        for raw_sheet in raw_sheets:
            processed_sheet = self._match_processed_to_raw(raw_sheet)

            metadata.append(asdict(ReplicationMetadata(
                test_identifier_number=str(test_id).strip() if test_id else None,
                test_start_date=test_start,
                test_end_date=test_end,
                replicate_label="XRD Diffractogram",
                raw_sheet_name=raw_sheet,
                processed_sheet_name=processed_sheet,
            )))

        return metadata

    # -----------------------------------------------------
    # Raw data
    # -----------------------------------------------------
    def extract_raw_data(self) -> List[Dict[str, Any]]:
        raw_sheets = self._find_raw_sheets()
        raw_blocks: List[Dict[str, Any]] = []

        for raw_sheet_name in raw_sheets:
            processed_sheet = self._match_processed_to_raw(raw_sheet_name)
            raw_blocks.append(self._extract_raw_block(raw_sheet_name, processed_sheet))

        return raw_blocks

    def _extract_raw_block(self, raw_sheet_name: str, processed_sheet_name: Optional[str]) -> Dict[str, Any]:
        ws = self.wb[raw_sheet_name]

        # Detect column layout from header row 1
        headers: Dict[int, str] = {}
        for col_idx in range(1, ws.max_column + 1):
            v = ws.cell(row=1, column=col_idx).value
            if v is not None:
                headers[col_idx] = str(v).strip()

        # Identify 2Theta column, mean column, and individual scan columns
        two_theta_col: Optional[int] = None
        mean_col: Optional[int] = None
        scan_cols: List[int] = []

        for col_idx, header in headers.items():
            header_lower = header.lower()
            if "2theta" in header_lower or "theta" in header_lower:
                two_theta_col = col_idx
            elif "mean" in header_lower:
                mean_col = col_idx
            elif header_lower.startswith("counts"):
                scan_cols.append(col_idx)

        if two_theta_col is None:
            two_theta_col = 1
        if mean_col is None:
            mean_col = 2

        scan_cols = sorted(scan_cols)

        spectrum_points: List[Dict[str, Any]] = []
        two_theta_vals: List[float] = []
        mean_vals: List[float] = []

        for row_idx in range(2, ws.max_row + 1):
            tt = self._safe_float(ws.cell(row=row_idx, column=two_theta_col).value)
            if tt is None:
                continue

            mean_count = self._safe_float(ws.cell(row=row_idx, column=mean_col).value)

            individual = []
            for sc in scan_cols:
                individual.append(self._safe_float(ws.cell(row=row_idx, column=sc).value))

            spectrum_points.append(asdict(XRDSpectrumPoint(
                two_theta_deg=tt,
                counts_mean=mean_count,
                counts_individual=individual,
            )))
            two_theta_vals.append(tt)
            if mean_count is not None:
                mean_vals.append(mean_count)

        result = XRDRawDataBlock(
            metric_name="XRD Diffractogram (2Theta vs Counts)",
            raw_sheet_name=raw_sheet_name,
            processed_sheet_name=processed_sheet_name,
            two_theta_unit="deg",
            counts_unit="counts",
            number_of_scans=len(scan_cols) if scan_cols else None,
            point_count=len(spectrum_points),
            min_two_theta_deg=min(two_theta_vals) if two_theta_vals else None,
            max_two_theta_deg=max(two_theta_vals) if two_theta_vals else None,
            min_counts_mean=min(mean_vals) if mean_vals else None,
            max_counts_mean=max(mean_vals) if mean_vals else None,
            spectrum_points=spectrum_points,
        )
        return asdict(result)

    # -----------------------------------------------------
    # Processed data (peak list)
    # -----------------------------------------------------
    def extract_processed_data(self, raw_blocks: List[Dict[str, Any]]) -> Dict[str, Any]:
        processed_sheets = self._find_processed_sheets()
        blocks: List[Dict[str, Any]] = []

        for ps_name in processed_sheets:
            pws = self.wb[ps_name]
            peaks: List[Dict[str, Any]] = []

            # Detect header row: look for "Pos." or "No." in first few rows
            header_row = None
            for r in range(1, min(6, pws.max_row + 1)):
                for c in range(1, min(8, pws.max_column + 1)):
                    v = pws.cell(row=r, column=c).value
                    if v is not None and ("pos" in str(v).lower() or "no." in str(v).lower()):
                        header_row = r
                        break
                if header_row:
                    break

            if header_row is None:
                header_row = 2  # fallback

            # Parse headers to find column mapping
            col_map: Dict[str, int] = {}
            for c in range(1, pws.max_column + 1):
                v = pws.cell(row=header_row, column=c).value
                if v is None:
                    continue
                v_lower = str(v).strip().lower()
                if v_lower.startswith("no"):
                    col_map["no"] = c
                elif "pos" in v_lower:
                    col_map["position"] = c
                elif "d-spacing" in v_lower or "d_spacing" in v_lower:
                    col_map["d_spacing"] = c
                elif "height" in v_lower:
                    col_map["height"] = c
                elif "fwhm" in v_lower:
                    col_map["fwhm"] = c
                elif "area" in v_lower:
                    col_map["area"] = c

            for row_idx in range(header_row + 1, pws.max_row + 1):
                peak_no = self._safe_int(pws.cell(row=row_idx, column=col_map.get("no", 1)).value)
                if peak_no is None:
                    continue

                peaks.append(asdict(XRDPeakEntry(
                    peak_number=peak_no,
                    position_2theta_deg=self._safe_float(
                        pws.cell(row=row_idx, column=col_map.get("position", 2)).value
                    ),
                    d_spacing_angstrom=self._safe_float(
                        pws.cell(row=row_idx, column=col_map.get("d_spacing", 3)).value
                    ),
                    height_counts=self._safe_float(
                        pws.cell(row=row_idx, column=col_map.get("height", 4)).value
                    ),
                    fwhm_left_2theta_deg=self._safe_float(
                        pws.cell(row=row_idx, column=col_map.get("fwhm", 5)).value
                    ),
                    area_counts_2theta=self._safe_float(
                        pws.cell(row=row_idx, column=col_map.get("area", 6)).value
                    ),
                )))

            blocks.append(asdict(XRDProcessedDataBlock(
                processed_sheet_name=ps_name,
                peak_count=len(peaks),
                peaks=peaks,
            )))

        return {
            "peak_lists": blocks,
            "processed_sheet_count": len(processed_sheets),
            "total_peaks_identified": sum(b.get("peak_count", 0) for b in blocks),
        }

    # -----------------------------------------------------
    # Final results
    # -----------------------------------------------------
    def extract_final_results(self) -> List[Dict[str, Any]]:
        final_sheets = self._find_final_results_sheets()

        if not final_sheets:
            self.parser_warnings.append({
                "type": "missing_sheet",
                "sheet": "Final results",
                "note": "No Final results sheet found in this XRD workbook.",
            })
            return []

        results: List[Dict[str, Any]] = []

        for fs_name in final_sheets:
            ws = self.wb[fs_name]

            # Header row 1: Results | Crystal structure | Other crystal forms | Other crystal forms concentration
            # Data row 2+
            for row_idx in range(2, ws.max_row + 1):
                crystal_structure = ws.cell(row=row_idx, column=2).value
                other_forms = ws.cell(row=row_idx, column=3).value
                other_conc = ws.cell(row=row_idx, column=4).value

                if all(v in (None, "") for v in [crystal_structure, other_forms, other_conc]):
                    continue

                results.append(asdict(XRDFinalResultBlock(
                    crystal_structure=str(crystal_structure).strip() if crystal_structure else None,
                    other_crystal_forms=str(other_forms).strip() if other_forms else None,
                    other_crystal_forms_concentration=str(other_conc).strip() if other_conc else None,
                )))

        return results

    # -----------------------------------------------------
    # Statistical analysis
    # -----------------------------------------------------
    def extract_statistical_analysis(self) -> Dict[str, Any]:
        return {
            "available": False,
            "notes": "No dedicated statistical analysis section was found in this XRD workbook.",
        }

    # -----------------------------------------------------
    # Parse all
    # -----------------------------------------------------
    def parse_all_data(self) -> Dict[str, Union[Dict, List]]:
        try:
            work_package_data = self.extract_work_package_data()
            material_data = self.extract_material_data()
            dispersion_data = self.extract_dispersion_data()
            instrumentation_data = self.extract_instrumentation_data()

            replication_metadata = self.extract_replication_metadata()
            raw_data = self.extract_raw_data()
            processed_data = self.extract_processed_data(raw_data)
            final_results = self.extract_final_results()
            statistical_analysis = self.extract_statistical_analysis()

            parsed_data = {
                "test_details": {
                    "work_package": work_package_data,
                    "material": material_data,
                    "cell_line": {},
                    "dispersion": dispersion_data,
                    "instrumentation": instrumentation_data,
                },
                "replication_metadata": replication_metadata,
                "replications": raw_data,
                "processed_data": processed_data,
                "final_results": final_results,
                "statistical_analysis": statistical_analysis,
            }

            if self.parser_warnings:
                parsed_data["parser_warnings"] = self.parser_warnings

            logger.info("FINAL XRD JSON:\n%s", json.dumps(parsed_data, indent=2, default=str))
            return parsed_data

        except Exception as e:
            logger.error("Error parsing XRD file: %s\n%s", e, traceback.format_exc())
            raise


def parse_excel_xrd(file_path: str, sheet_name: str = "Test Information") -> Dict[str, Union[Dict, List]]:
    try:
        parser = XRDParser(file_path, sheet_name)
        return parser.parse_all_data()
    except Exception as e:
        logger.error("Error in parse_excel_xrd: %s\n%s", e, traceback.format_exc())
        raise


if __name__ == "__main__":
    import sys

    file_path = sys.argv[1] if len(sys.argv) > 1 else "/mnt/user-data/uploads/WP2_XRD_17aR1.xlsx"
    parsed_data = parse_excel_xrd(file_path)

    print("=" * 70)
    print("XRD PARSER OUTPUT SUMMARY")
    print("=" * 70)

    wp = parsed_data["test_details"]["work_package"]
    print(f"\nWP: {wp['wp_name']}, Partner: {wp['partner']}")
    print(f"Test: {wp['test_acronym']} - {wp['full_test_name']}")
    print(f"Lead Scientists: {[s['name'] for s in wp['lead_scientists']]}")

    mat = parsed_data["test_details"]["material"]
    print(f"\nMaterial: {mat['material_name']} ({mat['material_identifier']})")
    print(f"Core Chemistry: {mat['core_chemistry']}")

    inst = parsed_data["test_details"]["instrumentation"]
    print(f"\nDiffractometer: {inst['diffractometer_model']}")
    print(f"Scans: {inst['number_of_scans']}, 2Theta Range: {inst['two_theta_range']}")

    print(f"\nReplication Metadata Count: {len(parsed_data['replication_metadata'])}")
    print(f"Raw Blocks Count: {len(parsed_data['replications'])}")
    for block in parsed_data["replications"]:
        print(f"  {block['metric_name']}: {block['point_count']} points, {block['number_of_scans']} scans")
        print(f"    2Theta range: {block['min_two_theta_deg']:.3f} - {block['max_two_theta_deg']:.3f} deg")

    proc = parsed_data["processed_data"]
    print(f"\nProcessed Data: {proc['total_peaks_identified']} peaks across {proc['processed_sheet_count']} sheet(s)")
    for pl in proc["peak_lists"]:
        print(f"  Sheet '{pl['processed_sheet_name']}': {pl['peak_count']} peaks")
        for p in pl["peaks"]:
            print(f"    Peak #{p['peak_number']}: 2θ={p['position_2theta_deg']:.4f}°, "
                  f"d={p['d_spacing_angstrom']:.5f}Å, h={p['height_counts']:.1f}")

    print(f"\nFinal Results Count: {len(parsed_data['final_results'])}")
    for fr in parsed_data["final_results"]:
        print(f"  Crystal: {fr['crystal_structure']}, Other forms: {fr['other_crystal_forms']} "
              f"({fr['other_crystal_forms_concentration']})")

    if "parser_warnings" in parsed_data:
        print(f"\nParser Warnings ({len(parsed_data['parser_warnings'])}):")
        for w in parsed_data["parser_warnings"]:
            print(f"  [{w['type']}] {w.get('note', '')}")

    print("\n--- JSON Preview ---")
    print(json.dumps(parsed_data, indent=2, default=str)[:5000])